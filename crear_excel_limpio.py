#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Crear RegistroDoc_PRO.xlsx - Excel limpio sin macros, sin forms, sin módulos
Solo estructura de datos pura
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def crear_excel_limpio():
    wb = Workbook()
    
    # Remover hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ===== PORTADA =====
    ws_portada = wb.create_sheet('Portada', 0)
    ws_portada['A1'] = 'REGISTRODOC PRO'
    ws_portada['A1'].font = Font(size=24, bold=True, color='FFFFFF')
    ws_portada['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws_portada['A2'] = 'Sistema de Registro Académico'
    ws_portada['A2'].font = Font(size=14)
    ws_portada['A4'] = f'Versión: 2.0'
    ws_portada['A5'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y")}'
    ws_portada['A7'] = 'Para: Instituciones Educativas en Panamá'
    
    # ===== MAESTRO (Estudiantes) =====
    ws_maestro = wb.create_sheet('MAESTRO')
    maestro_headers = ['ID', 'Cédula', 'Nombre', 'Grado', 'Email', 'Teléfono', 'Fecha Nacimiento', 'Dirección']
    ws_maestro.append(maestro_headers)
    
    # Formato headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws_maestro[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_maestro.column_dimensions['A'].width = 8
    ws_maestro.column_dimensions['B'].width = 12
    ws_maestro.column_dimensions['C'].width = 20
    ws_maestro.column_dimensions['D'].width = 10
    ws_maestro.column_dimensions['E'].width = 20
    ws_maestro.column_dimensions['F'].width = 15
    ws_maestro.column_dimensions['G'].width = 15
    ws_maestro.column_dimensions['H'].width = 25
    
    # ===== HOJAS DE CALIFICACIONES (PROM) =====
    materias = ['Inglés', 'Comercio', 'Agropecuaria', 'Hogar', 'Artística']
    grados = [7, 8, 9]
    
    for materia in materias:
        for grado in grados:
            sheet_name = f'{materia[:8].strip()}_{grado}°'
            ws = wb.create_sheet(sheet_name)
            
            headers = ['ID', 'Estudiante', 'Parcial 1', 'Parcial 2', 'Parcial 3', 
                      'Apreciación', 'Prueba', 'Promedio', 'Estado']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col].width = 12
    
    # ===== AUDITORIA =====
    ws_audit = wb.create_sheet('_Auditoria')
    audit_headers = ['ID', 'Usuario', 'Acción', 'Tabla', 'Detalles', 'Timestamp']
    ws_audit.append(audit_headers)
    
    for cell in ws_audit[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Guardar como .xlsx
    archivo_salida = 'RegistroDoc_PRO.xlsx'
    wb.save(archivo_salida)
    
    print(f"✅ Archivo creado: {archivo_salida}")
    print(f"   - Formato: .xlsx (sin macros)")
    print(f"   - Hojas: {len(wb.sheetnames)}")
    print(f"   - Estructura: Limpia y lista para agregar macros manualmente")
    print(f"\n📋 Próximos pasos:")
    print(f"   1. Abre {archivo_salida} en Excel")
    print(f"   2. Habilita el contenido cuando se solicite")
    print(f"   3. Guarda como .xlsm")
    print(f"   4. Ejecuta: python agregar_macros_y_formularios.py")
    
    return True

if __name__ == '__main__':
    crear_excel_limpio()
