#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agregar macros (.bas) y formularios al RegistroDoc_PRO.xlsx
Convierte a .xlsm y agrega el contenido de RD_Macros.bas
"""

import os
import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

def leer_archivo_bas():
    """Lee el contenido del archivo .bas"""
    if os.path.exists('RD_Macros.bas'):
        with open('RD_Macros.bas', 'r', encoding='utf-8') as f:
            return f.read()
    else:
        print("❌ Archivo RD_Macros.bas no encontrado")
        return None

def crear_vbaProject_xml():
    """Crea la estructura XML para los macros VBA"""
    xml_content = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Project xmlns="http://schemas.microsoft.com/office/appforoffice/11/2006/ovml">
 <ProjectHelpFileName bin="00000000"/>
 <HelpContextID bin="00000000"/>
 <Name>VBAProject</Name>
 <DocString></DocString>
 <BaseClass ClassType="GlobalNamespace">
  <Name>ThisWorkbook</Name>
  <Description></Description>
  <CodeModule Reference="{00020820-0000-0000-C000-000000000046}#0#0#C:\\Program Files\\Common Files\\Microsoft Shared\\VBA\\VBA7.1\\vbe7.dll#Visual Basic for Applications">
   <Code>{% code %}</Code>
  </CodeModule>
 </BaseClass>
 <References>
  <Reference Name="Excel" LibID="{00020813-0000-0000-C000-000000000046}" Version="1.9" SubLibFlags="0" LibFlags="3">
  </Reference>
  <Reference Name="stdole" LibID="{00020430-0000-0000-C000-000000000046}" Version="2.0" SubLibFlags="0" LibFlags="1">
  </Reference>
 </References>
</Project>'''
    return xml_content

def agregar_macros_basico():
    """Agrega macros de forma básica sin complejidad"""
    
    try:
        # Leer archivo .bas
        contenido_bas = leer_archivo_bas()
        if not contenido_bas:
            print("⚠️  Continuando sin macros...")
            return False
        
        # Cargar workbook
        archivo_origen = 'RegistroDoc_PRO.xlsx'
        archivo_destino = 'RegistroDoc_PRO.xlsm'
        
        print(f"📂 Cargando: {archivo_origen}")
        wb = load_workbook(archivo_origen)
        
        # Guardar como .xlsm
        wb.save(archivo_destino)
        print(f"✅ Archivo guardado como: {archivo_destino}")
        print(f"   (Ahora soporta macros)")
        
        # Información sobre macros
        print(f"\n📋 Información de macros:")
        print(f"   - Archivo .bas encontrado: RD_Macros.bas")
        print(f"   - Líneas de código: {len(contenido_bas.splitlines())}")
        print(f"   - Tamaño: {len(contenido_bas)} bytes")
        
        print(f"\n⚠️  PRÓXIMO PASO (Manual en Excel):")
        print(f"   1. Abre {archivo_destino} en Excel")
        print(f"   2. Ve a Developer > Visual Basic")  
        print(f"   3. File > Import File > Selecciona RD_Macros.bas")
        print(f"   4. Guarda el archivo")
        print(f"\n   O ejecuta:")
        print(f"   python agregar_macros_powershell.py")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return False

def agregar_macros_via_powershell():
    """Agrega macros usando PowerShell y COM de Excel"""
    
    print("🔧 Intentando agregar macros via PowerShell...")
    
    ps_script = '''
$ExcelApp = New-Object -ComObject Excel.Application
$ExcelApp.Visible = $false

$archivo = "C:\\RegistroDoc\\RegistroDoc_PRO.xlsm"
$basfile = "C:\\RegistroDoc\\RD_Macros.bas"

if (Test-Path $archivo) {
    $wb = $ExcelApp.Workbooks.Open($archivo)
    
    if (Test-Path $basfile) {
        $vbModule = $wb.VBProject.VBComponents.Import($basfile)
        Write-Host "✅ Módulo VBA importado exitosamente"
    }
    
    $wb.Save()
    $wb.Close()
    Write-Host "✅ Archivo guardado con macros"
} else {
    Write-Host "❌ Archivo no encontrado: $archivo"
}

$ExcelApp.Quit()
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($ExcelApp) | Out-Null
'''
    
    # Guardar script temporalmente
    ps_file = 'temp_agregar_macros.ps1'
    with open(ps_file, 'w', encoding='utf-8') as f:
        f.write(ps_script)
    
    # Ejecutar
    os.system(f'powershell -ExecutionPolicy Bypass -File {ps_file}')
    
    # Limpiar
    if os.path.exists(ps_file):
        os.remove(ps_file)

if __name__ == '__main__':
    print("=" * 60)
    print("AGREGAR MACROS Y FORMULARIOS A RegistroDoc_PRO")
    print("=" * 60)
    print()
    
    # Paso 1: Agregar macros básicamente
    exito = agregar_macros_basico()
    
    if exito:
        print()
        print("=" * 60)
        print("✅ LISTO PARA AGREGAR MACROS MANUALMENTE")
        print("=" * 60)
