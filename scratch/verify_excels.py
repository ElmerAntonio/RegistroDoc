import openpyxl
import os

def check_excel_file(filename):
    print(f"=== Checking {filename} ===")
    if not os.path.exists(filename):
        print(f"File {filename} does not exist!")
        return False
        
    try:
        wb = openpyxl.load_workbook(filename, data_only=False)
        errors_found = 0
        total_formulas = 0
        
        for name in wb.sheetnames:
            ws = wb[name]
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=row, column=col).value
                    if val is not None and isinstance(val, str) and val.startswith("="):
                        total_formulas += 1
                        if "#REF!" in val:
                            print(f"  [Error #REF!] Sheet: '{name}', Cell: {openpyxl.utils.get_column_letter(col)}{row}, Formula: {val}")
                            errors_found += 1
                        if "#NAME?" in val:
                            print(f"  [Error #NAME?] Sheet: '{name}', Cell: {openpyxl.utils.get_column_letter(col)}{row}, Formula: {val}")
                            errors_found += 1
                            
        print(f"Scanned {total_formulas} formulas. Found {errors_found} errors.")
        wb.close()
        return errors_found == 0
    except Exception as e:
        print(f"Error checking file: {e}")
        return False

if __name__ == "__main__":
    check_excel_file("Registro_2026.xlsx")
    check_excel_file("Registro_2026_FINAL_v2.xlsx")
    check_excel_file("Registro_Primaria.xlsx")
