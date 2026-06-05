import openpyxl

wb = openpyxl.load_workbook("Registro_Primaria.xlsx")
ws = wb["RESUMEN_1°"]

for r in [77, 78, 79, 119, 120, 121]:
    print(f"Row {r}:")
    for col in range(1, 15):
        val = ws.cell(row=r, column=col).value
        print(f"  Col {col} ({openpyxl.utils.get_column_letter(col)}): {val}")
wb.close()
