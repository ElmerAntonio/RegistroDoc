import openpyxl

def repair_premedia_file(filepath):
    print(f"=== Repairing Premedia File: {filepath} ===")
    wb = openpyxl.load_workbook(filepath)
    
    # We scan all sheets in the premedia workbook for "Planilla" and "8°"
    for sheet_name in wb.sheetnames:
        if "Planilla" in sheet_name and "8°" in sheet_name:
            ws = wb[sheet_name]
            # Extract subject name
            temp = sheet_name.replace("Planilla (", "")
            sub_name = temp.split(")")[0].strip()
            
            # Find the corresponding PROM sheet
            prom_sheet_name = f"PROM ({sub_name})"
            if prom_sheet_name not in wb.sheetnames:
                for s in wb.sheetnames:
                    if s.strip() == f"PROM ({sub_name})":
                        prom_sheet_name = s
                        break
            
            print(f"  Fixing sheet: {sheet_name} (Subject: {sub_name}, PROM sheet: {prom_sheet_name})")
            
            # Repair student rows 16 to 55 (indices 1 to 40)
            for r in range(16, 56):
                idx = r - 15
                maestro_row = idx + 4
                
                ws.cell(row=r, column=3).value = f'=IF(MAESTRO!D{maestro_row}="","",MAESTRO!D{maestro_row})'
                ws.cell(row=r, column=6).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('{prom_sheet_name}'!AF{maestro_row},\"\"))"
                ws.cell(row=r, column=7).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('{prom_sheet_name}'!BM{maestro_row},\"\"))"
                ws.cell(row=r, column=8).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('{prom_sheet_name}'!CU{maestro_row},\"\"))"
                ws.cell(row=r, column=9).value = f'=IF(MAESTRO!D{maestro_row}="","",IFERROR(TRUNC(AVERAGE(F{r},G{r},H{r}),1),""))'
                
                # J to O
                ws.cell(row=r, column=10).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('Asistencia (8°) '!BI{idx+2},\"\"))"
                ws.cell(row=r, column=11).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('Asistencia (8°) '!BJ{idx+2},\"\"))"
                ws.cell(row=r, column=12).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('Asistencia (8°) '!BI{idx+45},\"\"))"
                ws.cell(row=r, column=13).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('Asistencia (8°) '!BJ{idx+45},\"\"))"
                ws.cell(row=r, column=14).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('Asistencia (8°) '!BI{idx+88},\"\"))"
                ws.cell(row=r, column=15).value = f"=IF(MAESTRO!D{maestro_row}=\"\",\"\",IFERROR('Asistencia (8°) '!BJ{idx+88},\"\"))"
                
    wb.save(filepath)
    wb.close()
    print(f"{filepath} repaired successfully.")

def repair_registro_primaria():
    print("=== Repairing Registro_Primaria.xlsx ===")
    wb = openpyxl.load_workbook("Registro_Primaria.xlsx")
    
    # 1. Repair RESUMEN_1°
    ws_res = wb["RESUMEN_1°"]
    subjects = [
        ("Ingles", "Ingles"),
        ("Español", "Espa\u00f1ol"),
        ("Matemática", "Matem\u00e1tica"),
        ("C.Sociales", "C.Sociales"),
        ("C.Naturales", "C.Naturales"),
        ("Religión", "Religi\u00f3n"),
        ("Artística", "Art\u00edstica"),
        ("Tecnología", "Tecnolog\u00eda"),
        ("E.Física", "E.F\u00edsica")
    ]
    
    # Student 35 in Block 2 (Row 78)
    ws_res.cell(row=78, column=2).value = '=IF(\'MAESTRO 1\u00b0\'!B39="","",\'MAESTRO 1\u00b0\'!B39)'
    for c_idx, (sub_name, sheet_name) in enumerate(subjects, 3):
        ws_res.cell(row=78, column=c_idx).value = f"=IF('MAESTRO 1\u00b0'!B39=\"\",\"\",IFERROR('Planilla ({sheet_name} 1\u00b0) '!G50,\"\"))"
        
    # Student 35 in Block 3 (Row 120)
    ws_res.cell(row=120, column=2).value = '=IF(\'MAESTRO 1\u00b0\'!B39="","",\'MAESTRO 1\u00b0\'!B39)'
    for c_idx, (sub_name, sheet_name) in enumerate(subjects, 3):
        ws_res.cell(row=120, column=c_idx).value = f"=IF('MAESTRO 1\u00b0'!B39=\"\",\"\",IFERROR('Planilla ({sheet_name} 1\u00b0) '!H50,\"\"))"
        
    # 2. Repair Asistencia 1°
    ws_asis = wb["Asistencia 1\u00b0 "]
    # Block 1 (Trimestre 1): rows 37 to 42 (student 35 to 40)
    for r in range(37, 43):
        maestro_row = r + 2
        ws_asis.cell(row=r, column=2).value = f"=IF('MAESTRO 1\u00b0'!B{maestro_row}=\"\",\"\",'MAESTRO 1\u00b0'!B{maestro_row})"
        
    # Block 2 (Trimestre 2): rows 80 to 85 (student 35 to 40)
    for r in range(80, 86):
        maestro_row = r - 41
        ws_asis.cell(row=r, column=2).value = f"=IF('MAESTRO 1\u00b0'!B{maestro_row}=\"\",\"\",'MAESTRO 1\u00b0'!B{maestro_row})"
        
    # Block 3 (Trimestre 3): rows 123 to 128 (student 35 to 40)
    for r in range(123, 129):
        maestro_row = r - 84
        ws_asis.cell(row=r, column=2).value = f"=IF('MAESTRO 1\u00b0'!B{maestro_row}=\"\",\"\",'MAESTRO 1\u00b0'!B{maestro_row})"
        
    wb.save("Registro_Primaria.xlsx")
    wb.close()
    print("Registro_Primaria.xlsx repaired successfully.")

if __name__ == "__main__":
    repair_premedia_file("Registro_2026.xlsx")
    repair_premedia_file("Registro_2026_FINAL_v2.xlsx")
    repair_registro_primaria()
