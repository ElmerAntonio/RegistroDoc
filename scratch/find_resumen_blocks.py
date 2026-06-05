import openpyxl

wb = openpyxl.load_workbook("Registro_Primaria.xlsx")
ws = wb["RESUMEN_1°"]

for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val in ["No.", "No"]:
        print(f"Block header at row {r}, student 1 at row {r+2}")
wb.close()
