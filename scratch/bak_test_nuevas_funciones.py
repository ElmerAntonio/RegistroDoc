import pytest
import os
import openpyxl
import shutil
from rddata import DataEngine, validar_nota_meduca
from rdprint import encontrar_hoja_impresion

@pytest.fixture
def temp_workbooks(tmp_path):
    orig_primaria = "c:/RegistroDoc/Registro_Primaria.xlsx"
    orig_premedia = "c:/RegistroDoc/Registro_2026.xlsx"
    
    dest_primaria = tmp_path / "Registro_Primaria.xlsx"
    dest_premedia = tmp_path / "Registro_2026.xlsx"
    
    shutil.copy(orig_primaria, dest_primaria)
    shutil.copy(orig_premedia, dest_premedia)
    
    return {
        "primaria": str(dest_primaria),
        "premedia": str(dest_premedia)
    }

def test_config_propagation(temp_workbooks):
    """Test that configuring profile variables propagates to all sheets."""
    engine = DataEngine(ruta_excel=temp_workbooks["primaria"], modalidad="primaria")
    
    datos_config = {
        "escuela_nombre": "ESCUELA TEST",
        "escuela_region": "REGION TEST",
        "distrito": "DISTRITO TEST",
        "corregimiento": "CORR TEST",
        "zona_escolar": "1",
        "docente_nombre": "PROF. TESTER",
        "docente_cedula": "8-000-000",
        "ano_lectivo": "2027",
        "jornada": "Matutina",
        "director_nombre": "DIR TEST",
        "subdirector_nombre": "SUBDIR TEST",
        "coordinador_nombre": "COORD TEST",
        "telefono": "123",
        "correo": "test@test.com",
        "seguro_social": "999",
        "numero_posicion": "777"
    }
    
    engine.sincronizar_plantilla_maestra(datos_config)
    
    wb = openpyxl.load_workbook(temp_workbooks["primaria"])
    
    # Portada - docente is written to cell P14
    ws_portada = wb["Portada"]
    assert ws_portada["P14"].value == "PROF. TESTER"
    
    # Planilla - docente is written to cell L11 (row 11, col 12)
    # Note the trailing space in the sheet name
    ws_planilla = wb["Planilla (Ingles 1°) "]
    assert ws_planilla["L11"].value == "PROF. TESTER"
    
    wb.close()

def test_calculation_accuracy_bypass(temp_workbooks):
    """Test that obtaining grades averages bypasses openpyxl formula cache using pure Python calculations."""
    engine = DataEngine(ruta_excel=temp_workbooks["primaria"], modalidad="primaria")
    
    # 1. Registrar estudiante de prueba
    exito = engine.agregar_estudiante("1°", "ALUMNO TESTER", "8-123-456")
    assert exito
    
    # Obtener el nombre del estudiante registrado en el grado y su fila correspondientemente
    estudiantes = engine.obtener_estudiantes_completos("1°")
    assert len(estudiantes) > 0
    
    idx_tester = next(i for i, est in enumerate(estudiantes) if est["nombre"] == "ALUMNO TESTER")
    tester_row = 5 + idx_tester
    
    wb = openpyxl.load_workbook(temp_workbooks["primaria"])
    ws = wb["PROM (Ingles 1°)"]
    
    # Col 3 a 5 (Diaria/Parcial Trimestre 1)
    ws.cell(row=tester_row, column=3).value = 4.5
    ws.cell(row=tester_row, column=4).value = 3.5
    ws.cell(row=tester_row, column=5).value = 5.0
    
    # Apreciación Trimestre 1: Col 18 a 19
    ws.cell(row=tester_row, column=18).value = 4.0
    ws.cell(row=tester_row, column=19).value = 5.0
    
    wb.save(temp_workbooks["primaria"])
    wb.close()
    
    # Recargar el caché de DataEngine con los datos de las notas guardados en el disco
    engine._cargar_en_memoria()
    
    # Obtener promedios reales
    promedios = engine.obtener_promedios_reales("1°", "Ingles", "Trimestre 1")
    
    # Estudiante 1
    # promedio diario = (4.5 + 3.5 + 5.0) / 3 = 4.333
    # promedio apreciacion = (4.0 + 5.0) / 2 = 4.5
    # Calificación Trimestral = (4.33 + 4.5) / 2 = 4.41 -> rounded to 4.4
    assert promedios["ALUMNO TESTER"] == pytest.approx(4.4, abs=0.05)

def test_sheet_matching_robustness(temp_workbooks):
    """Test sheet matcher resolution logic under both modalities."""
    engine_prim = DataEngine(ruta_excel=temp_workbooks["primaria"], modalidad="primaria")
    engine_pre = DataEngine(ruta_excel=temp_workbooks["premedia"], modalidad="premedia")
    
    wb_prim = openpyxl.load_workbook(temp_workbooks["primaria"])
    sheet_name = engine_prim._encontrar_hoja_prom(wb_prim, "1°", "Ingles")
    assert sheet_name == "PROM (Ingles 1°)"
    wb_prim.close()
    
    wb_pre = openpyxl.load_workbook(temp_workbooks["premedia"])
    sheet_name = engine_pre._encontrar_hoja_prom(wb_pre, "7°", "Ingles")
    assert "PROM (Ingles 7" in sheet_name
    wb_pre.close()

def test_validation_rules():
    """Test custom MEDUCA notes validation rules."""
    assert validar_nota_meduca("4.5") == (True, 4.5, "OK")
    assert validar_nota_meduca("3") == (True, 3.0, "OK")
    assert validar_nota_meduca(5) == (True, 5.0, "OK")
    
    valido, _, _ = validar_nota_meduca("0.5")
    assert not valido
    valido, _, _ = validar_nota_meduca("5.1")
    assert not valido
    valido, _, _ = validar_nota_meduca("abc")
    assert not valido

def test_dashboard_dynamic_stats(temp_workbooks):
    """Test dashboard dynamic stats retrieval and rendering layout values."""
    engine_prim = DataEngine(ruta_excel=temp_workbooks["primaria"], modalidad="primaria")
    stats = engine_prim.get_dashboard_stats()
    assert "total" in stats
    assert "riesgo" in stats
    assert "honor" in stats
    assert "asistencia" in stats
    
    grados = engine_prim.obtener_grados_activos()
    assert len(grados) > 0

def test_formatear_reportes_excel_primaria(temp_workbooks):
    """Test that formatear_reportes_excel writes valid dynamic formulas without errors."""
    engine = DataEngine(ruta_excel=temp_workbooks["primaria"], modalidad="primaria")
    
    # 1. Agregar alumnos de prueba con sexo
    assert engine.agregar_estudiante("1°", "ALUMNO M", "8-123-456", "M")
    assert engine.agregar_estudiante("1°", "ALUMNA F", "8-789-012", "F")
    
    # 2. Formatear
    exito = engine.formatear_reportes_excel("1°")
    assert exito
    
    # 3. Verificar formulas escritas
    wb = openpyxl.load_workbook(temp_workbooks["primaria"])
    assert "REPORTE_DIRECCIÓN" in wb.sheetnames
    ws_dir = wb["REPORTE_DIRECCIÓN"]
    
    # Sexo count formulas are at row 8, column 3 (C) and 4 (D)
    assert ws_dir["C8"].value == "=COUNTIF('MAESTRO 1°'!G5:G48,\"M\")"
    assert ws_dir["D8"].value == "=COUNTIF('MAESTRO 1°'!G5:G48,\"F\")"
    
    # Approved list formulas are updated
    assert "REPORTE_APROBADOS" in wb.sheetnames
    ws_ap = wb["REPORTE_APROBADOS"]
    assert "ESTUDIANTES APROBADOS" in ws_ap["A4"].value
    
    wb.close()


def test_formatear_cedula_panamena():
    """Test that formatear_cedula_panamena correctly formats various Panamanian cedula inputs."""
    from dapp import EstudiantesFrame
    
    class DummyFrame:
        formatear_cedula_panamena = EstudiantesFrame.formatear_cedula_panamena
        
    df = DummyFrame()
    
    # 1. Standard format without hyphens
    assert df.formatear_cedula_panamena("8123456") == "8-123-456"
    assert df.formatear_cedula_panamena("10123456") == "10-123-456"
    
    # 2. Standard format with spaces or uppercase/lowercase mix
    assert df.formatear_cedula_panamena(" 8-123-456 ") == "8-123-456"
    
    # 3. Foreign/Special prefixes
    assert df.formatear_cedula_panamena("PE123456") == "PE-123-456"
    assert df.formatear_cedula_panamena("E123456") == "E-123-456"
    assert df.formatear_cedula_panamena("n123456") == "N-123-456"
    assert df.formatear_cedula_panamena("PI123456") == "PI-123-456"
    
    # 4. Foreign/Special custom hyphenated formats (must be preserved)
    assert df.formatear_cedula_panamena("15-329-874") == "15-329-874"
    assert df.formatear_cedula_panamena("PASS-123-456") == "PASS-123-456"
    assert df.formatear_cedula_panamena("99-ABC-88") == "99-ABC-88"


def test_get_standard_planilla_name():
    """Test get_standard_planilla_name normalization behavior."""
    from src.rddata import DataEngine
    engine = DataEngine("Registro_Primaria.xlsx", modalidad="primaria")
    
    # Test standard subjects mapping
    assert engine.get_standard_planilla_name("INGLES", "1°") == "Planilla (Ingles 1°) "
    assert engine.get_standard_planilla_name("ESPAÑOL", "1°") == "Planilla (Español 1°) "
    assert engine.get_standard_planilla_name("Matemática", "1°") == "Planilla (Matemática 1°) "
    assert engine.get_standard_planilla_name("C.SOCIALES", "1°") == "Planilla (C.Sociales 1°) "
    assert engine.get_standard_planilla_name("E. FISICA", "1°") == "Planilla (E.Física 1°) "
    
    # Test fallback title casing for unmapped subject
    assert engine.get_standard_planilla_name("Biologia", "1°") == "Planilla (Biologia 1°) "


def test_corregir_formulas_generales(tmp_path):
    """Test that _corregir_formulas_generales fixes DASHBOARD and GRAFICOS formulas correctly."""
    import shutil
    import openpyxl
    from src.rddata import DataEngine
    
    # Copy template files to tmp_path to avoid modifying actual workspace files during test
    temp_primaria = tmp_path / "Registro_Primaria_Test.xlsx"
    shutil.copy("Registro_Primaria.xlsx", temp_primaria)
    
    # Instantiate engine
    engine = DataEngine(str(temp_primaria), modalidad="primaria")
    
    # Load and break dashboard formulas intentionally
    wb = openpyxl.load_workbook(str(temp_primaria))
    ws_db = wb["DASHBOARD"]
    ws_db["B13"] = "broken formula"
    ws_db["B14"] = "broken formula 2"
    
    # Apply self-healing correction
    engine._corregir_formulas_generales(wb)
    
    # Check that they are restored
    assert "IFERROR" in ws_db["B13"].value
    assert "IFERROR" in ws_db["B14"].value
    
    wb.close()




