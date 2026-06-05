import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

files = [
    "Registro_2026.xlsx",
    "Registro_2026_FINAL_v2.xlsx",
    "Registro_Primaria.xlsx"
]

for filename in files:
    print(f"\n===== FILE: {filename} =====")
    try:
        wb = openpyxl.load_workbook(filename)
    except Exception as e:
        print(f"Error loading {filename}: {e}")
        continue
    
    # 1. Caratula/Portada
    for name in ["Caratula", "Portada"]:
        if name in wb.sheetnames:
            ws = wb[name]
            if filename == "Registro_Primaria.xlsx":
                print(f"  {name} O12:", ws["O12"].value)
                print(f"  {name} U18:", ws["U18"].value)
                print(f"  {name} AD20:", ws["AD20"].value)
            else:
                if "Q24" in ws:
                    print(f"  {name} Q24:", ws["Q24"].value)
                else:
                    # Let's search for Insturctor or Instructor in row 24
                    for col in range(1, 40):
                        val = ws.cell(row=24, column=col).value
                        if val and ("insturctor" in str(val).lower() or "instructor" in str(val).lower()):
                            print(f"  {name} Row 24 Col {col}:", val)

    # 2. Planillas Commerce/Agro 8°
    for name in ["Planilla (Comercio 8°)", "Planilla (Agropecuaria 8°)"]:
        if name in wb.sheetnames:
            ws = wb[name]
            print(f"  {name} J46:", ws["J46"].value)
            print(f"  {name} O46:", ws["O46"].value)

    # 3. GRAFICOS
    if "GRAFICOS" in wb.sheetnames:
        ws = wb["GRAFICOS"]
        print("  GRAFICOS I7:", ws["I7"].value)
        print("  GRAFICOS J7:", ws["J7"].value)

    # 4. DASHBOARD
    if "DASHBOARD" in wb.sheetnames:
        ws = wb["DASHBOARD"]
        print("  DASHBOARD B13:", ws["B13"].value)
        print("  DASHBOARD B14:", ws["B14"].value)

    # 5. RESUMEN_1°
    if "RESUMEN_1°" in wb.sheetnames:
        ws = wb["RESUMEN_1°"]
        ref_cells = []
        for r in range(5, 39):
            val = ws.cell(row=r, column=3).value # Col C (Ingles)
            if val and "#REF!" in str(val):
                ref_cells.append(f"C{r}")
        print(f"  RESUMEN_1° has {len(ref_cells)} C-column cells with #REF! (checked rows 5-38)")
        if ref_cells:
            print("    Sample #REF! cell:", ref_cells[0], ws[ref_cells[0]].value)
        else:
            print("    Sample Col C cell (row 5):", ws["C5"].value)
            print("    Sample Col C cell (row 44):", ws["C44"].value)

    wb.close()
