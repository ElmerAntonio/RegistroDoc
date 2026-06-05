import openpyxl

wb = openpyxl.load_workbook("Registro_Primaria.xlsx")
ws = wb["Asistencia 1° "]

# Check around row 37
print("=== Around Row 37 ===")
for r in range(35, 44):
    print(f"Row {r}: {ws.cell(row=r, column=2).value}")

# Check around row 80
print("=== Around Row 80 ===")
for r in range(78, 87):
    print(f"Row {r}: {ws.cell(row=r, column=2).value}")

# Check around row 123
print("=== Around Row 123 ===")
for r in range(121, 130):
    print(f"Row {r}: {ws.cell(row=r, column=2).value}")

wb.close()
