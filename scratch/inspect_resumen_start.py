import openpyxl

wb = openpyxl.load_workbook("Registro_Primaria.xlsx")
ws = wb["RESUMEN_1°"]

for r in range(40, 47):
    print(f"Row {r}:")
    for col in range(1, 6):
        val = ws.cell(row=r, column=col).value
        print(f"  Col {col} ({openpyxl.utils.get_column_letter(col)}): {val}")
wb.close()
