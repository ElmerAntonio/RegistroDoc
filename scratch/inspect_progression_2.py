import openpyxl

wb = openpyxl.load_workbook("Registro_2026.xlsx")
ws = wb["Planilla (Agropecuaria 8°) "]

for r in range(15, 26):
    print(f"Row {r}:")
    print(f"  Col 2 (B): {ws.cell(row=r, column=2).value}")
    print(f"  Col 3 (C): {ws.cell(row=r, column=3).value}")
    print(f"  Col 10 (J): {ws.cell(row=r, column=10).value}")
wb.close()
