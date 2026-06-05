import openpyxl

wb = openpyxl.load_workbook("Registro_2026.xlsx")
ws = wb["Planilla (Agropecuaria 8°) "]

for r in [45, 46]:
    print(f"Row {r}:")
    for col in range(1, 16):
        val = ws.cell(row=r, column=col).value
        print(f"  Col {col} ({openpyxl.utils.get_column_letter(col)}): {val}")
wb.close()
