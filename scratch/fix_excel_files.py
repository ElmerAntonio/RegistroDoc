import openpyxl

def get_standard_planilla_name(subject, grado):
    sub_map = {
        "INGLES": "Ingles",
        "ESPAÑOL": "Español",
        "MATEMÁTICA": "Matemática",
        "MATEMATICA": "Matemática",
        "C.SOCIALES": "C.Sociales",
        "C. NATURALES": "C.Naturales",
        "C.NATURALES": "C.Naturales",
        "RELIGIÓN": "Religión",
        "RELIGION": "Religión",
        "ARTÍSTICA": "Artística",
        "ARTISTICA": "Artística",
        "TECNOLOGIA": "Tecnología",
        "TECNOLOGÍA": "Tecnología",
        "E. FISICA": "E.Física",
        "E.FISICA": "E.Física",
    }
    sub_clean = subject.upper().strip()
    sub_title = sub_map.get(sub_clean, subject.title().strip())
    return f"Planilla ({sub_title} {grado}) "

def fix_premedia_file(filename):
    print(f"Repairing Premedia file: {filename}...")
    wb = openpyxl.load_workbook(filename)
    
    # 1. Caratula Q24 typo
    if "Caratula" in wb.sheetnames:
        ws = wb["Caratula"]
        val = ws["Q24"].value
        if val and "Insturctor" in str(val):
            ws["Q24"].value = str(val).replace("Insturctor", "Instructor")
            print("  Fixed Caratula Q24 typo.")

    # 2. Planillas Commerce/Agro 8° #REF!
    for sheet in ["Planilla (Comercio 8°)", "Planilla (Agropecuaria 8°)"]:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            fixed_count = 0
            for col in range(10, 16): # Columns J to O
                cell = ws.cell(row=46, column=col)
                val = cell.value
                if val and "MAESTRO!#REF!" in str(val):
                    cell.value = str(val).replace("MAESTRO!#REF!", "MAESTRO!D36")
                    fixed_count += 1
            print(f"  Fixed {fixed_count} #REF! formulas in {sheet}.")

    # 3. GRAFICOS I7, J7 compatibility
    if "GRAFICOS" in wb.sheetnames:
        ws = wb["GRAFICOS"]
        ws["I7"].value = '=IFERROR(INDEX(C6:C125,MATCH(MINIFS(E6:E125,E6:E125,">0"),E6:E125,0)),"")'
        ws["J7"].value = '=IFERROR(MINIFS(E6:E125,E6:E125,">0"),"")'
        print("  Fixed GRAFICOS I7 and J7 compatibility formulas.")

    # 4. DASHBOARD B13, B14
    if "DASHBOARD" in wb.sheetnames:
        ws = wb["DASHBOARD"]
        ws["B13"].value = '=IF(MAX(E6:E8)>0,"Reforzar grado "&IFERROR(INDEX(D6:D8,MATCH(MAX(E6:E8),E6:E8,0)),"")&" ("&MAX(E6:E8)&" estudiantes bajo 3.0)","✓ Sin alertas académicas")'
        ws["B14"].value = '=IF(MAX(H6:H8)>0,"Revisar ausencias en grado "&IFERROR(INDEX(G6:G8,MATCH(MAX(H6:H8),H6:H8,0)),"")&" ("&MAX(H6:H8)&" ausencias)","✓ Sin alertas de asistencia")'
        print("  Fixed DASHBOARD B13 and B14 IFERROR formulas.")

    wb.save(filename)
    wb.close()
    print(f"File {filename} saved successfully.\n")

def fix_primaria_file(filename):
    print(f"Repairing Primaria file: {filename}...")
    wb = openpyxl.load_workbook(filename)

    # 1. Portada O12, U18, AD20
    if "Portada" in wb.sheetnames:
        ws = wb["Portada"]
        ws["O12"].value = "PRIMARIA MULTIGRADO"
        ws["U18"].value = "Maestro de Grado"
        ws["AD20"].value = "Todas las materias"
        print("  Fixed Portada info (O12, U18, AD20).")

    # 2. DASHBOARD B13, B14
    if "DASHBOARD" in wb.sheetnames:
        ws = wb["DASHBOARD"]
        ws["B13"].value = '=IF(MAX(E6:E6)>0,"Reforzar grado "&IFERROR(INDEX(D6:D6,MATCH(MAX(E6:E6),E6:E6,0)),"")&" ("&MAX(E6:E6)&" estudiantes bajo 3.0)","✓ Sin alertas académicas")'
        ws["B14"].value = '=IF(MAX(H6:H6)>0,"Revisar ausencias en grado "&IFERROR(INDEX(G6:G6,MATCH(MAX(H6:H6),H6:H6,0)),"")&" ("&MAX(H6:H6)&" ausencias)","✓ Sin alertas de asistencia")'
        print("  Fixed DASHBOARD B13 and B14 IFERROR formulas.")

    # 3. RESUMEN_1° Formulas Rebuild
    if "RESUMEN_1°" in wb.sheetnames:
        ws = wb["RESUMEN_1°"]
        
        # Mapear materias por columna en T1
        cols_materias_t = {}
        for col in range(3, 12): # Columns C to K
            val = ws.cell(row=3, column=col).value
            if val:
                cols_materias_t[col] = str(val).strip()
                
        # Mapear materias por columna en ANUAL
        cols_materias_a = {}
        for col in range(16, 25): # Columns P to X
            val = ws.cell(row=3, column=col).value
            if val:
                cols_materias_a[col] = str(val).strip()
                
        # Reconstruir celdas de estudiantes
        for i in range(1, 35): # 34 students max
            r_t1 = 4 + i
            r_t2 = 43 + i
            r_t3 = 85 + i
            
            # Nombre formulas
            f_name = f"=IF('MAESTRO 1°'!B{4+i}=\"\",\"\",'MAESTRO 1°'!B{4+i})"
            ws.cell(row=r_t1, column=2).value = f_name
            ws.cell(row=r_t2, column=2).value = f_name
            ws.cell(row=r_t3, column=2).value = f_name
            ws.cell(row=r_t1, column=15).value = f_name
            
            # T1 materias
            for col, m_name in cols_materias_t.items():
                plan_sheet = get_standard_planilla_name(m_name, "1°")
                ws.cell(row=r_t1, column=col).value = f"=IF('MAESTRO 1°'!B{4+i}=\"\",\"\",IFERROR('{plan_sheet}'!F{15+i},\"\"))"
                
            # T2 materias
            for col, m_name in cols_materias_t.items():
                plan_sheet = get_standard_planilla_name(m_name, "1°")
                ws.cell(row=r_t2, column=col).value = f"=IF('MAESTRO 1°'!B{4+i}=\"\",\"\",IFERROR('{plan_sheet}'!G{15+i},\"\"))"
                
            # T3 materias
            for col, m_name in cols_materias_t.items():
                plan_sheet = get_standard_planilla_name(m_name, "1°")
                ws.cell(row=r_t3, column=col).value = f"=IF('MAESTRO 1°'!B{4+i}=\"\",\"\",IFERROR('{plan_sheet}'!H{15+i},\"\"))"
                
            # ANUAL materias
            for col, m_name in cols_materias_a.items():
                plan_sheet = get_standard_planilla_name(m_name, "1°")
                ws.cell(row=r_t1, column=col).value = f"=IF('MAESTRO 1°'!B{4+i}=\"\",\"\",IFERROR('{plan_sheet}'!I{15+i},\"\"))"

        print("  Rebuilt RESUMEN_1° student formulas successfully.")

    wb.save(filename)
    wb.close()
    print(f"File {filename} saved successfully.\n")

if __name__ == "__main__":
    fix_premedia_file("Registro_2026.xlsx")
    fix_premedia_file("Registro_2026_FINAL_v2.xlsx")
    fix_primaria_file("Registro_Primaria.xlsx")
