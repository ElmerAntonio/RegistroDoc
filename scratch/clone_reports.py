import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os

def copy_sheet(source_sheet, target_sheet):
    # Copy merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    # Copy row and column dimensions
    for col in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
    for row in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

    # Copy cell values, formats, styles
    for r in range(1, source_sheet.max_row + 1):
        for c in range(1, source_sheet.max_column + 1):
            source_cell = source_sheet.cell(row=r, column=c)
            target_cell = target_sheet.cell(row=r, column=c, value=source_cell.value)
            
            # Copy style properties
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)

def adapt_dashboard(ws, active_prom_sheets):
    # Adapts the DASHBOARD sheet for Primaria 1° Grade
    # Row 6: 1°
    ws.cell(row=6, column=1, value="Total estudiantes 1°")
    ws.cell(row=6, column=2, value="=IFERROR(COUNTA('MAESTRO 1°'!B5:B48),0)")
    ws.cell(row=6, column=4, value="1°")
    
    # Construct Riesgo Academico formula using actual PROM sheets
    if active_prom_sheets:
        countif_terms = [f"COUNTIF('{sh}'!AE5:AE48,\"<3\")" for sh in active_prom_sheets]
        formula_riesgo = f"=IFERROR(MAX({','.join(countif_terms)}),\"\")"
    else:
        formula_riesgo = ""
    ws.cell(row=6, column=5, value=formula_riesgo)
    
    ws.cell(row=6, column=7, value="1°")
    ws.cell(row=6, column=8, value="=IFERROR(SUMPRODUCT(('Asistencia 1° '!C3:BH46=\"-\")*1)+SUMPRODUCT(('Asistencia 1° '!C50:BH93=\"-\")*1)+SUMPRODUCT(('Asistencia 1° '!C97:BH140=\"-\")*1),\"\")")

    # Clear Rows 7 and 8 (used to be 8° and 9° in Premedia)
    for r in [7, 8]:
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell':
                cell.value = ""

    # Update Row 9 Total General
    ws.cell(row=9, column=1, value="Total general")
    ws.cell(row=9, column=2, value="=SUM(B6:B6)")  # Only sum 1° row
    
    # Update alerts in row 13 and 14
    ws.cell(row=13, column=2, value="=IF(MAX(E6:E6)>0,\"Reforzar grado \"&INDEX(D6:D6,MATCH(MAX(E6:E6),E6:E6,0))&\" (\"&MAX(E6:E6)&\" estudiantes bajo 3.0)\",\"✓ Sin alertas académicas\")")
    ws.cell(row=14, column=2, value="=IF(MAX(H6:H6)>0,\"Revisar ausencias en grado \"&INDEX(G6:G6,MATCH(MAX(H6:H6),H6:H6,0))&\" (\"&MAX(H6:H6)&\" ausencias)\",\"✓ Sin alertas de asistencia\")")

def adapt_reporte_direccion(ws, active_prom_sheets):
    # Title
    ws.cell(row=4, column=1, value="Docente: Elmer Tugri  |  Primaria Multigrado  |  Fecha: =TEXT(TODAY(),\"DD/MM/YYYY\")")
    
    # Table rows
    # Row 15: 1°
    ws.cell(row=15, column=1, value="1°")
    ws.cell(row=15, column=2, value="=IFERROR(COUNTA('MAESTRO 1°'!B5:B48),\"\")")
    ws.cell(row=15, column=3, value="=COUNTIF(RESUMEN_1°!L5:L48,\"✓ APROBADO\")")
    ws.cell(row=15, column=4, value="=COUNTIF(RESUMEN_1°!L5:L48,\"✗ REPROBADO\")")
    ws.cell(row=15, column=5, value="=IFERROR(ROUND(C15/B15*100,1)&\"%\",\"\")")
    
    # Promedio Ingles (PROM (Ingles 1°))
    ingles_sheet = next((sh for sh in active_prom_sheets if "ingles" in sh.lower()), None)
    if ingles_sheet:
        ws.cell(row=15, column=6, value=f"=IFERROR(ROUND(AVERAGEIF('{ingles_sheet}'!AE5:AE48,\"<>\"&\"\"),1),\"\")")
    else:
        ws.cell(row=15, column=6, value="")
        
    # Promedio Tecnologia (PROM (Tecnología 1°))
    tecnologia_sheet = next((sh for sh in active_prom_sheets if "tecnolog" in sh.lower()), None)
    if tecnologia_sheet:
        ws.cell(row=15, column=7, value=f"=IFERROR(ROUND(AVERAGEIF('{tecnologia_sheet}'!AE5:AE48,\"<>\"&\"\"),1),\"\")")
    else:
        ws.cell(row=15, column=7, value="")

    # Clear Row 16 and 17
    for r in [16, 17]:
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell':
                cell.value = ""
            
    # Update Total General (Row 11 in Matrícula table, wait, row 11 is B11)
    ws.cell(row=8, column=1, value="1°")
    ws.cell(row=8, column=2, value="=IFERROR(COUNTA('MAESTRO 1°'!B5:B48),\"\")")
    for r in [9, 10]:
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell':
                cell.value = ""
    ws.cell(row=11, column=2, value="=IFERROR(COUNTA('MAESTRO 1°'!B5:B48),\"\")")

def adapt_reporte_aprobados(ws):
    # Row 4 Title
    ws.cell(row=4, column=1, value="ESTUDIANTES APROBADOS — 1° GRADO")
    # Table headers: Row 5
    ws.cell(row=5, column=3, value="INGLÉS FINAL")
    ws.cell(row=5, column=4, value="TECNOLOGÍA FINAL")
    
    # Student rows (6 to 49 for 44 students)
    # The original sheet has 40 student rows. Let's extend it or adapt the existing 40 rows.
    # In Primaria, we can adapt rows 6 to 45 (which are student rows in target).
    for idx in range(40):
        r = 6 + idx
        student_excel_row = 5 + idx
        ws.cell(row=r, column=1, value=f'=IF(RESUMEN_1°!L{student_excel_row}="✓ APROBADO",{idx+1},"")')
        ws.cell(row=r, column=2, value=f'=IF(RESUMEN_1°!L{student_excel_row}="✓ APROBADO",RESUMEN_1°!B{student_excel_row},"")')
        ws.cell(row=r, column=3, value=f'=IF(RESUMEN_1°!L{student_excel_row}="✓ APROBADO",RESUMEN_1°!C{student_excel_row},"")')
        ws.cell(row=r, column=4, value=f'=IF(RESUMEN_1°!L{student_excel_row}="✓ APROBADO",RESUMEN_1°!J{student_excel_row},"")')
        ws.cell(row=r, column=5, value=f'=IF(RESUMEN_1°!L{student_excel_row}="✓ APROBADO",IFERROR(ROUND(AVERAGE(RESUMEN_1°!C{student_excel_row},RESUMEN_1°!J{student_excel_row}),1),""),"")')

    # Clear other tables (starts at Row 48 and Row 92)
    # Let's clear everything from Row 46 onwards to avoid confusion.
    for r in range(46, ws.max_row + 1):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell':
                cell.value = ""

def adapt_reporte_docente(ws, active_prom_sheets):
    # Headers
    ws.cell(row=5, column=2, value="1° GRADO")
    # Clear columns C and D
    for r in range(5, 13):
        for c in [3, 4]:
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell':
                cell.value = ""
        
    # Total column (E) becomes B (since others are empty, let's keep E pointing to B)
    ws.cell(row=6, column=2, value="=IFERROR(COUNTA('MAESTRO 1°'!B5:B48),\"\")")
    ws.cell(row=7, column=2, value="=COUNTIF(RESUMEN_1°!L5:L48,\"✓ APROBADO\")")
    ws.cell(row=8, column=2, value="=COUNTIF(RESUMEN_1°!L5:L48,\"✗ REPROBADO\")")
    ws.cell(row=9, column=2, value="=COUNTIF('Asistencia 1° '!BI3:BI46,\">5\")")
    
    # Promedio Ingles
    ingles_sheet = next((sh for sh in active_prom_sheets if "ingles" in sh.lower()), None)
    if ingles_sheet:
        ws.cell(row=10, column=2, value=f"=IFERROR(ROUND(AVERAGEIF('{ingles_sheet}'!AE5:AE48,\"<>\"&\"\"),1),\"\")")
    else:
        ws.cell(row=10, column=2, value="")
        
    # Promedio Tecnologia
    tecnologia_sheet = next((sh for sh in active_prom_sheets if "tecnolog" in sh.lower()), None)
    if tecnologia_sheet:
        ws.cell(row=11, column=2, value=f"=IFERROR(ROUND(AVERAGEIF('{tecnologia_sheet}'!AE5:AE48,\"<>\"&\"\"),1),\"\")")
    else:
        ws.cell(row=11, column=2, value="")

    # Update Totals (col E) to reference only col B
    ws.cell(row=6, column=5, value="=B6")
    ws.cell(row=7, column=5, value="=B7")
    ws.cell(row=8, column=5, value="=B8")
    ws.cell(row=9, column=5, value="=B9")
    ws.cell(row=10, column=5, value="=B10")
    ws.cell(row=11, column=5, value="=B11")

def main():
    source_path = "Registro_2026_FINAL_v2.xlsx"
    target_paths = ["Registro_Primaria.xlsx", "docs/Registro Primaria-back.xlsx"]
    
    wb_src = openpyxl.load_workbook(source_path)
    
    for t_path in target_paths:
        if not os.path.exists(t_path):
            print(f"File not found: {t_path}")
            continue
            
        print(f"Processing target: {t_path}")
        wb_tgt = openpyxl.load_workbook(t_path)
        
        # Get active PROM sheets in target
        active_prom_sheets = [sh for sh in wb_tgt.sheetnames if "PROM" in sh.upper()]
        print(f"Active PROM sheets in {t_path}: {active_prom_sheets}")
        
        for sh_name in ["DASHBOARD", "REPORTE_DIRECCIÓN", "REPORTE_APROBADOS", "REPORTE_DOCENTE"]:
            if sh_name in wb_tgt.sheetnames:
                del wb_tgt[sh_name]
                
            new_sheet = wb_tgt.create_sheet(title=sh_name)
            copy_sheet(wb_src[sh_name], new_sheet)
            
            if sh_name == "DASHBOARD":
                adapt_dashboard(new_sheet, active_prom_sheets)
            elif sh_name == "REPORTE_DIRECCIÓN":
                adapt_reporte_direccion(new_sheet, active_prom_sheets)
            elif sh_name == "REPORTE_APROBADOS":
                adapt_reporte_aprobados(new_sheet)
            elif sh_name == "REPORTE_DOCENTE":
                adapt_reporte_docente(new_sheet, active_prom_sheets)
                
            print(f"Copied and adapted sheet {sh_name}")
            
        wb_tgt.save(t_path)
        wb_tgt.close()
        print(f"Saved {t_path}")
        
    wb_src.close()
    print("Done!")

if __name__ == "__main__":
    main()
