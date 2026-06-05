import openpyxl

wb = openpyxl.load_workbook("Registro_2026.xlsx")
ws = wb["Planilla (Agropecuaria 8°) "]

for r in range(5, 15):
    print(f"Row {r}:")
    print(f"  Col 3 (C): {ws.cell(row=r, column=3).value}")
    print(f"  Col 10 (J): {ws.cell(row=r, column=10).value}")
wb.close()
