import time
import os
import openpyxl
from src.rddata import DataEngine

# Create dummy Excel file
dummy_path = "dummy_benchmark.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ASISTENCIA_12A"

# Add dummy dates in header row (row 5)
ws.cell(row=5, column=5, value="01-03")
ws.cell(row=5, column=6, value="02-03")
ws.cell(row=5, column=7, value="03-03")

wb.save(dummy_path)
wb.close()

# --- Unoptimized Engine (Simulate bypassing cache) ---
class UnoptimizedDataEngine(DataEngine):
    def obtener_fechas_asistencia_unoptimized(self, grado, trimestre):
        if not os.path.exists(self.ruta): return []
        wb = openpyxl.load_workbook(self.ruta, data_only=True)
        hoja = self._encontrar_hoja_asistencia(wb, grado)
        if not hoja:
            wb.close()
            return []
        ws = wb[hoja]
        mapa_trimestres = {"Trimestre 1": 2, "Trimestre 2": 45, "Trimestre 3": 88}
        fila_fechas = mapa_trimestres.get(trimestre, 2)

        fechas = []
        for col in range(5, 45):
            val = ws.cell(row=fila_fechas, column=col).value
            if val is not None:
                fechas.append(str(val))

        wb.close()
        return fechas

engine_unopt = UnoptimizedDataEngine(dummy_path)
engine_unopt._cargar_en_memoria()

# Benchmark Unoptimized
iterations = 100
start_time = time.time()
for _ in range(iterations):
    fechas = engine_unopt.obtener_fechas_asistencia_unoptimized("12°A", "Trimestre 1")
end_time_unopt = time.time()
time_unopt = end_time_unopt - start_time


# --- Optimized Engine ---
engine_opt = DataEngine(dummy_path)
engine_opt._cargar_en_memoria()

# Benchmark Optimized
start_time = time.time()
for _ in range(iterations):
    # Call the original method that we are optimizing, but right now it doesn't bypass cache. Let's make sure it does.
    fechas = engine_opt.obtener_fechas_asistencia("12°A", "Trimestre 1")
end_time_opt = time.time()
time_opt = end_time_opt - start_time

print(f"Time for {iterations} calls (Unoptimized): {time_unopt:.4f} seconds")
print(f"Time for {iterations} calls (Optimized): {time_opt:.4f} seconds")
if time_unopt > 0:
    print(f"Improvement: {(time_unopt - time_opt) / time_unopt * 100:.2f}%")

# Clean up
if engine_unopt._wb_cache:
    engine_unopt._wb_cache.close()
if engine_opt._wb_cache:
    engine_opt._wb_cache.close()
os.remove(dummy_path)
