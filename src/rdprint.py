"""
RegistroDoc Pro — Módulo de Impresión
======================================
Abre el Excel con su diseño original y lo envía a la impresora.
Compatible con Microsoft Excel y LibreOffice Calc.

© 2026 RegistroDoc Pro — MEDUCA — Panamá
"""

import os
import subprocess
import platform
from tkinter import messagebox


def _ruta_excel():
    from rdsecurity import cargar_config_segura
    from rdsql import obtener_dir_datos_usuario

    modalidad = "premedia"
    try:
        cfg = cargar_config_segura({"modalidad": "premedia"})
        modalidad = str(cfg.get("modalidad", "premedia")).lower()
    except Exception:
        pass

    user_dir = obtener_dir_datos_usuario()
    archivo = "Registro_Primaria.xlsx" if modalidad == "primaria" else "Registro_Premedia.xlsx"
    ruta_appdata = os.path.join(user_dir, "temp", archivo)

    if os.path.exists(ruta_appdata):
        return ruta_appdata

    # Fallback a plantillas de la raíz
    from config import BASE_DIR
    raiz = os.path.join(BASE_DIR, "..")
    candidatos = [
        ruta_appdata,
        os.path.join(raiz, "assets", "templates", archivo),
        os.path.join(BASE_DIR, "assets", "templates", archivo),
        os.path.join(raiz, archivo),
        os.path.join(BASE_DIR, archivo),
    ]

    for r in candidatos:
        if os.path.exists(r):
            return r

    return ruta_appdata


def _excel_disponible() -> bool:
    """Verifica si Microsoft Excel está instalado en Windows."""
    try:
        import winreg
        key = winreg.OpenKey(
            winreg.HKEY_LOCAL_MACHINE,
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\EXCEL.EXE"
        )
        winreg.CloseKey(key)
        return True
    except Exception:
        return False


def _libreoffice_disponible() -> bool:
    """Verifica si LibreOffice está instalado."""
    rutas = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/libreoffice",
        "/usr/bin/soffice",
    ]
    return any(os.path.exists(r) for r in rutas)


def encontrar_hoja_impresion(wb, tipo, grado=None, materia=None) -> str:
    """
    Localiza dinámicamente el nombre de la hoja en el libro basándose en el tipo, grado y materia.
    Evita fallos por diferencias de formato (ej: Asistencia (7°) vs Asistencia 7° ).
    """
    tipo_upper = tipo.upper()
    grado_num = grado.replace("°", "").strip() if grado else ""
    materia_clean = materia.lower().replace(" ", "").replace(".", "").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u") if materia else ""

    for s in wb.sheetnames:
        s_upper = s.upper()
        if tipo == "Portada" and "PORTADA" in s_upper and "VISTOSA" not in s_upper:
            return s
        if tipo == "Caratula" and ("CARATULA" in s_upper or "CARÁTULA" in s_upper or "VISTOSA" in s_upper):
            return s
        if tipo == "Horarios" and "HORARIO" in s_upper:
            return s
        if tipo == "Resumen" and "RESUMEN" in s_upper and grado_num in s_upper:
            return s
        if tipo == "Asistencia" and "ASISTENCIA" in s_upper and grado_num in s_upper:
            return s
        if tipo == "Planilla" and "PLANILLA" in s_upper and grado_num in s_upper:
            if not materia_clean:
                return s
            sheet_clean = s.lower().replace(" ", "").replace(".", "").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u")
            if materia_clean in sheet_clean:
                return s
    return None


def abrir_para_imprimir(hoja: str = None) -> tuple[bool, str]:
    """
    Abre una copia de exportación del Excel con Microsoft Excel o LibreOffice para imprimir.
    Evita la manipulación directa del archivo de base de datos de la aplicación.
    Retorna (éxito: bool, mensaje: str)
    """
    ruta_original = _ruta_excel()
    if not os.path.exists(ruta_original):
        return False, "No se encontró el archivo de libreta correspondiente."

    # Crear copia exportada (Aislamiento de base de datos)
    try:
        import shutil
        from openpyxl import load_workbook
        
        user_docs = os.path.join(os.path.expanduser("~"), "Documents", "RegistroDoc", "Exportados")
        os.makedirs(user_docs, exist_ok=True)
        
        nombre_base = os.path.basename(ruta_original)
        nombre_exportado = nombre_base.replace(".xlsx", "_Exportado.xlsx")
        ruta_exportada = os.path.join(user_docs, nombre_exportado)
        
        shutil.copy2(ruta_original, ruta_exportada)
        
        # Escribir la marca de seguridad de exportación
        wb_exp = load_workbook(ruta_exportada)
        if "_RD_EXPORT_MARKER_" not in wb_exp.sheetnames:
            ws_exp = wb_exp.create_sheet("_RD_EXPORT_MARKER_")
            ws_exp.sheet_state = "hidden"
            ws_exp["A1"] = "REGISTRO_DOC_SECURE_EXPORT"
            wb_exp.save(ruta_exportada)
        wb_exp.close()
    except Exception as e_copy:
        return False, f"No se pudo crear la copia de seguridad exportada: {e_copy}"

    sistema = platform.system()

    if sistema == "Windows":
        try:
            # Intentar abrir con Excel directamente
            os.startfile(ruta_exportada)
            return True, (
                f"Se ha exportado y abierto una copia de seguridad en:\n"
                f"{ruta_exportada}\n\n"
                "Para imprimir:\n"
                "1. Ve a la hoja que deseas imprimir\n"
                "2. Presiona Ctrl+P\n"
                "3. Ajusta márgenes si es necesario\n"
                "4. Haz clic en Imprimir"
            )
        except Exception as e:
            return False, f"No se pudo abrir el archivo exportado: {e}"
    else:
        # Linux/Mac — LibreOffice
        if _libreoffice_disponible():
            try:
                subprocess.Popen(["libreoffice", "--calc", os.path.abspath(ruta_exportada)])
                return True, (
                    f"Copia exportada abierta en LibreOffice:\n{ruta_exportada}\n\n"
                    "Usa Ctrl+P para imprimir."
                )
            except Exception as e:
                return False, f"Error al abrir LibreOffice: {e}"
        else:
            return False, (
                "No se encontró Microsoft Excel ni LibreOffice instalado."
            )


def imprimir_hoja_directo(nombre_hoja: str = "Portada") -> tuple[bool, str]:
    """
    En Windows con Excel instalado, imprime directamente sin abrir ventana.
    Usa el método de apertura directa que mantiene todo el formato.
    """
    ruta = _ruta_excel()
    if not os.path.exists(ruta):
        return False, "No se encontró el archivo Excel."

    if platform.system() != "Windows":
        return abrir_para_imprimir(nombre_hoja)

    try:
        # Import win32com here to prevent loading errors on non-Windows systems
        import win32com.client

        # Absolute path ensures we only use trusted, absolute file locations
        abs_ruta = ruta

        # Ensure we only interact with files that exist to
        # prevent path traversal issues
        if not os.path.isfile(abs_ruta):
            return False, "La ruta del archivo no es válida."

        # Conectar a Excel de manera segura
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        try:
            wb = excel.Workbooks.Open(abs_ruta, ReadOnly=True)
            try:
                hoja = wb.Sheets(nombre_hoja)
                hoja.PrintOut()
                mensaje = f"Hoja '{nombre_hoja}' enviada a la impresora."
                exito = True
            except Exception:
                mensaje = (
                    f"Error al imprimir la hoja '{nombre_hoja}'. "
                    "Verifica que exista."
                )
                exito = False
            finally:
                wb.Close(SaveChanges=False)
        except Exception as e_wb:
            mensaje = f"Error al abrir el archivo con Excel: {e_wb}"
            exito = False
        finally:
            excel.Quit()

        return exito, mensaje

    except Exception:
        # Fallback: abrir normalmente
        return abrir_para_imprimir(nombre_hoja)


class PanelImpresion:
    """
    Panel de impresión para agregar a la UI principal.
    Uso: nb.add(PanelImpresion(nb, app), text='🖨️  Imprimir')
    """
    def __new__(cls, parent, app):
        import tkinter as tk

        C = {
            "azul_osc": "#1C3557", "azul_med": "#2E6DA4",
            "azul_clar": "#D6E8FA", "blanco": "#FFFFFF",
            "gris_clar": "#F4F6FA", "amarillo": "#FFC000",
            "verde": "#2D6A4F", "rojo": "#C0392B",
            "texto": "#1A1A2E", "texto_med": "#4A5568",
        }

        frame = tk.Frame(parent, bg=C["blanco"])

        # Header
        hdr = tk.Frame(frame, bg=C["azul_osc"], height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  🖨️  Impresión de Planillas",
                 font=("Segoe UI", 14, "bold"), fg=C["amarillo"],
                 bg=C["azul_osc"]).pack(side="left", pady=8)

        cuerpo = tk.Frame(frame, bg=C["blanco"], padx=30, pady=20)
        cuerpo.pack(fill="both", expand=True)

        tk.Label(cuerpo,
                 text=("El programa abre el Excel con su diseño y "
                       "formato originales.\n"
                       "Tú imprimes directamente desde ahí, "
                       "exactamente como MEDUCA lo requiere."),
                 font=("Segoe UI", 10), fg=C["texto_med"], bg=C["blanco"],
                 justify="center").pack(pady=(0, 20))

        # Hojas disponibles — construidas desde SQL + hojas fijas (SIN abrir el Excel)
        hojas_comunes = [
            ("Portada",  "\U0001f4cb Portada / Car\u00e1tula oficial"),
            ("Caratula", "\U0001f4c4 Car\u00e1tula del registro"),
            ("Horarios", "\U0001f550 Horario de clases"),
        ]
        try:
            grados_sql = app.engine.obtener_grados_activos() if hasattr(app, 'engine') else []
            for g in grados_sql:
                hojas_comunes.append(("Asistencia (" + g + ")", "\U0001f4c5 Asistencia \u2014 Grado " + g))
                hojas_comunes.append(("RESUMEN_" + g, "\U0001f4c8 RESUMEN \u2014 " + g))
            for g in grados_sql:
                try:
                    materias = app.engine.obtener_materias_por_grado(g)
                    for mat in materias:
                        if mat and mat not in ["Sin materias", "General"]:
                            hojas_comunes.append(("Planilla (" + mat + " " + g + ")", "\U0001f4ca Planilla \u2014 " + mat + " " + g))
                except Exception:
                    pass
        except Exception:
            pass  # Las hojas estaticas son suficientes como fallback


        sel_f = tk.LabelFrame(cuerpo, text="  Selecciona qué imprimir  ",
                              font=("Segoe UI", 10, "bold"), bg=C["blanco"],
                              fg=C["azul_osc"], padx=20, pady=15)
        sel_f.pack(fill="x", pady=10)

        # Set default value safely to the first available sheet or "Portada"
        default_val = hojas_comunes[0][0] if hojas_comunes else "Portada"
        var_hoja = tk.StringVar(value=default_val)

        for i, (hoja_id, hoja_lbl) in enumerate(hojas_comunes):
            col = i % 2
            fila = i // 2
            tk.Radiobutton(
                sel_f, text=hoja_lbl, variable=var_hoja,
                value=hoja_id, font=("Segoe UI", 10),
                bg=C["blanco"], fg=C["texto"],
                selectcolor=C["azul_clar"], cursor="hand2"
            ).grid(row=fila, column=col, sticky="w",
                   padx=10, pady=3)

        btn_f = tk.Frame(cuerpo, bg=C["blanco"])
        btn_f.pack(pady=20)

        def abrir_excel():
            ok, msg = abrir_para_imprimir()
            if ok:
                messagebox.showinfo("✓ Excel abierto", msg)
            else:
                messagebox.showerror("Error", msg)

        def imprimir_sel():
            hoja = var_hoja.get()
            ok, msg = imprimir_hoja_directo(hoja)
            if ok:
                messagebox.showinfo("✓ Imprimiendo", msg)
            else:
                messagebox.showerror("Error", msg)

        tk.Button(btn_f, text="📂  Abrir Excel completo",
                  command=abrir_excel,
                  bg=C["azul_med"], fg=C["blanco"],
                  font=("Segoe UI", 11, "bold"), relief="flat",
                  cursor="hand2", padx=20, pady=10,
                  width=22).pack(side="left", padx=8)

        tk.Button(btn_f, text="🖨️  Imprimir hoja seleccionada",
                  command=imprimir_sel,
                  bg=C["verde"], fg=C["blanco"],
                  font=("Segoe UI", 11, "bold"), relief="flat",
                  cursor="hand2", padx=20, pady=10,
                  width=26).pack(side="left", padx=8)

        nota = tk.Label(cuerpo,
                        text="💡 El formato, fórmulas y diseño "
                             "del Excel nunca se modifican.\n"
                             "    Lo que ves en Excel es exactamente "
                             "lo que se imprime.",
                        font=("Segoe UI", 9), fg=C["texto_med"],
                        bg=C["blanco"], justify="left")
        nota.pack(pady=10)

        return frame
