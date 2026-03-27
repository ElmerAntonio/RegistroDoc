"""
RegistroDoc Pro v3.0 — Sistema de Registro Académico
Diseñado para docentes de Premedia Multigrado — Panamá 2026
Interfaz moderna, simple, pensada para maestros
© 2026 Elmer Tugri — Todos los derechos reservados
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.simpledialog as simpledialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os, json, shutil, re, datetime, copy

# ─────────────────────────────────────────────────────────
#  SEGURIDAD
# ─────────────────────────────────────────────────────────
try:
    from src.rdsecurity import (
        verificar_licencia, activar_licencia_completa,
        verificar_bloqueo, registrar_auditoria,
        guardar_config_segura, cargar_config_segura,
        verificar_integridad_excel, actualizar_hash_excel,
        leer_auditoria, validar_nota_meduca,
    )
    SEGURIDAD_ACTIVA = True
except ImportError:
    SEGURIDAD_ACTIVA = False
    def verificar_licencia(): return True, {"cedula": "DEMO"}
    def activar_licencia_completa(c, k): return True, "Demo"
    def verificar_bloqueo(): return False, ""
    def registrar_auditoria(*a): pass
    def guardar_config_segura(cfg): 
        with open("config.json","w",encoding="utf-8") as f: json.dump(cfg,f,ensure_ascii=False,indent=2)
    def cargar_config_segura(d): 
        if os.path.exists("config.json"):
            try:
                with open("config.json","r",encoding="utf-8") as f: c=json.load(f); d.update(c); return d
            except: pass
        return dict(d)
    def verificar_integridad_excel(r): return True, ""
    def actualizar_hash_excel(r): pass
    def leer_auditoria(): return []
    def validar_nota_meduca(v):
        try:
            n = float(str(v).replace(",",".")); return (1<=n<=5), n, "OK"
        except: return False, 0, "Número inválido"

# ─────────────────────────────────────────────────────────
#  CONFIGURACIÓN POR DEFECTO
# ─────────────────────────────────────────────────────────
CONFIG_DEFAULT = {
    "docente_nombre":"","docente_cedula":"","docente_posicion":"",
    "docente_ss":"","docente_telefono":"","docente_correo":"",
    "docente_condicion":"Instructor Vocacional",
    "escuela_nombre":"","escuela_director":"","escuela_subdirector":"",
    "escuela_coordinador":"","escuela_regional":"","escuela_distrito":"",
    "escuela_corregimiento":"","escuela_zona":"","escuela_telefono":"",
    "anio_lectivo":"2026","jornada":"Matutina",
    "area_curricular":"PREMEDIA MULTIGRADO","asignaturas":"","grados":"7°, 8° y 9°",
    "logo_path":"",
    "materias":{
        "7°":["Inglés","Hogar y Desarrollo","Agropecuaria"],
        "8°":["Inglés","Comercio","Agropecuaria"],
        "9°":["Inglés","Comercio","Agropecuaria","Artística"],
    },
    "horario":[
        {"periodo":"I",     "hora":"8:00 - 8:35",   "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"II",    "hora":"8:35 - 9:10",   "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"III",   "hora":"9:10 - 9:45",   "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"IV",    "hora":"9:45 - 10:20",  "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"RECESO","hora":"10:20 - 10:40", "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"V",     "hora":"10:40 - 11:15", "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"VI",    "hora":"11:15 - 11:50", "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"VII",   "hora":"11:50 - 12:25", "lun":"","mar":"","mie":"","jue":"","vie":""},
        {"periodo":"VIII",  "hora":"12:25 - 1:00",  "lun":"","mar":"","mie":"","jue":"","vie":""},
    ]
}

def cargar_config():
    return cargar_config_segura(CONFIG_DEFAULT)

def guardar_config(cfg):
    guardar_config_segura(cfg)

# ─────────────────────────────────────────────────────────
#  EXCEL — RUTAS Y MAPEO
# ─────────────────────────────────────────────────────────
def ruta_excel():
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "Registro_2026.xlsx")

COL_MAESTRO = {"7°":2, "8°":4, "9°":6}
FILA_EST_BASE = 5
HOJA_ASIST = {"7°":"Asistencia (7°)","8°":"Asistencia (8°)","9°":"Asistencia (9°)"}

# Estructura PROM — 3 trimestres
PROM_STRUCT = {
    1: {"parc":{"cols":range(3,16), "fecha":4,"desc":42},
        "aprec":{"cols":range(18,27),"fecha":4,"desc":42},
        "prueb":{"cols":range(28,30),"fecha":4,"desc":42},
        "p_parc":16,"p_aprec":27,"p_prueb":30,"cal":31,"obs":32},
    2: {"parc":{"cols":range(37,50),"fecha":4,"desc":42},
        "aprec":{"cols":range(51,61),"fecha":4,"desc":42},
        "prueb":{"cols":range(62,64),"fecha":4,"desc":42},
        "p_parc":50,"p_aprec":61,"p_prueb":64,"cal":65,"obs":66},
    3: {"parc":{"cols":range(71,84),"fecha":4,"desc":42},
        "aprec":{"cols":range(85,95),"fecha":4,"desc":42},
        "prueb":{"cols":range(96,98),"fecha":4,"desc":42},
        "p_parc":84,"p_aprec":95,"p_prueb":98,"cal":99,"obs":100},
}

# Planilla: F=T1, G=T2, H=T3, I=Final, J-O=asistencias
PLAN_COLS = {"T1":6,"T2":7,"T3":8,"Final":9,
             "AusT1":10,"TarT1":11,"AusT2":12,"TarT2":13,"AusT3":14,"TarT3":15}

MAPA_PROM = {
    ("7°","Inglés"):             "PROM (Ingles 7°)",
    ("7°","Hogar y Desarrollo"): "PROM (Hogar y Desarrollo 7°)",
    ("7°","Agropecuaria"):       "PROM (Agropecuaria 7°)",
    ("8°","Inglés"):             "PROM (Ingles 8°)",
    ("8°","Comercio"):           "PROM (Comercio 8°)",
    ("8°","Agropecuaria"):       "PROM (Agropecuaria 8°)",
    ("9°","Inglés"):             "PROM (Ingles 9°)",
    ("9°","Comercio"):           "PROM (Comercio 9°)",
    ("9°","Agropecuaria"):       "PROM (Agropecuaria 9°)",
    ("9°","Artística"):          "PROM (Artistica 9°)",
}
MAPA_PLAN = {
    ("7°","Inglés"):             "Planilla (Ingles 7°) ",
    ("7°","Hogar y Desarrollo"): "Planilla (Hogar y Desarrollo 7)",
    ("7°","Agropecuaria"):       "Planilla (Agropecuaria 7°)",
    ("8°","Inglés"):             "Planilla (Ingles 8°)",
    ("8°","Comercio"):           "Planilla (Comercio 8°)",
    ("8°","Agropecuaria"):       "Planilla (Agropecuaria 8°) ",
    ("9°","Inglés"):             "Planilla (Ingles 9°) ",
    ("9°","Comercio"):           "Planilla (Comercio 9°)",
    ("9°","Agropecuaria"):       "Planilla (Agropecuaria 9°)",
    ("9°","Artística"):          "Planilla (Artistica 9°)",
}

def hoja_prom(g, m):
    return MAPA_PROM.get((g,m), f"PROM ({m} {g})")

def hoja_plan(g, m):
    return MAPA_PLAN.get((g,m), f"Planilla ({m} {g})")

def guardar_wb(wb):
    r = ruta_excel()
    shutil.copy2(r, r.replace(".xlsx","_bak.xlsx"))
    wb.save(r)
    actualizar_hash_excel(r)

def wb_rw():
    return load_workbook(ruta_excel())

def wb_ro():
    return load_workbook(ruta_excel(), read_only=True, data_only=True)

def obtener_estudiantes(grado):
    wb = wb_ro()
    ws = wb["MAESTRO"]
    col = COL_MAESTRO.get(grado, 2)
    res = []
    for i in range(40):
        v = ws.cell(FILA_EST_BASE+i, col).value
        if v and str(v).strip() and not str(v).strip().startswith("="):
            res.append((i+1, str(v).strip()))
    wb.close()
    return res

def guardar_estudiante(grado, pos, nombre):
    wb = wb_rw()
    wb["MAESTRO"].cell(FILA_EST_BASE+pos-1, COL_MAESTRO.get(grado,2)).value = nombre or None
    guardar_wb(wb)
    registrar_auditoria("ESTUDIANTE","GUARDAR",f"Grado {grado} pos {pos}")

def leer_notas(grado, materia, tri, num_est):
    hn = hoja_prom(grado, materia)
    wb = wb_ro()
    if hn not in wb.sheetnames:
        wb.close(); return {}
    ws = wb[hn]
    fila = FILA_EST_BASE + num_est - 1
    s = PROM_STRUCT[tri]
    res = {}
    for comp, key in [("parc","parciales"),("aprec","apreciacion"),("prueb","pruebas")]:
        cols = list(s[comp]["cols"])
        res[key] = {
            "notas":  [ws.cell(fila,c).value for c in cols],
            "fechas": [ws.cell(s[comp]["fecha"],c).value for c in cols],
            "descs":  [ws.cell(s[comp]["desc"],c).value for c in cols],
        }
    for k,col in [("prom_parc",s["p_parc"]),("prom_aprec",s["p_aprec"]),
                  ("prom_prueb",s["p_prueb"]),("cal",s["cal"]),("obs",s["obs"])]:
        res[k] = ws.cell(fila, col).value
    wb.close()
    return res

def guardar_nota_excel(grado, materia, tri, num_est, comp_key, idx, valor, fecha="", desc=""):
    ok, nota_f, msg = validar_nota_meduca(str(valor))
    if not ok:
        raise ValueError(msg)
    hn = hoja_prom(grado, materia)
    wb = wb_rw()
    if hn not in wb.sheetnames:
        wb.close(); return False
    ws = wb[hn]
    fila = FILA_EST_BASE + num_est - 1
    s = PROM_STRUCT[tri]
    comp_map = {"parciales":"parc","apreciacion":"aprec","pruebas":"prueb"}
    comp = comp_map.get(comp_key, comp_key)
    cols = list(s[comp]["cols"])
    if idx >= len(cols):
        wb.close(); return False
    col = cols[idx]
    ws.cell(fila, col).value = nota_f
    if fecha:
        ws.cell(s[comp]["fecha"], col).value = fecha
    if desc:
        ws.cell(s[comp]["desc"], col).value = desc
    guardar_wb(wb)
    registrar_auditoria("NOTA","GUARDAR",f"{grado} {materia} T{tri} est{num_est}")
    return True

def guardar_obs_excel(grado, materia, tri, num_est, texto):
    hn = hoja_prom(grado, materia)
    wb = wb_rw()
    if hn not in wb.sheetnames:
        wb.close(); return
    wb[hn].cell(FILA_EST_BASE+num_est-1, PROM_STRUCT[tri]["obs"]).value = texto
    guardar_wb(wb)

def marcar_asist(grado, num_est, col_dia, estado):
    wb = wb_rw()
    ws = wb[HOJA_ASIST[grado]]
    ws.cell(3+num_est-1, col_dia).value = None if estado=="P" else estado
    guardar_wb(wb)
    registrar_auditoria("ASISTENCIA","MARCAR",f"{grado} est{num_est}")

def leer_resumen(cfg, grado):
    mats = cfg.get("materias",{}).get(grado,[])
    ests = obtener_estudiantes(grado)
    if not ests: return []
    wb = wb_ro()
    res = []
    for num_est, nombre in ests:
        fila_plan = 16 + num_est - 1
        entry = {"nombre":nombre,"materias":{}}
        for mat in mats:
            hp = hoja_plan(grado, mat)
            if hp not in wb.sheetnames: continue
            ws = wb[hp]
            t1 = ws.cell(fila_plan, 6).value
            t2 = ws.cell(fila_plan, 7).value
            t3 = ws.cell(fila_plan, 8).value
            fn = ws.cell(fila_plan, 9).value
            entry["materias"][mat] = {
                "T1": _to_float(t1), "T2": _to_float(t2),
                "T3": _to_float(t3), "Final": _to_float(fn)
            }
        res.append(entry)
    wb.close()
    return res

def _to_float(v):
    try: return round(float(str(v).replace(",",".")),1)
    except: return None

def actualizar_portada(cfg):
    if not os.path.exists(ruta_excel()):
        return False, "No se encontró el Excel"
    wb = wb_rw()
    escrituras = [
        ("Portada",8,16,cfg["anio_lectivo"]),
        ("Portada",14,16,cfg["docente_nombre"]),
        ("Portada",16,8,cfg["docente_cedula"]),
        ("Portada",16,35,cfg["docente_posicion"]),
        ("Portada",18,21,cfg["docente_condicion"]),
        ("Portada",20,8,cfg["jornada"]),
        ("Portada",20,30,cfg["asignaturas"]),
        ("Portada",22,19,cfg["escuela_director"]),
        ("Portada",24,20,cfg["escuela_subdirector"]),
        ("Portada",26,20,cfg["escuela_coordinador"]),
        ("Portada",28,17,cfg["escuela_nombre"]),
        ("Portada",30,6,cfg["escuela_telefono"]),
        ("Portada",30,32,cfg["docente_correo"]),
        ("Portada",32,13,cfg["escuela_regional"]),
        ("Portada",34,12,cfg["escuela_distrito"]),
        ("Portada",36,17,cfg["escuela_corregimiento"]),
        ("Portada",38,15,cfg["escuela_zona"]),
        ("Caratula",23,17,f"Prof. {cfg['docente_nombre']}"),
        ("Caratula",24,17,cfg["docente_condicion"]),
        ("PORTADA_VISTOSA",9,3,cfg["escuela_nombre"]),
        ("PORTADA_VISTOSA",10,3,cfg["docente_nombre"]),
        ("PORTADA_VISTOSA",11,3,cfg["anio_lectivo"]),
        ("Horarios",2,15,cfg["docente_nombre"]),
        ("Horarios",4,8,cfg["jornada"]),
        ("Horarios",6,9,cfg["asignaturas"]),
    ]
    for hoja,fila,col,valor in escrituras:
        if hoja in wb.sheetnames:
            wb[hoja].cell(row=fila,column=col).value = valor
    # Horario
    if "Horarios" in wb.sheetnames:
        wsh = wb["Horarios"]
        dc = {"lun":15,"mar":16,"mie":17,"jue":18,"vie":19}
        for i,p in enumerate(cfg.get("horario",[])):
            fi = 11+i
            wsh.cell(fi,4).value = p.get("periodo","")
            wsh.cell(fi,10).value = p.get("hora","")
            for d,c in dc.items():
                wsh.cell(fi,c).value = p.get(d,"")
    # Actualizar encabezados de PROM
    for grado, mats in cfg.get("materias",{}).items():
        for mat in mats:
            hn = hoja_prom(grado, mat)
            if hn in wb.sheetnames:
                ws = wb[hn]
                # Fila 1 encabezado (si existe)
                if ws.cell(1,1).value:
                    ws.cell(1,1).value = f"REGISTRO DE CALIFICACIONES — {mat.upper()} {grado} — AÑO LECTIVO {cfg['anio_lectivo']}"
            hp = hoja_plan(grado, mat)
            if hp in wb.sheetnames:
                ws = wb[hp]
                # Actualizar campos editables en planilla
                if ws.cell(5,2).value is not None or ws.cell(5,4).value is not None:
                    ws.cell(5,4).value  = cfg["anio_lectivo"]
                if ws.cell(11,12).value is not None:
                    ws.cell(11,12).value = grado
                if ws.cell(11,4).value is not None:
                    ws.cell(11,4).value = mat.upper()
                if ws.cell(11,12).value is not None:
                    ws.cell(11,12).value = grado
                if ws.cell(5,2).value is not None:
                    ws.cell(5,2).value = cfg["escuela_nombre"]
                if ws.cell(6,2).value is not None:
                    ws.cell(6,2).value = f"{cfg['escuela_corregimiento']}, {cfg['escuela_regional']}"
                if ws.cell(10,4).value is not None:
                    ws.cell(10,4).value = cfg["escuela_director"]
                if ws.cell(11,12).value is not None:
                    ws.cell(11,12).value = grado
    guardar_wb(wb)
    registrar_auditoria("CONFIG","PORTADA_ACTUALIZADA","OK")
    return True, "✓ Portada y encabezados actualizados"

def agregar_materia_excel(grado, materia, cfg):
    """Copia el formato de PROM y Planilla para una nueva materia."""
    if not os.path.exists(ruta_excel()):
        return False, "No se encontró el Excel"
    wb = wb_rw()

    nuevo_prom = f"PROM ({materia} {grado})"
    nuevo_plan = f"Planilla ({materia} {grado})"
    mcol = COL_MAESTRO.get(grado, 2)
    mcol_letra = get_column_letter(mcol)
    asist_hoja = HOJA_ASIST.get(grado, f"Asistencia ({grado})")

    # Copiar hoja PROM modelo
    modelo_prom_nom = hoja_prom(grado, cfg["materias"].get(grado,[None])[0])
    if not modelo_prom_nom or modelo_prom_nom not in wb.sheetnames:
        # Usar cualquier PROM del mismo grado
        for h in wb.sheetnames:
            if h.startswith(f"PROM (") and grado in h:
                modelo_prom_nom = h; break

    if modelo_prom_nom and modelo_prom_nom in wb.sheetnames and nuevo_prom not in wb.sheetnames:
        ws_src = wb[modelo_prom_nom]
        ws_new = wb.copy_worksheet(ws_src)
        ws_new.title = nuevo_prom
        # Actualizar referencias MAESTRO
        for row in ws_new.iter_rows():
            for cell in row:
                v = str(cell.value) if cell.value else ""
                if "MAESTRO!" in v:
                    # ya apuntan al col correcto si copiamos del mismo grado
                    pass
        wb.save(ruta_excel())

    # Copiar hoja Planilla modelo
    modelo_plan_nom = hoja_plan(grado, cfg["materias"].get(grado,[None])[0])
    if not modelo_plan_nom or modelo_plan_nom not in wb.sheetnames:
        for h in wb.sheetnames:
            if h.startswith("Planilla (") and grado in h:
                modelo_plan_nom = h; break

    if modelo_plan_nom and modelo_plan_nom in wb.sheetnames and nuevo_plan not in wb.sheetnames:
        wb2 = load_workbook(ruta_excel())
        ws_src2 = wb2[modelo_plan_nom]
        ws_new2 = wb2.copy_worksheet(ws_src2)
        ws_new2.title = nuevo_plan
        # Apuntar al nuevo PROM
        for row in ws_new2.iter_rows():
            for cell in row:
                v = str(cell.value) if cell.value else ""
                if v.startswith("=") and "PROM (" in v:
                    cell.value = v.replace(modelo_prom_nom, nuevo_prom)
        # Actualizar asistencia
        for row in ws_new2.iter_rows():
            for cell in row:
                v = str(cell.value) if cell.value else ""
                if v.startswith("=") and "Asistencia" in v:
                    cell.value = re.sub(r"'Asistencia \([^)]+\)'", f"'{asist_hoja}'", v)
        wb2.save(ruta_excel())

    # Actualizar MAPA dinámico
    MAPA_PROM[(grado, materia)] = nuevo_prom
    MAPA_PLAN[(grado, materia)] = nuevo_plan
    registrar_auditoria("CONFIG","NUEVA_MATERIA",f"{grado} {materia}")
    return True, f"✓ Materia '{materia}' agregada para {grado}"

# ─────────────────────────────────────────────────────────
#  TEMA VISUAL
# ─────────────────────────────────────────────────────────
C = {
    "bg":        "#0F1923",
    "panel":     "#1A2638",
    "card":      "#1E2D42",
    "borde":     "#253650",
    "azul":      "#2563EB",
    "azul_hover":"#1D4ED8",
    "azul_clar": "#3B82F6",
    "verde":     "#059669",
    "verde_clar":"#10B981",
    "rojo":      "#DC2626",
    "rojo_clar": "#EF4444",
    "amarillo":  "#F59E0B",
    "naranja":   "#EA580C",
    "texto":     "#F1F5F9",
    "texto_sec": "#94A3B8",
    "texto_ter": "#64748B",
    "input":     "#253650",
    "input_bord":"#3B82F6",
    "ok":        "#D1FAE5",
    "error":     "#FEE2E2",
    "warn":      "#FEF3C7",
}

def boton(parent, txt, cmd=None, color=None, w=16, h=1, grande=False, **kw):
    c = color or C["azul"]
    fs = 11 if grande else 10
    return tk.Button(parent, text=txt, command=cmd, bg=c, fg=C["texto"],
                     font=("Segoe UI",fs,"bold"), relief="flat", cursor="hand2",
                     padx=14, pady=8 if grande else 6,
                     activebackground=C["azul_hover"], activeforeground=C["texto"],
                     width=w, height=h, bd=0, **kw)

def entrada(parent, var=None, w=28, ph="", **kw):
    e = tk.Entry(parent, textvariable=var, font=("Segoe UI",10),
                 bg=C["input"], fg=C["texto"], insertbackground=C["texto"],
                 relief="flat", bd=0, width=w, **kw)
    if ph and not var:
        e.insert(0, ph)
        e.config(fg=C["texto_ter"])
        def on_focus_in(ev):
            if e.get() == ph:
                e.delete(0,"end"); e.config(fg=C["texto"])
        def on_focus_out(ev):
            if not e.get():
                e.insert(0,ph); e.config(fg=C["texto_ter"])
        e.bind("<FocusIn>",on_focus_in)
        e.bind("<FocusOut>",on_focus_out)
    return e

def tarjeta(parent, **kw):
    return tk.Frame(parent, bg=C["card"], bd=0, **kw)

def etiqueta(parent, txt, grande=False, color=None, bg=None, bold=False, **kw):
    c = color or C["texto"]; b = bg or C["card"]
    fs = 14 if grande else 10
    fw = "bold" if bold else "normal"
    return tk.Label(parent, text=txt, font=("Segoe UI",fs,fw),
                    fg=c, bg=b, **kw)

def header_panel(parent, ico, titulo, subtitulo="", bg=None):
    bg = bg or C["panel"]
    f = tk.Frame(parent, bg=bg, pady=0)
    f.pack(fill="x")
    inner = tk.Frame(f, bg=bg, padx=20, pady=14)
    inner.pack(fill="x")
    tk.Label(inner, text=f"{ico}  {titulo}", font=("Segoe UI",16,"bold"),
             fg=C["texto"], bg=bg).pack(side="left")
    if subtitulo:
        tk.Label(inner, text=subtitulo, font=("Segoe UI",9),
                 fg=C["texto_sec"], bg=bg).pack(side="left", padx=(10,0), pady=(4,0))
    # Línea divisora
    tk.Frame(parent, bg=C["azul"], height=2).pack(fill="x")

def scroll_frame(parent):
    canvas = tk.Canvas(parent, bg=C["bg"], highlightthickness=0)
    sc = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=sc.set)
    sc.pack(side="right", fill="y")
    canvas.pack(fill="both", expand=True)
    frame = tk.Frame(canvas, bg=C["bg"])
    cw = canvas.create_window((0,0), window=frame, anchor="nw")
    frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)),"units"))
    return frame

def combo(parent, var, vals, w=15):
    style = ttk.Style()
    style.configure("Dark.TCombobox",
        fieldbackground=C["input"], background=C["input"],
        foreground=C["texto"], arrowcolor=C["texto_sec"])
    cb = ttk.Combobox(parent, textvariable=var, values=vals,
                      state="readonly", font=("Segoe UI",10), width=w,
                      style="Dark.TCombobox")
    return cb

# ─────────────────────────────────────────────────────────
#  VENTANA DE ACTIVACIÓN
# ─────────────────────────────────────────────────────────
class VentanaActivacion(tk.Toplevel):
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("Activar RegistroDoc Pro")
        self.geometry("480x500")
        self.configure(bg=C["bg"])
        self.resizable(False,False)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", lambda: parent.destroy())
        self._build()

    def _build(self):
        # Header con logo
        hdr = tk.Frame(self, bg=C["panel"], pady=30)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📚", font=("Segoe UI",40), fg=C["amarillo"], bg=C["panel"]).pack()
        tk.Label(hdr, text="RegistroDoc Pro", font=("Segoe UI",20,"bold"),
                 fg=C["texto"], bg=C["panel"]).pack()
        tk.Label(hdr, text="Sistema de Registro Académico — Panamá 2026",
                 font=("Segoe UI",9), fg=C["texto_sec"], bg=C["panel"]).pack(pady=(4,0))
        tk.Frame(self, bg=C["azul"], height=2).pack(fill="x")

        # Formulario
        card = tk.Frame(self, bg=C["bg"], padx=35, pady=25)
        card.pack(fill="both", expand=True)

        tk.Label(card, text="🔐  Activa tu licencia para comenzar",
                 font=("Segoe UI",12,"bold"), fg=C["texto"], bg=C["bg"]).pack(pady=(0,20))

        # Cédula
        f1 = tk.Frame(card, bg=C["bg"]); f1.pack(fill="x", pady=5)
        tk.Label(f1, text="N° Cédula del docente", font=("Segoe UI",9),
                 fg=C["texto_sec"], bg=C["bg"]).pack(anchor="w")
        self.var_ced = tk.StringVar()
        ent_ced = entrada(f1, self.var_ced, w=44)
        ent_ced.pack(fill="x", ipady=8, pady=(2,0))

        # Código
        f2 = tk.Frame(card, bg=C["bg"]); f2.pack(fill="x", pady=5)
        tk.Label(f2, text="Código de activación  (RD-XXXXX-XXXXX-XXXXX-XXXXX)",
                 font=("Segoe UI",9), fg=C["texto_sec"], bg=C["bg"]).pack(anchor="w")
        self.var_cod = tk.StringVar()
        ent_cod = entrada(f2, self.var_cod, w=44)
        ent_cod.pack(fill="x", ipady=8, pady=(2,0))

        self.lbl_err = tk.Label(card, text="", font=("Segoe UI",9),
                                 fg=C["rojo_clar"], bg=C["bg"], wraplength=400)
        self.lbl_err.pack(pady=8)

        self.btn_act = boton(card, "✔  Activar programa", self._activar,
                              color=C["verde"], w=36, grande=True)
        self.btn_act.pack(fill="x", pady=5)

        tk.Label(card, text="Para obtener tu código, contacta al proveedor del programa.",
                 font=("Segoe UI",8), fg=C["texto_ter"], bg=C["bg"]).pack(pady=5)

    def _activar(self):
        bloq, msg_b = verificar_bloqueo()
        if bloq:
            self.lbl_err.config(text=msg_b); return
        self.btn_act.config(state="disabled", text="Verificando...")
        self.update()
        ok, msg = activar_licencia_completa(self.var_ced.get(), self.var_cod.get())
        if ok:
            self.destroy(); self.callback()
        else:
            self.lbl_err.config(text=msg)
            self.btn_act.config(state="normal", text="✔  Activar programa")

# ─────────────────────────────────────────────────────────
#  PANEL INICIO / DASHBOARD
# ─────────────────────────────────────────────────────────
class PanelInicio(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app = app
        self._build()

    def _build(self):
        header_panel(self, "🏠", "Panel Principal",
                     "Vista general del registro académico", bg=C["panel"])
        body = scroll_frame(self)

        cfg = cargar_config()
        nombre = cfg.get("docente_nombre","") or "Docente"
        escuela = cfg.get("escuela_nombre","") or "—"

        # Bienvenida
        bienvenida = tarjeta(body, padx=20, pady=15)
        bienvenida.pack(fill="x", padx=15, pady=(15,5))
        tk.Label(bienvenida, text=f"Bienvenido(a), {nombre}",
                 font=("Segoe UI",15,"bold"), fg=C["texto"], bg=C["card"]).pack(anchor="w")
        tk.Label(bienvenida, text=f"🏫  {escuela}  |  📅  Año lectivo {cfg.get('anio_lectivo','2026')}",
                 font=("Segoe UI",10), fg=C["texto_sec"], bg=C["card"]).pack(anchor="w", pady=(4,0))

        # Tarjetas de resumen por grado
        tk.Label(body, text="  Matrícula actual", font=("Segoe UI",11,"bold"),
                 fg=C["texto_sec"], bg=C["bg"]).pack(anchor="w", padx=15, pady=(15,5))

        grid_f = tk.Frame(body, bg=C["bg"]); grid_f.pack(fill="x", padx=15)
        grados = list(cfg.get("materias",{}).keys()) or ["7°","8°","9°"]
        colores = [C["azul"], C["verde"], C["amarillo"], C["naranja"]]
        for idx, grado in enumerate(grados):
            ests = obtener_estudiantes(grado)
            col_grado = colores[idx % len(colores)]
            card = tk.Frame(grid_f, bg=col_grado, padx=20, pady=15, relief="flat")
            card.grid(row=0, column=idx, padx=5, sticky="ew")
            grid_f.columnconfigure(idx, weight=1)
            tk.Label(card, text=f"Grado {grado}", font=("Segoe UI",10,"bold"),
                     fg=C["texto"], bg=col_grado).pack()
            tk.Label(card, text=str(len(ests)), font=("Segoe UI",28,"bold"),
                     fg=C["texto"], bg=col_grado).pack()
            tk.Label(card, text="estudiantes", font=("Segoe UI",9),
                     fg=C["texto"], bg=col_grado).pack()

        # Alertas
        self._mostrar_alertas(body, cfg)

        # Accesos rápidos
        tk.Label(body, text="  Acceso rápido", font=("Segoe UI",11,"bold"),
                 fg=C["texto_sec"], bg=C["bg"]).pack(anchor="w", padx=15, pady=(20,5))
        atajos = tk.Frame(body, bg=C["bg"]); atajos.pack(fill="x", padx=15, pady=(0,20))
        acciones = [
            ("📝  Ingresar Notas",     C["azul"],    lambda: self.app.ir_tab("notas")),
            ("📅  Marcar Asistencia",  C["verde"],   lambda: self.app.ir_tab("asistencia")),
            ("👤  Gestionar Lista",    C["naranja"],  lambda: self.app.ir_tab("estudiantes")),
            ("🖨️  Imprimir Planilla",  C["amarillo"], lambda: self.app.ir_tab("imprimir")),
        ]
        for i,(txt,col,cmd) in enumerate(acciones):
            b = boton(atajos, txt, cmd, color=col, w=20, grande=True)
            b.grid(row=0, column=i, padx=5)
            atajos.columnconfigure(i, weight=1)

    def _mostrar_alertas(self, parent, cfg):
        alertas = []
        grados = list(cfg.get("materias",{}).keys()) or ["7°","8°","9°"]
        try:
            wb = wb_ro()
            for grado in grados:
                ws_a = wb[HOJA_ASIST[grado]]
                for i in range(40):
                    aus = ws_a.cell(3+i, 61).value
                    nombre = ws_a.cell(3+i, 2).value
                    if aus and nombre and str(nombre).strip():
                        try:
                            n_aus = int(float(str(aus)))
                            if n_aus >= 5:
                                alertas.append((grado, str(nombre).strip(), n_aus))
                        except: pass
            wb.close()
        except: pass

        if not alertas and not cfg.get("docente_nombre"):
            alertas_panel = tarjeta(parent, padx=15, pady=12)
            alertas_panel.pack(fill="x", padx=15, pady=5)
            tk.Label(alertas_panel,
                     text="⚠️  Configura tu perfil antes de comenzar → ve a Configuración",
                     font=("Segoe UI",10), fg=C["amarillo"], bg=C["card"]).pack(anchor="w")
            return

        if alertas:
            tk.Label(parent, text="  ⚠️  Alertas de asistencia",
                     font=("Segoe UI",11,"bold"), fg=C["amarillo"], bg=C["bg"]).pack(
                         anchor="w", padx=15, pady=(15,5))
            alerta_card = tarjeta(parent, padx=15, pady=10)
            alerta_card.pack(fill="x", padx=15)
            for grado, nombre, aus in alertas[:8]:
                f = tk.Frame(alerta_card, bg=C["card"])
                f.pack(fill="x", pady=2)
                tk.Label(f, text=f"  🔴  {nombre}",
                         font=("Segoe UI",10), fg=C["texto"], bg=C["card"]).pack(side="left")
                tk.Label(f, text=f"Grado {grado}  |  {aus} ausencias",
                         font=("Segoe UI",9), fg=C["rojo_clar"], bg=C["card"]).pack(side="right")
            if len(alertas) > 8:
                tk.Label(alerta_card, text=f"  ... y {len(alertas)-8} más con alertas",
                         font=("Segoe UI",9), fg=C["texto_ter"], bg=C["card"]).pack(anchor="w")

# ─────────────────────────────────────────────────────────
#  PANEL ESTUDIANTES
# ─────────────────────────────────────────────────────────
class PanelEstudiantes(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app = app; self._ests = []
        self._build()

    def _build(self):
        header_panel(self,"👤","Gestión de Estudiantes","Agrega, edita o elimina estudiantes")

        sel = tk.Frame(self, bg=C["panel"], padx=15, pady=10)
        sel.pack(fill="x")
        tk.Label(sel, text="Grado:", font=("Segoe UI",10,"bold"),
                 fg=C["texto_sec"], bg=C["panel"]).pack(side="left")
        cfg = cargar_config()
        grados = list(cfg.get("materias",{}).keys()) or ["7°","8°","9°"]
        self.var_grado = tk.StringVar(value=grados[0] if grados else "7°")
        for g in grados:
            tk.Radiobutton(sel, text=f"  {g}  ", variable=self.var_grado, value=g,
                           font=("Segoe UI",11,"bold"), bg=C["panel"],
                           fg=C["texto"], selectcolor=C["azul"],
                           activebackground=C["panel"],
                           command=self._cargar).pack(side="left", padx=4)
        boton(sel,"🔄 Actualizar",self._cargar,w=12).pack(side="right")

        cuerpo = tk.Frame(self, bg=C["bg"]); cuerpo.pack(fill="both", expand=True, padx=12, pady=10)

        # Lista
        iz = tk.Frame(cuerpo, bg=C["card"], padx=10, pady=10)
        iz.pack(side="left", fill="both", expand=True, padx=(0,8))
        tk.Label(iz, text="Lista de estudiantes", font=("Segoe UI",10,"bold"),
                 fg=C["texto_sec"], bg=C["card"]).pack(anchor="w", pady=(0,8))
        self.lista = tk.Listbox(iz, font=("Consolas",11), bg=C["input"],
                                 fg=C["texto"], selectbackground=C["azul"],
                                 selectforeground=C["texto"], relief="flat",
                                 activestyle="none", highlightthickness=0)
        sc = ttk.Scrollbar(iz, orient="vertical", command=self.lista.yview)
        self.lista.config(yscrollcommand=sc.set)
        sc.pack(side="right", fill="y")
        self.lista.pack(fill="both", expand=True)
        self.lista.bind("<<ListboxSelect>>", self._selec)
        self.lbl_tot = tk.Label(iz, text="Total: 0", font=("Segoe UI",9),
                                 fg=C["texto_ter"], bg=C["card"])
        self.lbl_tot.pack(pady=(6,0))

        # Formulario
        de = tk.Frame(cuerpo, bg=C["card"], padx=15, pady=15, width=280)
        de.pack(side="right", fill="y"); de.pack_propagate(False)
        tk.Label(de, text="Agregar / Editar estudiante", font=("Segoe UI",10,"bold"),
                 fg=C["texto"], bg=C["card"]).pack(anchor="w", pady=(0,12))

        for etq, attr in [("Apellido (s):","e_ap"),("Nombre (s):","e_nm"),("Posición en lista (1-40):","e_pos")]:
            tk.Label(de, text=etq, font=("Segoe UI",9), fg=C["texto_sec"], bg=C["card"]).pack(anchor="w")
            e = entrada(de, w=28)
            e.pack(fill="x", ipady=7, pady=(2,10))
            setattr(self, attr, e)

        boton(de,"💾  Guardar estudiante",self._guardar,color=C["verde"],w=26).pack(fill="x",pady=3)
        boton(de,"🗑  Eliminar seleccionado",self._eliminar,color=C["rojo"],w=26).pack(fill="x",pady=3)
        boton(de,"✖  Limpiar campos",self._limpiar,color=C["texto_ter"],w=26).pack(fill="x",pady=3)

        self._cargar()

    def _cargar(self):
        self.lista.delete(0,"end")
        self._ests = obtener_estudiantes(self.var_grado.get())
        for pos,nombre in self._ests:
            self.lista.insert("end",f"  {pos:2d}.  {nombre}")
        self.lbl_tot.config(text=f"Total: {len(self._ests)} estudiantes registrados")

    def _selec(self,_):
        s = self.lista.curselection()
        if not s or s[0] >= len(self._ests): return
        pos,nombre = self._ests[s[0]]
        partes = nombre.split(",",1)
        self._limpiar()
        self.e_ap.insert(0,partes[0].strip())
        self.e_nm.insert(0,partes[1].strip() if len(partes)>1 else "")
        self.e_pos.insert(0,str(pos))

    def _guardar(self):
        ap=self.e_ap.get().strip(); nm=self.e_nm.get().strip(); pos=self.e_pos.get().strip()
        if not ap or not nm:
            messagebox.showwarning("Datos incompletos","Escribe el apellido y el nombre del estudiante."); return
        if not pos.isdigit() or not (1<=int(pos)<=40):
            messagebox.showwarning("Posición inválida","La posición debe ser un número entre 1 y 40."); return
        guardar_estudiante(self.var_grado.get(), int(pos), f"{ap.upper()}, {nm.title()}")
        messagebox.showinfo("✓ Guardado","Estudiante guardado correctamente.")
        self._cargar(); self._limpiar()

    def _eliminar(self):
        s = self.lista.curselection()
        if not s: messagebox.showwarning("Selecciona","Haz clic en un estudiante de la lista."); return
        pos,nombre = self._ests[s[0]]
        if messagebox.askyesno("Confirmar eliminación",f"¿Eliminar a:\n{nombre}?\n\nEsta acción no se puede deshacer."):
            guardar_estudiante(self.var_grado.get(), pos, "")
            self._cargar()

    def _limpiar(self):
        for e in [self.e_ap,self.e_nm,self.e_pos]: e.delete(0,"end")

# ─────────────────────────────────────────────────────────
#  PANEL NOTAS
# ─────────────────────────────────────────────────────────
class PanelNotas(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app = app
        self.entradas={}; self.fechas={}; self.descs={}
        self._build()

    def _build(self):
        header_panel(self,"📝","Ingreso de Calificaciones","Escala MEDUCA: 1.0 a 5.0")

        sel = tk.Frame(self, bg=C["panel"], padx=12, pady=8)
        sel.pack(fill="x")

        cfg = cargar_config()
        grados = list(cfg.get("materias",{}).keys()) or ["7°","8°","9°"]

        def mk_cmb(var,vals,w=8):
            return combo(sel, var, vals, w)

        tk.Label(sel,text="Grado:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        self.var_g = tk.StringVar(value=grados[0])
        mk_cmb(self.var_g,grados,6).pack(side="left",padx=(3,12))
        self.var_g.trace("w",lambda *_:self._upd_mat())

        tk.Label(sel,text="Materia:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        self.var_mat=tk.StringVar()
        self.cmb_mat=mk_cmb(self.var_mat,[],18)
        self.cmb_mat.pack(side="left",padx=(3,12))
        self.var_mat.trace("w",lambda *_:self._upd_est())

        tk.Label(sel,text="Trimestre:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        self.var_t=tk.StringVar(value="1")
        mk_cmb(self.var_t,["1 — Primer Trimestre","2 — Segundo Trimestre","3 — Tercer Trimestre"],22).pack(side="left",padx=(3,12))
        self.var_t.trace("w",lambda *_:self._upd_notas())

        tk.Label(sel,text="Estudiante:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        self.var_est=tk.StringVar()
        self.cmb_est=mk_cmb(self.var_est,[],28)
        self.cmb_est.pack(side="left",padx=(3,12))
        self.var_est.trace("w",lambda *_:self._upd_notas())

        boton(sel,"💾 Guardar todo",self._guardar_todo,color=C["verde"],w=14).pack(side="right",padx=4)

        # Área de notas con scroll
        cnv = tk.Canvas(self, bg=C["bg"], highlightthickness=0)
        scr = ttk.Scrollbar(self, orient="vertical", command=cnv.yview)
        cnv.configure(yscrollcommand=scr.set)
        scr.pack(side="right",fill="y"); cnv.pack(fill="both",expand=True,padx=6,pady=4)
        self.fn = tk.Frame(cnv, bg=C["bg"])
        cw = cnv.create_window((0,0),window=self.fn,anchor="nw")
        self.fn.bind("<Configure>",lambda e:cnv.configure(scrollregion=cnv.bbox("all")))
        cnv.bind("<Configure>",lambda e:cnv.itemconfig(cw,width=e.width))
        cnv.bind_all("<MouseWheel>",lambda e:cnv.yview_scroll(int(-1*(e.delta/120)),"units"))

        self._upd_mat()

    def _upd_mat(self):
        g = self.var_g.get()
        cfg = cargar_config()
        mats = cfg.get("materias",{}).get(g,[])
        self.cmb_mat["values"] = mats
        if mats: self.var_mat.set(mats[0])
        self._upd_est()

    def _upd_est(self):
        ests = obtener_estudiantes(self.var_g.get())
        self._ests = ests
        vals = [f"{p}.  {n}" for p,n in ests]
        self.cmb_est["values"] = vals
        if vals: self.var_est.set(vals[0])
        self._upd_notas()

    def _upd_notas(self,*_):
        for w in self.fn.winfo_children(): w.destroy()
        self.entradas={}; self.fechas={}; self.descs={}
        g=self.var_g.get(); mat=self.var_mat.get()
        tri_s=self.var_t.get(); est_s=self.var_est.get()
        if not all([g,mat,tri_s,est_s]): return
        try:
            tri = int(tri_s.split("—")[0].strip()) if "—" in tri_s else int(tri_s)
            num_est = int(est_s.split(".")[0])
        except: return

        datos = leer_notas(g, mat, tri, num_est)
        if not datos:
            tk.Label(self.fn,text="No hay datos para esta selección.",
                     font=("Segoe UI",11),bg=C["bg"],fg=C["texto_sec"]).pack(pady=30); return

        COMP = {
            "parciales":  ("📋  NOTAS DIARIAS / PARCIALES","Hasta 13 notas  ·  Escala 1.0 – 5.0"),
            "apreciacion":("⭐  APRECIACIÓN",               "Hasta 9 notas  ·  Escala 1.0 – 5.0"),
            "pruebas":    ("📝  PRUEBAS / EXÁMENES",        "Hasta 2 notas  ·  Escala 1.0 – 5.0"),
        }

        for comp,(titulo,hint) in COMP.items():
            # Cabecera de sección
            sf = tk.Frame(self.fn, bg=C["panel"]); sf.pack(fill="x", pady=(12,0), padx=6)
            tk.Label(sf, text=f"  {titulo}", font=("Segoe UI",11,"bold"),
                     fg=C["texto"], bg=C["panel"], pady=8).pack(side="left")
            tk.Label(sf, text=hint, font=("Segoe UI",9),
                     fg=C["texto_ter"], bg=C["panel"]).pack(side="right", padx=12)

            # Encabezados columnas
            enc = tk.Frame(self.fn, bg=C["borde"]); enc.pack(fill="x", padx=6)
            for txt,w in [("#",3),("NOTA (1.0-5.0)",13),("FECHA",14),("DESCRIPCIÓN DE LA EVALUACIÓN",0)]:
                tk.Label(enc, text=txt, width=w, font=("Segoe UI",9,"bold"),
                         fg=C["texto_sec"], bg=C["borde"], pady=4,
                         anchor="w" if w==0 else "center").pack(side="left",padx=3)

            notas = datos[comp]["notas"]
            fechas= datos[comp]["fechas"]
            descs = datos[comp]["descs"]

            for i, nota in enumerate(notas):
                bg_f = C["card"] if i%2==0 else C["panel"]
                ff = tk.Frame(self.fn, bg=bg_f); ff.pack(fill="x", padx=6)

                tk.Label(ff, text=str(i+1), width=3, font=("Segoe UI",10),
                         fg=C["texto_ter"], bg=bg_f).pack(side="left",padx=4)

                # Entrada de nota con validación visual en tiempo real
                en = tk.Entry(ff, width=11, font=("Segoe UI",11),
                               bg=C["input"], fg=C["texto"],
                               insertbackground=C["texto"],
                               relief="flat", bd=0, justify="center")
                en.insert(0, str(nota) if nota!="" and nota is not None else "")
                en.pack(side="left", padx=6, pady=3, ipady=5)

                def on_key(ev, e=en):
                    v = e.get().strip()
                    if not v: e.config(bg=C["input"]); return
                    ok,_,_ = validar_nota_meduca(v)
                    e.config(bg="#1A3A2A" if ok else "#3A1A1A")
                en.bind("<KeyRelease>", on_key)
                self.entradas[(comp,i)] = en

                ef = tk.Entry(ff, width=13, font=("Segoe UI",10),
                               bg=C["input"], fg=C["texto"],
                               insertbackground=C["texto"], relief="flat", bd=0)
                ef.insert(0, str(fechas[i]) if fechas[i] else "")
                ef.pack(side="left", padx=4, ipady=5)
                self.fechas[(comp,i)] = ef

                ed = tk.Entry(ff, font=("Segoe UI",10), bg=C["input"], fg=C["texto"],
                               insertbackground=C["texto"], relief="flat", bd=0)
                ed.insert(0, str(descs[i]) if descs[i] else "")
                ed.pack(side="left", fill="x", expand=True, padx=4, ipady=5)
                self.descs[(comp,i)] = ed

        # Promedios calculados
        sf2 = tk.Frame(self.fn, bg=C["panel"]); sf2.pack(fill="x", pady=(12,0), padx=6)
        tk.Label(sf2, text="  📊  PROMEDIOS  —  calculados automáticamente por el Excel",
                 font=("Segoe UI",10,"bold"), fg=C["amarillo"], bg=C["panel"], pady=8).pack(side="left")

        pf = tk.Frame(self.fn, bg=C["card"]); pf.pack(fill="x", padx=6)
        for i,(etq,key) in enumerate([
            ("Promedio Parciales","prom_parc"),
            ("Promedio Apreciación","prom_aprec"),
            ("Promedio Pruebas","prom_prueb"),
            ("⭐  CALIFICACIÓN TRIMESTRAL","cal")
        ]):
            val = datos.get(key)
            try:
                vf = round(float(str(val).replace(",",".")),1)
                color = C["verde_clar"] if vf>=3.0 else C["rojo_clar"]; txt=str(vf)
            except: color=C["texto_ter"]; txt="—"
            cf = tk.Frame(pf, bg=C["card"], padx=18, pady=12); cf.grid(row=0,column=i,padx=4,sticky="ew")
            pf.columnconfigure(i,weight=1)
            tk.Label(cf,text=etq,font=("Segoe UI",8),fg=C["texto_sec"],bg=C["card"]).pack()
            tk.Label(cf,text=txt,font=("Segoe UI",20,"bold"),fg=color,bg=C["card"]).pack()
            if vf if 'vf' in dir() else False:
                estado = "APROBADO ✓" if vf>=3 else "REPROBADO ✗"
                tk.Label(cf,text=estado,font=("Segoe UI",8,"bold"),
                         fg=color,bg=C["card"]).pack()

        # Observaciones
        sf3 = tk.Frame(self.fn, bg=C["panel"]); sf3.pack(fill="x", pady=(10,0), padx=6)
        tk.Label(sf3, text="  💬  OBSERVACIONES DEL ESTUDIANTE",
                 font=("Segoe UI",10,"bold"), fg=C["texto"], bg=C["panel"], pady=8).pack(side="left")
        of = tk.Frame(self.fn, bg=C["card"], padx=8, pady=8); of.pack(fill="x", padx=6)
        self.ent_obs = tk.Entry(of, font=("Segoe UI",11), bg=C["input"], fg=C["texto"],
                                 insertbackground=C["texto"], relief="flat", bd=0)
        self.ent_obs.insert(0, str(datos.get("obs") or ""))
        self.ent_obs.pack(fill="x", ipady=7)

        # Botón guardar al final
        bf = tk.Frame(self.fn, bg=C["bg"], pady=12); bf.pack(fill="x", padx=6)
        boton(bf,"💾  Guardar todas las notas de este estudiante",
              self._guardar_todo, color=C["verde"], w=46, grande=True).pack()

    def _guardar_todo(self):
        g=self.var_g.get(); mat=self.var_mat.get(); est_s=self.var_est.get()
        if not est_s: messagebox.showwarning("Selecciona","Selecciona un estudiante."); return
        try:
            tri_s=self.var_t.get()
            tri = int(tri_s.split("—")[0].strip()) if "—" in tri_s else int(tri_s)
            num_est = int(est_s.split(".")[0])
        except: return

        errores=[]; guardadas=0
        for (comp,idx),ent in self.entradas.items():
            val = ent.get().strip()
            if not val: continue
            ok,nota_f,msg = validar_nota_meduca(val)
            if not ok: errores.append(f"• Nota {idx+1} de {comp}: {msg}"); continue
            fe = self.fechas.get((comp,idx)); de = self.descs.get((comp,idx))
            try:
                guardar_nota_excel(g,mat,tri,num_est,comp,idx,nota_f,
                                   fe.get().strip() if fe else "",
                                   de.get().strip() if de else "")
                guardadas+=1
            except Exception as e:
                errores.append(f"• Error al guardar nota {idx+1}: {e}")

        obs = self.ent_obs.get().strip() if hasattr(self,"ent_obs") else ""
        if obs: guardar_obs_excel(g,mat,tri,num_est,obs)

        if errores:
            messagebox.showwarning("Algunas notas no se guardaron","\n".join(errores))
        if guardadas > 0:
            messagebox.showinfo("✓ Guardado",f"{guardadas} nota(s) guardadas correctamente.")
        self._upd_notas()

# ─────────────────────────────────────────────────────────
#  PANEL ASISTENCIA
# ─────────────────────────────────────────────────────────
class PanelAsistencia(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app = app; self._build()

    def _build(self):
        header_panel(self,"📅","Control de Asistencia","Registra presencia, ausencia y tardanza")

        sel = tk.Frame(self, bg=C["panel"], padx=12, pady=8)
        sel.pack(fill="x")
        tk.Label(sel,text="Grado:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        cfg = cargar_config()
        grados = list(cfg.get("materias",{}).keys()) or ["7°","8°","9°"]
        self.var_g = tk.StringVar(value=grados[0])
        for g in grados:
            tk.Radiobutton(sel,text=f"  {g}  ",variable=self.var_g,value=g,
                           font=("Segoe UI",11,"bold"),bg=C["panel"],
                           fg=C["texto"],selectcolor=C["azul"],
                           activebackground=C["panel"],
                           command=self._cargar).pack(side="left",padx=4)

        tk.Label(sel,text="  Trimestre:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left",padx=(15,0))
        self.var_tri = tk.StringVar(value="1")
        combo(sel, self.var_tri, ["1 — T1  (col 3-60)","2 — T2  (col 46+)","3 — T3  (col 89+)"], 18).pack(side="left",padx=(3,15))
        self.var_tri.trace("w", lambda *_: self._actualizar_col_base())

        tk.Label(sel,text="Día (col):",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        self.var_col = tk.StringVar(value="3")
        tk.Entry(sel,textvariable=self.var_col,width=6,font=("Segoe UI",10),
                  bg=C["input"],fg=C["texto"],insertbackground=C["texto"],
                  relief="flat",bd=0).pack(side="left",padx=(3,15),ipady=4)

        tk.Label(sel,text="Estado:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        self.var_est = tk.StringVar(value="-")
        for lt,val,clr in [("✓ Presente","P",C["verde"]),("✗ Ausente","-",C["rojo"]),("T Tardanza","T",C["amarillo"])]:
            tk.Radiobutton(sel,text=lt,variable=self.var_est,value=val,
                           font=("Segoe UI",10,"bold"),bg=C["panel"],fg=clr,
                           selectcolor=C["panel"],activebackground=C["panel"]).pack(side="left",padx=5)

        boton(sel,"💾 Marcar",self._marcar,color=C["verde"],w=10).pack(side="right",padx=4)
        boton(sel,"🔄",self._cargar,color=C["borde"],w=4).pack(side="right")

        # Tabla
        tf = tk.Frame(self, bg=C["bg"]); tf.pack(fill="both", expand=True, padx=8, pady=6)
        cols = ("pos","nombre","hoy","aus_t1","tard_t1","aus_t2","tard_t2","aus_t3","tard_t3")

        st = ttk.Style()
        st.configure("Dark.Treeview",
            background=C["card"], fieldbackground=C["card"],
            foreground=C["texto"], rowheight=30,
            font=("Segoe UI",10))
        st.configure("Dark.Treeview.Heading",
            background=C["panel"], foreground=C["texto_sec"],
            font=("Segoe UI",9,"bold"), relief="flat")
        st.map("Dark.Treeview", background=[("selected",C["azul"])])

        self.tree = ttk.Treeview(tf, columns=cols, show="headings",
                                  height=22, style="Dark.Treeview")
        for c,lbl_t,w in [("pos","#",40),("nombre","APELLIDO, NOMBRE",260),
                           ("hoy","HOY",90),("aus_t1","Aus T1",70),("tard_t1","Tar T1",70),
                           ("aus_t2","Aus T2",70),("tard_t2","Tar T2",70),
                           ("aus_t3","Aus T3",70),("tard_t3","Tar T3",70)]:
            self.tree.heading(c,text=lbl_t)
            self.tree.column(c,width=w,anchor="w" if c=="nombre" else "center")

        sc = ttk.Scrollbar(tf,orient="vertical",command=self.tree.yview)
        self.tree.configure(yscrollcommand=sc.set)
        sc.pack(side="right",fill="y"); self.tree.pack(fill="both",expand=True)

        self.lbl_inf = tk.Label(self,text="",font=("Segoe UI",9),
                                 fg=C["texto_ter"],bg=C["bg"])
        self.lbl_inf.pack(pady=3)
        self._cargar()

    def _actualizar_col_base(self):
        tri_s = self.var_tri.get()
        tri = int(tri_s.split("—")[0].strip()) if "—" in tri_s else int(tri_s)
        bases = {1: 3, 2: 46, 3: 89}
        # Ajustar col base al inicio del trimestre
        col_actual = int(self.var_col.get()) if self.var_col.get().isdigit() else 3
        base = bases.get(tri, 3)
        if col_actual < base or col_actual > base + 58:
            self.var_col.set(str(base))
        self._cargar()

    def _cargar(self):
        g = self.var_g.get()
        try: col = int(self.var_col.get())
        except: col = 3
        self.tree.delete(*self.tree.get_children())
        ests = obtener_estudiantes(g)
        self._ests = ests
        try:
            wb = wb_ro()
            ws = wb[HOJA_ASIST[g]]
            for pos,nombre in ests:
                # T1 filas 3-41, T2 46-85, T3 89-128
                fila_t1 = 3+pos-1; fila_t2=46+pos-1; fila_t3=89+pos-1
                hoy = ws.cell(fila_t1, col).value if 3<=col<=60 else ws.cell(fila_t2,col).value if 46<=col<=105 else ws.cell(fila_t3,col).value
                hoy_txt = "✓ Presente" if not hoy else ("✗ Ausente" if hoy=="-" else "T Tardanza")
                aus1=ws.cell(fila_t1,61).value or 0; tar1=ws.cell(fila_t1,62).value or 0
                aus2=ws.cell(fila_t2,61).value or 0; tar2=ws.cell(fila_t2,62).value or 0
                aus3=ws.cell(fila_t3,61).value or 0; tar3=ws.cell(fila_t3,62).value or 0
                tag = "aus" if hoy=="-" else "tar" if hoy=="T" else ""
                self.tree.insert("","end",
                    values=(pos,nombre,hoy_txt,aus1,tar1,aus2,tar2,aus3,tar3),
                    tags=(tag,))
            self.tree.tag_configure("aus", background="#3A1A1A")
            self.tree.tag_configure("tar", background="#3A2E0A")
            wb.close()
        except: pass
        self.lbl_inf.config(
            text=f"Grado {g} · {len(ests)} estudiantes · Columna actual: {col}  "
                 f"(T1=cols 3-60 · T2=cols 46-105 · T3=cols 89-148)")

    def _marcar(self):
        sel = self.tree.selection()
        if not sel: messagebox.showwarning("Selecciona","Haz clic en uno o varios estudiantes."); return
        g = self.var_g.get()
        try: col=int(self.var_col.get())
        except: messagebox.showwarning("Columna","Ingresa un número de columna válido."); return
        estado=self.var_est.get()
        for item in sel:
            pos = int(self.tree.item(item)["values"][0])
            marcar_asist(g, pos, col, estado)
        messagebox.showinfo("✓",f"Asistencia marcada para {len(sel)} estudiante(s).")
        self._cargar()

# ─────────────────────────────────────────────────────────
#  PANEL RESUMEN
# ─────────────────────────────────────────────────────────
class PanelResumen(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app = app; self._build()

    def _build(self):
        header_panel(self,"📊","Resumen de Calificaciones","Notas finales y estado por grado")

        sel = tk.Frame(self, bg=C["panel"], padx=12, pady=8)
        sel.pack(fill="x")
        tk.Label(sel,text="Grado:",font=("Segoe UI",9,"bold"),fg=C["texto_sec"],bg=C["panel"]).pack(side="left")
        cfg = cargar_config()
        grados = list(cfg.get("materias",{}).keys()) or ["7°","8°","9°"]
        self.var_g = tk.StringVar(value=grados[0])
        for g in grados:
            tk.Radiobutton(sel,text=f"  {g}  ",variable=self.var_g,value=g,
                           font=("Segoe UI",11,"bold"),bg=C["panel"],
                           fg=C["texto"],selectcolor=C["azul"],
                           activebackground=C["panel"],
                           command=self._cargar).pack(side="left",padx=4)
        boton(sel,"🔄 Generar resumen",self._cargar,w=18).pack(side="right")

        self.tbl_f = tk.Frame(self, bg=C["bg"])
        self.tbl_f.pack(fill="both", expand=True, padx=8, pady=5)
        self.lbl_stats = tk.Label(self,text="",font=("Segoe UI",9),
                                   fg=C["texto_sec"],bg=C["bg"])
        self.lbl_stats.pack(pady=3)

    def _cargar(self):
        for w in self.tbl_f.winfo_children(): w.destroy()
        cfg = cargar_config()
        g = self.var_g.get()
        mats = cfg.get("materias",{}).get(g,[])
        datos = leer_resumen(cfg, g)

        cols = ["nombre"]
        for m in mats:
            for s in ["_T1","_T2","_T3","_F"]: cols.append(f"{m[:8]}{s}")
        cols.append("estado")

        st = ttk.Style()
        st.configure("Dark.Treeview",
            background=C["card"],fieldbackground=C["card"],
            foreground=C["texto"],rowheight=28,font=("Segoe UI",10))
        st.configure("Dark.Treeview.Heading",
            background=C["panel"],foreground=C["texto_sec"],
            font=("Segoe UI",9,"bold"),relief="flat")
        st.map("Dark.Treeview",background=[("selected",C["azul"])])

        tree = ttk.Treeview(self.tbl_f,columns=cols,show="headings",
                             height=22,style="Dark.Treeview")
        tree.heading("nombre",text="NOMBRE"); tree.column("nombre",width=200,anchor="w")
        for m in mats:
            for s,ls in [("_T1","T1"),("_T2","T2"),("_T3","T3"),("_F","Final")]:
                c=f"{m[:8]}{s}"
                tree.heading(c,text=f"{m[:7]}\n{ls}"); tree.column(c,width=62,anchor="center")
        tree.heading("estado",text="Estado"); tree.column("estado",width=110,anchor="center")

        sc_v=ttk.Scrollbar(self.tbl_f,orient="vertical",command=tree.yview)
        sc_h=ttk.Scrollbar(self.tbl_f,orient="horizontal",command=tree.xview)
        tree.configure(yscrollcommand=sc_v.set,xscrollcommand=sc_h.set)
        sc_v.pack(side="right",fill="y"); sc_h.pack(side="bottom",fill="x")
        tree.pack(fill="both",expand=True)

        apro=repr_=sin=0
        for est in datos:
            fila=[est["nombre"]]; proms=[]
            for m in mats:
                info=est["materias"].get(m,{})
                for k in ["T1","T2","T3","Final"]:
                    v=info.get(k)
                    fila.append(str(v) if v is not None else "—")
                if info.get("Final"): proms.append(info["Final"])
            if proms:
                pg=sum(proms)/len(proms)
                e_=("✓ APROBADO" if pg>=3 else "✗ REPROBADO")
                if pg>=3: apro+=1
                else: repr_+=1
            else: e_="Sin notas"; sin+=1
            fila.append(e_)
            tag="apr" if "APROBADO" in e_ else "rep" if "REPROBADO" in e_ else ""
            tree.insert("","end",values=fila,tags=(tag,))

        tree.tag_configure("apr",background="#0A2A1A")
        tree.tag_configure("rep",background="#2A0A0A")
        self.lbl_stats.config(
            text=f"Grado {g}  ·  Total: {len(datos)}   ✓ Aprobados: {apro}   "
                 f"✗ Reprobados: {repr_}   · Sin notas: {sin}")

# ─────────────────────────────────────────────────────────
#  PANEL IMPRESIÓN
# ─────────────────────────────────────────────────────────
class PanelImpresion(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app = app; self._build()

    def _build(self):
        header_panel(self,"🖨️","Imprimir Planillas","Abre el Excel con tus datos para imprimir")

        body = scroll_frame(self)

        # Info
        info = tarjeta(body, padx=20, pady=15)
        info.pack(fill="x", padx=15, pady=15)
        tk.Label(info, text="El Excel se abre con TODO tu diseño y datos originales.",
                 font=("Segoe UI",12,"bold"), fg=C["texto"], bg=C["card"]).pack(anchor="w")
        tk.Label(info, text="Presiona Ctrl+P en Excel para imprimir. El formato oficial de MEDUCA se mantiene intacto.",
                 font=("Segoe UI",10), fg=C["texto_sec"], bg=C["card"]).pack(anchor="w",pady=(4,0))

        # Botón principal
        boton(body,"📂  Abrir Excel completo para imprimir",
              self._abrir_todo, color=C["azul"], w=40, grande=True).pack(pady=10, padx=15)

        # Hojas por sección
        cfg = cargar_config()
        secciones = [
            ("📋  Portadas y carátulas", [
                ("Portada","📋 Portada oficial MEDUCA"),
                ("Caratula","📄 Carátula del registro"),
                ("Horarios","🕐 Horario de clases"),
            ]),
            ("📅  Asistencia por grado", [
                (f"Asistencia ({g})", f"📅 Asistencia — Grado {g}")
                for g in cfg.get("materias",{}).keys()
            ]),
        ]
        # Planillas por materia
        for grado, mats in cfg.get("materias",{}).items():
            items = []
            for mat in mats:
                hp = hoja_plan(grado, mat)
                items.append((hp, f"📊 {mat} — Grado {grado}"))
            secciones.append((f"📊  Planillas — Grado {grado}", items))

        for titulo, hojas in secciones:
            sec = tarjeta(body, padx=15, pady=12)
            sec.pack(fill="x", padx=15, pady=4)
            tk.Label(sec, text=titulo, font=("Segoe UI",10,"bold"),
                     fg=C["amarillo"], bg=C["card"]).pack(anchor="w", pady=(0,8))
            grid = tk.Frame(sec, bg=C["card"]); grid.pack(fill="x")
            for i,(hoja_id,hoja_lbl) in enumerate(hojas):
                boton(grid, hoja_lbl, lambda h=hoja_id: self._abrir_hoja(h),
                      color=C["borde"], w=30).grid(row=i//2, column=i%2, padx=4, pady=2, sticky="ew")
            grid.columnconfigure(0, weight=1); grid.columnconfigure(1, weight=1)

    def _abrir_todo(self):
        r = ruta_excel()
        if not os.path.exists(r):
            messagebox.showerror("Error","No se encontró el Excel."); return
        try:
            os.startfile(r)
            messagebox.showinfo("✓ Excel abierto",
                "El Excel se abrió.\n\nPara imprimir:\n"
                "1. Ve a la hoja que quieres imprimir\n"
                "2. Presiona Ctrl + P\n"
                "3. Ajusta márgenes si es necesario\n"
                "4. Imprime")
        except Exception as e:
            messagebox.showerror("Error",f"No se pudo abrir el Excel:\n{e}")

    def _abrir_hoja(self, nombre_hoja):
        """En Windows: abre Excel en la hoja específica via VBScript."""
        r = ruta_excel()
        if not os.path.exists(r):
            messagebox.showerror("Error","No se encontró el Excel."); return
        import platform
        if platform.system() == "Windows":
            try:
                vbs = f"""
Set xl = CreateObject("Excel.Application")
xl.Visible = True
Set wb = xl.Workbooks.Open("{r.replace(chr(92),chr(92)*2)}")
On Error Resume Next
wb.Sheets("{nombre_hoja}").Activate
"""
                tmp = os.path.join(os.environ.get("TEMP","."), "rd_open.vbs")
                with open(tmp,"w") as f: f.write(vbs)
                import subprocess
                subprocess.Popen(["cscript","//Nologo",tmp])
                return
            except: pass
        os.startfile(r)

# ─────────────────────────────────────────────────────────
#  PANEL CONFIGURACIÓN
# ─────────────────────────────────────────────────────────
class PanelConfig(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app=app; self.vars={}; self.horario_entries=[]; self._cfg_materias={}
        self._build()

    def _build(self):
        header_panel(self,"⚙️","Configuración","Datos del docente, escuela, materias y horario")
        body = scroll_frame(self)
        cfg = cargar_config()
        self._llenar_campos(body, cfg)
        self._build_materias(body, cfg)
        self._build_horario(body, cfg)
        self._build_botones(body)

    def _campo(self, parent, etq, clave, ancho=30):
        f = tk.Frame(parent, bg=C["card"]); f.pack(fill="x", pady=2)
        tk.Label(f, text=etq, width=26, anchor="w", font=("Segoe UI",9),
                 fg=C["texto_sec"], bg=C["card"]).pack(side="left")
        cfg = cargar_config()
        v = tk.StringVar(value=str(cfg.get(clave,"")))
        self.vars[clave] = v
        e = entrada(f, v, w=ancho)
        e.pack(side="left", fill="x", expand=True, ipady=6, padx=(0,4))

    def _sec(self, parent, ico, titulo):
        f = tk.Frame(parent, bg=C["panel"], pady=0)
        f.pack(fill="x", padx=12, pady=(14,3))
        tk.Label(f, text=f"  {ico}  {titulo}", font=("Segoe UI",10,"bold"),
                 fg=C["amarillo"], bg=C["panel"], pady=8).pack(side="left")

    def _llenar_campos(self, parent, cfg):
        self._sec(parent,"👤","DATOS DEL DOCENTE")
        card = tarjeta(parent, padx=15, pady=10)
        card.pack(fill="x", padx=12)
        f = tk.Frame(card, bg=C["card"]); f.pack(fill="x")
        iz=tk.Frame(f,bg=C["card"]); iz.pack(side="left",fill="both",expand=True)
        de=tk.Frame(f,bg=C["card"]); de.pack(side="right",fill="both",expand=True)
        for et,cl in [("Nombre completo:","docente_nombre"),("N° Cédula:","docente_cedula"),
                      ("N° Posición:","docente_posicion"),("N° S.S.:","docente_ss")]:
            self._campo_en(iz, et, cl)
        for et,cl in [("Teléfono:","docente_telefono"),("Correo electrónico:","docente_correo"),
                      ("Condición:","docente_condicion"),("Jornada:","jornada")]:
            self._campo_en(de, et, cl)

        self._sec(parent,"🏫","DATOS DE LA ESCUELA")
        card2 = tarjeta(parent, padx=15, pady=10)
        card2.pack(fill="x", padx=12)
        f2 = tk.Frame(card2,bg=C["card"]); f2.pack(fill="x")
        iz2=tk.Frame(f2,bg=C["card"]); iz2.pack(side="left",fill="both",expand=True)
        de2=tk.Frame(f2,bg=C["card"]); de2.pack(side="right",fill="both",expand=True)
        for et,cl in [("Centro educativo:","escuela_nombre"),("Director(a):","escuela_director"),
                      ("Subdirector(a):","escuela_subdirector"),("Coordinador(a):","escuela_coordinador"),
                      ("Teléfono escuela:","escuela_telefono")]:
            self._campo_en(iz2, et, cl)
        for et,cl in [("Regional MEDUCA:","escuela_regional"),("Distrito:","escuela_distrito"),
                      ("Corregimiento:","escuela_corregimiento"),("Zona Escolar N°:","escuela_zona")]:
            self._campo_en(de2, et, cl)

        self._sec(parent,"📚","DATOS ACADÉMICOS")
        card3 = tarjeta(parent, padx=15, pady=10)
        card3.pack(fill="x", padx=12)
        f3=tk.Frame(card3,bg=C["card"]); f3.pack(fill="x")
        iz3=tk.Frame(f3,bg=C["card"]); iz3.pack(side="left",fill="both",expand=True)
        de3=tk.Frame(f3,bg=C["card"]); de3.pack(side="right",fill="both",expand=True)
        for et,cl in [("Año lectivo:","anio_lectivo"),("Grados:","grados"),
                      ("Área curricular:","area_curricular")]:
            self._campo_en(iz3, et, cl)
        self._campo_en(de3, "Asignaturas:", "asignaturas")

        # Logo
        self._sec(parent,"🖼️","LOGO DE LA ESCUELA")
        card4 = tarjeta(parent, padx=15, pady=10)
        card4.pack(fill="x", padx=12)
        fl = tk.Frame(card4, bg=C["card"]); fl.pack(fill="x")
        cfg2 = cargar_config()
        self.vars["logo_path"] = tk.StringVar(value=cfg2.get("logo_path",""))
        entrada(fl, self.vars["logo_path"], w=50).pack(side="left", ipady=6)
        boton(fl,"📂 Seleccionar",self._sel_logo,w=14).pack(side="left",padx=8)

    def _campo_en(self, parent, etq, clave, ancho=24):
        f = tk.Frame(parent, bg=C["card"]); f.pack(fill="x", pady=2, padx=4)
        tk.Label(f, text=etq, width=22, anchor="w", font=("Segoe UI",9),
                 fg=C["texto_sec"], bg=C["card"]).pack(side="left")
        cfg = cargar_config()
        v = tk.StringVar(value=str(cfg.get(clave,"")))
        self.vars[clave] = v
        entrada(f, v, w=ancho).pack(side="left", fill="x", expand=True, ipady=6)

    def _sel_logo(self):
        r = filedialog.askopenfilename(title="Logo",
            filetypes=[("Imágenes","*.png *.jpg *.jpeg"),("Todos","*.*")])
        if r: self.vars["logo_path"].set(r)

    def _build_materias(self, parent, cfg):
        self._sec(parent,"📋","MATERIAS Y GRADOS")
        card = tarjeta(parent, padx=15, pady=12)
        card.pack(fill="x", padx=12)
        tk.Label(card, text="Agrega nuevos grados y materias. El programa creará automáticamente las hojas en el Excel.",
                 font=("Segoe UI",9), fg=C["texto_sec"], bg=C["card"]).pack(anchor="w", pady=(0,8))
        self.mat_frame = tk.Frame(card, bg=C["card"])
        self.mat_frame.pack(fill="x")
        self._cfg_materias = {g:list(m) for g,m in cfg.get("materias",{}).items()}
        self._render_mats()

        # Agregar nuevo grado
        ng_f = tk.Frame(card, bg=C["card"]); ng_f.pack(fill="x", pady=(10,0))
        tk.Label(ng_f, text="Nuevo grado:", font=("Segoe UI",9),
                 fg=C["texto_sec"], bg=C["card"]).pack(side="left")
        self.ent_nuevo_grado = entrada(ng_f, w=8)
        self.ent_nuevo_grado.pack(side="left", padx=6, ipady=5)
        boton(ng_f,"+ Agregar grado",self._add_grado,color=C["azul_clar"],w=16).pack(side="left")

    def _render_mats(self):
        for w in self.mat_frame.winfo_children(): w.destroy()
        for grado, mats in self._cfg_materias.items():
            gf = tk.LabelFrame(self.mat_frame, text=f"  Grado {grado}  ",
                                font=("Segoe UI",9,"bold"), bg=C["card"],
                                fg=C["azul_clar"], padx=10, pady=6, bd=1,
                                relief="solid", highlightbackground=C["borde"])
            gf.pack(side="left", fill="y", padx=6, pady=4, anchor="n")
            for i,mat in enumerate(mats):
                mf = tk.Frame(gf, bg=C["card"]); mf.pack(fill="x", pady=1)
                e = tk.Entry(mf, font=("Segoe UI",9), bg=C["input"], fg=C["texto"],
                              insertbackground=C["texto"], relief="flat", bd=0, width=18)
                e.insert(0,mat); e.pack(side="left", ipady=4)
                def on_blur(ev,g=grado,idx=i,ent=e):
                    if idx<len(self._cfg_materias[g]):
                        self._cfg_materias[g][idx]=ent.get()
                e.bind("<FocusOut>",on_blur)
                tk.Button(mf,text="✕",font=("Segoe UI",8),fg=C["rojo"],
                          bg=C["card"],relief="flat",cursor="hand2",bd=0,
                          command=lambda g=grado,idx=i:self._del_mat(g,idx)
                          ).pack(side="left",padx=2)
            boton(gf,"+ Materia",lambda g=grado:self._add_mat(g),
                  color=C["borde"],w=14).pack(pady=(6,0))

    def _add_grado(self):
        g = self.ent_nuevo_grado.get().strip()
        if not g: messagebox.showwarning("Falta dato","Escribe el nombre del grado (ej: 10°)"); return
        if g in self._cfg_materias:
            messagebox.showwarning("Ya existe",f"El grado {g} ya está registrado."); return
        self._cfg_materias[g] = []
        self._render_mats()
        self.ent_nuevo_grado.delete(0,"end")
        messagebox.showinfo("✓",f"Grado {g} agregado. Ahora agrega sus materias.")

    def _add_mat(self, grado):
        n = simpledialog.askstring("Nueva materia",
                f"Nombre de la nueva materia para Grado {grado}:", parent=self)
        if n and n.strip():
            self._cfg_materias[grado].append(n.strip())
            self._render_mats()

    def _del_mat(self, grado, idx):
        mat = self._cfg_materias[grado][idx]
        if messagebox.askyesno("Eliminar",f"¿Eliminar '{mat}' de Grado {grado}?"):
            self._cfg_materias[grado].pop(idx)
            self._render_mats()

    def _build_horario(self, parent, cfg):
        self._sec(parent,"🕐","HORARIO DE CLASES")
        hf = tarjeta(parent, padx=12, pady=10)
        hf.pack(fill="x", padx=12)
        tabla = tk.Frame(hf, bg=C["card"]); tabla.pack(fill="x")
        for col,txt,w in [(0,"Período",8),(1,"Hora",13),(2,"Lunes",13),(3,"Martes",13),
                           (4,"Miércoles",13),(5,"Jueves",13),(6,"Viernes",13)]:
            tk.Label(tabla,text=txt,font=("Segoe UI",9,"bold"),bg=C["borde"],
                     fg=C["texto_sec"],width=w,relief="flat",pady=5).grid(
                         row=0,column=col,padx=1,pady=1,sticky="ew")
        self.horario_entries=[]
        for fi,per in enumerate(cfg.get("horario",CONFIG_DEFAULT["horario"])):
            fila_e=[]
            for col,key,w in [(0,"periodo",8),(1,"hora",13),(2,"lun",13),(3,"mar",13),
                               (4,"mie",13),(5,"jue",13),(6,"vie",13)]:
                e=tk.Entry(tabla,font=("Segoe UI",9),bg=C["input"],fg=C["texto"],
                            insertbackground=C["texto"],relief="flat",bd=0,width=w)
                e.insert(0,per.get(key,""))
                e.grid(row=fi+1,column=col,padx=1,pady=1,sticky="ew",ipady=4)
                fila_e.append((key,e))
            self.horario_entries.append(fila_e)

    def _build_botones(self, parent):
        tk.Frame(parent,bg=C["borde"],height=1).pack(fill="x",padx=12,pady=10)
        bf = tk.Frame(parent, bg=C["bg"]); bf.pack(fill="x", padx=12, pady=8)
        boton(bf,"💾  Guardar configuración",self._guardar,color=C["verde"],w=26,grande=True).pack(side="left",padx=4)
        boton(bf,"📄  Aplicar al Excel",self._act_excel,color=C["azul"],w=20,grande=True).pack(side="left",padx=4)
        boton(bf,"↺  Recargar",self._recargar,color=C["borde"],w=12).pack(side="left",padx=4)

    def _guardar(self):
        cfg = cargar_config()
        for clave,var in self.vars.items(): cfg[clave]=var.get().strip()
        cfg["horario"] = [{key:e.get().strip() for key,e in fe} for fe in self.horario_entries]
        cfg["materias"] = self._cfg_materias
        guardar_config(cfg); self.app.cfg=cfg
        messagebox.showinfo("✓ Guardado","Configuración guardada correctamente.")

    def _act_excel(self):
        self._guardar()
        cfg = cargar_config()
        # Agregar nuevas materias al Excel si no existen
        for grado, mats in cfg.get("materias",{}).items():
            for mat in mats:
                if hoja_prom(grado,mat) not in load_workbook(ruta_excel(), read_only=True).sheetnames:
                    ok, msg = agregar_materia_excel(grado, mat, cfg)
                    if not ok:
                        messagebox.showwarning("Advertencia",msg)
        ok,msg = actualizar_portada(cfg)
        if ok: messagebox.showinfo("✓ Excel actualizado",msg)
        else:  messagebox.showerror("Error",msg)

    def _recargar(self):
        for w in self.winfo_children(): w.destroy()
        self.__init__(self.master,self.app)

# ─────────────────────────────────────────────────────────
#  PANEL AUDITORÍA
# ─────────────────────────────────────────────────────────
class PanelAuditoria(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=C["bg"])
        self.app=app; self._build()

    def _build(self):
        header_panel(self,"🔍","Log de Auditoría","ISO 27001 — Registro de todas las acciones")
        sel = tk.Frame(self,bg=C["panel"],padx=12,pady=8); sel.pack(fill="x")
        boton(sel,"🔄 Actualizar",self._cargar,w=12).pack(side="left")
        self.lbl_tot=tk.Label(sel,text="",font=("Segoe UI",9),fg=C["texto_sec"],bg=C["panel"])
        self.lbl_tot.pack(side="left",padx=10)
        tk.Label(sel,text="Log cifrado — solo lectura",font=("Segoe UI",8),
                 fg=C["texto_ter"],bg=C["panel"]).pack(side="right")

        tf=tk.Frame(self,bg=C["bg"]); tf.pack(fill="both",expand=True,padx=8,pady=5)
        cols=("ts","cat","accion","detalle")

        st=ttk.Style()
        st.configure("Dark.Treeview",
            background=C["card"],fieldbackground=C["card"],
            foreground=C["texto"],rowheight=26,font=("Segoe UI",9))
        st.configure("Dark.Treeview.Heading",
            background=C["panel"],foreground=C["texto_sec"],
            font=("Segoe UI",9,"bold"),relief="flat")

        self.tree=ttk.Treeview(tf,columns=cols,show="headings",height=28,style="Dark.Treeview")
        for c,lbl_t,w in [("ts","Fecha/Hora UTC",160),("cat","Categoría",100),
                           ("accion","Acción",160),("detalle","Detalle",400)]:
            self.tree.heading(c,text=lbl_t); self.tree.column(c,width=w,anchor="w")
        sc=ttk.Scrollbar(tf,orient="vertical",command=self.tree.yview)
        self.tree.configure(yscrollcommand=sc.set)
        sc.pack(side="right",fill="y"); self.tree.pack(fill="both",expand=True)
        self._cargar()

    def _cargar(self):
        self.tree.delete(*self.tree.get_children())
        eventos=leer_auditoria()
        for ev in reversed(eventos):
            cat=ev.get("cat","")
            tag="seg" if cat=="SEGURIDAD" else "acc" if cat in ("ACCESO","LICENCIA") else ""
            self.tree.insert("","end",
                values=(ev.get("ts",""),cat,ev.get("accion",""),ev.get("detalle","")),
                tags=(tag,))
        self.tree.tag_configure("seg",background="#2A0A0A")
        self.tree.tag_configure("acc",background="#0A1A2A")
        self.lbl_tot.config(text=f"{len(eventos)} eventos registrados")

# ─────────────────────────────────────────────────────────
#  APLICACIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("RegistroDoc Pro v3.0 — Panamá 2026")
        self.geometry("1200x750")
        self.configure(bg=C["bg"])
        self.resizable(True,True)
        self.minsize(1000,640)
        self.cfg = cargar_config()

        # Aplicar tema oscuro a ttk
        st = ttk.Style(self)
        st.theme_use("clam")
        st.configure("TScrollbar", background=C["borde"], troughcolor=C["bg"],
                      arrowcolor=C["texto_sec"])
        st.configure("TNotebook", background=C["bg"], borderwidth=0)

        activado, datos_lic = verificar_licencia()
        if not activado:
            self.withdraw()
            self.after(200, lambda: VentanaActivacion(self, self._post_act))
        else:
            self._construir_ui(datos_lic)

    def _post_act(self):
        activado, datos_lic = verificar_licencia()
        if activado:
            self.deiconify(); self._construir_ui(datos_lic)
        else:
            self.destroy()

    def _construir_ui(self, datos_lic):
        # Verificar Excel
        if not os.path.exists(ruta_excel()):
            tk.Label(self,
                text="⚠  No se encontró Registro_2026.xlsx\nColócalo en la misma carpeta que app.py",
                font=("Segoe UI",14), fg=C["rojo_clar"], bg=C["bg"],
                justify="center").pack(expand=True)
            return

        # Verificar integridad
        ok, msg_int = verificar_integridad_excel(ruta_excel())
        if not ok:
            messagebox.showwarning("⚠ Advertencia de integridad",
                msg_int+"\n\nEl evento fue registrado en el log.")

        # Barra lateral
        sidebar = tk.Frame(self, bg=C["panel"], width=220)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # Logo en sidebar
        logo_f = tk.Frame(sidebar, bg=C["panel"], pady=20)
        logo_f.pack(fill="x")
        tk.Label(logo_f, text="📚", font=("Segoe UI",32), fg=C["amarillo"],
                 bg=C["panel"]).pack()
        tk.Label(logo_f, text="RegistroDoc Pro", font=("Segoe UI",12,"bold"),
                 fg=C["texto"], bg=C["panel"]).pack()
        tk.Label(logo_f, text="v3.0 · Panamá 2026", font=("Segoe UI",8),
                 fg=C["texto_ter"], bg=C["panel"]).pack()
        tk.Frame(sidebar, bg=C["borde"], height=1).pack(fill="x", padx=10)

        # Info docente
        cfg = cargar_config()
        dn = cfg.get("docente_nombre","") or datos_lic.get("cedula","")
        tk.Label(sidebar, text=dn[:22] if dn else "Docente",
                 font=("Segoe UI",9), fg=C["texto_sec"], bg=C["panel"],
                 wraplength=180).pack(pady=(8,0))
        tk.Frame(sidebar, bg=C["borde"], height=1).pack(fill="x", padx=10, pady=8)

        # Área de contenido
        content = tk.Frame(self, bg=C["bg"])
        content.pack(side="right", fill="both", expand=True)

        # Paneles
        self._panels = {}
        self._panel_actual = None

        MENU = [
            ("inicio",      "🏠", "Inicio",        PanelInicio),
            ("estudiantes", "👤", "Estudiantes",    PanelEstudiantes),
            ("notas",       "📝", "Notas",          PanelNotas),
            ("asistencia",  "📅", "Asistencia",     PanelAsistencia),
            ("resumen",     "📊", "Resumen",        PanelResumen),
            ("imprimir",    "🖨️", "Imprimir",       PanelImpresion),
            ("config",      "⚙️", "Configuración",  PanelConfig),
            ("auditoria",   "🔍", "Auditoría",      PanelAuditoria),
        ]

        self._btns_menu = {}
        for key, ico, nombre, PanelClass in MENU:
            panel = PanelClass(content, self)
            panel.place(relwidth=1, relheight=1)
            self._panels[key] = panel

            btn = tk.Button(sidebar, text=f"  {ico}  {nombre}",
                           font=("Segoe UI",11), fg=C["texto_sec"], bg=C["panel"],
                           relief="flat", cursor="hand2", anchor="w", padx=15, pady=10,
                           activebackground=C["borde"], activeforeground=C["texto"],
                           command=lambda k=key: self.ir_tab(k))
            btn.pack(fill="x")
            self._btns_menu[key] = btn

        # Espaciador
        tk.Frame(sidebar, bg=C["panel"]).pack(fill="both", expand=True)

        # Info en sidebar inferior
        tk.Frame(sidebar, bg=C["borde"], height=1).pack(fill="x", padx=10, pady=4)
        if SEGURIDAD_ACTIVA:
            tk.Label(sidebar, text="🔒 AES-256 · ISO 27001",
                     font=("Segoe UI",8), fg=C["verde_clar"], bg=C["panel"]).pack(pady=4)

        self.ir_tab("inicio")

    def ir_tab(self, key):
        if self._panel_actual:
            self._panels[self._panel_actual].lower()
            if self._panel_actual in self._btns_menu:
                self._btns_menu[self._panel_actual].config(
                    bg=C["panel"], fg=C["texto_sec"])
        if key in self._panels:
            self._panels[key].lift()
            self._panel_actual = key
            if key in self._btns_menu:
                self._btns_menu[key].config(bg=C["borde"], fg=C["texto"])

if __name__ == "__main__":
    app = App()
    app.mainloop()
