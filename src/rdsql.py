"""
RegistroDoc Pro — Módulo de Persistencia y Base de Datos SQL
===========================================================
Implementa la gestión de SQLite3 cifrada en reposo con AES-256,
cumpliendo con ISO/IEC 27001, ISO/IEC 25010 y Ley 81 de Panamá.

© 2026 RegistroDoc Pro
"""

import os
import sqlite3
import shutil
import platform
import logging
import threading
from rdsecurity import cifrar, descifrar, _hw_fingerprint

logger = logging.getLogger("rdsql")

def obtener_dir_datos_usuario() -> str:
    """Devuelve la ruta absoluta del directorio de datos del usuario."""
    if "PYTEST_CURRENT_TEST" in os.environ:
        import tempfile
        temp_base = os.path.join(tempfile.gettempdir(), "RegistroDoc_test_run")
        os.makedirs(temp_base, exist_ok=True)
        os.makedirs(os.path.join(temp_base, "data"), exist_ok=True)
        os.makedirs(os.path.join(temp_base, "temp"), exist_ok=True)
        os.makedirs(os.path.join(temp_base, "backups"), exist_ok=True)
        return temp_base

    if platform.system() == "Windows":
        base = os.environ.get("LOCALAPPDATA", os.path.expanduser("~\\AppData\\Local"))
    else:
        base = os.path.expanduser("~/.local/share")
    ruta = os.path.join(base, "RegistroDoc")
    os.makedirs(ruta, exist_ok=True)
    os.makedirs(os.path.join(ruta, "data"), exist_ok=True)
    os.makedirs(os.path.join(ruta, "temp"), exist_ok=True)
    os.makedirs(os.path.join(ruta, "backups"), exist_ok=True)
    return ruta

class ThreadSafeConnection:
    def __init__(self, conn, lock):
        self._conn = conn
        self._lock = lock

    def cursor(self):
        with self._lock:
            return ThreadSafeCursor(self._conn.cursor(), self._lock)

    def commit(self):
        with self._lock:
            return self._conn.commit()

    def rollback(self):
        with self._lock:
            return self._conn.rollback()

    def close(self):
        with self._lock:
            return self._conn.close()

    def execute(self, *args, **kwargs):
        with self._lock:
            return self._conn.execute(*args, **kwargs)

    def executemany(self, *args, **kwargs):
        with self._lock:
            return self._conn.executemany(*args, **kwargs)

    def executescript(self, *args, **kwargs):
        with self._lock:
            return self._conn.executescript(*args, **kwargs)

    def __getattr__(self, name):
        return getattr(self._conn, name)


class ThreadSafeCursor:
    def __init__(self, cursor, lock):
        self._cursor = cursor
        self._lock = lock

    def execute(self, *args, **kwargs):
        with self._lock:
            self._cursor.execute(*args, **kwargs)
            return self

    def executemany(self, *args, **kwargs):
        with self._lock:
            self._cursor.executemany(*args, **kwargs)
            return self

    def executescript(self, *args, **kwargs):
        with self._lock:
            self._cursor.executescript(*args, **kwargs)
            return self

    def fetchone(self):
        with self._lock:
            return self._cursor.fetchone()

    def fetchall(self):
        with self._lock:
            return self._cursor.fetchall()

    def fetchmany(self, *args, **kwargs):
        with self._lock:
            return self._cursor.fetchmany(*args, **kwargs)

    def __iter__(self):
        return self

    def __next__(self):
        with self._lock:
            return next(self._cursor)

    def __getattr__(self, name):
        return getattr(self._cursor, name)


class SQLDatabaseManager:
    """
    Controlador de ciclo de vida de la Base de Datos SQLite.
    Descifra el archivo al iniciar y lo mantiene sincronizado/cifrado.
    """
    def __init__(self, key: bytes = None):
        self.db_lock = threading.RLock()
        self._save_in_progress = False  # Previene ejecuciones simultáneas de guardar_cifrado
        self._save_lock = threading.Lock()  # Lock no-reentrante para guardar_cifrado
        self.user_dir = obtener_dir_datos_usuario()
        self.db_enc_path = os.path.join(self.user_dir, "data", "registro.db.enc")
        self.db_temp_path = os.path.join(self.user_dir, "temp", "sqlite_temp.db")

        # Derivación dinámica de la clave: Hardware Fingerprint + Docente Cédula
        import hashlib
        self.raw_hw_key = key if key else _hw_fingerprint()

        # ── Estrategia de recuperación multi-fuente para la cédula ──────────
        # Fuente 1: rd_token.bin (preferida, creado en activación)
        # Fuente 2: config.enc   (fallback robusto, siempre disponible)
        # Fuente 3: hw solo      (último recurso para bases nuevas/sin cédula)
        cedula_token = ""
        cedula_config = ""

        try:
            from rdsecurity import _leer_cedula_token
            token_data = _leer_cedula_token()
            cedula_token = token_data.get("hint", "").strip()
        except Exception:
            pass

        try:
            from rdsecurity import cargar_config_segura, _guardar_cedula_token
            cfg = cargar_config_segura({})
            cedula_config = cfg.get("docente_cedula", "").strip()
            # Si el token faltaba pero config.enc tiene la cédula → recrear token automáticamente
            if not cedula_token and cedula_config:
                try:
                    _guardar_cedula_token(cedula_config)
                    cedula_token = cedula_config
                    logger.info("SQLDatabaseManager: rd_token.bin recreado automáticamente desde config.enc")
                except Exception as e:
                    logger.warning(f"SQLDatabaseManager: No se pudo recrear rd_token.bin: {e}")
        except Exception:
            pass

        # Cédula efectiva: token tiene prioridad, luego config
        cedula_efectiva = cedula_token or cedula_config

        if cedula_efectiva:
            self.key = hashlib.sha256(self.raw_hw_key + cedula_efectiva.encode("utf-8")).digest()
        else:
            self.key = self.raw_hw_key

        # ── Claves candidatas para auto-healing ─────────────────────────────
        self.claves_candidatas = [self.key]
        # Añadir hw+cedula_config si difiere de la clave principal (ej. token y config divergen)
        if cedula_config and cedula_config != cedula_efectiva:
            alt_key = hashlib.sha256(self.raw_hw_key + cedula_config.encode("utf-8")).digest()
            if alt_key not in self.claves_candidatas:
                self.claves_candidatas.append(alt_key)
        # Añadir hw+cedula_token si difiere
        if cedula_token and cedula_token != cedula_efectiva:
            tok_key = hashlib.sha256(self.raw_hw_key + cedula_token.encode("utf-8")).digest()
            if tok_key not in self.claves_candidatas:
                self.claves_candidatas.append(tok_key)
        # Añadir hw solo como último recurso
        if self.raw_hw_key not in self.claves_candidatas:
            self.claves_candidatas.append(self.raw_hw_key)

        # Derivar la clave de hardware antigua (v1.x) para retrocompatibilidad
        try:
            import subprocess
            import platform
            import uuid
            old_comps = []
            if platform.system() == "Windows":
                creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
                # 1. UUID BIOS
                try:
                    r = subprocess.check_output(
                        ["wmic", "csproduct", "get", "UUID"],
                        stderr=subprocess.DEVNULL, timeout=3,
                        creationflags=creationflags
                    ).decode().strip().split("\n")
                    uid_old = r[-1].strip() if len(r) > 1 else ""
                    if uid_old and uid_old != "UUID":
                        old_comps.append(uid_old)
                except Exception:
                    pass
                # 2. Volume Serial
                try:
                    r = subprocess.check_output(
                        ["cmd.exe", "/c", "vol", "C:"],
                        stderr=subprocess.DEVNULL, timeout=3,
                        creationflags=creationflags
                    ).decode()
                    serial_old = re.search(r"[0-9A-F]{4}-[0-9A-F]{4}", r)
                    if serial_old:
                        old_comps.append(serial_old.group())
                except Exception:
                    pass
            old_comps.append(platform.node())
            old_comps.append(os.environ.get("USERNAME", "") or os.environ.get("USER", ""))
            try:
                mac_old = hex(uuid.getnode())
                old_comps.append(mac_old)
            except Exception:
                pass
                
            seed_old = "|".join(old_comps).encode("utf-8")
            old_hw_key = hashlib.sha256(seed_old).digest()
            
            # Añadir old_hw_key solo
            if old_hw_key not in self.claves_candidatas:
                self.claves_candidatas.append(old_hw_key)
            # Añadir old_hw_key + cedula
            if cedula_efectiva:
                old_key_ced = hashlib.sha256(old_hw_key + cedula_efectiva.encode("utf-8")).digest()
                if old_key_ced not in self.claves_candidatas:
                    self.claves_candidatas.append(old_key_ced)
            if cedula_config and cedula_config != cedula_efectiva:
                old_key_cfg = hashlib.sha256(old_hw_key + cedula_config.encode("utf-8")).digest()
                if old_key_cfg not in self.claves_candidatas:
                    self.claves_candidatas.append(old_key_cfg)
        except Exception as e:
            logger.warning(f"SQLDatabaseManager: Error calculando la clave de hardware antigua: {e}")

        self.conn = None

    def _validar_proceso_ejecutor(self) -> bool:
        """
        Verifica que el script/ejecutable que realiza la llamada sea el principal de RegistroDoc
        y que no esté siendo ejecutado bajo un depurador o inyectado por herramientas de análisis.
        """
        if os.environ.get("REGISTRODOC_DEV_MODE") == "1":
            return True
        try:
            from anti_analysis import detect_debugger
            if detect_debugger():
                return False
            import sys
            main_file = os.path.basename(sys.argv[0]).lower()
            # Permitir: scripts .py directos, lanzadores de RegistroDoc, pytest y __main__
            if main_file.endswith(".py"):
                return True
            if not any(x in main_file for x in ["python", "registrodoc", "pytest", "__main__.py"]):
                return False
        except Exception:
            pass
        return True

    def encriptar_campo(self, texto: str) -> str:
        """Encripta un valor a nivel de columna de forma rápida usando AES-256-GCM sin KDF lento."""
        if not texto:
            return ""
        try:
            import hashlib
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            col_key = hashlib.sha256(self.key + b"columna_datos_sensibles").digest()
            aesgcm = AESGCM(col_key)
            nonce = os.urandom(12)
            cifrado = aesgcm.encrypt(nonce, texto.encode("utf-8"), None)
            return (nonce + cifrado).hex()
        except Exception:
            return texto

    def desencriptar_campo(self, cifrado_hex: str) -> str:
        """Descifra un valor a nivel de columna usando AES-256-GCM con soporte retrocompatible y fallback de llaves."""
        if not cifrado_hex:
            return ""
        
        claves_a_probar = []
        if hasattr(self, "key") and self.key:
            claves_a_probar.append(self.key)
        if hasattr(self, "claves_candidatas") and self.claves_candidatas:
            for k in self.claves_candidatas:
                if k not in claves_a_probar:
                    claves_a_probar.append(k)
        if hasattr(self, "raw_hw_key") and self.raw_hw_key and self.raw_hw_key not in claves_a_probar:
            claves_a_probar.append(self.raw_hw_key)
            
        for candidate_key in claves_a_probar:
            try:
                import hashlib
                from cryptography.hazmat.primitives.ciphers.aead import AESGCM
                col_key = hashlib.sha256(candidate_key + b"columna_datos_sensibles").digest()
                blob = bytes.fromhex(cifrado_hex)
                
                # Soporte retrocompatible para formato lento anterior (comienza con magic bytes RD26)
                if len(blob) >= 4 and blob[:4] == b"RD26":
                    descifrado = descifrar(blob, col_key)
                    return descifrado.decode("utf-8")
                    
                if len(blob) < 12:
                    continue
                nonce = blob[:12]
                cifrado = blob[12:]
                aesgcm = AESGCM(col_key)
                descifrado = aesgcm.decrypt(nonce, cifrado, None)
                return descifrado.decode("utf-8")
            except Exception:
                continue
                
        return cifrado_hex

    def inicializar_tablas(self, conn: sqlite3.Connection):
        """Crea la estructura de tablas, relaciones, triggers e índices."""
        cursor = conn.cursor()
        
        # 1. Tabla de configuración
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS configuracion (
            clave TEXT PRIMARY KEY,
            valor TEXT NOT NULL
        );
        """)

        # 2. Tabla de grados
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS grados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            seccion TEXT NOT NULL,
            modalidad TEXT CHECK(modalidad IN ('primaria', 'premedia')) NOT NULL,
            UNIQUE(nombre, seccion, modalidad)
        );
        """)

        # 3. Tabla de estudiantes
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS estudiantes (
            id TEXT PRIMARY KEY,
            nombre TEXT NOT NULL,
            cedula TEXT,
            sexo TEXT,
            grado_id INTEGER NOT NULL,
            estado TEXT CHECK(estado IN ('Activo', 'Retirado')) DEFAULT 'Activo',
            fecha_retiro TEXT,
            motivo_retiro TEXT,
            nombre_acudiente TEXT,
            FOREIGN KEY (grado_id) REFERENCES grados(id) ON UPDATE CASCADE ON DELETE RESTRICT
        );
        """)

        # Migración automática: agregar columnas de retiro a BDs existentes
        columnas_existentes = {row[1] for row in cursor.execute('PRAGMA table_info(estudiantes);').fetchall()}
        migraciones_retiro = [
            ('fecha_retiro',     'ALTER TABLE estudiantes ADD COLUMN fecha_retiro TEXT;'),
            ('motivo_retiro',    'ALTER TABLE estudiantes ADD COLUMN motivo_retiro TEXT;'),
            ('nombre_acudiente', 'ALTER TABLE estudiantes ADD COLUMN nombre_acudiente TEXT;'),
        ]
        for col_name, sql in migraciones_retiro:
            if col_name not in columnas_existentes:
                try:
                    cursor.execute(sql)
                except Exception:
                    pass

        # 4. Tabla de materias
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS materias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            grado_id INTEGER NOT NULL,
            FOREIGN KEY (grado_id) REFERENCES grados(id) ON UPDATE CASCADE ON DELETE CASCADE
        );
        """)

        # 5. Tabla de horario escolar
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS horario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bloque_orden INTEGER NOT NULL DEFAULT 0,
            dia TEXT CHECK(dia IN ('lunes', 'martes', 'miercoles', 'jueves', 'viernes')) NOT NULL,
            bloque_hora TEXT NOT NULL,
            materia_texto TEXT DEFAULT '',
            materia_id INTEGER,
            grado_id INTEGER,
            FOREIGN KEY (materia_id) REFERENCES materias(id) ON UPDATE CASCADE ON DELETE SET NULL,
            FOREIGN KEY (grado_id) REFERENCES grados(id) ON UPDATE CASCADE ON DELETE SET NULL
        );
        """)

        # Migración automática: agregar columnas faltantes a horario
        cols_horario = {row[1] for row in cursor.execute('PRAGMA table_info(horario);').fetchall()}
        for col_name, col_sql in [
            ('materia_texto', 'ALTER TABLE horario ADD COLUMN materia_texto TEXT DEFAULT \'\';'),
            ('bloque_orden',  'ALTER TABLE horario ADD COLUMN bloque_orden INTEGER NOT NULL DEFAULT 0;'),
        ]:
            if col_name not in cols_horario:
                try:
                    cursor.execute(col_sql)
                except Exception:
                    pass

        # 6. Tabla de calificaciones (notas)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS notas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estudiante_id TEXT NOT NULL,
            materia_id INTEGER NOT NULL,
            trimestre INTEGER CHECK(trimestre IN (1, 2, 3)) NOT NULL,
            tipo TEXT CHECK(tipo IN ('Diaria / Parcial', 'Apreciación', 'Examen')) NOT NULL,
            descripcion TEXT NOT NULL,
            valor REAL CHECK(valor >= 1.0 AND valor <= 5.0),
            puntos_obtenidos REAL,
            puntos_maximos REAL,
            fecha TEXT NOT NULL,
            FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id) ON UPDATE CASCADE ON DELETE CASCADE,
            FOREIGN KEY (materia_id) REFERENCES materias(id) ON UPDATE CASCADE ON DELETE CASCADE
        );
        """)

        # 7. Tabla de asistencia
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS asistencia (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estudiante_id TEXT NOT NULL,
            fecha TEXT NOT NULL,
            estado TEXT CHECK(estado IN ('.', '-', 'T', 'E')) NOT NULL,
            trimestre INTEGER,
            motivo TEXT,
            FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id) ON UPDATE CASCADE ON DELETE CASCADE
        );
        """)

        # 8. Tabla de observaciones cualitativas
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS observaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estudiante_id TEXT NOT NULL,
            trimestre INTEGER CHECK(trimestre IN (1, 2, 3)) NOT NULL,
            comentario TEXT NOT NULL,
            FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id) ON UPDATE CASCADE ON DELETE CASCADE
        );
        """)

        # 9. Tabla de hábitos y actitudes
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS habitos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estudiante_id TEXT NOT NULL,
            trimestre INTEGER CHECK(trimestre IN (1, 2, 3)) NOT NULL,
            criterio_codigo TEXT NOT NULL,
            nota TEXT CHECK(nota IN ('S', 'R', 'X', '-')) NOT NULL,
            frecuencia TEXT DEFAULT 'Trimestral',
            periodo TEXT DEFAULT '',
            FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id) ON UPDATE CASCADE ON DELETE CASCADE
        );
        """)

        # 10. Tabla de tareas programadas
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS tareas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            descripcion TEXT NOT NULL,
            fecha_limite TEXT NOT NULL,
            estado TEXT CHECK(estado IN ('Pendiente', 'Completada')) DEFAULT 'Pendiente',
            materia_id INTEGER,
            FOREIGN KEY (materia_id) REFERENCES materias(id) ON UPDATE CASCADE ON DELETE CASCADE
        );
        """)

        # 11. Tabla de bitácora de auditoría
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now')),
            accion TEXT NOT NULL,
            detalle TEXT NOT NULL
        );
        """)

        # 12. Creación de Índices
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_estudiantes_grado ON estudiantes(grado_id);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_notas_est_mat ON notas(estudiante_id, materia_id, trimestre);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_asistencia_est_fecha ON asistencia(estudiante_id, fecha);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_habitos_est ON habitos(estudiante_id, trimestre);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_observaciones_est ON observaciones(estudiante_id);")

        # 13. Migración/Actualización de columnas en la tabla estudiantes si fuera necesario
        cursor.execute("PRAGMA table_info(estudiantes);")
        columnas_existentes = [row[1] for row in cursor.fetchall()]
        if columnas_existentes:
            if "cedula" not in columnas_existentes:
                try:
                    cursor.execute("ALTER TABLE estudiantes ADD COLUMN cedula TEXT;")
                    logger.info("Migración: Columna 'cedula' agregada exitosamente a la tabla estudiantes.")
                except Exception as ex:
                    logger.error(f"Error al agregar columna 'cedula': {ex}")
            if "sexo" not in columnas_existentes:
                try:
                    cursor.execute("ALTER TABLE estudiantes ADD COLUMN sexo TEXT;")
                    logger.info("Migración: Columna 'sexo' agregada exitosamente a la tabla estudiantes.")
                except Exception as ex:
                    logger.error(f"Error al agregar columna 'sexo': {ex}")

        # 14. Migración de la restricción CHECK de la tabla asistencia
        cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='asistencia';")
        asistencia_sql_row = cursor.fetchone()
        if asistencia_sql_row:
            asistencia_sql = asistencia_sql_row[0]
            if "'.'" not in asistencia_sql or "'A'" in asistencia_sql:
                try:
                    logger.info("Migración: Actualizando la tabla asistencia para admitir únicamente '.', '-', 'T', 'E'...")
                    
                    # 1. Check if old temp table exists, drop it
                    cursor.execute("DROP TABLE IF EXISTS asistencia_old;")
                    
                    # 2. Rename current table
                    cursor.execute("ALTER TABLE asistencia RENAME TO asistencia_old;")
                    
                    # 3. Create clean table
                    cursor.execute("""
                    CREATE TABLE asistencia (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        estudiante_id TEXT NOT NULL,
                        fecha TEXT NOT NULL,
                        estado TEXT CHECK(estado IN ('.', '-', 'T', 'E')) NOT NULL,
                        trimestre INTEGER,
                        motivo TEXT,
                        FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id) ON UPDATE CASCADE ON DELETE CASCADE
                    );
                    """)
                    
                    # 4. Migrate and clean old records
                    # Check if old table has trimestre column
                    cursor.execute("PRAGMA table_info(asistencia_old);")
                    old_cols = [r[1] for r in cursor.fetchall()]
                    has_trim = "trimestre" in old_cols
                    
                    if has_trim:
                        cursor.execute("SELECT id, estudiante_id, fecha, estado, trimestre, motivo FROM asistencia_old;")
                        filas = cursor.fetchall()
                    else:
                        cursor.execute("SELECT id, estudiante_id, fecha, estado, motivo FROM asistencia_old;")
                        filas = [r + (None,) for r in cursor.fetchall()] # append None for trim
                        
                    for f in filas:
                        if has_trim:
                            old_id, est_id, fecha, estado, trim, motivo = f
                        else:
                            old_id, est_id, fecha, estado, motivo, trim = f
                            
                        if estado == 'A':
                            estado_nuevo = '.'
                        elif estado in ('F', 'FJ'):
                            estado_nuevo = '-'
                        elif estado in ('.', '-', 'T', 'E'):
                            estado_nuevo = estado
                        else:
                            estado_nuevo = '.'
                        
                        cursor.execute("""
                        INSERT INTO asistencia (id, estudiante_id, fecha, estado, trimestre, motivo)
                        VALUES (?, ?, ?, ?, ?, ?);
                        """, (old_id, est_id, fecha, estado_nuevo, trim, motivo))
                    
                    # 5. Clean up old table
                    cursor.execute("DROP TABLE asistencia_old;")
                    cursor.execute("DROP INDEX IF EXISTS idx_asistencia_est_fecha;")
                    cursor.execute("CREATE INDEX idx_asistencia_est_fecha ON asistencia(estudiante_id, fecha);")
                    logger.info("Migración: Tabla asistencia migrada exitosamente a la restricción limpia de MEDUCA.")
                except Exception as ex:
                    logger.error(f"Error al migrar la tabla asistencia: {ex}")
                    try: cursor.execute("DROP TABLE IF EXISTS asistencia;")
                    except: pass
                    try: cursor.execute("ALTER TABLE asistencia_old RENAME TO asistencia;")
                    except: pass

        # Migración de columna trimestre en la tabla asistencia
        cursor.execute("PRAGMA table_info(asistencia);")
        asistencia_cols = [row[1] for row in cursor.fetchall()]
        if asistencia_cols:
            if "trimestre" not in asistencia_cols:
                try:
                    cursor.execute("ALTER TABLE asistencia ADD COLUMN trimestre INTEGER;")
                    logger.info("Migración: Columna 'trimestre' agregada exitosamente a la tabla asistencia.")
                except Exception as ex:
                    logger.error(f"Error al agregar columna 'trimestre': {ex}")

        # 15. Migración de columnas frecuencia y periodo en la tabla habitos
        cursor.execute("PRAGMA table_info(habitos);")
        habitos_cols = [row[1] for row in cursor.fetchall()]
        if habitos_cols:
            if "frecuencia" not in habitos_cols:
                try:
                    cursor.execute("ALTER TABLE habitos ADD COLUMN frecuencia TEXT DEFAULT 'Trimestral';")
                    logger.info("Migración: Columna 'frecuencia' agregada exitosamente a la tabla habitos.")
                except Exception as ex:
                    logger.error(f"Error al agregar columna 'frecuencia': {ex}")
            if "periodo" not in habitos_cols:
                try:
                    cursor.execute("ALTER TABLE habitos ADD COLUMN periodo TEXT DEFAULT '';")
                    logger.info("Migración: Columna 'periodo' agregada exitosamente a la tabla habitos.")
                except Exception as ex:
                    logger.error(f"Error al agregar columna 'periodo': {ex}")

        # 16. Deduplicación de materias y creación de índice único
        try:
            cursor.execute("SELECT nombre, grado_id, MIN(id) FROM materias GROUP BY nombre, grado_id")
            canonical = cursor.fetchall()
            for nombre, grado_id, min_id in canonical:
                cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?", (nombre, grado_id))
                all_ids = [r[0] for r in cursor.fetchall()]
                dup_ids = [i for i in all_ids if i != min_id]
                if dup_ids:
                    placeholders = ",".join(["?"] * len(dup_ids))
                    cursor.execute(f"UPDATE notas SET materia_id = ? WHERE materia_id IN ({placeholders})", (min_id, *dup_ids))
                    cursor.execute(f"UPDATE horario SET materia_id = ? WHERE materia_id IN ({placeholders})", (min_id, *dup_ids))
                    cursor.execute(f"UPDATE tareas SET materia_id = ? WHERE materia_id IN ({placeholders})", (min_id, *dup_ids))
                    cursor.execute(f"DELETE FROM materias WHERE id IN ({placeholders})", dup_ids)
            
            cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_materias_nombre_grado ON materias (nombre, grado_id);")
            logger.info("Migración: Deduplicación de materias completada y creado índice único.")
        except Exception as ex:
            logger.error(f"Error en migración de deduplicación de materias: {ex}")

        conn.commit()

    def conectar(self) -> sqlite3.Connection:
        """
        Descifra la base de datos a su ubicación temporal y abre una conexión.
        Si no existe, inicializa una nueva base de datos.
        """
        if self.conn:
            return ThreadSafeConnection(self.conn, self.db_lock)

        # 1. Validar proceso ejecutor (Process-Binding Protection)
        if not self._validar_proceso_ejecutor():
            logger.critical("Intento de acceso no autorizado a la base de datos (proceso no permitido).")
            raise PermissionError("Acceso denegado: Proceso no autorizado a enlazar con la base de datos.")

        # Asegurar limpieza previa si quedó algún remanente huérfano por caída de energía
        if os.path.exists(self.db_temp_path) and not os.path.exists(self.db_enc_path):
            try:
                os.remove(self.db_temp_path)
            except Exception:
                pass

        # 2. Descifrar si existe la base cifrada
        descifrado_ok = False
        rekey_needed = False

        # Las claves candidatas ya fueron construidas robustamente en __init__
        # Inicialización defensiva para mocks/tests que reemplazan __init__
        import hashlib
        if not hasattr(self, 'raw_hw_key'):
            self.raw_hw_key = self.key if hasattr(self, 'key') else b''
        if not hasattr(self, 'claves_candidatas'):
            self.claves_candidatas = [self.key] if hasattr(self, 'key') else []

        claves_candidatas = list(self.claves_candidatas)

        # Candidata extra: re-leer el token en tiempo de conexión por si fue recreado
        try:
            from rdsecurity import _leer_cedula_token
            token_data = _leer_cedula_token()
            token_cedula = token_data.get("hint", "").strip()
            if token_cedula and self.raw_hw_key:
                token_key = hashlib.sha256(self.raw_hw_key + token_cedula.encode("utf-8")).digest()
                if token_key not in claves_candidatas:
                    claves_candidatas.insert(0, token_key)  # Prioridad máxima
        except Exception:
            pass

        self.claves_candidatas = claves_candidatas

        if os.path.exists(self.db_enc_path):
            # Intentar descifrar con cada una
            for candidate in claves_candidatas:
                try:
                    with open(self.db_enc_path, "rb") as f:
                        cifrado = f.read()
                    descifrado = descifrar(cifrado, candidate)
                    with open(self.db_temp_path, "wb") as f:
                        f.write(descifrado)
                    
                    # Verificar integridad del SQLite descifrado
                    test_conn = sqlite3.connect(self.db_temp_path, check_same_thread=False, timeout=30.0)
                    try:
                        test_conn.execute("SELECT name FROM sqlite_master;")
                    finally:
                        test_conn.close()
                    
                    descifrado_ok = True
                    if candidate != self.key:
                        rekey_needed = True
                    break
                except Exception:
                    if os.path.exists(self.db_temp_path):
                        try:
                            os.remove(self.db_temp_path)
                        except Exception:
                            pass
                    continue

            if not descifrado_ok:
                logger.error("Error descifrando base de datos: Todas las llaves candidatas fallaron. Iniciando diagnóstico y reparación...")
                # Buscar si existe algún backup válido para restaurar
                dir_backups = os.path.join(self.user_dir, "backups")
                backups_existentes = []
                if os.path.exists(dir_backups):
                    backups_existentes = sorted(
                        [os.path.join(dir_backups, f) for f in os.listdir(dir_backups) if f.startswith("registro_db_backup_") and f.endswith(".db.enc")],
                        key=os.path.getmtime,
                        reverse=True
                    )
                
                restaurado = False
                for bak_path in backups_existentes:
                    for candidate in claves_candidatas:
                        try:
                            with open(bak_path, "rb") as f:
                                cifrado = f.read()
                            descifrado = descifrar(cifrado, candidate)
                            with open(self.db_temp_path, "wb") as f:
                                f.write(descifrado)
                            
                            # Intentar abrir conexión para verificar que no esté corrupta
                            test_conn = sqlite3.connect(self.db_temp_path, check_same_thread=False, timeout=30.0)
                            try:
                                test_conn.execute("SELECT name FROM sqlite_master;")
                            finally:
                                test_conn.close()
                            
                            # Si llegó aquí, restaurar este backup como el archivo oficial
                            shutil.copy2(bak_path, self.db_enc_path)
                            descifrado_ok = True
                            restaurado = True
                            logger.info(f"Base de datos recuperada exitosamente desde el respaldo: {bak_path}")
                            break
                        except Exception:
                            if os.path.exists(self.db_temp_path):
                                try:
                                    os.remove(self.db_temp_path)
                                except Exception:
                                    pass
                            continue
                    if restaurado:
                        break

                if not descifrado_ok:
                    # Si falla por manipulación o corrupción y no hay backups, generar respaldo de emergencia y reiniciar
                    bak_err = os.path.join(self.user_dir, "data", "registro_corrupted.bak")
                    try:
                        shutil.move(self.db_enc_path, bak_err)
                    except Exception:
                        pass

        # Conectar a la base de datos temporal
        self.conn = sqlite3.connect(self.db_temp_path, check_same_thread=False, timeout=30.0)
        self.conn.execute("PRAGMA foreign_keys = ON;")
        self.conn.execute("PRAGMA journal_mode = DELETE;")
        
        # Inicializar tablas si es necesario
        self.inicializar_tablas(self.conn)
        
        # Intentar auto-reparar la corrupción de nombres/cédulas en hex
        try:
            if self._autocubrir_y_reparar_hex_corrupcion(self.conn):
                rekey_needed = True
        except Exception as e:
            logger.error(f"SQLDatabaseManager: Error durante la auto-reparación: {e}")
            
        # Intentar fusionar registros de registro_plano.db de forma automática y única
        self._revisar_y_fusionar_registro_plano(self.conn)
        
        # Guardar una versión cifrada limpia si es nueva o si requiere re-keying
        if not os.path.exists(self.db_enc_path) or rekey_needed:
            self.guardar_cifrado()
            
        return ThreadSafeConnection(self.conn, self.db_lock)

    def _autocubrir_y_reparar_hex_corrupcion(self, conn) -> bool:
        """
        Detecta si los nombres de estudiantes están corruptos (cifrados como hex strings legibles).
        Si es así, busca copias de seguridad de Excel no corruptas y restaura los nombres reales
        tanto en los archivos de trabajo de Excel como en la base de datos SQLite.
        Mantiene intactos los géneros y todas las calificaciones/asistencias actuales.
        """
        import re
        import openpyxl
        import shutil
        
        cursor = conn.cursor()
        
        # 1. Obtener estudiantes de la BD para verificar si hay corrupción
        try:
            cursor.execute("SELECT id, nombre, cedula FROM estudiantes;")
            estudiantes_db = cursor.fetchall()
        except Exception:
            return False # Si no se puede consultar, abortar
            
        if not estudiantes_db:
            return False # Sin estudiantes, nada que hacer
            
        corruptos_count = 0
        for est_id, nombre_enc, cedula_enc in estudiantes_db:
            dec_nombre = self.desencriptar_campo(nombre_enc)
            if dec_nombre and re.match(r"^[0-9a-fA-F]{30,}$", dec_nombre):
                corruptos_count += 1
                
        if corruptos_count == 0:
            logger.info("SQLDatabaseManager: Verificación de consistencia de nombres exitosa (sin corrupción).")
            return False
            
        logger.warning(f"SQLDatabaseManager: Detectada corrupción de nombres en {corruptos_count} estudiantes. Iniciando auto-reparación...")
        
        # 2. Determinar la modalidad y los grados configurados en la base de datos
        try:
            cursor.execute("SELECT id, nombre, modalidad FROM grados;")
            grados_info = cursor.fetchall()
        except Exception:
            grados_info = []
            
        if not grados_info:
            return False
            
        # Agrupar por modalidad
        modalidad_grados = {}
        for g_id, g_nombre, g_modalidad in grados_info:
            if g_modalidad not in modalidad_grados:
                modalidad_grados[g_modalidad] = []
            modalidad_grados[g_modalidad].append((g_id, g_nombre))
            
        # 3. Buscar archivos de respaldo Excel que contengan nombres válidos (no hex)
        directorios_respaldo = [
            os.path.join(self.user_dir, "temp", "Respaldos_Locales"),
            os.path.join(self.user_dir, "temp", "Respaldos_Auto"),
            os.path.join(self.user_dir, "temp"),
            r"c:\RegistroDoc\Respaldos_Locales",
            r"c:\RegistroDoc\Respaldos_Auto",
            r"c:\RegistroDoc"
        ]
        
        candidatos = []
        for d in directorios_respaldo:
            if os.path.exists(d):
                try:
                    for f in os.listdir(d):
                        if f.endswith(".xlsx") and not f.startswith("~$") and "test" not in f.lower() and "cache" not in f.lower():
                            path = os.path.join(d, f)
                            candidatos.append((path, os.path.getmtime(path)))
                except Exception:
                    pass
                    
        # Ordenar por fecha de modificación descendente (más nuevos primero)
        candidatos.sort(key=lambda x: x[1], reverse=True)
        
        # Encontrar respaldos válidos por modalidad
        backup_por_modalidad = {}
        for path, _ in candidatos:
            try:
                # Comprobar si el archivo contiene nombres reales y no hex
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                nombre_maestro = None
                for s in wb.sheetnames:
                    if "MAESTRO" in s.upper():
                        nombre_maestro = s
                        break
                if not nombre_maestro:
                    wb.close()
                    continue
                    
                ws = wb[nombre_maestro]
                tiene_hex = False
                tiene_nombres_reales = False
                
                # Escanear columnas buscando nombres reales
                for col in range(1, 40):
                    for r in range(5, 46):
                        val = ws.cell(row=r, column=col).value
                        if val:
                            val_str = str(val).strip()
                            if re.match(r"^[0-9a-fA-F]{30,}$", val_str):
                                tiene_hex = True
                                break
                            elif len(val_str) > 2 and any(c.isalpha() for c in val_str):
                                tiene_nombres_reales = True
                    if tiene_hex:
                        break
                        
                wb.close()
                if tiene_nombres_reales and not tiene_hex:
                    ws_type = "premedia"
                    if "primaria" in path.lower() or "primaria" in nombre_maestro.lower():
                        ws_type = "primaria"
                    elif "premedia" in path.lower() or "premedia" in nombre_maestro.lower():
                        ws_type = "premedia"
                    else:
                        ws_type = "premedia"
                        
                    if ws_type not in backup_por_modalidad:
                        backup_por_modalidad[ws_type] = path
                        logger.info(f"SQLDatabaseManager: Encontrado respaldo válido para {ws_type}: {path}")
                        
                    if len(backup_por_modalidad) >= len(modalidad_grados):
                        break
            except Exception:
                pass
                
        # 4. Extraer el mapeo de nombres y cédulas
        student_recovery_map = {}
        for mod, g_list in modalidad_grados.items():
            path_bak = backup_por_modalidad.get(mod)
            if not path_bak:
                path_bak = next(iter(backup_por_modalidad.values()), None)
                
            if not path_bak:
                logger.error(f"SQLDatabaseManager: No se encontró respaldo válido para la modalidad {mod}.")
                continue
                
            try:
                wb_bak = openpyxl.load_workbook(path_bak, data_only=True)
                nombre_maestro = None
                for s in wb_bak.sheetnames:
                    if "MAESTRO" in s.upper():
                        nombre_maestro = s
                        break
                ws_m_bak = wb_bak[nombre_maestro]
                
                for grado_id, grado_nombre in g_list:
                    col_nom = None
                    grado_limpio = grado_nombre.replace("°", "").strip()
                    for r_h in [3, 4]:
                        for c in range(1, 40):
                            val = str(ws_m_bak.cell(row=r_h, column=c).value or "").strip()
                            if grado_limpio in val or grado_nombre in val:
                                if "NOMBRE" in str(ws_m_bak.cell(row=4, column=c).value or "").upper():
                                    col_nom = c
                                    break
                                elif "NOMBRE" in str(ws_m_bak.cell(row=4, column=c+1).value or "").upper():
                                    col_nom = c+1
                                    break
                        if col_nom:
                            break
                            
                    if not col_nom:
                        continue
                        
                    ws_planilla_bak = None
                    if mod == "premedia":
                        for sheet in wb_bak.sheetnames:
                            if "Planilla" in sheet and grado_nombre.replace("°", "") in sheet:
                                ws_planilla_bak = wb_bak[sheet]
                                break
                                
                    for r in range(5, 46):
                        nom = ws_m_bak.cell(row=r, column=col_nom).value
                        if nom:
                            student_id = grado_id * 100 + (r - 4)
                            cedula = ""
                            if mod == "primaria":
                                cedula = ws_m_bak.cell(row=r, column=5).value
                            elif ws_planilla_bak:
                                fila_plan = 15 + (r - 4)
                                cedula = ws_planilla_bak.cell(row=fila_plan, column=5).value
                                
                            student_recovery_map[student_id] = {
                                "nombre": str(nom).strip(),
                                "cedula": str(cedula).strip() if cedula else ""
                            }
                wb_bak.close()
            except Exception as e:
                logger.error(f"SQLDatabaseManager: Error extrayendo datos del respaldo {path_bak}: {e}")
                
        if not student_recovery_map:
            logger.error("SQLDatabaseManager: No se pudieron extraer estudiantes válidos de los respaldos.")
            return False
            
        # 5. Sobrescribir en el Excel activo (en AppData/temp y en la raíz)
        excel_repaired = False
        for mod, g_list in modalidad_grados.items():
            active_excel_name = "Registro_Premedia.xlsx" if mod == "premedia" else "Registro_Primaria.xlsx"
            active_excel_path = os.path.join(self.user_dir, "temp", active_excel_name)
            root_excel_path = os.path.join(r"c:\RegistroDoc", active_excel_name)
            
            if os.path.exists(active_excel_path):
                try:
                    shutil.copy2(active_excel_path, active_excel_path + ".corrupted.bak")
                    
                    wb_active = openpyxl.load_workbook(active_excel_path)
                    nombre_maestro = None
                    for s in wb_active.sheetnames:
                        if "MAESTRO" in s.upper():
                            nombre_maestro = s
                            break
                    ws_m_act = wb_active[nombre_maestro]
                    
                    for grado_id, grado_nombre in g_list:
                        col_nom = None
                        grado_limpio = grado_nombre.replace("°", "").strip()
                        for r_h in [3, 4]:
                            for c in range(1, 40):
                                val = str(ws_m_act.cell(row=r_h, column=c).value or "").strip()
                                if grado_limpio in val or grado_nombre in val:
                                    if "NOMBRE" in str(ws_m_act.cell(row=4, column=c).value or "").upper():
                                        col_nom = c
                                        break
                                    elif "NOMBRE" in str(ws_m_act.cell(row=4, column=c+1).value or "").upper():
                                        col_nom = c+1
                                        break
                            if col_nom:
                                break
                                
                        if not col_nom:
                            continue
                            
                        ws_planilla_act = None
                        if mod == "premedia":
                            for sheet in wb_active.sheetnames:
                                if "Planilla" in sheet and grado_nombre.replace("°", "") in sheet:
                                    ws_planilla_act = wb_active[sheet]
                                    break
                                    
                        for r in range(5, 46):
                            student_id = grado_id * 100 + (r - 4)
                            if student_id in student_recovery_map:
                                recovered = student_recovery_map[student_id]
                                ws_m_act.cell(row=r, column=col_nom).value = recovered["nombre"]
                                if mod == "primaria":
                                    ws_m_act.cell(row=r, column=5).value = recovered["cedula"]
                                elif ws_planilla_act:
                                    fila_plan = 15 + (r - 4)
                                    ws_planilla_act.cell(row=fila_plan, column=5).value = recovered["cedula"]
                                    
                    wb_active.save(active_excel_path)
                    wb_active.close()
                    excel_repaired = True
                    logger.info(f"SQLDatabaseManager: Excel activo reparado con éxito: {active_excel_path}")
                except Exception as e:
                    logger.error(f"SQLDatabaseManager: Error reparando el Excel activo {active_excel_name}: {e}")
                    
            if excel_repaired and os.path.exists(root_excel_path):
                try:
                    shutil.copy2(root_excel_path, root_excel_path + ".corrupted.bak")
                    shutil.copy2(active_excel_path, root_excel_path)
                    logger.info(f"SQLDatabaseManager: Excel raíz en {root_excel_path} actualizado.")
                except Exception as e:
                    logger.warning(f"SQLDatabaseManager: No se pudo copiar a la raíz: {e}")
                    
        # 6. Actualizar la base de datos SQLite con los nombres y cédulas cifrados correctamente
        try:
            updated_db_count = 0
            for student_id, recovered in student_recovery_map.items():
                nombre_cifrado = self.encriptar_campo(recovered["nombre"])
                cedula_cifrada = self.encriptar_campo(recovered["cedula"])
                cursor.execute(
                    "UPDATE estudiantes SET nombre = ?, cedula = ? WHERE id = ?;",
                    (nombre_cifrado, cedula_cifrada, str(student_id))
                )
                updated_db_count += 1
            conn.commit()
            logger.info(f"SQLDatabaseManager: Base de datos SQLite reparada ({updated_db_count} estudiantes).")
            return True
        except Exception as e:
            logger.error(f"SQLDatabaseManager: Error actualizando estudiantes en BD: {e}")
            return False

    def _revisar_y_fusionar_registro_plano(self, conn):
        plano_path = r"c:\RegistroDoc\registro_plano.db"
        if not os.path.exists(plano_path):
            return
            
        logger.info("Detectada base de datos plana registro_plano.db. Iniciando fusión...")
        try:
            conn_plano = sqlite3.connect(plano_path)
            cursor_pla = conn_plano.cursor()
            
            # Obtener datos de asistencia en plano
            cursor_pla.execute("SELECT estudiante_id, fecha, estado, trimestre FROM asistencia;")
            asis_pla = cursor_pla.fetchall()
            conn_plano.close()
            
            # Obtener datos actuales
            cursor_act = conn.cursor()
            cursor_act.execute("SELECT estudiante_id, fecha, estado, trimestre FROM asistencia;")
            asis_act = set(cursor_act.fetchall())
            
            merged_count = 0
            for est_id, fecha, estado, trim in asis_pla:
                # Comprobar si ya existe el registro (normalizando fechas invertidas MM-DD y DD-MM)
                swapped_fecha = fecha
                if len(fecha) == 5 and '-' in fecha:
                    parts = fecha.split('-')
                    swapped_fecha = f"{parts[1]}-{parts[0]}"
                    
                has_match = False
                for act_est, act_f, act_est_val, act_trim in asis_act:
                    if act_est == est_id and (act_f == fecha or act_f == swapped_fecha):
                        has_match = True
                        break
                        
                if not has_match:
                    cursor_act.execute(
                        "INSERT INTO asistencia (estudiante_id, fecha, estado, trimestre) VALUES (?, ?, ?, ?);",
                        (est_id, fecha, estado, trim)
                    )
                    merged_count += 1
            
            if merged_count > 0:
                conn.commit()
                # Cifrar y guardar inmediatamente
                self.guardar_cifrado()
                logger.info(f"Fusión automática completada. Se recuperaron {merged_count} registros.")
            
            # Renombrar archivo a .bak para no repetir
            bak_path = plano_path + "_procesado.bak"
            if os.path.exists(bak_path):
                try: os.remove(bak_path)
                except: pass
            os.rename(plano_path, bak_path)
            logger.info("Archivo registro_plano.db renombrado exitosamente.")
            
        except Exception as e:
            logger.error(f"Error en la fusión de registro_plano.db: {e}", exc_info=True)

    def guardar_cifrado(self):
        """Lee el archivo temporal SQLite activo, lo cifra y guarda en disco.
        Llamar directamente solo al cerrar la app. En ejecucion normal, usar _schedule_save() en DataEngine.
        """
        with self.db_lock:  # RLock garantiza exclusion mutua
            if not os.path.exists(self.db_temp_path):
                return
                
            try:
                # Asegurar volcado de transacciones pendientes a disco
                if self.conn:
                    self.conn.commit()
                    # Pequeño checkpoint de SQLite para sincronizar el WAL/journal a disco
                    self.conn.execute("PRAGMA wal_checkpoint(FULL);")
                    
                # Leer bytes del SQLite temporal y cifrarlos
                with open(self.db_temp_path, "rb") as f:
                    datos = f.read()
                cifrado = cifrar(datos, self.key)
                
                # Guardar de forma atómica para evitar corrupción si se corta la corriente
                db_enc_tmp = self.db_enc_path + ".tmp"
                with open(db_enc_tmp, "wb") as f:
                    f.write(cifrado)
                shutil.move(db_enc_tmp, self.db_enc_path)
            except Exception as e:
                logger.error(f"Error guardando base cifrada: {e}")

    def cerrar(self):
        """Cierra la conexión, realiza el guardado cifrado final y limpia el archivo temporal."""
        with self.db_lock:
            if os.path.exists(self.db_temp_path):
                self.guardar_cifrado()

            if self.conn:
                try:
                    self.conn.close()
                except Exception:
                    pass
                self.conn = None

            if os.path.exists(self.db_temp_path):
                # Destruir archivo temporal de forma segura (wiping antes de eliminar)
                try:
                    tamanio = os.path.getsize(self.db_temp_path)
                    with open(self.db_temp_path, "wb") as f:
                        f.write(os.urandom(tamanio))
                    os.remove(self.db_temp_path)
                except Exception:
                    # Fallback simple
                    try:
                        os.remove(self.db_temp_path)
                    except Exception:
                        pass

    def registrar_auditoria(self, accion: str, detalle: str):
        """Registra un evento de auditoría interna de forma rápida."""
        if not self.conn:
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO auditoria (accion, detalle) VALUES (?, ?);", (accion, detalle))
            self.conn.commit()
            self.guardar_cifrado()
        except Exception:
            pass

    def respaldar_base_datos(self) -> bool:
        """Crea un respaldo del archivo SQLite cifrado actual en la carpeta de backups."""
        if not os.path.exists(self.db_enc_path):
            return False
        try:
            import datetime
            import shutil
            dir_backups = os.path.join(self.user_dir, "backups")
            ahora = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            nombre_bak = f"registro_db_backup_{ahora}.db.enc"
            ruta_bak = os.path.join(dir_backups, nombre_bak)
            
            # Copiar archivo cifrado
            shutil.copy2(self.db_enc_path, ruta_bak)
            
            # Mantener máximo 3 backups rotativos
            archivos = sorted(
                [os.path.join(dir_backups, f) for f in os.listdir(dir_backups) if f.startswith("registro_db_backup_") and f.endswith(".db.enc")],
                key=os.path.getmtime
            )
            while len(archivos) > 3:
                try:
                    os.remove(archivos.pop(0))
                except Exception:
                    pass
            return True
        except Exception as e:
            logger.error(f"Error al respaldar base de datos: {e}")
            return False
