import openpyxl
import os

class ExcelCleaner:
    def limpiar_todo(self, ruta, modalidad):
        if not os.path.exists(ruta): 
            return False
            
        try:
            wb = openpyxl.load_workbook(ruta)
            
            # 1. Limpiar MAESTRO (Nombres y Cédulas)
            if "MAESTRO" in wb.sheetnames:
                ws = wb["MAESTRO"]
                for r in range(5, 51):
                    for c in range(1, 15):
                        celda = ws.cell(row=r, column=c)
                        if type(celda).__name__ != 'MergedCell':
                            celda.value = None

            # 2. Limpiar hojas de trabajo
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sheet_upper = sheet.upper()
                
                # PROM: Respeta las celdas combinadas y fórmulas
                if "PROM" in sheet_upper:
                    for r in range(4, 51):
                        for c in range(3, 100):
                            celda = ws.cell(row=r, column=c)
                            if type(celda).__name__ != 'MergedCell' and celda.data_type != 'f':
                                celda.value = None
                                
                # PLANILLA: Limpia notas, respeta fórmulas
                elif "PLANILLA" in sheet_upper:
                    for r in range(15, 60):
                        for c in range(5, 20):
                            celda = ws.cell(row=r, column=c)
                            if type(celda).__name__ != 'MergedCell' and celda.data_type != 'f':
                                celda.value = None
                                
                # ASISTENCIA: Limpieza con bisturí
                elif "ASISTENCIA" in sheet_upper:
                    # Limpiar fechas (Fila 2). Solo borra si la celda tiene exactamente 5 letras y un guion "DD-MM"
                    for c in range(3, 100):
                        celda = ws.cell(row=2, column=c)
                        val = str(celda.value or "").strip()
                        if type(celda).__name__ != 'MergedCell' and len(val) == 5 and "-" in val:
                            celda.value = None
                            
                    # Limpiar asistencia (P, A, T). NO TOCA fórmulas ni la palabra OBSERVACIONES
                    for r in range(3, 60):
                        for c in range(3, 100):
                            celda = ws.cell(row=r, column=c)
                            if type(celda).__name__ != 'MergedCell' and celda.data_type != 'f':
                                if str(celda.value).strip().upper() != "OBSERVACIONES":
                                    celda.value = None

            wb.save(ruta)
            wb.close()
            return True
            
        except Exception as e:
            print(f"Error crítico al limpiar el archivo: {e}")
            return False