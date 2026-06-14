import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import openpyxl

try:
    import docx
    DOCX_DISPONIBLE = True
except ImportError:
    DOCX_DISPONIBLE = False

from rdprint import abrir_para_imprimir, imprimir_hoja_directo

def parsear_archivo_estudiantes(ruta_archivo) -> list[dict]:
    """
    Parsea estudiantes (nombres y cédulas) desde un archivo Excel, Word o TXT.
    Retorna una lista de dicts: [{'nombre': '...', 'cedula': '...'}]
    """
    estudiantes = []
    ext = os.path.splitext(ruta_archivo)[1].lower()
    
    # Regex para Cédula Panameña: ej. 8-123-4567, PE-33-44, 4-PI-777-88
    cedula_pattern = re.compile(
        r'\b(?:(?:PE|PI|E|\d+)-[A-Z0-9]+-\d+-\d+|(?:PE|PI|E|\d+)-\d+-\d+)\b', 
        re.IGNORECASE
    )

    if ext == ".xlsx":
        # Verificar marca de aislamiento
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True)
        if "_RD_EXPORT_MARKER_" in wb.sheetnames:
            wb.close()
            raise PermissionError("SEGURIDAD: No se permite cargar o importar datos de un archivo Excel exportado por el sistema.")
        
        ws = wb.active
        # Buscar en cada fila
        for r in range(1, ws.max_row + 1):
            fila_valores = [str(ws.cell(row=r, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
            fila_str = " ".join(fila_valores)
            cedulas_encontradas = cedula_pattern.findall(fila_str)
            cedula_ref = cedulas_encontradas[0] if cedulas_encontradas else ""
            
            # Buscar nombre: primer elemento no vacío que sea texto largo y no sea cédula ni cabecera
            nombre_candidato = ""
            for val in fila_valores:
                val_clean = val.strip()
                if not val_clean:
                    continue
                if val_clean.isdigit() or val_clean == cedula_ref:
                    continue
                if val_clean.upper() in ["N°", "NO.", "ESTUDIANTE", "NOMBRE", "CEDULA", "CÉDULA", "LISTADO", "REGISTRO"]:
                    continue
                if len(val_clean) > 4:
                    nombre_candidato = val_clean
                    break
            
            if nombre_candidato:
                estudiantes.append({
                    "nombre": nombre_candidato,
                    "cedula": cedula_ref
                })
        wb.close()

    elif ext == ".docx":
        if not DOCX_DISPONIBLE:
            raise ImportError("La librería python-docx no está instalada.")
        doc = docx.Document(ruta_archivo)
        for table in doc.tables:
            for row in table.rows:
                fila_valores = [cell.text.strip() for cell in row.cells]
                fila_str = " ".join(fila_valores)
                cedulas_encontradas = cedula_pattern.findall(fila_str)
                cedula_ref = cedulas_encontradas[0] if cedulas_encontradas else ""
                
                nombre_candidato = ""
                for val in fila_valores:
                    val_clean = val.strip()
                    if not val_clean:
                        continue
                    if val_clean.isdigit() or val_clean == cedula_ref:
                        continue
                    if val_clean.upper() in ["N°", "NO.", "ESTUDIANTE", "NOMBRE", "CEDULA", "CÉDULA"]:
                        continue
                    if len(val_clean) > 4:
                        nombre_candidato = val_clean
                        break
                        
                if nombre_candidato:
                    estudiantes.append({
                        "nombre": nombre_candidato,
                        "cedula": cedula_ref
                    })
                    
    elif ext in [".txt", ".csv"]:
        with open(ruta_archivo, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line_str = line.strip()
                if not line_str:
                    continue
                cedulas_encontradas = cedula_pattern.findall(line_str)
                cedula_ref = cedulas_encontradas[0] if cedulas_encontradas else ""
                
                temp_name = line_str
                if cedula_ref:
                    temp_name = temp_name.replace(cedula_ref, "")
                # Remover números de orden/id iniciales
                temp_name = re.sub(r'^\d+[\s\.\,\-\)\t]+', '', temp_name).strip()
                
                temp_upper = temp_name.upper()
                if temp_upper in ["ESTUDIANTE", "NOMBRE", "CEDULA", "CÉDULA", "LISTADO", "REGISTRO", "NOMBRE Y APELLIDO", "NOMBRES Y APELLIDOS"]:
                    continue
                if temp_upper.startswith("NOMBRE ") or temp_upper.startswith("CEDULA ") or temp_upper.startswith("ESTUDIANTE "):
                    continue
                    
                if len(temp_name) > 4:
                    estudiantes.append({
                        "nombre": temp_name,
                        "cedula": cedula_ref
                    })

    # Eliminar duplicados manteniendo orden
    vistos = set()
    estudiantes_unicos = []
    for est in estudiantes:
        key = (est["nombre"].upper(), est["cedula"])
        if key not in vistos:
            vistos.add(key)
            estudiantes_unicos.append(est)
            
    return estudiantes_unicos


class ImpresionFrame(ctk.CTkFrame):
    def __init__(self, master, engine, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.engine = engine

        # Paleta de colores
        self.C = {
            "azul_osc": "#1E293B",
            "azul_med": "#3B82F6",
            "verde": "#10B981",
            "card_bg": "#FFFFFF" if ctk.get_appearance_mode() == "Light" else "#1E293B",
            "texto": "#0F172A" if ctk.get_appearance_mode() == "Light" else "#F8FAFC",
            "texto_sec": "#64748B" if ctk.get_appearance_mode() == "Light" else "#94A3B8"
        }

        # Tabview principal
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=20, pady=10)

        self.tab_impresion = self.tabview.add("🖨️ Centro de Impresión y Exportación")
        self.tab_importacion = self.tabview.add("📥 Importador Masivo de Estudiantes")

        self.crear_interfaz_impresion()
        self.crear_interfaz_importacion()

    def crear_interfaz_impresion(self):
        # Título Principal
        title_lbl = ctk.CTkLabel(
            self.tab_impresion, text="🖨️ Centro de Impresión y Exportación",
            font=("Outfit", 24, "bold"), text_color=self.C["texto"]
        )
        title_lbl.pack(anchor="w", padx=30, pady=(15, 5))

        desc_lbl = ctk.CTkLabel(
            self.tab_impresion, text="Seleccione qué reporte desea visualizar, imprimir o abrir directamente en Microsoft Excel.",
            font=("Inter", 13), text_color=self.C["texto_sec"]
        )
        desc_lbl.pack(anchor="w", padx=30, pady=(0, 15))

        # Contenedor Grid (2 Columnas)
        grid_frame = ctk.CTkFrame(self.tab_impresion, fg_color="transparent")
        grid_frame.pack(fill="both", expand=True, padx=30, pady=5)
        grid_frame.columnconfigure(0, weight=1, uniform="grid")
        grid_frame.columnconfigure(1, weight=1, uniform="grid")
        grid_frame.rowconfigure(0, weight=1)

        # Panel Izquierdo: Configuración
        config_card = ctk.CTkFrame(
            grid_frame, fg_color=self.C["card_bg"], corner_radius=15, 
            border_width=1, border_color="#E2E8F0" if ctk.get_appearance_mode() == "Light" else "#334155"
        )
        config_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=10)

        # Panel Derecho: Acciones y Resumen
        actions_card = ctk.CTkFrame(
            grid_frame, fg_color=self.C["card_bg"], corner_radius=15, 
            border_width=1, border_color="#E2E8F0" if ctk.get_appearance_mode() == "Light" else "#334155"
        )
        actions_card.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=10)

        # --- PANEL IZQUIERDO: CONFIGURACIÓN ---
        ctk.CTkLabel(
            config_card, text="⚙️ Parámetros del Reporte", 
            font=("Outfit", 16, "bold"), text_color=self.C["texto"]
        ).pack(anchor="w", padx=25, pady=(20, 15))

        # 1. Tipo de Reporte
        ctk.CTkLabel(
            config_card, text="1. Seleccione el Tipo de Reporte:", 
            font=("Inter", 12, "bold"), text_color=self.C["texto"]
        ).pack(anchor="w", padx=25, pady=(5, 5))
        
        self.report_type_var = ctk.StringVar(value="Portada")
        self.report_types = [
            ("📋 Portada Oficial", "Portada"),
            ("📄 Carátula del Registro", "Caratula"),
            ("🕐 Horario Escolar", "Horarios"),
            ("📊 Planilla de Calificaciones", "Planilla"),
            ("📅 Asistencia de Estudiantes", "Asistencia"),
            ("📝 Resumen de Calificaciones", "Resumen"),
            ("📚 Libro de Registro Completo", "Libro Completo"),
            ("📝 Plantilla en Blanco (Calificar a Mano)", "Plantilla Blanco")
        ]

        self.report_menu = ctk.CTkOptionMenu(
            config_card, values=[t[0] for t in self.report_types],
            command=self._on_report_type_change,
            fg_color="#3B82F6", button_color="#2563EB", button_hover_color="#1D4ED8"
        )
        self.report_menu.pack(fill="x", padx=25, pady=(0, 15))

        # 2. Grado (Condicional)
        self.grado_label = ctk.CTkLabel(
            config_card, text="2. Seleccione el Grado:", 
            font=("Inter", 12, "bold"), text_color=self.C["texto"]
        )
        self.grado_label.pack(anchor="w", padx=25, pady=(5, 5))
        
        self.grados_activos = self.engine.obtener_grados_activos()
        self.grado_menu = ctk.CTkOptionMenu(
            config_card, values=self.grados_activos if self.grados_activos else ["No hay grados"],
            command=self._on_grado_change
        )
        self.grado_menu.pack(fill="x", padx=25, pady=(0, 15))

        # 3. Asignatura (Condicional)
        self.materia_label = ctk.CTkLabel(
            config_card, text="3. Seleccione la Asignatura:", 
            font=("Inter", 12, "bold"), text_color=self.C["texto"]
        )
        self.materia_label.pack(anchor="w", padx=25, pady=(5, 5))

        self.materia_menu = ctk.CTkOptionMenu(config_card, values=["General"])
        self.materia_menu.pack(fill="x", padx=25, pady=(0, 15))

        # 4. Periodo (Condicional)
        self.periodo_label = ctk.CTkLabel(
            config_card, text="4. Seleccione el Periodo:", 
            font=("Inter", 12, "bold"), text_color=self.C["texto"]
        )
        self.periodo_label.pack(anchor="w", padx=25, pady=(5, 5))
        self.periodo_menu = ctk.CTkOptionMenu(
            config_card, values=["Trimestre 1", "Trimestre 2", "Trimestre 3", "Anual"],
            command=self._actualizar_info
        )
        self.periodo_menu.pack(fill="x", padx=25, pady=(0, 20))

        # --- PANEL DERECHO: ACCIONES ---
        ctk.CTkLabel(
            actions_card, text="🚀 Acciones de Impresión", 
            font=("Outfit", 16, "bold"), text_color=self.C["texto"]
        ).pack(anchor="w", padx=25, pady=(20, 15))

        # Info Box
        info_frame = ctk.CTkFrame(
            actions_card, fg_color="#EFF6FF" if ctk.get_appearance_mode() == "Light" else "#1E293B", 
            corner_radius=10
        )
        info_frame.pack(fill="x", padx=25, pady=(0, 25))
        
        self.info_lbl = ctk.CTkLabel(
            info_frame, text="Se enviará a imprimir la hoja correspondiente al reporte seleccionado.",
            font=("Inter", 12), text_color="#1E40AF" if ctk.get_appearance_mode() == "Light" else "#93C5FD",
            wraplength=250, justify="left"
        )
        self.info_lbl.pack(padx=15, pady=15)

        # Botones de Acción
        self.btn_open = ctk.CTkButton(
            actions_card, text="📂 Abrir en Microsoft Excel", font=("Inter", 13, "bold"),
            fg_color="#3B82F6", hover_color="#2563EB", text_color="#FFFFFF",
            height=45, command=self._abrir_excel
        )
        self.btn_open.pack(fill="x", padx=25, pady=8)

        self.btn_print = ctk.CTkButton(
            actions_card, text="🖨️ Enviar a Impresora", font=("Inter", 13, "bold"),
            fg_color="#10B981", hover_color="#059669", text_color="#FFFFFF",
            height=45, command=self._imprimir_directo
        )
        self.btn_print.pack(fill="x", padx=25, pady=8)

        self._on_report_type_change()

    def crear_interfaz_importacion(self):
        scroll_imp = ctk.CTkScrollableFrame(self.tab_importacion, fg_color="transparent")
        scroll_imp.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(
            scroll_imp, text="📥 Carga Masiva de Roster de Estudiantes",
            font=("Outfit", 20, "bold"), text_color=self.C["texto"]
        ).pack(anchor="w", padx=20, pady=(10, 5))

        ctk.CTkLabel(
            scroll_imp, text="Importe su lista de alumnos directamente desde un archivo de Excel (.xlsx), Word (.docx) o Texto (.txt). El sistema detectará automáticamente los nombres y cédulas panameñas.",
            font=("Inter", 12), text_color=self.C["texto_sec"], wraplength=700, justify="left"
        ).pack(anchor="w", padx=20, pady=(0, 20))

        ctrl_frame = ctk.CTkFrame(scroll_imp, fg_color=self.C["card_bg"], corner_radius=12)
        ctrl_frame.pack(fill="x", padx=20, pady=10, ipady=15)

        ctk.CTkLabel(
            ctrl_frame, text="1. Seleccione el Grado Destino:",
            font=("Inter", 12, "bold"), text_color=self.C["texto"]
        ).pack(anchor="w", padx=20, pady=(15, 5))

        self.import_grado_menu = ctk.CTkOptionMenu(
            ctrl_frame, values=self.grados_activos if self.grados_activos else ["No hay grados"],
            width=250
        )
        self.import_grado_menu.pack(anchor="w", padx=20, pady=(0, 15))

        ctk.CTkLabel(
            ctrl_frame, text="2. Cargue el archivo con el listado:",
            font=("Inter", 12, "bold"), text_color=self.C["texto"]
        ).pack(anchor="w", padx=20, pady=(5, 5))

        self.btn_select_file = ctk.CTkButton(
            ctrl_frame, text="📁 Seleccionar Archivo (Excel / Word / TXT)",
            font=("Inter", 12, "bold"), fg_color="#3B82F6", hover_color="#2563EB",
            command=self._cargar_archivo_importacion,
            width=350, height=40
        )
        self.btn_select_file.pack(anchor="w", padx=20, pady=(0, 15))

        self.lbl_archivo_cargado = ctk.CTkLabel(
            ctrl_frame, text="Ningún archivo seleccionado.",
            font=("Inter", 11, "italic"), text_color=self.C["texto_sec"]
        )
        self.lbl_archivo_cargado.pack(anchor="w", padx=20, pady=(0, 5))

        self.preview_frame = ctk.CTkFrame(scroll_imp, fg_color="transparent")
        self.preview_frame.pack(fill="both", expand=True, padx=20, pady=10)

        self.importar_estudiantes_detectados = []
        self.checklist_vars = {}

    def _cargar_archivo_importacion(self):
        tipos = [
            ("Archivos Soportados", "*.xlsx;*.docx;*.txt;*.csv"),
            ("Microsoft Excel", "*.xlsx"),
            ("Microsoft Word", "*.docx"),
            ("Archivos de Texto", "*.txt;*.csv"),
            ("Todos los archivos", "*.*")
        ]
        ruta = filedialog.askopenfilename(title="Seleccionar Listado de Estudiantes", filetypes=tipos)
        if not ruta:
            return

        self.lbl_archivo_cargado.configure(text=f"Archivo: {os.path.basename(ruta)}")

        try:
            estudiantes = parsear_archivo_estudiantes(ruta)
            self.importar_estudiantes_detectados = estudiantes
            self._actualizar_preview_importacion()
        except PermissionError as pe:
            messagebox.showerror("Seguridad", str(pe))
            self.lbl_archivo_cargado.configure(text="Ningún archivo seleccionado (Rechazado).")
            self._limpiar_preview()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo: {e}")
            self._limpiar_preview()

    def _actualizar_preview_importacion(self):
        self._limpiar_preview()
        
        if not self.importar_estudiantes_detectados:
            ctk.CTkLabel(
                self.preview_frame, 
                text="⚠️ No se detectaron estudiantes en el archivo. Verifique el formato e inténtelo de nuevo.",
                font=("Inter", 12, "bold"), text_color="#EF4444"
            ).pack(pady=20)
            return

        header_row = ctk.CTkFrame(self.preview_frame, fg_color=self.C["azul_osc"], corner_radius=6)
        header_row.pack(fill="x", pady=(0, 5))
        
        ctk.CTkLabel(header_row, text="Importar", font=("Inter", 11, "bold"), text_color="#FFFFFF", width=80).pack(side="left", padx=10)
        ctk.CTkLabel(header_row, text="Estudiante", font=("Inter", 11, "bold"), text_color="#FFFFFF", width=250, anchor="w").pack(side="left", padx=10)
        ctk.CTkLabel(header_row, text="Cédula (Editable)", font=("Inter", 11, "bold"), text_color="#FFFFFF", width=180, anchor="w").pack(side="left", padx=10)

        self.checklist_vars = {}
        
        list_scroll = ctk.CTkScrollableFrame(self.preview_frame, height=250, fg_color=self.C["card_bg"], corner_radius=8)
        list_scroll.pack(fill="both", expand=True)

        for idx, est in enumerate(self.importar_estudiantes_detectados):
            row_frame = ctk.CTkFrame(list_scroll, fg_color="transparent")
            row_frame.pack(fill="x", pady=2)

            var = ctk.BooleanVar(value=True)
            self.checklist_vars[idx] = var
            
            cb = ctk.CTkCheckBox(row_frame, text="", variable=var, width=80)
            cb.pack(side="left", padx=10)

            lbl_nom = ctk.CTkLabel(row_frame, text=est["nombre"], width=250, anchor="w", text_color=self.C["texto"])
            lbl_nom.pack(side="left", padx=10)

            entry_ced_var = ctk.StringVar(value=est["cedula"])
            entry_ced = ctk.CTkEntry(row_frame, textvariable=entry_ced_var, width=180, font=("Inter", 11))
            entry_ced.pack(side="left", padx=10)
            est["cedula_var"] = entry_ced_var

        btn_frame = ctk.CTkFrame(self.preview_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=15)

        self.btn_importar_confirm = ctk.CTkButton(
            btn_frame, text=f"📥 Confirmar Importación",
            font=("Inter", 13, "bold"), fg_color="#10B981", hover_color="#059669",
            command=self._confirmar_importacion_estudiantes
        )
        self.btn_importar_confirm.pack(side="right")
        
        def toggle_all():
            val = self.var_toggle_all.get()
            for v in self.checklist_vars.values():
                v.set(val)
                
        self.var_toggle_all = ctk.BooleanVar(value=True)
        cb_all = ctk.CTkCheckBox(btn_frame, text="Seleccionar Todos", variable=self.var_toggle_all, command=toggle_all)
        cb_all.pack(side="left", padx=10)

    def _limpiar_preview(self):
        for w in self.preview_frame.winfo_children():
            w.destroy()

    def _confirmar_importacion_estudiantes(self):
        grado = self.import_grado_menu.get()
        if not grado or grado == "No hay grados":
            messagebox.showwarning("Atención", "Seleccione un grado destino válido.")
            return

        seleccionados = []
        for idx, est in enumerate(self.importar_estudiantes_detectados):
            if self.checklist_vars[idx].get():
                ced_val = est.get("cedula_var").get().strip() if "cedula_var" in est else est.get("cedula", "")
                seleccionados.append({
                    "nombre": est["nombre"],
                    "cedula": ced_val
                })

        if not seleccionados:
            messagebox.showwarning("Atención", "No ha seleccionado ningún estudiante para importar.")
            return

        confirm = messagebox.askyesno(
            "Confirmar Importación",
            f"¿Desea registrar los {len(seleccionados)} estudiantes seleccionados en el grado {grado}?\n\n"
            "Esto los insertará directamente en la base de datos y la libreta de Excel."
        )
        if not confirm:
            return

        exito_count = 0
        error_count = 0
        for est in seleccionados:
            res = self.engine.agregar_estudiante(grado, est["nombre"], est["cedula"])
            if res:
                exito_count += 1
            else:
                error_count += 1

        messagebox.showinfo(
            "✓ Importación Finalizada",
            f"Proceso completado:\n"
            f"• Estudiantes registrados con éxito: {exito_count}\n"
            f"• No importados (duplicados o límite excedido): {error_count}"
        )
        self._limpiar_preview()
        self.lbl_archivo_cargado.configure(text="Ningún archivo seleccionado.")

    def _on_report_type_change(self, *args):
        label = self.report_menu.get()
        tipo = "Portada"
        for t in self.report_types:
            if t[0] == label:
                tipo = t[1]
                break
        
        self.report_type_var.set(tipo)

        if tipo in ["Portada", "Caratula", "Horarios", "Libro Completo"]:
            self.grado_label.pack_forget()
            self.grado_menu.pack_forget()
            self.materia_label.pack_forget()
            self.materia_menu.pack_forget()
            self.periodo_label.pack_forget()
            self.periodo_menu.pack_forget()
        elif tipo == "Asistencia":
            self.grado_label.pack(anchor="w", padx=25, pady=(5, 5))
            self.grado_menu.pack(fill="x", padx=25, pady=(0, 15))
            self.materia_label.pack_forget()
            self.materia_menu.pack_forget()
            self.periodo_label.pack(anchor="w", padx=25, pady=(5, 5))
            self.periodo_menu.pack(fill="x", padx=25, pady=(0, 20))
            self.periodo_menu.configure(values=["Trimestre 1", "Trimestre 2", "Trimestre 3"])
            self.periodo_menu.set("Trimestre 1")
        elif tipo == "Resumen":
            self.grado_label.pack(anchor="w", padx=25, pady=(5, 5))
            self.grado_menu.pack(fill="x", padx=25, pady=(0, 15))
            self.materia_label.pack_forget()
            self.materia_menu.pack_forget()
            self.periodo_label.pack(anchor="w", padx=25, pady=(5, 5))
            self.periodo_menu.pack(fill="x", padx=25, pady=(0, 20))
            self.periodo_menu.configure(values=["Trimestre 1", "Anual"])
            self.periodo_menu.set("Trimestre 1")
        elif tipo == "Planilla":
            self.grado_label.pack(anchor="w", padx=25, pady=(5, 5))
            self.grado_menu.pack(fill="x", padx=25, pady=(0, 15))
            self.materia_label.pack(anchor="w", padx=25, pady=(5, 5))
            self.materia_menu.pack(fill="x", padx=25, pady=(0, 15))
            self.periodo_label.pack_forget()
            self.periodo_menu.pack_forget()
            self._on_grado_change()
        elif tipo == "Plantilla Blanco":
            self.grado_label.pack(anchor="w", padx=25, pady=(5, 5))
            self.grado_menu.pack(fill="x", padx=25, pady=(0, 15))
            self.materia_label.pack_forget()
            self.materia_menu.pack_forget()
            self.periodo_label.pack_forget()
            self.periodo_menu.pack_forget()

        self._actualizar_info()

    def _on_grado_change(self, *args):
        tipo = self.report_type_var.get()
        if tipo == "Planilla":
            grado = self.grado_menu.get()
            materias = self.engine.obtener_materias_por_grado(grado)
            self.materia_menu.configure(values=materias if materias else ["General"])
            if materias:
                self.materia_menu.set(materias[0])
            else:
                self.materia_menu.set("General")
        self._actualizar_info()

    def _actualizar_info(self, *args):
        tipo = self.report_type_var.get()
        if tipo == "Libro Completo":
            msg = "Se abrirá o imprimirá el libro de calificaciones completo con todas sus hojas."
        elif tipo in ["Portada", "Caratula", "Horarios"]:
            msg = f"Se procesará la hoja general de '{tipo}' del registro."
        elif tipo == "Plantilla Blanco":
            grado = self.grado_menu.get()
            msg = f"Se generará una plantilla en blanco con los nombres del grado {grado} para calificar a mano."
        else:
            grado = self.grado_menu.get()
            periodo = self.periodo_menu.get() if tipo != "Planilla" else "N/A"
            msg = f"Reporte: {tipo}\nGrado: {grado}\n"
            if tipo == "Planilla":
                msg += f"Asignatura: {self.materia_menu.get()}"
            else:
                msg += f"Periodo: {periodo}"
        self.info_lbl.configure(text=msg)

    def _resolver_nombre_hoja(self):
        tipo = self.report_type_var.get()
        if tipo in ["Portada", "Caratula", "Horarios"]:
            return tipo

        if tipo == "Libro Completo":
            return None

        grado = self.grado_menu.get()
        materia = self.materia_menu.get() if tipo == "Planilla" else None
        
        if not os.path.exists(self.engine.ruta):
            return None
        
        wb = openpyxl.load_workbook(self.engine.ruta, read_only=True)
        from rdprint import encontrar_hoja_impresion
        nombre_hoja = encontrar_hoja_impresion(wb, tipo, grado, materia)
        wb.close()
        return nombre_hoja

    def _abrir_excel(self):
        tipo = self.report_type_var.get()
        if tipo == "Plantilla Blanco":
            self._generar_y_abrir_plantilla_blanco()
            return

        hoja = self._resolver_nombre_hoja()
        ok, msg = abrir_para_imprimir(hoja)
        if ok:
            messagebox.showinfo("✓ Éxito", msg)
        else:
            messagebox.showerror("Error", msg)

    def _imprimir_directo(self):
        tipo = self.report_type_var.get()
        if tipo == "Plantilla Blanco":
            self._generar_y_abrir_plantilla_blanco()
            return

        hoja = self._resolver_nombre_hoja()
        if not hoja:
            if self.report_type_var.get() == "Libro Completo":
                ok, msg = abrir_para_imprimir(None)
                if ok:
                    messagebox.showinfo("✓ Éxito", msg)
                else:
                    messagebox.showerror("Error", msg)
                return
            messagebox.showerror("Error", "No se encontró la hoja correspondiente en el archivo Excel.")
            return

        ok, msg = imprimir_hoja_directo(hoja)
        if ok:
            messagebox.showinfo("✓ Enviado a Impresora", msg)
        else:
            messagebox.showerror("Error", msg)

    def _generar_y_abrir_plantilla_blanco(self):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import tempfile
        
        grado = self.grado_menu.get()
        estudiantes = self.engine.obtener_estudiantes_completos(grado)
        if not estudiantes:
            messagebox.showwarning("Atención", "No se encontraron estudiantes para el grado seleccionado.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Auxiliar"
        
        ws.views.sheetView[0].showGridLines = True
        
        font_title = Font(name="Calibri", size=16, bold=True, color="1B365D")
        font_subtitle = Font(name="Calibri", size=11, bold=True, color="475569")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Calibri", size=11)
        font_bold = Font(name="Calibri", size=11, bold=True)
        
        fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws.merge_cells("A1:N1")
        ws["A1"] = "MINISTERIO DE EDUCACIÓN — REGISTRO AUXILIAR"
        ws["A1"].font = font_title
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:N2")
        ws["A2"] = f"Grado: {grado}  |  Trimestre: _______  |  Asignatura: ___________________________"
        ws["A2"].font = font_subtitle
        ws["A2"].alignment = Alignment(horizontal="center")
        
        headers = ["N°", "Estudiante (Apellido, Nombre)", "C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9", "C10", "Promedio", "Observaciones"]
        ws.row_dimensions[4].height = 28
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = text
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        current_row = 5
        for idx, est in enumerate(estudiantes, 1):
            ws.row_dimensions[current_row].height = 20
            
            cell_num = ws.cell(row=current_row, column=1, value=idx)
            cell_num.font = font_bold
            cell_num.alignment = Alignment(horizontal="center")
            cell_num.border = thin_border
            
            cell_name = ws.cell(row=current_row, column=2, value=est["nombre"])
            cell_name.font = font_body
            cell_name.alignment = Alignment(horizontal="left", vertical="center")
            cell_name.border = thin_border
            
            if idx % 2 == 0:
                cell_num.fill = fill_zebra
                cell_name.fill = fill_zebra
                
            for col_idx in range(3, 15):
                cell_g = ws.cell(row=current_row, column=col_idx)
                cell_g.border = thin_border
                if idx % 2 == 0:
                    cell_g.fill = fill_zebra
                    
            current_row += 1
            
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        for col_idx in range(3, 13):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 6
        ws.column_dimensions["M"].width = 10
        ws.column_dimensions["N"].width = 25
        
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"Plantilla_Limpia_{grado.replace('°','')}.xlsx")
        try:
            wb.save(file_path)
            if os.name == 'nt':
                os.startfile(file_path)
            else:
                import subprocess
                subprocess.call(['open', file_path])
            messagebox.showinfo("✓ Éxito", f"Se generó la plantilla en blanco con éxito para el grado {grado}.\n\nSe abrirá en Excel para que la pueda imprimir.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar o abrir el archivo: {e}")
