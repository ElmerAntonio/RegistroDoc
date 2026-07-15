import os
import re
import logging
import threading
from rdsecurity import validar_nota_meduca, sanitizar_entrada
import openpyxl
from openpyxl.styles import Alignment, Font
from utils.cleaner import ExcelCleaner

logger = logging.getLogger("rddata")

def _contar_estudiantes_excel(ruta_xls, modalidad):
    """Cuenta el número de estudiantes en un archivo Excel de forma rápida y de sólo lectura."""
    try:
        wb = openpyxl.load_workbook(ruta_xls, data_only=True)
        nombre_maestro = None
        for sheet in wb.sheetnames:
            if "MAESTRO" in sheet.upper():
                nombre_maestro = sheet
                break
        if not nombre_maestro:
            wb.close()
            return 0
        ws = wb[nombre_maestro]
        cant = 0
        columnas = [2, 4, 6] if modalidad == "premedia" else [2]
        for col in columnas:
            for r in range(5, 46):
                if ws.cell(row=r, column=col).value:
                    cant += 1
        wb.close()
        return cant
    except Exception:
        return 0

class DataEngine:
    def __init__(self, ruta_excel, modalidad="premedia"):
        self.modalidad = modalidad.lower()
        
        nombre_base = os.path.basename(ruta_excel)
        if nombre_base in ("Registro_Premedia.xlsx", "Registro_Primaria.xlsx"):
            # Redireccionar el archivo de trabajo a un directorio temporal privado de AppData
            import shutil
            from rdsql import obtener_dir_datos_usuario
            user_dir = obtener_dir_datos_usuario()
            active_excel = os.path.join(user_dir, "temp", nombre_base)
            
            # Copiar si no existe, o si el de la raíz tiene más estudiantes, o es más reciente y no vacío
            if os.path.exists(ruta_excel):
                if not os.path.exists(active_excel):
                    shutil.copy2(ruta_excel, active_excel)
                else:
                    cant_orig = _contar_estudiantes_excel(ruta_excel, self.modalidad)
                    cant_temp = _contar_estudiantes_excel(active_excel, self.modalidad)
                    if cant_orig > cant_temp:
                        shutil.copy2(ruta_excel, active_excel)
                    else:
                        mtime_origen = os.path.getmtime(ruta_excel)
                        mtime_destino = os.path.getmtime(active_excel)
                        if mtime_origen > mtime_destino + 2.0 and cant_orig > 0:
                            shutil.copy2(ruta_excel, active_excel)
            
            self.ruta = active_excel
        else:
            self.ruta = ruta_excel
        self.cleaner = ExcelCleaner()
        self.fila_desc = 39 if self.modalidad == "primaria" else 42
        self._wb_cache = None
        self._wb_formulas_cache = None
        self._last_mtime = 0.0
        
        # Conexión a SQLite cifrada
        from rdsql import SQLDatabaseManager
        self.db_manager = SQLDatabaseManager()
        self.db_conn = self.db_manager.conectar()

        # --- Debounce de guardado cifrado ---
        # El commit() a SQLite es INMEDIATO (datos nunca se pierden en memoria).
        # El cifrado al .enc ocurre máximo cada 5 segundos en un hilo de fondo,
        # evitando congelar la UI cuando el docente guarda rápidamente.
        self._save_timer = None
        self._save_timer_lock = threading.Lock()
        self._SAVE_DEBOUNCE_SECS = 5.0
        
        self._cargar_en_memoria()
        self.sincronizar_excel_a_sql()

    # ─── Guardado diferido (debounce) ────────────────────────────────────────
    def _schedule_save(self):
        """Programa un guardado cifrado en background con debounce de 5 segundos.
        El commit a SQLite ya ocurrió; esto solo persiste el .enc en disco.
        Si se llama varias veces seguidas, solo ejecuta UNA escritura al final.
        """
        with self._save_timer_lock:
            if self._save_timer is not None:
                self._save_timer.cancel()  # Cancelar el timer anterior
            self._save_timer = threading.Timer(self._SAVE_DEBOUNCE_SECS, self._flush_save)
            self._save_timer.daemon = True
            self._save_timer.start()

    def _flush_save(self):
        """Ejecuta el guardado cifrado real en el hilo del timer (fondo)."""
        with self._save_timer_lock:
            self._save_timer = None
        try:
            self.db_manager.guardar_cifrado()
        except Exception as e:
            logger.warning(f"[debounce] Error en guardado cifrado: {e}")

    def flush_save_sync(self):
        """Fuerza un guardado cifrado inmediato y sincrono (usar solo al cerrar la app)."""
        with self._save_timer_lock:
            if self._save_timer is not None:
                self._save_timer.cancel()
                self._save_timer = None
        self.db_manager.guardar_cifrado()


    def sincronizar_excel_a_sql(self, force_sync=False):
        """Migra de forma unidireccional y robusta los datos desde Excel a SQLite."""
        if not os.path.exists(self.ruta) and not self._wb_cache:
            return

        cursor = self.db_conn.cursor()

        # Eliminar registros huérfanos del bug de colisión de IDs (IDs menores a 100)
        try:
            cursor.execute("DELETE FROM notas WHERE CAST(estudiante_id AS INTEGER) < 100;")
            cursor.execute("DELETE FROM asistencia WHERE CAST(estudiante_id AS INTEGER) < 100;")
            cursor.execute("DELETE FROM estudiantes WHERE CAST(id AS INTEGER) < 100;")
            self.db_conn.commit()
        except Exception as e:
            logger.error(f"Error limpiando IDs de estudiantes inválidos: {e}")

        # Asegurar carga
        if not self._wb_cache:
            try:
                self._cargar_en_memoria()
            except Exception:
                return

        # 1. Cargar metadatos / Datos Generales
        generales = self.obtener_datos_generales()
        if generales:
            for k, v in generales.items():
                cursor.execute(
                    "INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?);",
                    (k, str(v))
                )

        # Sincronizar metadatos extra
        try:
            # Sincronizar ruta_excel
            cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('ruta_excel', ?);", (self.ruta,))
            
            # Sincronizar trimestre_activo
            from utils.date_helpers import obtener_trimestre_actual
            trimestre = obtener_trimestre_actual()
            cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('trimestre_activo', ?);", (trimestre,))
            
            # Sincronizar otros desde config.enc
            from rdsecurity import cargar_config_segura
            cfg = cargar_config_segura({})
            for k in ["modalidad", "tema", "logo_escuela_path"]:
                if k in cfg:
                    cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?);", (k, str(cfg[k])))
                elif k == "modalidad":
                    cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('modalidad', ?);", (self.modalidad,))
            self.db_conn.commit()
        except Exception as e:
            logger.error(f"Error guardando metadatos extra en SQLite: {e}")

        # 2. Obtener grados del archivo Excel para limpieza y sincronización
        grados_excel = []
        wb_check = self._wb_cache
        should_close_check = False
        if not wb_check:
            try:
                import openpyxl
                wb_check = openpyxl.load_workbook(self.ruta, read_only=True)
                should_close_check = True
            except Exception:
                pass
        if wb_check:
            for s in wb_check.sheetnames:
                if s.upper().startswith("RESUMEN_"):
                    grados_excel.append(s.split("_", 1)[1].strip())
            if should_close_check:
                wb_check.close()

        # Limpiar grados obsoletos solo si la BD está completamente vacía (primera sincronización)
        # Cuando ya hay datos, SQL es la fuente de verdad y no eliminamos grados basados en el Excel
        cursor.execute("SELECT COUNT(*) FROM notas;")
        bd_vacia = cursor.fetchone()[0] == 0
        if grados_excel and bd_vacia:
            try:
                placeholders = ",".join("?" for _ in grados_excel)
                cursor.execute(f"SELECT id FROM grados WHERE modalidad = ? AND nombre NOT IN ({placeholders});", (self.modalidad, *grados_excel))
                obsolete_ids = [r[0] for r in cursor.fetchall()]
                if obsolete_ids:
                    for obs_id in obsolete_ids:
                        cursor.execute("DELETE FROM estudiantes WHERE grado_id = ?;", (obs_id,))
                        cursor.execute("DELETE FROM materias WHERE grado_id = ?;", (obs_id,))
                        cursor.execute("DELETE FROM grados WHERE id = ?;", (obs_id,))
                    self.db_conn.commit()
            except Exception as e:
                logger.error(f"Error limpiando grados obsoletos de SQLite: {e}")


        # Registrar los nuevos grados de Excel en SQLite antes de consultar
        if grados_excel:
            try:
                for g in grados_excel:
                    cursor.execute("INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES (?, 'A', ?);", (g, self.modalidad))
                self.db_conn.commit()
            except Exception as e:
                logger.error(f"Error registrando grados iniciales: {e}")

        # Ahora sí, obtener la lista final de grados activos de SQLite
        grados = self.obtener_grados_activos()

        # Verificar si ya está poblado con estudiantes, notas y asistencia
        cursor.execute("SELECT COUNT(*) FROM estudiantes;")
        ya_poblado = cursor.fetchone()[0] > 0

        cursor.execute("SELECT COUNT(*) FROM notas;")
        tiene_notas = cursor.fetchone()[0] > 0

        cursor.execute("SELECT COUNT(*) FROM asistencia;")
        tiene_asistencia = cursor.fetchone()[0] > 0

        if ya_poblado and not force_sync:
            # Sincronización parcial en cada inicio: actualiza nombres, cédulas y sexos.
            # NO borra notas ni asistencia (SQL es la fuente de verdad post-primera-sync).
            for g in grados:
                cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (g, self.modalidad))
                grado_row = cursor.fetchone()
                if grado_row:
                    grado_id = grado_row[0]
                    estudiantes = self._obtener_estudiantes_desde_excel(g, wb=self._wb_cache)
                    active_ids = {str(est["id"]) for est in estudiantes}

                    # Cargar IDs de estudiantes registrados en la DB para este grado
                    cursor.execute("SELECT id FROM estudiantes WHERE grado_id = ? AND estado = 'Activo';", (grado_id,))
                    db_ids = {row[0] for row in cursor.fetchall()}

                    # Solo marcar como Retirado si el Excel TÍene datos para este grado
                    # y el alumno no aparece en él. Si Excel está vacío para el grado,
                    # SQL es la fuente de verdad — no tocamos el estado.
                    if active_ids:  # Excel sí tiene alumnos para este grado
                        for old_id in db_ids - active_ids:
                            cursor.execute("UPDATE estudiantes SET estado = 'Retirado' WHERE id = ? AND grado_id = ?;", (old_id, grado_id))

                    for est in estudiantes:
                        nombre_cifrado = self.db_manager.encriptar_campo(est["nombre"])
                        cedula_cifrada = self.db_manager.encriptar_campo(est["cedula"])
                        sexo_cifrado = self.db_manager.encriptar_campo(est["sexo"])
                        
                        # Evitar sobrescribir el estado de estudiantes retirados
                        cursor.execute("SELECT id FROM estudiantes WHERE id = ?;", (str(est["id"]),))
                        if cursor.fetchone():
                            cursor.execute(
                                "UPDATE estudiantes SET nombre = ?, cedula = ?, grado_id = ? WHERE id = ?;",
                                (nombre_cifrado, cedula_cifrada, grado_id, str(est["id"]))
                            )
                        else:
                            cursor.execute(
                                "INSERT INTO estudiantes (id, nombre, cedula, sexo, grado_id, estado) VALUES (?, ?, ?, ?, ?, 'Activo');",
                                (str(est["id"]), nombre_cifrado, cedula_cifrada, sexo_cifrado, grado_id)
                            )

                    # Sincronización incremental de notas desde Excel para este grado (Autocuración)
                    materias = self.obtener_materias_por_grado(g, wb=self._wb_cache)
                    for mat in materias:
                        if mat == "Sin materias registradas":
                            continue
                        cursor.execute(
                            "INSERT OR IGNORE INTO materias (nombre, grado_id) VALUES (?, ?);",
                            (mat, grado_id)
                        )
                        self.db_conn.commit()
                        
                        cursor.execute(
                            "SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;",
                            (mat, grado_id)
                        )
                        materia_id = cursor.fetchone()[0]
                        
                        for trim in (1, 2, 3):
                            trim_str = f"Trimestre {trim}"
                            for tipo in ("Diaria / Parcial", "Apreciación", "Examen"):
                                descripciones = self.obtener_descripciones_notas(g, mat, trim_str, tipo, wb=self._wb_cache)
                                if not descripciones:
                                    continue
                                
                                wb = self._wb_cache
                                nombre_hoja = self._encontrar_hoja_prom(wb, g, mat)
                                if not nombre_hoja:
                                    continue
                                ws = wb[nombre_hoja]
                                rango = self._obtener_rango_columnas(ws, trim_str, tipo)
                                if rango == (None, None):
                                    continue
                                
                                col_inicio, col_fin = rango
                                for c in range(col_inicio, col_fin + 1):
                                    val_desc = ws.cell(row=self.fila_desc, column=c).value
                                    if not val_desc:
                                        continue
                                    
                                    fecha_val = ws.cell(row=4, column=c).value or "01-01"
                                    desc_limpia = str(val_desc).replace('\n', ' ').strip()
                                    
                                    for est in estudiantes:
                                        fila_excel = 4 + (int(est["id"]) % 100)
                                        val_nota = ws.cell(row=fila_excel, column=c).value
                                        if val_nota is not None:
                                            try:
                                                cursor.execute(
                                                    """SELECT id, valor FROM notas 
                                                    WHERE estudiante_id = ? AND materia_id = ? AND trimestre = ? AND tipo = ? AND descripcion = ?;""",
                                                    (str(est["id"]), materia_id, trim, tipo, desc_limpia)
                                                )
                                                existing_note = cursor.fetchone()
                                                if existing_note:
                                                    note_id, current_val = existing_note
                                                    if current_val != float(val_nota):
                                                        cursor.execute(
                                                            "UPDATE notas SET valor = ? WHERE id = ?;",
                                                            (float(val_nota), note_id)
                                                        )
                                                else:
                                                    cursor.execute(
                                                        """INSERT INTO notas 
                                                        (estudiante_id, materia_id, trimestre, tipo, descripcion, valor, fecha) 
                                                        VALUES (?, ?, ?, ?, ?, ?, ?);""",
                                                        (str(est["id"]), materia_id, trim, tipo, desc_limpia, float(val_nota), str(fecha_val))
                                                    )
                                            except Exception:
                                                pass

                    # Si no tiene asistencia en la BD, importar asistencia desde Excel para este grado (Autocuración)
                    cursor.execute("""
                        SELECT COUNT(*) FROM asistencia 
                        WHERE estudiante_id IN (SELECT id FROM estudiantes WHERE grado_id = ?);
                    """, (grado_id,))
                    tiene_asistencia_grado = cursor.fetchone()[0] > 0
                    
                    if not tiene_asistencia_grado:
                        for trim in (1, 2, 3):
                            trim_str = f"Trimestre {trim}"
                            fechas = self.obtener_fechas_asistencia(g, trim_str, wb=self._wb_cache)
                            for f_val in fechas:
                                asistencia_col = self.buscar_asistencia_existente(g, trim_str, f_val, wb=self._wb_cache)
                                if asistencia_col:
                                    cursor.execute("SELECT id, nombre FROM estudiantes WHERE grado_id = ?;", (grado_id,))
                                    est_map = {}
                                    for r_est in cursor.fetchall():
                                        dec_name = self.db_manager.desencriptar_campo(r_est[1]).strip().lower()
                                        est_map[dec_name] = r_est[0]
                                    
                                    valid_student_ids = set(est_map.values())
                                    dict_asis = asistencia_col.get("asistencia", asistencia_col) if isinstance(asistencia_col, dict) and "asistencia" in asistencia_col else asistencia_col
                                    if isinstance(dict_asis, dict):
                                        for est_key, estado in dict_asis.items():
                                            est_id = None
                                            if str(est_key).isdigit():
                                                est_id = str(grado_id * 100 + int(est_key))
                                            else:
                                                est_nom_clean = str(est_key).strip().lower()
                                                est_id = est_map.get(est_nom_clean)
                                                
                                            if est_id and est_id in valid_student_ids:
                                                est_status = str(estado).strip()
                                                if est_status not in ('.', '-', 'T', 'E'):
                                                    if est_status.lower() in ('p', 'asiste', 'presente', '1', '✓'):
                                                        est_status = '.'
                                                    elif est_status.lower() in ('f', 'falta', 'ausente', '0', 'x'):
                                                        est_status = '-'
                                                    elif est_status.lower() in ('t', 'tardanza', 'tarde'):
                                                        est_status = 'T'
                                                    elif est_status.lower() in ('e', 'excusa', 'justificado'):
                                                        est_status = 'E'
                                                    else:
                                                        est_status = '.'
                                                
                                                cursor.execute(
                                                    """INSERT OR REPLACE INTO asistencia 
                                                    (estudiante_id, fecha, estado, trimestre) VALUES (?, ?, ?, ?);""",
                                                    (est_id, f_val, est_status, trim)
                                                )
            self.db_conn.commit()
            self._schedule_save()
            return


        # Sincronización completa: solo cuando no hay estudiantes en la BD (primera ejecución).
        # IMPORTANTE: NO borrar notas ni asistencia si ya_poblado es True pero tiene_notas es False.
        # Esto evita el ciclo vicioso donde la sync borra notas recién guardadas.
        if not bd_vacia and ya_poblado:
            # Hay estudiantes pero no hay notas: sincronizar metadatos y estudiantes sin tocar datos.
            pass
        elif not ya_poblado:
            # BD realmente vacía: sincronizar todo desde Excel.
            cursor.execute("DELETE FROM notas;")
            cursor.execute("DELETE FROM asistencia;")
            cursor.execute("DELETE FROM horario;")
            self.db_conn.commit()
        for g in grados:
            # Encontrar modalidad activa
            cursor.execute(
                "INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES (?, ?, ?);",
                (g, "A", self.modalidad)
            )
            self.db_conn.commit()
            
            # Obtener el grado_id recién creado/existente
            cursor.execute(
                "SELECT id FROM grados WHERE nombre = ? AND seccion = ? AND modalidad = ?;",
                (g, "A", self.modalidad)
            )
            grado_id = cursor.fetchone()[0]
            
            # 3. Estudiantes (Cifrados)
            estudiantes = self._obtener_estudiantes_desde_excel(g, wb=self._wb_cache)
            for est in estudiantes:
                nombre_cifrado = self.db_manager.encriptar_campo(est["nombre"])
                cedula_cifrada = self.db_manager.encriptar_campo(est["cedula"])
                sexo_cifrado = self.db_manager.encriptar_campo(est["sexo"])
                
                # Evitar sobrescribir el estado de estudiantes retirados
                cursor.execute("SELECT id FROM estudiantes WHERE id = ?;", (str(est["id"]),))
                if cursor.fetchone():
                    cursor.execute(
                        "UPDATE estudiantes SET nombre = ?, cedula = ?, grado_id = ? WHERE id = ?;",
                        (nombre_cifrado, cedula_cifrada, grado_id, str(est["id"]))
                    )
                else:
                    cursor.execute(
                        "INSERT INTO estudiantes (id, nombre, cedula, sexo, grado_id, estado) VALUES (?, ?, ?, ?, ?, 'Activo');",
                        (str(est["id"]), nombre_cifrado, cedula_cifrada, sexo_cifrado, grado_id)
                    )
            
            # 4. Materias
            materias = self.obtener_materias_por_grado(g, wb=self._wb_cache)
            for mat in materias:
                if mat == "Sin materias registradas":
                    continue
                cursor.execute(
                    "INSERT OR IGNORE INTO materias (nombre, grado_id) VALUES (?, ?);",
                    (mat, grado_id)
                )
                self.db_conn.commit()
                
                # Obtener materia_id
                cursor.execute(
                    "SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;",
                    (mat, grado_id)
                )
                materia_id = cursor.fetchone()[0]
                
                # 5. Notas
                for trim in (1, 2, 3):
                    trim_str = f"Trimestre {trim}"
                    for tipo in ("Diaria / Parcial", "Apreciación", "Examen"):
                        descripciones = self.obtener_descripciones_notas(g, mat, trim_str, tipo, wb=self._wb_cache)
                        if not descripciones:
                            continue
                        
                        wb = self._wb_cache
                        nombre_hoja = self._encontrar_hoja_prom(wb, g, mat)
                        if not nombre_hoja:
                            continue
                        ws = wb[nombre_hoja]
                        rango = self._obtener_rango_columnas(ws, trim_str, tipo)
                        if rango == (None, None):
                            continue
                        
                        col_inicio, col_fin = rango
                        for c in range(col_inicio, col_fin + 1):
                            val_desc = ws.cell(row=self.fila_desc, column=c).value
                            if not val_desc:
                                continue
                            
                            # Obtener fecha
                            fecha = ws.cell(row=4, column=c).value or "01-01"
                            desc_limpia = str(val_desc).replace('\n', ' ').strip()
                            
                            for est in estudiantes:
                                fila_excel = 4 + (int(est["id"]) % 100)
                                val_nota = ws.cell(row=fila_excel, column=c).value
                                if val_nota is not None:
                                    try:
                                        cursor.execute(
                                            """INSERT INTO notas 
                                            (estudiante_id, materia_id, trimestre, tipo, descripcion, valor, fecha) 
                                            VALUES (?, ?, ?, ?, ?, ?, ?);""",
                                            (str(est["id"]), materia_id, trim, tipo, desc_limpia, float(val_nota), str(fecha))
                                        )
                                    except Exception:
                                        pass
            
            # 6. Asistencia
            for trim in (1, 2, 3):
                trim_str = f"Trimestre {trim}"
                fechas = self.obtener_fechas_asistencia(g, trim_str, wb=self._wb_cache)
                for f_val in fechas:
                    asistencia_col = self.buscar_asistencia_existente(g, trim_str, f_val, wb=self._wb_cache)
                    if asistencia_col:
                        cursor.execute("SELECT id, nombre FROM estudiantes WHERE grado_id = ?;", (grado_id,))
                        est_map = {}
                        for r_est in cursor.fetchall():
                            dec_name = self.db_manager.desencriptar_campo(r_est[1]).strip().lower()
                            est_map[dec_name] = r_est[0]
                        
                        valid_student_ids = set(est_map.values())

                        # Manejar formato de Excel (diccionario con clave 'asistencia') o formato SQLite (dict directo)
                        dict_asis = asistencia_col.get("asistencia", asistencia_col) if isinstance(asistencia_col, dict) and "asistencia" in asistencia_col else asistencia_col
                        if isinstance(dict_asis, dict):
                            for est_key, estado in dict_asis.items():
                                est_id = None
                                if str(est_key).isdigit():
                                    est_id = str(grado_id * 100 + int(est_key))
                                else:
                                    est_nom_clean = str(est_key).strip().lower()
                                    est_id = est_map.get(est_nom_clean)
                                    
                                if est_id and est_id in valid_student_ids:
                                    est_status = str(estado).strip()
                                    if est_status not in ('.', '-', 'T', 'E'):
                                        if est_status.lower() in ('p', 'asiste', 'presente', '1', '✓'):
                                            est_status = '.'
                                        elif est_status.lower() in ('f', 'falta', 'ausente', '0', 'x'):
                                            est_status = '-'
                                        elif est_status.lower() in ('t', 'tardanza', 'tarde'):
                                            est_status = 'T'
                                        elif est_status.lower() in ('e', 'excusa', 'justificado'):
                                            est_status = 'E'
                                        else:
                                            est_status = '.'
                                    
                                    cursor.execute(
                                        """INSERT OR REPLACE INTO asistencia 
                                        (estudiante_id, fecha, estado, trimestre) VALUES (?, ?, ?, ?);""",
                                        (est_id, f_val, est_status, trim)
                                    )
        
        # 7. Horario — sincronizar Excel → SQL guardando texto libre por día/bloque
        try:
            horario_excel = self.obtener_horario_desde_excel()  # Siempre lee del Excel
            if any(any(v for k, v in s.items() if k != "horas") for s in horario_excel):
                # Solo reemplazar si el Excel tiene datos reales
                cursor.execute("DELETE FROM horario;")
                for idx, slot in enumerate(horario_excel):
                    bloque_hora = slot.get("horas", f"Período {idx+1}")
                    for dia in ("lunes", "martes", "miercoles", "jueves", "viernes"):
                        texto = slot.get(dia, "").strip()
                        cursor.execute(
                            """INSERT INTO horario (bloque_orden, dia, bloque_hora, materia_texto)
                               VALUES (?, ?, ?, ?);""",
                            (idx, dia, bloque_hora, texto)
                        )
        except Exception as e:
            logger.warning(f"[sync] Error sincronizando horario: {e}")

        self.db_conn.commit()
        self._schedule_save()



    def _safe_set_value(self, ws, row, column, value):
        celda = ws.cell(row=row, column=column)
        # Even if the type is not MergedCell, it could be the top-left cell of a merged range.
        for merged_range in list(ws.merged_cells.ranges):
            if celda.coordinate in merged_range:
                ws.unmerge_cells(str(merged_range))
                break
        ws.cell(row=row, column=column).value = value

    def _safe_clear_value(self, ws, row, column):
        celda = ws.cell(row=row, column=column)
        if type(celda).__name__ == 'MergedCell':
            return # Don't clear if it's merged, or we could unmerge. Let's just ignore or unmerge? The code was doing: if type != MergedCell: cell.value = None. So if it's Merged, just skip.
        if celda.data_type != 'f':
            celda.value = None

    def _cargar_en_memoria(self):
        if self._wb_cache is not None:
            try:
                self._wb_cache.close()
            except Exception:
                pass
            self._wb_cache = None
        if self._wb_formulas_cache is not None:
            try:
                self._wb_formulas_cache.close()
            except Exception:
                pass
            self._wb_formulas_cache = None

        if os.path.exists(self.ruta):
            # Tarea 22: Asegurar aislamiento de importación de Excel (prohibición estricta de lectura de archivos Excel exportados externos).
            try:
                wb_temp = openpyxl.load_workbook(self.ruta, read_only=True)
                if "_RD_EXPORT_MARKER_" in wb_temp.sheetnames:
                    wb_temp.close()
                    raise PermissionError("SEGURIDAD: No se permite cargar, modificar o importar un archivo Excel exportado por el sistema.")
                wb_temp.close()
            except PermissionError as pe:
                raise pe
            except Exception:
                pass

            try:
                self._wb_cache = openpyxl.load_workbook(self.ruta, data_only=True)
                self._wb_formulas_cache = openpyxl.load_workbook(self.ruta, data_only=False)
                self._last_mtime = os.path.getmtime(self.ruta)
            except PermissionError as pe:
                raise pe
            except Exception:
                self._wb_cache = None
                self._wb_formulas_cache = None
                self._last_mtime = 0.0
    def _verificar_y_recargar_cache(self):
        if os.path.exists(self.ruta):
            try:
                current_mtime = os.path.getmtime(self.ruta)
                if current_mtime != self._last_mtime:
                    self._cargar_en_memoria()
            except Exception:
                pass

    def _parse_cell_ref(self, cell_ref):
        match = re.match(r"^([A-Z]+)([0-9]+)$", cell_ref)
        if not match:
            return 1, 1
        col_str = match.group(1)
        row = int(match.group(2))
        col = 0
        for char in col_str:
            col = col * 26 + (ord(char) - ord('A') + 1)
        return row, col

    def _resolver_celda(self, sheet_name, row, col, wb_data=None, wb_formulas=None):
        if wb_data is None:
            wb_data = self._wb_cache
        if wb_formulas is None:
            wb_formulas = self._wb_formulas_cache

        if not wb_data or not wb_formulas:
            return None

        if sheet_name not in wb_formulas or sheet_name not in wb_data:
            return None

        ws_f = wb_formulas[sheet_name]
        ws_d = wb_data[sheet_name]
        val = ws_f.cell(row=row, column=col).value
        if val is None:
            return None
        
        if not str(val).startswith("="):
            return ws_d.cell(row=row, column=col).value

        formula = str(val)
        
        # Strip IF wrapper if present e.g. =IF(MAESTRO!D5="","",...)
        if formula.startswith("=IF(") and ',"",' in formula:
            parts = formula.split(',"",', 1)
            actual_part = parts[1]
            if actual_part.endswith(")"):
                actual_part = actual_part[:-1]
            formula = "=" + actual_part

        # 1. Resolve sheet references inside IFERROR or direct sheet references
        ref_match = re.search(r"IFERROR\('?([^']+)'?!([A-Z]+[0-9]+)", formula)
        if ref_match:
            target_sheet = ref_match.group(1)
            target_cell = ref_match.group(2)
            t_row, t_col = self._parse_cell_ref(target_cell)
            return self._resolver_celda(target_sheet, t_row, t_col, wb_data, wb_formulas)

        direct_ref = re.search(r"([A-Za-z0-9_]+)!([A-Z]+[0-9]+)", formula)
        if direct_ref:
            target_sheet = direct_ref.group(1)
            target_cell = direct_ref.group(2)
            t_row, t_col = self._parse_cell_ref(target_cell)
            return self._resolver_celda(target_sheet, t_row, t_col, wb_data, wb_formulas)

        # 2. Resolve average formulas in PROM sheets
        if "AVERAGE" in formula or "PROMEDIO" in formula:
            args_match = re.search(r"AVERAGE\(([^)]+)\)", formula)
            if args_match:
                args_str = args_match.group(1)
                parts = [p.strip() for p in args_str.split(",")]
                vals = []
                for part in parts:
                    if ":" in part:
                        start_cell, end_cell = part.split(":")
                        r_start, c_start = self._parse_cell_ref(start_cell)
                        r_end, c_end = self._parse_cell_ref(end_cell)
                        for r_idx in range(r_start, r_end + 1):
                            for c_idx in range(c_start, c_end + 1):
                                val_d = self._resolver_celda(sheet_name, r_idx, c_idx, wb_data, wb_formulas)
                                if val_d is not None and isinstance(val_d, (int, float)):
                                    vals.append(val_d)
                    else:
                        r_ref, c_ref = self._parse_cell_ref(part)
                        val_d = self._resolver_celda(sheet_name, r_ref, c_ref, wb_data, wb_formulas)
                        if val_d is not None and isinstance(val_d, (int, float)):
                            vals.append(val_d)
                
                if vals:
                    avg = sum(vals) / len(vals)
                    if "TRUNC" in formula:
                        return int(avg * 10) / 10.0
                    return round(avg, 2)
                return None

        # Fallback
        return ws_d.cell(row=row, column=col).value

    def _save_wb(self, wb):
        import shutil
        import tempfile
        import datetime

        # 1. Backup de seguridad antes de guardar
        bak_ruta = self.ruta.replace(".xlsx", "_bak.xlsx")
        if os.path.exists(self.ruta):
            try:
                shutil.copy2(self.ruta, bak_ruta)
                
                # Respaldo histórico local automatizado (offline)
                dir_base = os.path.dirname(os.path.abspath(self.ruta))
                dir_respaldos = os.path.join(dir_base, "Respaldos_Locales")
                if not os.path.exists(dir_respaldos):
                    os.makedirs(dir_respaldos)
                
                # Crear nombre con timestamp
                ahora = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                nombre_base = os.path.basename(self.ruta).replace(".xlsx", "")
                nombre_respaldo = f"{nombre_base}_respaldo_{ahora}.xlsx"
                shutil.copy2(self.ruta, os.path.join(dir_respaldos, nombre_respaldo))
                
                # Auto-limpieza: mantener solo los últimos 3 respaldos
                respaldos = sorted(
                    [os.path.join(dir_respaldos, f) for f in os.listdir(dir_respaldos) if f.endswith(".xlsx")],
                    key=os.path.getmtime
                )
                while len(respaldos) > 3:
                    viejo = respaldos.pop(0)
                    try: os.remove(viejo)
                    except: pass
            except Exception:
                pass

        # 2. Guardar a través de archivo temporal de forma atómica
        dir_base = os.path.dirname(os.path.abspath(self.ruta))
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=dir_base)
        os.close(fd)

        try:
            # Insertar automáticamente el logo de la escuela en la portada
            try:
                self._insertar_logo_escuela(wb)
            except Exception:
                pass
                
            wb.save(temp_path)
            # Reemplazar de forma atómica
            if os.path.exists(self.ruta):
                os.replace(temp_path, self.ruta)
            else:
                shutil.move(temp_path, self.ruta)
        except Exception as e:
            if os.path.exists(temp_path):
                try: os.remove(temp_path)
                except: pass
            raise e

    def _insertar_logo_escuela(self, wb):
        try:
            # Excluir las plantillas limpias originales de la raíz del proyecto
            from config import BASE_DIR
            raiz_proyecto = os.path.abspath(os.path.join(BASE_DIR, ".."))
            ruta_absoluta = os.path.abspath(self.ruta)
            if os.path.dirname(ruta_absoluta).lower() == raiz_proyecto.lower():
                return

            from utils.footer_utils import get_school_logo_path
            logo_path = get_school_logo_path()
            if not logo_path or not os.path.exists(logo_path):
                return
            
            # Buscar la hoja de portada/caratula
            nombre_hoja = None
            for name in wb.sheetnames:
                if "PORTADA" in name.upper() or "CARATULA" in name.upper() or "CARÁTULA" in name.upper() or "VISTOSA" in name.upper():
                    nombre_hoja = name
                    break
            
            if nombre_hoja:
                ws = wb[nombre_hoja]
                # Limpiar imágenes previas para evitar acumulación
                ws._images = []
                
                from openpyxl.drawing.image import Image as OpenpyxlImage
                img = OpenpyxlImage(logo_path)
                img.width = 90
                img.height = 90
                # Colocar en la esquina superior derecha
                ws.add_image(img, "H1")
        except Exception:
            pass

        # Actualizar mtime para evitar recarga de caché redundante por nuestra propia escritura
        try:
            self._last_mtime = os.path.getmtime(self.ruta)
        except Exception:
            self._last_mtime = 0.0


    def reiniciar_libreta(self):
        result = self.cleaner.limpiar_todo(self.ruta, self.modalidad)
        if result:
            self._cargar_en_memoria()
        return result

    def realizar_cierre_ano_lectivo(self, promover_estudiantes=True) -> tuple[bool, str]:
        """
        Cierra el año lectivo actual:
        1. Crea un respaldo en Documentos/RegistroDoc/Historial_Academico/ de la base de datos cifrada y el Excel.
        2. Promueve cohortes al grado superior (o los gradúa/retira) si se solicita.
        3. Borra calificaciones, asistencia, hábitos, observaciones y tareas.
        4. Limpia el archivo Excel de notas.
        5. Incrementa el año lectivo en la configuración.
        """
        try:
            from rdsecurity import cargar_config_segura, guardar_config_segura
            import datetime
            import shutil

            # 1. Obtener información de año actual
            cfg = cargar_config_segura({})
            ano_actual = cfg.get("ano_lectivo", "2026").strip()
            
            # Directorio de respaldo
            dir_historial = os.path.join(os.path.expanduser("~"), "Documents", "RegistroDoc", "Historial_Academico", f"Historial_Ano_{ano_actual}")
            os.makedirs(dir_historial, exist_ok=True)
            
            # Copiar archivos
            # Base de datos
            db_manager_enc_path = os.path.join(self.db_manager.user_dir, "data", "registro.db.enc")
            if os.path.exists(db_manager_enc_path):
                shutil.copy2(db_manager_enc_path, os.path.join(dir_historial, "registro.db.enc"))
            # Excel
            if os.path.exists(self.ruta):
                shutil.copy2(self.ruta, os.path.join(dir_historial, os.path.basename(self.ruta)))
                
            # 2. Procesar base de datos
            cursor = self.db_conn.cursor()
            
            # Promoción de estudiantes
            if promover_estudiantes:
                promotion_map_primaria = {"1°": "2°", "2°": "3°", "3°": "4°", "4°": "5°", "5°": "6°"}
                promotion_map_premedia = {"7°": "8°", "8°": "9°"}
                
                cursor.execute("SELECT id, nombre, seccion, modalidad FROM grados;")
                grados_dict = {row[0]: (row[1], row[2], row[3]) for row in cursor.fetchall()}
                
                cursor.execute("SELECT id, nombre, grado_id FROM estudiantes WHERE estado = 'Activo';")
                estudiantes = cursor.fetchall()
                
                for est_id, est_nombre, g_id in estudiantes:
                    if g_id not in grados_dict:
                        continue
                    g_nombre, g_sec, g_mod = grados_dict[g_id]
                    g_clean = g_nombre.strip()
                    
                    target_name = None
                    if g_mod == "primaria" and g_clean in promotion_map_primaria:
                        target_name = promotion_map_primaria[g_clean]
                    elif g_mod == "premedia" and g_clean in promotion_map_premedia:
                        target_name = promotion_map_premedia[g_clean]
                        
                    if target_name:
                        # Buscar si ya existe el grado de destino
                        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = ? AND modalidad = ?;", (target_name, g_sec, g_mod))
                        row_target = cursor.fetchone()
                        if row_target:
                            target_g_id = row_target[0]
                            cursor.execute("UPDATE estudiantes SET grado_id = ? WHERE id = ?;", (target_g_id, est_id))
                        else:
                            cursor.execute("INSERT INTO grados (nombre, seccion, modalidad) VALUES (?, ?, ?);", (target_name, g_sec, g_mod))
                            new_g_id = cursor.lastrowid
                            cursor.execute("UPDATE estudiantes SET grado_id = ? WHERE id = ?;", (new_g_id, est_id))
                    else:
                        # Estudiantes de último año se gradúan (se marcan como Retirado)
                        cursor.execute("UPDATE estudiantes SET estado = 'Retirado' WHERE id = ?;", (est_id,))
            
            # Borrar datos del año anterior
            cursor.execute("DELETE FROM notas;")
            cursor.execute("DELETE FROM asistencia;")
            cursor.execute("DELETE FROM observaciones;")
            cursor.execute("DELETE FROM habitos;")
            cursor.execute("DELETE FROM tareas;")
            
            # Incrementar año académico
            try:
                nuevo_ano = str(int(ano_actual) + 1)
            except Exception:
                nuevo_ano = "2027"
                
            cfg["ano_lectivo"] = nuevo_ano
            guardar_config_segura(cfg)
            
            cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('ano_lectivo', ?);", (nuevo_ano,))
            cursor.execute("INSERT INTO auditoria (accion, detalle) VALUES (?, ?);", ("CIERRE_ANO_LECTIVO", f"Cierre del año lectivo {ano_actual}. Promoción: {promover_estudiantes}. Nuevo año: {nuevo_ano}"))
            
            self.db_conn.commit()
            
            # 3. Limpiar Excel
            self.reiniciar_libreta()
            
            return True, f"Cierre de año lectivo {ano_actual} exitoso.\n\nSe creó un respaldo seguro en:\n{dir_historial}\n\nLos estudiantes activos han sido actualizados y el sistema se configuró para el año {nuevo_ano}."
        except Exception as e:
            return False, f"Error al procesar el cierre de año lectivo: {e}"

    def _encontrar_hoja_maestro(self, wb):
        if "MAESTRO" in wb.sheetnames:
            return "MAESTRO"
        for s in wb.sheetnames:
            if "MAESTRO" in s.upper():
                return s
        return wb.sheetnames[0] if wb.sheetnames else "MAESTRO"

    def _obtener_columna_nombres(self, grado, wb=None):
        if wb is None:
            self._verificar_y_recargar_cache()
        if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None: return 2
        should_close = not bool(self._wb_cache) if wb is None else False
        if wb is None:
            wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)

        nombre_maestro = self._encontrar_hoja_maestro(wb)
        ws = wb[nombre_maestro]
        grado_limpio = grado.replace("°", "").strip()
        col = 2
        for r in [3, 4]:
            for c in range(1, 40):
                val = str(ws.cell(row=r, column=c).value or "").strip()
                if grado_limpio in val or grado in val:
                    if "NOMBRE" in str(ws.cell(row=4, column=c).value or "").upper(): col = c
                    elif "NOMBRE" in str(ws.cell(row=4, column=c+1).value or "").upper(): col = c+1
                    else: col = c
                    break
            if col != 2: break
        if should_close: wb.close()
        
        if col != 2: return col
        if self.modalidad == "primaria": return 2
        if "7" in grado: return 2
        if "8" in grado: return 4
        if "9" in grado: return 6
        return 2

    # --- LECTOR EXACTO DE CELDAS ---
    def obtener_datos_generales(self, wb=None):
        datos = {
            "docente_nombre": "", "docente_cedula": "", "seguro_social": "", "numero_posicion": "",
            "condicion_nombramiento": "", "escuela_nombre": "", "escuela_region": "", "distrito": "",
            "corregimiento": "", "zona_escolar": "", "director_nombre": "", "subdirector_nombre": "",
            "coordinador_nombre": "", "telefono": "", "correo": "", "ano_lectivo": "2026",
            "jornada": "", "fecha_t1": "", "fecha_t2": "", "fecha_t3": ""
        }
        
        # Intentar cargar desde SQLite primero
        has_sqlite = False
        if hasattr(self, "db_conn") and self.db_conn is not None:
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("SELECT count(*) FROM sqlite_master WHERE type='table' AND name='configuracion';")
                if cursor.fetchone()[0] > 0:
                    cursor.execute("SELECT clave, valor FROM configuracion;")
                    rows = cursor.fetchall()
                    if rows:
                        has_sqlite = True
                        for k, v in rows:
                            if k in datos:
                                datos[k] = str(v)
            except Exception as e:
                logger.error(f"Error cargando generales desde SQLite: {e}")

        # Fallback a Excel si SQLite no tiene datos de docente
        if not datos["docente_nombre"] and not datos["docente_cedula"]:
            if wb is None:
                self._verificar_y_recargar_cache()
            if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None: return datos
            should_close = not bool(self._wb_cache) if wb is None else False
            if wb is None:
                wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
            
            # 1. EXTRACCIÓN EN "PORTADA"
            if "Portada" in wb.sheetnames:
                ws = wb["Portada"]
                datos["docente_nombre"] = str(ws["P14"].value or ws["Q14"].value or ws["R14"].value or "").strip()
                datos["docente_cedula"] = str(ws["H16"].value or "").strip()
                datos["seguro_social"] = str(ws["AA16"].value or "").strip()
                datos["numero_posicion"] = str(ws["AQ16"].value or "").strip()
                datos["condicion_nombramiento"] = str(ws["U18"].value or "").strip()
                datos["jornada"] = str(ws["H20"].value or "").strip()
                datos["director_nombre"] = str(ws["S22"].value or "").strip()
                datos["subdirector_nombre"] = str(ws["T24"].value or "").strip()
                datos["coordinador_nombre"] = str(ws["O26"].value or "").strip()
                datos["escuela_nombre"] = str(ws["Q28"].value or "").strip()
                datos["telefono"] = str(ws["F30"].value or "").strip()
                datos["correo"] = str(ws["AF30"].value or "").strip()
                datos["escuela_region"] = str(ws["M32"].value or "").strip()
                datos["distrito"] = str(ws["L34"].value or "").strip()
                datos["corregimiento"] = str(ws["Q36"].value or "").strip()
                datos["zona_escolar"] = str(ws["O38"].value or "").strip()
                datos["ano_lectivo"] = str(ws["P8"].value or "2026").strip()

            # 2. EXTRACCIÓN DE FECHAS PURAS (C1, C44, C87) Limpiando guiones
            for sheet in wb.sheetnames:
                if "ASISTENCIA" in sheet.upper():
                    ws = wb[sheet]
                    def extraer_fecha(celda):
                        return DataEngine._procesar_texto_asistencia(ws[celda].value)
                    
                    datos["fecha_t1"] = extraer_fecha("C1")
                    datos["fecha_t2"] = extraer_fecha("C44")
                    datos["fecha_t3"] = extraer_fecha("C87")
                    break
                    
            if should_close: wb.close()
            
            # Si leímos de Excel, guardar estos en la base de datos SQLite de forma inmediata si es posible
            if hasattr(self, "db_conn") and self.db_conn is not None:
                try:
                    cursor = self.db_conn.cursor()
                    for k, v in datos.items():
                        cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?);", (k, str(v)))
                    self.db_conn.commit()
                except Exception as e:
                    logger.error(f"Error persistiendo fallback de config a SQLite: {e}")

        return datos

    @staticmethod
    def _procesar_texto_asistencia(valor):
        """Lógica pura para extraer la fecha de una celda de asistencia."""
        if not valor:
            return ""
        val = str(valor).upper()
        if "ASISTENCIA DEL" in val:
            # Corta la palabra, limpia guiones y espacios dobles
            fecha = val.replace("ASISTENCIA DEL", "").replace("_", " ").strip()
            return " ".join(fecha.split())
        return ""

    def obtener_horario(self, wb=None):
        """Retorna el horario. Lee SQL primero; si está vacío, cae a Excel y sincroniza."""
        # 1. Intentar desde SQL
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("""
                SELECT bloque_orden, dia, bloque_hora, materia_texto
                FROM horario
                ORDER BY bloque_orden, dia;
            """)
            rows = cursor.fetchall()

            if rows:
                # Reconstruir la estructura de 8 slots
                horario = [{"horas": "", "lunes": "", "martes": "", "miercoles": "",
                            "jueves": "", "viernes": ""} for _ in range(8)]
                for bloque_orden, dia, bloque_hora, materia_texto in rows:
                    idx = int(bloque_orden) if bloque_orden is not None else 0
                    if 0 <= idx < 8:
                        horario[idx]["horas"] = bloque_hora or ""
                        if dia in horario[idx]:
                            horario[idx][dia] = materia_texto or ""
                # Solo devolver SQL si tiene al menos un dato útil
                if any(any(v for k, v in s.items() if k != "horas") for s in horario):
                    return horario
        except Exception as e:
            logger.warning(f"[horario] Error leyendo SQL: {e}")

        # 2. Fallback al Excel (rescata datos existentes y los sincroniza)
        horario_excel = self.obtener_horario_desde_excel(wb)
        tiene_datos = any(any(v for k, v in s.items() if k != "horas") for s in horario_excel)
        if tiene_datos:
            # Sincronizar al SQL para que la próxima vez no necesite Excel
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("DELETE FROM horario;")
                for idx, slot in enumerate(horario_excel):
                    bloque_hora = slot.get("horas", f"Período {idx+1}")
                    for dia in ("lunes", "martes", "miercoles", "jueves", "viernes"):
                        texto = slot.get(dia, "").strip()
                        cursor.execute(
                            """INSERT INTO horario (bloque_orden, dia, bloque_hora, materia_texto)
                               VALUES (?, ?, ?, ?);""",
                            (idx, dia, bloque_hora, texto)
                        )
                self.db_conn.commit()
                self._schedule_save()
                logger.info("[horario] Horario rescatado del Excel y guardado en SQL.")
            except Exception as e:
                logger.warning(f"[horario] Error al rescatar horario a SQL: {e}")
        return horario_excel

    def obtener_horario_desde_excel(self, wb=None):
        """Lee el horario directamente del Excel. Uso interno para sincronización."""
        if wb is None:
            self._verificar_y_recargar_cache()
        horario = [{"horas": "", "lunes": "", "martes": "", "miercoles": "",
                    "jueves": "", "viernes": ""} for _ in range(8)]
        if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None:
            return horario
        should_close = not bool(self._wb_cache) if wb is None else False
        if wb is None:
            wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
        hoja_horario = next((s for s in wb.sheetnames if "HORARIO" in s.upper()), None)

        if hoja_horario:
            ws = wb[hoja_horario]
            idx = 0
            for r in range(10, 25):
                if idx >= 8:
                    break
                fila_texto = "".join(str(ws.cell(row=r, column=c).value or "").upper() for c in range(1, 10))
                if "RECESO" in fila_texto:
                    continue
                es_periodo = any(
                    str(ws.cell(row=r, column=c).value or "").strip().upper()
                    in ["I", "II", "III", "IV", "V", "VI", "VII", "VIII"]
                    for c in range(1, 8)
                )
                if es_periodo:
                    def atrapar_texto(col_base):
                        for offset in range(3):
                            v = str(ws.cell(row=r, column=col_base + offset).value or "").strip()
                            if v and v != "None":
                                return v
                        return ""
                    horario[idx]["horas"]      = atrapar_texto(10)
                    horario[idx]["lunes"]      = atrapar_texto(15)
                    horario[idx]["martes"]     = atrapar_texto(16)
                    horario[idx]["miercoles"]  = atrapar_texto(17)
                    horario[idx]["jueves"]     = atrapar_texto(18)
                    horario[idx]["viernes"]    = atrapar_texto(19)
                    idx += 1

        if should_close:
            wb.close()
        return horario

    def guardar_horario(self, datos_horario):
        """Guarda el horario en SQL (fuente de verdad) y dispara guardado cifrado.
        También actualiza el Excel para mantener compatibilidad visual.
        """
        # 1. Guardar en SQL
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("DELETE FROM horario;")
            for idx, slot in enumerate(datos_horario):
                bloque_hora = slot.get("horas", f"Período {idx+1}")
                for dia in ("lunes", "martes", "miercoles", "jueves", "viernes"):
                    texto = slot.get(dia, "").strip()
                    cursor.execute(
                        """INSERT INTO horario (bloque_orden, dia, bloque_hora, materia_texto)
                           VALUES (?, ?, ?, ?);""",
                        (idx, dia, bloque_hora, texto)
                    )
            self.db_conn.commit()
            self._schedule_save()
            sql_ok = True
        except Exception as e:
            logger.error(f"[guardar_horario] Error SQL: {e}")
            sql_ok = False

        # 2. También actualizar Excel (compatibilidad con reportes Word)
        if os.path.exists(self.ruta):
            try:
                wb = openpyxl.load_workbook(self.ruta)
                hoja_horario = next((s for s in wb.sheetnames if "HORARIO" in s.upper()), None)
                if hoja_horario:
                    ws = wb[hoja_horario]
                    idx = 0
                    for r in range(10, 25):
                        if idx >= len(datos_horario):
                            break
                        fila_texto = "".join(str(ws.cell(row=r, column=c).value or "").upper() for c in range(1, 10))
                        if "RECESO" in fila_texto:
                            continue
                        es_periodo = any(
                            str(ws.cell(row=r, column=c).value or "").strip().upper()
                            in ["I", "II", "III", "IV", "V", "VI", "VII", "VIII"]
                            for c in range(1, 8)
                        )
                        if es_periodo:
                            h = datos_horario[idx]
                            mapeo = {10: "horas", 15: "lunes", 16: "martes",
                                     17: "miercoles", 18: "jueves", 19: "viernes"}
                            for col, llave in mapeo.items():
                                try:
                                    ws.cell(row=r, column=col).value = h.get(llave, "")
                                except Exception:
                                    pass
                            idx += 1
                self._save_wb(wb)
                wb.close()
                self._cargar_en_memoria()
            except Exception as e:
                logger.warning(f"[guardar_horario] No se pudo actualizar Excel: {e}")

        return sql_ok

    def sincronizar_plantilla_maestra(self, config):
        # Guardar en SQLite configuracion
        try:
            cursor = self.db_conn.cursor()
            for k, v in config.items():
                if isinstance(v, (list, dict)):
                    v_str = json.dumps(v, ensure_ascii=False)
                else:
                    v_str = str(v)
                cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?);", (k, v_str))
            
            # También persistir ruta_excel, trimestre_activo, modalidad
            cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('ruta_excel', ?);", (self.ruta,))
            from utils.date_helpers import obtener_trimestre_actual
            trimestre = obtener_trimestre_actual()
            cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('trimestre_activo', ?);", (trimestre,))
            cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('modalidad', ?);", (self.modalidad,))
            
            self.db_conn.commit()
            self._schedule_save()
        except Exception as e:
            logger.error(f"Error guardando config en SQLite en sincronizar_plantilla_maestra: {e}")

        if not os.path.exists(self.ruta): return False
        wb = openpyxl.load_workbook(self.ruta)
        
        docente = str(config.get("docente_nombre", "DOCENTE")).upper()
        docente_titulo = str(config.get("docente_nombre", "Docente")).title()
        cedula = str(config.get("docente_cedula", ""))
        ss = str(config.get("seguro_social", ""))
        pos = str(config.get("numero_posicion", ""))
        condicion = str(config.get("condicion_nombramiento", "")).title()
        ano = str(config.get("ano_lectivo", "2026"))
        escuela = str(config.get("escuela_nombre", "ESCUELA")).upper()
        region = str(config.get("escuela_region", "REGIÓN")).title()
        distrito = str(config.get("distrito", "")).upper()
        corregimiento = str(config.get("corregimiento", "")).upper()
        zona = str(config.get("zona_escolar", ""))
        director = str(config.get("director_nombre", "")).title()
        subdirector = str(config.get("subdirector_nombre", "")).title()
        coordinador = str(config.get("coordinador_nombre", "")).title()
        jornada = str(config.get("jornada", "MATUTINA")).upper()
        telefono = str(config.get("telefono", ""))
        correo = str(config.get("correo", ""))
        ft1 = str(config.get("fecha_t1", "")).upper()
        ft2 = str(config.get("fecha_t2", "")).upper()
        ft3 = str(config.get("fecha_t3", "")).upper()

        grados_activos = self.obtener_grados_activos()
        str_grados = " - ".join(grados_activos)
        todas_materias = []
        for g in grados_activos: todas_materias.extend(self.obtener_materias_por_grado(g))
        str_materias = " - ".join(sorted(list(set(todas_materias))))

        # 1. INYECCIÓN EXACTA EN "PORTADA"
        if "Portada" in wb.sheetnames:
            ws = wb["Portada"]
            ws["R10"] = str_grados       # <--- ¡AGREGA ESTA LÍNEA AQUÍ!
            ws["P14"] = docente
            ws["H16"] = cedula
            ws["AA16"] = ss
            ws["AQ16"] = pos
            ws["U18"] = condicion
            ws["H20"] = jornada.capitalize()
            ws["AD20"] = str_materias
            ws["S22"] = director
            ws["T24"] = subdirector
            ws["O26"] = coordinador
            ws["L34"] = distrito
            ws["O38"] = zona
            ws["AF30"] = correo
            ws["Q28"] = escuela
            ws["F30"] = telefono
            ws["M32"] = region
            ws["Q36"] = corregimiento
            ws["J38"] = zona
            ws["P8"] = ano

        # 2. INYECCIÓN EXACTA EN "HORARIO"
        for sheet in wb.sheetnames:
            if "HORARIO" in sheet.upper():
                ws = wb[sheet]
                ws["O2"] = docente
                ws["H4"] = jornada
                ws["I6"] = str_materias
                ws["O8"] = str_grados

        # 3. CARÁTULA Y PORTADA VISTOSA 
        for sheet in wb.sheetnames:
            if "CARATULA" in sheet.upper() or "CARÁTULA" in sheet.upper() or "VISTOSA" in sheet.upper():
                ws = wb[sheet]
                for r in range(1, 40):
                    for c in range(1, 10):
                        try:
                            texto = str(ws.cell(row=r, column=c).value or "")
                            if not texto.strip(): continue
                            modificado = False
                            
                            if "Prof." in texto or "Profesor" in texto:
                                texto = re.sub(r'Prof\.\s*.*', f"Prof. {docente_titulo}", texto)
                                modificado = True
                            if "Instructor" in texto or "Pre-Media" in texto:
                                texto = f" {condicion}"
                                modificado = True
                            if "Grupo:" in texto:
                                texto = f"Grupo: {str_grados}"
                                modificado = True
                            if "Ano lectivo" in texto or "Año lectivo" in texto:
                                ws.cell(row=r, column=c+1).value = ano
                                
                            if "CERRO CACICON" in texto.upper() or "ESCUELA DE EJEMPLO" in texto.upper():
                                texto = re.sub(r'CERRO CACICON|CERRO CACICÓN|ESCUELA DE EJEMPLO', escuela, texto, flags=re.IGNORECASE)
                                modificado = True
                            if "ELMER TUGRI" in texto.upper() or "JUAN PÉREZ" in texto.upper() or "JUAN PEREZ" in texto.upper():
                                texto = re.sub(r'ELMER TUGRI|JUAN PÉREZ|JUAN PEREZ', docente_titulo, texto, flags=re.IGNORECASE)
                                modificado = True
                            
                            if modificado: ws.cell(row=r, column=c).value = texto
                        except Exception as e:
                            print(f"[!] Error in sincronizar_plantilla_maestra (carátula) at row {r}, col {c}: {e}")

        # 4. INYECCIÓN DE FECHAS EN ASISTENCIA (Solo escribe en la celda)
        for sheet in wb.sheetnames:
            if "ASISTENCIA" in sheet.upper():
                ws = wb[sheet]
                if ft1: ws["C1"] = f"ASISTENCIA DEL {ft1}"
                if ft2: ws["C44"] = f"ASISTENCIA DEL {ft2}"
                if ft3: ws["C87"] = f"ASISTENCIA DEL {ft3}"
                
                for r in range(1, 6):
                    for c in range(1, 10):
                        try:
                            val = str(ws.cell(row=r, column=c).value or "").upper()
                            if "JORNADA:" in val and "ASIGNATURA:" in val:
                                ws.cell(row=r, column=c).value = f"JORNADA:           {jornada}                  ASIGNATURA:      Todas (Multigrado)"
                            if "AÑO:" in val and "AULA:" in val:
                                g_match = re.search(r'\((.*?)\)', sheet)
                                aula = g_match.group(1) if g_match else "Multigrado"
                                ws.cell(row=r, column=c).value = f"AÑO: {ano}     AULA:      {aula}                   NOMBRE DEL PROF./A CONSEJERO/A:      {docente}"
                        except Exception as e:
                            print(f"[!] Error in sincronizar_plantilla_maestra (asistencia) at row {r}, col {c}: {e}")

        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True

    def obtener_grados_activos(self, wb=None):
        if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None:
            return []
        cursor = self.db_conn.cursor()
        cursor.execute("SELECT DISTINCT nombre FROM grados WHERE modalidad = ? ORDER BY nombre;", (self.modalidad,))
        rows = cursor.fetchall()
        if rows:
            return [r[0] for r in rows]

        # Si la base de datos está vacía, escanear Excel
        should_close = False
        if wb is None:
            if self._wb_cache:
                wb = self._wb_cache
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(self.ruta, read_only=True)
                    should_close = True
                except Exception:
                    return []

        grados = []
        for s in wb.sheetnames:
            if s.upper().startswith("RESUMEN_"):
                g_name = s.split("_", 1)[1].strip()
                grados.append(g_name)

        if should_close:
            wb.close()

        # Insertar grados iniciales en SQLite
        try:
            for g in grados:
                cursor.execute("INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES (?, 'A', ?);", (g, self.modalidad))
            self.db_conn.commit()
        except Exception as e:
            logger.error(f"Error insertando grados escaneados: {e}")

        return sorted(list(set(grados)))

    def obtener_materias_por_grado(self, grado, wb=None):
        if wb is not None or not os.path.basename(self.ruta).startswith("Registro_"):
            if wb is None:
                self._verificar_y_recargar_cache()
            if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None:
                return []
            should_close = not bool(self._wb_cache) if wb is None else False
            if wb is None:
                wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
            materias = []
            grado_num = grado.replace("°", "")
            for sheet in wb.sheetnames:
                if "PROM" in sheet.upper():
                    if self.modalidad == "premedia" and grado_num in sheet:
                        mat = sheet.upper().replace("PROM", "").replace("(", "").replace(")", "").replace(grado, "").replace(grado_num, "").replace("°", "").strip()
                        materias.append(mat.title())
                    elif self.modalidad == "primaria":
                        mat = sheet.upper().replace("PROM", "").replace("(", "").replace(")", "").strip()
                        materias.append(mat.title())
            if should_close: wb.close()
            return sorted(list(set(materias))) if materias else ["Sin materias registradas"]

        cursor = self.db_conn.cursor()
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        if not row_g:
            return ["Sin materias registradas"]
        grado_id = row_g[0]
        
        cursor.execute("SELECT DISTINCT nombre FROM materias WHERE grado_id = ? ORDER BY nombre;", (grado_id,))
        rows = cursor.fetchall()
        return [r[0] for r in rows] if rows else ["Sin materias registradas"]

    def _obtener_estudiantes_desde_excel(self, grado, wb=None):
        """Extrae la lista de estudiantes directamente del archivo Excel (sin consultar SQL)."""
        if wb is None:
            self._verificar_y_recargar_cache()
        if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None:
            return []
        should_close = not bool(self._wb_cache) if wb is None else False
        if wb is None:
            wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)

        try:
            nombre_maestro = self._encontrar_hoja_maestro(wb)
            ws_m = wb[nombre_maestro]
            col_nom = self._obtener_columna_nombres(grado, wb=wb)
            ws_planilla = None
            if self.modalidad == "premedia":
                for sheet in wb.sheetnames:
                    if "Planilla" in sheet and grado.replace("°", "") in sheet:
                        ws_planilla = wb[sheet]
                        break
            grado_id = 0
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
                row_g = cursor.fetchone()
                if row_g:
                    grado_id = row_g[0]
            except Exception:
                pass
            if not grado_id:
                if "7" in grado: grado_id = 2
                elif "8" in grado: grado_id = 3
                elif "9" in grado: grado_id = 4
                else: grado_id = 1

            estudiantes = []
            for r in range(5, 46):
                nom = ws_m.cell(row=r, column=col_nom).value
                if nom:
                    cedula = ""
                    if self.modalidad == "primaria":
                        cedula = ws_m.cell(row=r, column=5).value
                    elif ws_planilla:
                        fila_plan = 15 + (r - 4)
                        cedula = ws_planilla.cell(row=fila_plan, column=5).value
                    estudiantes.append({
                        "id": grado_id * 100 + (r - 4),
                        "nombre": str(nom).strip(),
                        "cedula": str(cedula).strip() if cedula else "",
                        "sexo": "M"
                    })
            return estudiantes
        finally:
            if should_close:
                wb.close()

    def obtener_estudiantes_completos(self, grado, wb=None):
        # ── Caché en memoria con invalidación por version de BD ──
        db_version = 0
        try:
            if hasattr(self, 'db_conn') and self.db_conn:
                raw_conn = getattr(self.db_conn, '_conn', self.db_conn)
                db_version = raw_conn.total_changes
        except Exception:
            pass

        if not hasattr(self, '_cache_estudiantes'):
            self._cache_estudiantes = {}
            self._cache_est_version = -1

        if self._cache_est_version == db_version and grado in self._cache_estudiantes:
            return self._cache_estudiantes[grado]

        # Si la versión cambió, limpiar todo el caché
        if self._cache_est_version != db_version:
            self._cache_estudiantes.clear()
            self._cache_est_version = db_version

        # 1. Intentar cargar desde SQLite y descifrar columnas
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = ? AND modalidad = ?;", (grado, "A", self.modalidad))
            row_g = cursor.fetchone()
            if row_g:
                grado_id = row_g[0]
                cursor.execute("SELECT id, nombre, cedula, sexo FROM estudiantes WHERE grado_id = ? AND estado = 'Activo' ORDER BY CAST(id AS INTEGER);", (grado_id,))
                rows = cursor.fetchall()
                if rows:
                    estudiantes = []
                    for r in rows:
                        est_id = int(r[0]) if str(r[0]).isdigit() else r[0]
                        nom_desc = self.db_manager.desencriptar_campo(r[1])
                        ced_desc = self.db_manager.desencriptar_campo(r[2])
                        sex_desc = self.db_manager.desencriptar_campo(r[3]) or "M"
                        estudiantes.append({
                            "id": est_id,
                            "nombre": nom_desc,
                            "cedula": ced_desc,
                            "sexo": sex_desc
                        })
                    self._cache_estudiantes[grado] = estudiantes
                    return estudiantes
        except Exception as e:
            logger.error(f"Error cargando estudiantes desde SQL: {e}")

        # Fallback a Excel
        resultado = self._obtener_estudiantes_desde_excel(grado, wb=wb)
        self._cache_estudiantes[grado] = resultado
        return resultado

    def agregar_estudiante(self, grado, nombre, cedula="", sexo="M"):
        if not os.path.exists(self.ruta): return False

        # Sanitizar entrada para evitar espacios raros, saltos de línea y tabulaciones (Tarea 13)
        nombre = sanitizar_entrada(nombre)
        cedula = sanitizar_entrada(cedula)

        # Validar limites
        max_estudiantes = 34 if self.modalidad == "primaria" else 36
        actuales = self.obtener_estudiantes_completos(grado)
        if len(actuales) >= max_estudiantes:
            return False

        wb = openpyxl.load_workbook(self.ruta)
        ws_m = wb[self._encontrar_hoja_maestro(wb)]
        col_nom = self._obtener_columna_nombres(grado)
        fila_vacia = None
        for r in range(5, 5 + max_estudiantes):
            if not ws_m.cell(row=r, column=col_nom).value:
                fila_vacia = r
                break
        if not fila_vacia:
            wb.close()
            return False 
        ws_m.cell(row=fila_vacia, column=col_nom).value = nombre
        if self.modalidad == "primaria": ws_m.cell(row=fila_vacia, column=5).value = cedula
        else:
            num_grado = "".join(c for c in grado if c.isdigit())
            id_est = fila_vacia - 4
            for sheet in wb.sheetnames:
                if "Planilla" in sheet and num_grado in sheet:
                    ws_p = wb[sheet]
                    ws_p.cell(row=15+id_est, column=5).value = cedula
        self._save_wb(wb)
        wb.close()

        # Sincronizar a SQLite
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES (?, 'A', ?);", (grado, self.modalidad))
            self.db_conn.commit()
            cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
            row_g = cursor.fetchone()
            if row_g:
                grado_id = row_g[0]
                est_id = str(grado_id * 100 + (fila_vacia - 4))
                nombre_cifrado = self.db_manager.encriptar_campo(nombre)
                cedula_cifrada = self.db_manager.encriptar_campo(cedula)
                sexo_cifrado = self.db_manager.encriptar_campo(sexo)
                cursor.execute(
                    "INSERT OR REPLACE INTO estudiantes (id, nombre, cedula, sexo, grado_id, estado) VALUES (?, ?, ?, ?, ?, 'Activo');",
                    (est_id, nombre_cifrado, cedula_cifrada, sexo_cifrado, grado_id)
                )
                self.db_conn.commit()
                self._schedule_save()
        except Exception as e:
            logger.error(f"Error guardando estudiante en SQLite: {e}")

        self._cargar_en_memoria()
        return True

    def guardar_cambios_estudiantes(self, grado, datos_modificados):
        if not os.path.exists(self.ruta): return False
        wb = openpyxl.load_workbook(self.ruta)
        ws_m = wb[self._encontrar_hoja_maestro(wb)]
        col_nom = self._obtener_columna_nombres(grado)
        
        # Sanitizar datos modificados
        datos_sanos = {}
        for id_est, datos in datos_modificados.items():
            datos_sanos[id_est] = {
                "nombre": sanitizar_entrada(datos["nombre"]),
                "cedula": sanitizar_entrada(datos.get("cedula", "")),
                "sexo": datos.get("sexo", "M")
            }

        for id_est, datos in datos_sanos.items():
            fila = 4 + (int(id_est) % 100)
            ws_m.cell(row=fila, column=col_nom).value = datos["nombre"]
            if self.modalidad == "primaria": ws_m.cell(row=fila, column=5).value = datos["cedula"]
            else:
                num_grado = "".join(c for c in grado if c.isdigit())
                for sheet in wb.sheetnames:
                    if "Planilla" in sheet and num_grado in sheet:
                        ws_p = wb[sheet]
                        ws_p.cell(row=15+(int(id_est) % 100), column=5).value = datos["cedula"]
        self._save_wb(wb)
        wb.close()

        # Sincronizar a SQLite
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
            row_g = cursor.fetchone()
            if row_g:
                grado_id = row_g[0]
                for id_est, datos in datos_sanos.items():
                    nombre_cifrado = self.db_manager.encriptar_campo(datos["nombre"])
                    cedula_cifrada = self.db_manager.encriptar_campo(datos.get("cedula", ""))
                    sexo_cifrado = self.db_manager.encriptar_campo(datos.get("sexo", "M"))
                    cursor.execute(
                        "INSERT OR REPLACE INTO estudiantes (id, nombre, cedula, sexo, grado_id, estado) VALUES (?, ?, ?, ?, ?, 'Activo');",
                        (str(id_est), nombre_cifrado, cedula_cifrada, sexo_cifrado, grado_id)
                    )
                self.db_conn.commit()
                self._schedule_save()
        except Exception as e:
            logger.error(f"Error actualizando estudiantes en SQLite: {e}")

        self._cargar_en_memoria()
        return True

    def limpiar_acentos(self, texto):
        if not texto:
            return ""
        texto = str(texto).upper()
        for k, v in {"Á":"A", "É":"E", "Í":"I", "Ó":"O", "Ú":"U", "Ü":"U"}.items():
            texto = texto.replace(k, v)
        return texto

    # ─── Retiro de Estudiantes ──────────────────────────────────────────────

    def retirar_estudiante(self, est_id, motivo: str, nombre_acudiente: str = "") -> tuple:
        """Marca un estudiante como Retirado en SQLite y programa el guardado cifrado.
        Retorna (True, nombre_descifrado) si tuvo éxito, (False, mensaje_error) si falló.
        """
        try:
            cursor = self.db_conn.cursor()
            # Verificar que el estudiante existe y está activo
            cursor.execute(
                "SELECT nombre FROM estudiantes WHERE id = ? AND estado = 'Activo';",
                (str(est_id),)
            )
            row = cursor.fetchone()
            if not row:
                return False, "Estudiante no encontrado o ya retirado."

            nombre_cifrado = row[0]
            nombre = self.db_manager.desencriptar_campo(nombre_cifrado)

            # Encriptar el nombre del acudiente antes de guardar
            acudiente_cifrado = self.db_manager.encriptar_campo(nombre_acudiente) if nombre_acudiente else ""

            from datetime import date
            fecha_retiro = date.today().strftime("%Y-%m-%d")

            cursor.execute(
                """UPDATE estudiantes
                   SET estado = 'Retirado',
                       fecha_retiro = ?,
                       motivo_retiro = ?,
                       nombre_acudiente = ?
                   WHERE id = ?;""",
                (fecha_retiro, motivo.strip(), acudiente_cifrado, str(est_id))
            )
            self.db_conn.commit()
            self._schedule_save()
            logger.info(f"Estudiante id={est_id} retirado. Motivo: {motivo}")
            return True, nombre

        except Exception as e:
            logger.error(f"Error retirando estudiante id={est_id}: {e}")
            return False, str(e)

    def reactivar_estudiante(self, est_id) -> tuple:
        """Reactiva un estudiante retirado, limpiando sus datos de retiro.
        Retorna (True, nombre) o (False, error).
        """
        try:
            cursor = self.db_conn.cursor()
            cursor.execute(
                "SELECT nombre FROM estudiantes WHERE id = ? AND estado = 'Retirado';",
                (str(est_id),)
            )
            row = cursor.fetchone()
            if not row:
                return False, "Estudiante no encontrado o ya está activo."

            nombre = self.db_manager.desencriptar_campo(row[0])
            cursor.execute(
                """UPDATE estudiantes
                   SET estado = 'Activo',
                       fecha_retiro = NULL,
                       motivo_retiro = NULL,
                       nombre_acudiente = NULL
                   WHERE id = ?;""",
                (str(est_id),)
            )
            self.db_conn.commit()
            self._schedule_save()
            logger.info(f"Estudiante id={est_id} reactivado.")
            return True, nombre

        except Exception as e:
            logger.error(f"Error reactivando estudiante id={est_id}: {e}")
            return False, str(e)

    def obtener_estudiantes_retirados(self, grado: str = None) -> list:
        """Retorna lista de estudiantes retirados, descifrados.
        Si grado se especifica, filtra por ese grado. Ordenados por fecha_retiro DESC.
        """
        try:
            cursor = self.db_conn.cursor()
            if grado:
                cursor.execute(
                    "SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;",
                    (grado, self.modalidad)
                )
                row_g = cursor.fetchone()
                if not row_g:
                    return []
                grado_id = row_g[0]
                cursor.execute(
                    """SELECT e.id, e.nombre, e.cedula, e.sexo, e.fecha_retiro,
                              e.motivo_retiro, e.nombre_acudiente, g.nombre as grado_nombre
                       FROM estudiantes e
                       JOIN grados g ON e.grado_id = g.id
                       WHERE e.estado = 'Retirado' AND e.grado_id = ?
                       ORDER BY e.fecha_retiro DESC;""",
                    (grado_id,)
                )
            else:
                cursor.execute(
                    """SELECT e.id, e.nombre, e.cedula, e.sexo, e.fecha_retiro,
                              e.motivo_retiro, e.nombre_acudiente, g.nombre as grado_nombre
                       FROM estudiantes e
                       JOIN grados g ON e.grado_id = g.id
                       WHERE e.estado = 'Retirado'
                       ORDER BY e.fecha_retiro DESC;"""
                )

            resultado = []
            for r in cursor.fetchall():
                resultado.append({
                    "id":               r[0],
                    "nombre":           self.db_manager.desencriptar_campo(r[1]),
                    "cedula":           self.db_manager.desencriptar_campo(r[2]) if r[2] else "",
                    "sexo":             self.db_manager.desencriptar_campo(r[3]) if r[3] else "",
                    "fecha_retiro":     r[4] or "",
                    "motivo_retiro":    r[5] or "",
                    "nombre_acudiente": self.db_manager.desencriptar_campo(r[6]) if r[6] else "",
                    "grado":            r[7] or "",
                })
            return resultado

        except Exception as e:
            logger.error(f"Error obteniendo estudiantes retirados: {e}")
            return []

    def obtener_distribucion_sexo_estado(self, grado: str = None) -> dict:
        """Retorna conteo de estudiantes por sexo y estado para gráficos.
        Resultado: {'masculino': N, 'femenino': N, 'retirados': N, 'total': N}
        """
        try:
            cursor = self.db_conn.cursor()
            if grado:
                cursor.execute(
                    "SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;",
                    (grado, self.modalidad)
                )
                row_g = cursor.fetchone()
                grado_filter = f"AND e.grado_id = {row_g[0]}" if row_g else ""
            else:
                grado_filter = ""

            cursor.execute(f"""
                SELECT e.sexo, e.estado, COUNT(*) as total
                FROM estudiantes e
                WHERE 1=1 {grado_filter}
                GROUP BY e.sexo, e.estado;
            """)

            masc = 0
            fem  = 0
            retirados = 0
            for row in cursor.fetchall():
                sexo_enc = row[0] or ""
                estado   = row[1] or "Activo"
                conteo   = row[2]
                sexo_dec = self.db_manager.desencriptar_campo(sexo_enc).upper() if sexo_enc else ""

                if estado == "Retirado":
                    retirados += conteo
                elif sexo_dec in ("F", "FEM", "FEMENINO"):
                    fem += conteo
                else:
                    masc += conteo

            total = masc + fem + retirados
            return {"masculino": masc, "femenino": fem, "retirados": retirados, "total": total}

        except Exception as e:
            logger.error(f"Error obteniendo distribucion sexo/estado: {e}")
            return {"masculino": 0, "femenino": 0, "retirados": 0, "total": 0}

    def _calcular_promedios_sql(self, grado, materia, trimestre):
        cursor = self.db_conn.cursor()
        
        # 1. Obtener los estudiantes del grado
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        if not row_g:
            return {}
        grado_id = row_g[0]
        
        cursor.execute("SELECT id, nombre FROM estudiantes WHERE grado_id = ? ORDER BY CAST(id AS INTEGER);", (grado_id,))
        estudiantes = []
        for r in cursor.fetchall():
            dec_name = self.db_manager.desencriptar_campo(r[1])
            estudiantes.append((r[0], dec_name))
        if not estudiantes:
            return {}
            
        results = {}
        
        # 2. Si se pide una materia específica
        if materia and materia not in ["Sin materias", "No hay materias", "General", "Todas las Materias"]:
            # Obtener materia_id
            cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
            row_m = cursor.fetchone()
            if not row_m:
                return {}
            materia_id = row_m[0]
            
            # Si el trimestre es "Anual" o "FINAL" o "PROMEDIO"
            if str(trimestre).upper() in ["ANUAL", "FINAL", "PROMEDIO", "TODOS", "TODOS LOS TRIMESTRES"]:
                for est_id, est_nombre in estudiantes:
                    trim_notas = []
                    for t in [1, 2, 3]:
                        prom_t = self._calcular_promedio_trimestre_materia_sql(est_id, materia_id, t)
                        if prom_t is not None:
                            trim_notas.append(prom_t)
                    if trim_notas:
                        results[est_nombre] = round(sum(trim_notas) / len(trim_notas), 1)
            else:
                if isinstance(trimestre, int):
                    trim_num = trimestre
                else:
                    try:
                        trim_num = int(str(trimestre).replace("Trimestre ", ""))
                    except ValueError:
                        trim_num = 1
                for est_id, est_nombre in estudiantes:
                    prom = self._calcular_promedio_trimestre_materia_sql(est_id, materia_id, trim_num)
                    if prom is not None:
                        results[est_nombre] = prom
        else:
            # Promedio general de todas las materias del grado
            cursor.execute("SELECT id, nombre FROM materias WHERE grado_id = ?;", (grado_id,))
            materias_del_grado = cursor.fetchall()
            if not materias_del_grado:
                return {}
                
            tech_mats = []
            norm_mats = []
            for m_id, m_nom in materias_del_grado:
                m_upper = self.limpiar_acentos(m_nom).upper()
                if "HOGAR" in m_upper or "DESARROLLO" in m_upper or "AGRO" in m_upper or "COMERCIO" in m_upper:
                    tech_mats.append(m_id)
                else:
                    norm_mats.append(m_id)
                    
            for est_id, est_nombre in estudiantes:
                norm_notas = []
                for m_id in norm_mats:
                    prom_m = None
                    if str(trimestre).upper() in ["ANUAL", "FINAL", "PROMEDIO", "TODOS", "TODOS LOS TRIMESTRES"]:
                        trim_vals = []
                        for t in [1, 2, 3]:
                            p_t = self._calcular_promedio_trimestre_materia_sql(est_id, m_id, t)
                            if p_t is not None:
                                trim_vals.append(p_t)
                        if trim_vals:
                            prom_m = round(sum(trim_vals) / len(trim_vals), 1)
                    else:
                        if isinstance(trimestre, int):
                            trim_num = trimestre
                        else:
                            try:
                                trim_num = int(str(trimestre).replace("Trimestre ", ""))
                            except ValueError:
                                trim_num = 1
                        prom_m = self._calcular_promedio_trimestre_materia_sql(est_id, m_id, trim_num)
                    if prom_m is not None:
                        norm_notas.append(prom_m)
                        
                tech_notas = []
                for m_id in tech_mats:
                    prom_m = None
                    if str(trimestre).upper() in ["ANUAL", "FINAL", "PROMEDIO", "TODOS", "TODOS LOS TRIMESTRES"]:
                        trim_vals = []
                        for t in [1, 2, 3]:
                            p_t = self._calcular_promedio_trimestre_materia_sql(est_id, m_id, t)
                            if p_t is not None:
                                trim_vals.append(p_t)
                        if trim_vals:
                            prom_m = round(sum(trim_vals) / len(trim_vals), 1)
                    else:
                        if isinstance(trimestre, int):
                            trim_num = trimestre
                        else:
                            try:
                                trim_num = int(str(trimestre).replace("Trimestre ", ""))
                            except ValueError:
                                trim_num = 1
                        prom_m = self._calcular_promedio_trimestre_materia_sql(est_id, m_id, trim_num)
                    if prom_m is not None:
                        tech_notas.append(prom_m)
                        
                tec_prom = sum(tech_notas) / len(tech_notas) if tech_notas else None
                
                comb_notas = list(norm_notas)
                if tec_prom is not None:
                    comb_notas.append(tec_prom)
                    
                if comb_notas:
                    results[est_nombre] = round(sum(comb_notas) / len(comb_notas), 2)
                    
        return results

    def _calcular_promedio_trimestre_materia_sql(self, estudiante_id, materia_id, trimestre):
        cursor = self.db_conn.cursor()
        cursor.execute("""
            SELECT tipo, valor 
            FROM notas 
            WHERE estudiante_id = ? AND materia_id = ? AND trimestre = ?;
        """, (estudiante_id, materia_id, trimestre))
        rows = cursor.fetchall()
        if not rows:
            return None
            
        diarias = []
        apreciaciones = []
        examenes = []
        
        for tipo, valor in rows:
            if valor is None:
                continue
            if tipo == "Diaria / Parcial":
                diarias.append(valor)
            elif tipo == "Apreciación":
                apreciaciones.append(valor)
            elif tipo == "Examen":
                examenes.append(valor)
                
        areas = []
        if diarias:
            areas.append(sum(diarias) / len(diarias))
        if apreciaciones:
            areas.append(sum(apreciaciones) / len(apreciaciones))
        if examenes:
            areas.append(sum(examenes) / len(examenes))
            
        if not areas:
            return None
            
        return round(sum(areas) / len(areas), 1)

    def obtener_promedios_reales(self, grado, materia, trimestre, wb=None):
        return self._calcular_promedios_sql(grado, materia, trimestre)

    def obtener_notas_estudiante(self, nombre_estudiante, grado, trimestre, wb=None):
        nombre_est_clean = nombre_estudiante.strip().lower()
        trim_str = f"Trimestre {trimestre}" if isinstance(trimestre, int) else trimestre
        materias = self.obtener_materias_por_grado(grado)
        notas_estudiante = {}
        for mat in materias:
            if mat in ["Sin materias registradas", "Sin materias", "No hay materias", "General"]:
                continue
            proms = self.obtener_promedios_reales(grado, mat, trim_str)
            for est_nom, nota in proms.items():
                if est_nom.strip().lower() == nombre_est_clean:
                    notas_estudiante[mat] = nota
                    break
        return notas_estudiante

    def obtener_promedios_reales_bulk(self, grado, materias, trimestre, wb=None):
        resultados = {}
        for m in materias:
            resultados[m] = self.obtener_promedios_reales(grado, m, trimestre)
        return resultados

    def obtener_historial_real(self, grado, materia, nombre_estudiante, wb=None):
        cursor = self.db_conn.cursor()
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        if not row_g:
            return [3.0, 3.0]
        grado_id = row_g[0]
        
        cursor.execute("SELECT id, nombre FROM estudiantes WHERE grado_id = ?;", (grado_id,))
        est_id = None
        for r in cursor.fetchall():
            dec_name = self.db_manager.desencriptar_campo(r[1])
            if dec_name.strip().lower() == nombre_estudiante.strip().lower():
                est_id = r[0]
                break
        if not est_id:
            return [3.0, 3.0]
        
        historial = []
        if materia and materia not in ["Sin materias", "No hay materias", "General"]:
            cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
            row_m = cursor.fetchone()
            if not row_m:
                return [3.0, 3.0]
            materia_id = row_m[0]
            
            for t in [1, 2, 3]:
                prom = self._calcular_promedio_trimestre_materia_sql(est_id, materia_id, t)
                if prom is not None:
                    historial.append(prom)
        else:
            for t in [1, 2, 3]:
                proms = self.obtener_promedios_reales(grado, "General", f"Trimestre {t}")
                if nombre_estudiante in proms:
                    historial.append(proms[nombre_estudiante])
                    
        return historial if len(historial) >= 2 else [3.0, 3.0] # Fallback to avoid math errors in scipy

    def obtener_datos_reportes(self, grado, trimestre="Todos", wb=None):
        """
        Genera datos de reportes docente/aprobados/dirección EXCLUSIVAMENTE desde SQLite.
        El parámetro wb se mantiene por compatibilidad de firma pero se ignora.
        """
        datos = {"docente": [], "aprobados": [], "direccion": []}
        try:
            cursor = self.db_conn.cursor()

            # Obtener grado_id
            cursor.execute(
                "SELECT id FROM grados WHERE nombre = ? AND modalidad = ? LIMIT 1;",
                (grado, self.modalidad)
            )
            row_g = cursor.fetchone()
            if not row_g:
                return datos
            grado_id = row_g[0]

            # Obtener estudiantes activos del grado
            cursor.execute(
                "SELECT id, nombre, cedula FROM estudiantes WHERE grado_id = ? AND estado = 'Activo' ORDER BY nombre;",
                (grado_id,)
            )
            estudiantes_raw = cursor.fetchall()
            if not estudiantes_raw:
                return datos

            # Obtener los promedios reales calculados correctamente por materia y área
            # Si trimestre es "Todos", pasamos "Anual"
            trim_query = "Anual" if trimestre == "Todos" else trimestre
            promedios_reales = self.obtener_promedios_reales(grado, None, trim_query)

            # Calcular promedio por estudiante consultando notas en SQL
            for est_id, est_nombre, est_cedula in estudiantes_raw:
                # Descifrar nombre y cédula
                nom = self.db_manager.desencriptar_campo(est_nombre) if est_nombre else ""
                ced = self.db_manager.desencriptar_campo(est_cedula) if est_cedula else ""

                # Obtener el promedio real de la estructura
                prom_val = promedios_reales.get(nom, None)

                if prom_val is not None:
                    prom_str = f"{prom_val:.2f}"
                    estado = "APROBADO" if prom_val >= 3.0 else "REPROBADO"
                else:
                    prom_str = "—"
                    estado = "SIN NOTAS"

                datos["docente"].append([nom, ced, prom_str])
                datos["aprobados"].append([nom, estado])
                datos["direccion"].append([nom, ced, prom_str, estado])

        except Exception as e:
            logger.error(f"Error generando datos de reportes desde SQL: {e}", exc_info=True)

        return datos

    def obtener_cuadro_honor_general(self, wb=None):
        """
        Calcula el cuadro de honor EXCLUSIVAMENTE desde SQLite.
        El parámetro wb se mantiene por compatibilidad de firma pero se ignora.
        """

        resultados = []
        try:
            cursor = self.db_conn.cursor()

            # Una sola query: para cada estudiante activo obtener promedio anual por trimestre
            # usando las notas almacenadas en SQL.
            # Estructura: por cada estudiante, promedio de (avg_T1 + avg_T2 + avg_T3) que tengan notas.
            cursor.execute("""
                SELECT
                    e.id,
                    e.nombre,
                    g.nombre AS grado,
                    n.trimestre,
                    AVG(n.valor) AS prom_trim
                FROM estudiantes e
                JOIN grados g ON g.id = e.grado_id
                JOIN notas n  ON n.estudiante_id = e.id
                WHERE e.estado = 'Activo'
                  AND n.valor IS NOT NULL
                GROUP BY e.id, n.trimestre
                ORDER BY e.id, n.trimestre;
            """)
            rows = cursor.fetchall()

            # Agrupar por estudiante
            from collections import defaultdict
            est_trims = defaultdict(lambda: {"nombre": "", "grado": "", "trims": []})
            for est_id, est_nombre_enc, grado_nombre, trim, prom_trim in rows:
                nom = self.db_manager.desencriptar_campo(est_nombre_enc) if est_nombre_enc else ""
                est_trims[est_id]["nombre"] = nom
                est_trims[est_id]["grado"] = grado_nombre
                est_trims[est_id]["trims"].append(prom_trim)

            for est_id, info in est_trims.items():
                trims = [v for v in info["trims"] if v is not None]
                if not trims:
                    continue
                prom_general = round(sum(trims) / len(trims), 2)
                if 4.5 <= prom_general <= 5.0:
                    resultados.append({
                        "nombre": info["nombre"],
                        "grado": info["grado"],
                        "promedio": prom_general
                    })

        except Exception as e:
            logger.error(f"Error calculando cuadro de honor desde SQL: {e}", exc_info=True)

        # Ordenar de mayor a menor promedio
        resultados = sorted(resultados, key=lambda x: x["promedio"], reverse=True)
        return resultados



    def get_dashboard_stats(self, wb=None):
        """
        Calcula estadísticas del dashboard EXCLUSIVAMENTE desde SQLite.
        El parámetro wb se mantiene por compatibilidad pero se ignora — ya no se lee Excel.
        """
        from utils.date_helpers import obtener_trimestre_actual
        trimestre_actual = obtener_trimestre_actual()

        # Valores por defecto
        total = 0
        riesgo = 0
        excusas = 0
        tareas_sin_nota = 0
        total_asist_dias = 0
        total_asist_ausencias = 0

        try:
            cursor = self.db_conn.cursor()

            # 1. Total de estudiantes activos
            cursor.execute("SELECT COUNT(*) FROM estudiantes WHERE estado = 'Activo';")
            total = cursor.fetchone()[0] or 0

            # 2. Alumnos en riesgo (promedio < 3.0 en trimestre actual, fallback a Anual)
            try:
                cursor.execute("SELECT nombre FROM grados WHERE modalidad = ?;", (self.modalidad,))
                grados_activos = [r[0] for r in cursor.fetchall()]
                riesgo = 0
                for g in grados_activos:
                    proms = self.obtener_promedios_reales(g, None, trimestre_actual)
                    if not proms:
                        proms = self.obtener_promedios_reales(g, None, "Anual")
                    for p in proms.values():
                        if p is not None and p < 3.0:
                            riesgo += 1
            except Exception:
                riesgo = 0

            # 3. Asistencia del trimestre actual
            try:
                trim_num = int(trimestre_actual.replace("Trimestre ", ""))
                cursor.execute("""
                    SELECT estado, COUNT(*) FROM asistencia
                    WHERE trimestre = ?
                    GROUP BY estado;
                """, (trim_num,))
                rows_asist = cursor.fetchall()
                for estado, cnt in rows_asist:
                    total_asist_dias += cnt
                    if estado == '-':
                        total_asist_ausencias += cnt
                    elif estado == 'E':
                        excusas += cnt
            except Exception:
                pass

            # 4. Hábitos y actitudes
            s_count = 0
            r_count = 0
            x_count = 0
            cursor.execute("SELECT nota, COUNT(*) FROM habitos GROUP BY nota;")
            for nota, cnt in cursor.fetchall():
                if nota == "S":
                    s_count = cnt
                elif nota == "R":
                    r_count = cnt
                elif nota == "X":
                    x_count = cnt

        except Exception as e:
            logger.error(f"Error cargando dashboard stats desde SQL: {e}")

        # 5. Cuadro de Honor (ya usa SQL internamente)
        try:
            cuadro_honor = self.obtener_cuadro_honor_general()
            honor_cant = len(cuadro_honor)
            best_student = cuadro_honor[0]["nombre"] if honor_cant > 0 else "N/A"
            best_prom = cuadro_honor[0]["promedio"] if honor_cant > 0 else 0.0
        except Exception:
            honor_cant = 0
            best_student = "N/A"
            best_prom = 0.0

        asistencia_pct = "—"
        if total_asist_dias > 0:
            presentes = total_asist_dias - total_asist_ausencias
            pct = (presentes / total_asist_dias) * 100
            asistencia_pct = f"{pct:.1f}%"

        honor = f"{best_student} ({best_prom:.2f})" if best_prom > 0.0 else "N/A"

        return {
            "total": total,
            "riesgo": riesgo,
            "honor": honor,
            "honor_cant": honor_cant,
            "asistencia": asistencia_pct,
            "tareas_sin_nota": tareas_sin_nota,
            "excusas": excusas,
            "habitos": {
                "S": s_count,
                "R": r_count,
                "X": x_count
            }
        }


    def _encontrar_hoja_prom(self, wb, grado, materia):
        materia_clean = materia.lower().replace(" ", "").replace(".", "")
        for sheet in wb.sheetnames:
            if "PROM" in sheet.upper():
                sheet_clean = sheet.lower().replace(" ", "").replace(".", "")
                if materia_clean in sheet_clean:
                    if self.modalidad == "primaria" or grado.replace("°","") in sheet: return sheet
        return None

    def _obtener_rango_columnas(self, ws, trimestre, tipo_nota):
        textos_busqueda = {"Diaria / Parcial": "PARCIAL", "Apreciación": "APRECIACIÓN", "Examen": "PRUEBA"}
        texto_b = textos_busqueda[tipo_nota]
        ocurrencias = []
        for c in range(1, 200):
            val = ws.cell(row=3, column=c).value
            if val and texto_b in str(val).upper(): ocurrencias.append(c)
        if not ocurrencias: return None, None
        idx_trimestre = int(trimestre.replace("Trimestre ", "")) - 1
        if idx_trimestre >= len(ocurrencias): return None, None
        col_inicio = ocurrencias[idx_trimestre]
        col_fin = col_inicio
        for c in range(col_inicio + 1, col_inicio + 25):
            val = ws.cell(row=3, column=c).value
            if val and ("PROMEDIO" in str(val).upper() or "CALIFICACIÓN" in str(val).upper() or "NOTAS" in str(val).upper() or "PRUEBAS" in str(val).upper()):
                break
            col_fin = c
        return col_inicio, col_fin

    def guardar_columna_notas(self, grado, materia, trimestre, tipo_nota, fecha, desc, dic_notas):
        from utils.date_helpers import normalizar_fecha_a_mm_dd
        trim_num = int(str(trimestre).replace("Trimestre ", ""))
        fecha = normalizar_fecha_a_mm_dd(fecha, trimestre=trim_num)
        if os.path.basename(self.ruta).startswith("Registro_"):
            if not self._wb_cache:
                self._verificar_y_recargar_cache()
            wb = self._wb_cache
            nombre_hoja = self._encontrar_hoja_prom(wb, grado, materia)
            if not nombre_hoja:
                return False, f"No se encontró la hoja PROM para {materia} {grado}"
            ws = wb[nombre_hoja]
            rango = self._obtener_rango_columnas(ws, trimestre, tipo_nota)
            if rango == (None, None):
                return False, "No se encontró el bloque en el Excel."
            col_inicio, col_fin = rango
            limite_columnas = col_fin - col_inicio + 1
            
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = ? AND modalidad = ?;", (grado, "A", self.modalidad))
            row_g = cursor.fetchone()
            if not row_g:
                return False, "Grado no encontrado."
            grado_id = row_g[0]
            
            cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
            row_m = cursor.fetchone()
            if not row_m:
                return False, "Materia no encontrada."
            materia_id = row_m[0]
            
            cursor.execute("""
                SELECT COUNT(DISTINCT descripcion) FROM notas
                WHERE materia_id = ? AND trimestre = ? AND tipo = ?;
            """, (materia_id, trim_num, tipo_nota))
            columnas_ocupadas = cursor.fetchone()[0]
            
            if columnas_ocupadas >= limite_columnas:
                if tipo_nota == "Examen": return False, "Límite alcanzado: Solo hay espacio para 2 exámenes."
                return False, "El bloque de notas está lleno."
                
            if tipo_nota == "Examen": desc = f"Examen {columnas_ocupadas + 1} ({fecha})"
            
            for id_estudiante, val_nota in dic_notas.items():
                cursor.execute(
                    """INSERT OR REPLACE INTO notas (estudiante_id, materia_id, trimestre, tipo, descripcion, valor, fecha)
                    VALUES (?, ?, ?, ?, ?, ?, ?);""",
                    (str(id_estudiante), materia_id, trim_num, tipo_nota, desc, float(val_nota) if val_nota is not None else None, str(fecha))
                )
            self.db_conn.commit()
            self._schedule_save()
            self.actualizar_resumen(grado)
            return True, ""

        if not os.path.exists(self.ruta): return False, "El archivo Excel no existe."
        wb = openpyxl.load_workbook(self.ruta)
        nombre_hoja = self._encontrar_hoja_prom(wb, grado, materia)
        if not nombre_hoja:
            wb.close()
            return False, f"No se encontró la hoja PROM para {materia} {grado}"
        ws = wb[nombre_hoja]
        rango = self._obtener_rango_columnas(ws, trimestre, tipo_nota)
        if rango == (None, None):
            wb.close()
            return False, "No se encontró el bloque en el Excel."
        col_inicio, col_fin = rango
        col_vacia = None
        columnas_ocupadas = 0
        for c in range(col_inicio, col_fin + 1):
            if not ws.cell(row=self.fila_desc, column=c).value:
                col_vacia = c
                break
            else: columnas_ocupadas += 1
        if not col_vacia:
            wb.close()
            if tipo_nota == "Examen": return False, "Límite alcanzado: Solo hay espacio para 2 exámenes."
            return False, "El bloque de notas está lleno en el Excel."
            
        if tipo_nota == "Examen": desc = f"Examen {columnas_ocupadas + 1} ({fecha})"
        try: ws.cell(row=4, column=col_vacia).value = fecha
        except AttributeError:
            if tipo_nota != "Examen": desc = f"{fecha} - {desc}"
        try:
            cell_desc = ws.cell(row=self.fila_desc, column=col_vacia)
            cell_desc.value = desc
            cell_desc.alignment = Alignment(wrapText=True, horizontal='center', vertical='center', textRotation=90)
            cell_desc.font = Font(name='Calibri', size=11, bold=True)
        except AttributeError: pass
        for id_estudiante, nota in dic_notas.items():
            fila_excel = 4 + (int(id_estudiante) % 100)
            try: ws.cell(row=fila_excel, column=col_vacia).value = nota
            except AttributeError: pass
            
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = ? AND modalidad = ?;", (grado, "A", self.modalidad))
            row_g = cursor.fetchone()
            if row_g:
                grado_id = row_g[0]
                cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
                row_m = cursor.fetchone()
                if row_m:
                    materia_id = row_m[0]
                    trim_num = int(trimestre.replace("Trimestre ", ""))
                    for id_estudiante, val_nota in dic_notas.items():
                        cursor.execute(
                            """INSERT OR REPLACE INTO notas (estudiante_id, materia_id, trimestre, tipo, descripcion, valor, fecha)
                            VALUES (?, ?, ?, ?, ?, ?, ?);""",
                            (str(id_estudiante), materia_id, trim_num, tipo_nota, desc, float(val_nota) if val_nota is not None else None, str(fecha))
                        )
                    self.db_conn.commit()
                    self._schedule_save()
        except Exception as e:
            logger.error(f"Error guardando notas en SQLite: {e}")

        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        self.actualizar_resumen(grado)
        return True, ""

    def obtener_descripciones_notas(self, grado, materia, trimestre, tipo_nota, wb=None):
        if wb is not None or not os.path.basename(self.ruta).startswith("Registro_"):
            if wb is None:
                self._verificar_y_recargar_cache()
            if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None: return []
            should_close = not bool(self._wb_cache) if wb is None else False
            if wb is None:
                wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
            nombre_hoja = self._encontrar_hoja_prom(wb, grado, materia)
            if not nombre_hoja:
                if should_close: wb.close()
                return []
            ws = wb[nombre_hoja]
            rango = self._obtener_rango_columnas(ws, trimestre, tipo_nota)
            if rango == (None, None):
                if should_close: wb.close()
                return []
            col_inicio, col_fin = rango
            descripciones = []
            for c in range(col_inicio, col_fin + 1):
                try:
                    val = ws.cell(row=self.fila_desc, column=c).value
                    if val:
                        desc_limpia = str(val).replace('\n', ' ').strip()
                        descripciones.append(desc_limpia)
                except AttributeError: continue
            if should_close: wb.close()
            return descripciones

        cursor = self.db_conn.cursor()
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        if not row_g:
            return []
        grado_id = row_g[0]
        
        cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
        row_m = cursor.fetchone()
        if not row_m:
            return []
        materia_id = row_m[0]
        
        trim_num = int(str(trimestre).replace("Trimestre ", ""))
        cursor.execute("""
            SELECT DISTINCT descripcion FROM notas 
            WHERE materia_id = ? AND trimestre = ? AND tipo = ?;
        """, (materia_id, trim_num, tipo_nota))
        rows = cursor.fetchall()
        return [r[0] for r in rows if r[0]]

    def buscar_notas_por_descripcion_exacta(self, grado, materia, trimestre, tipo_nota, descripcion, wb=None):
        cursor = self.db_conn.cursor()
        
        # Obtener grado_id
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        if not row_g:
            return None
        grado_id = row_g[0]
        
        # Obtener materia_id
        cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
        row_m = cursor.fetchone()
        if not row_m:
            return None
        materia_id = row_m[0]
        
        trim_num = int(trimestre.replace("Trimestre ", ""))
        
        # Buscar notas coincidentes en la base de datos
        cursor.execute("""
            SELECT estudiante_id, valor, fecha 
            FROM notas 
            WHERE materia_id = ? AND trimestre = ? AND tipo = ? AND descripcion = ?;
        """, (materia_id, trim_num, tipo_nota, descripcion))
        rows = cursor.fetchall()
        
        if not rows:
            return None
            
        notas = {}
        fecha_nota = ""
        for est_id, val, fecha in rows:
            notas[int(est_id)] = val
            if not fecha_nota and fecha:
                fecha_nota = fecha
                
        return {
            "columna": descripcion,
            "notas": notas,
            "fecha": fecha_nota or datetime.date.today().strftime("%d-%m")
        }

    def actualizar_notas_existentes(self, grado, materia, columna, dic_notas, nueva_fecha=None):
        if os.path.basename(self.ruta).startswith("Registro_"):
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
            row_g = cursor.fetchone()
            if not row_g:
                return False
            grado_id = row_g[0]
            
            cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
            row_m = cursor.fetchone()
            if not row_m:
                return False
            materia_id = row_m[0]
            
            desc_limpia = None
            if isinstance(columna, str):
                desc_limpia = columna.replace('\n', ' ').strip()
            else:
                if not self._wb_cache:
                    self._verificar_y_recargar_cache()
                nombre_hoja = self._encontrar_hoja_prom(self._wb_cache, grado, materia)
                if nombre_hoja:
                    ws = self._wb_cache[nombre_hoja]
                    desc = ws.cell(row=self.fila_desc, column=columna).value
                    if desc:
                        desc_limpia = str(desc).replace('\n', ' ').strip()
            
            if not desc_limpia:
                return False
                
            # Calcular nueva fecha y descripción si aplica
            nueva_desc = desc_limpia
            fecha_final = None
            if nueva_fecha:
                from utils.date_helpers import normalizar_fecha_a_mm_dd
                cursor.execute("SELECT trimestre FROM notas WHERE materia_id = ? AND descripcion = ? LIMIT 1;", (materia_id, desc_limpia))
                r_trim = cursor.fetchone()
                trim_val = r_trim[0] if r_trim else 1
                fecha_final = normalizar_fecha_a_mm_dd(nueva_fecha, trimestre=trim_val)
                if desc_limpia.startswith("Examen "):
                    import re
                    nueva_desc = re.sub(r'\([\d\-/\s]+\)$', f'({fecha_final})', desc_limpia)
            
            for id_estudiante, val_nota in dic_notas.items():
                cursor.execute(
                    """SELECT id FROM notas WHERE estudiante_id = ? AND materia_id = ? AND descripcion = ?;""",
                    (str(id_estudiante), materia_id, desc_limpia)
                )
                row_n = cursor.fetchone()
                if row_n:
                    cursor.execute(
                        "UPDATE notas SET valor = ? WHERE id = ?;",
                        (float(val_nota) if val_nota is not None else None, row_n[0])
                    )
                else:
                    cursor.execute(
                        """INSERT INTO notas (estudiante_id, materia_id, trimestre, tipo, descripcion, valor, fecha)
                        VALUES (?, ?, 1, 'Diaria / Parcial', ?, ?, ?);""",
                        (str(id_estudiante), materia_id, desc_limpia, float(val_nota) if val_nota is not None else None, fecha_final or datetime.date.today().strftime("%m-%d"))
                    )
            
            if fecha_final:
                cursor.execute(
                    "UPDATE notas SET fecha = ?, descripcion = ? WHERE materia_id = ? AND descripcion = ?;",
                    (str(fecha_final), nueva_desc, materia_id, desc_limpia)
                )
                
            self.db_conn.commit()
            self._schedule_save()
            self.actualizar_resumen(grado)
            return True

        if not os.path.exists(self.ruta): return False
        nombre_hoja = self._encontrar_hoja_prom(self._wb_cache, grado, materia) if self._wb_cache else None
        wb = openpyxl.load_workbook(self.ruta)
        if not nombre_hoja:
            nombre_hoja = self._encontrar_hoja_prom(wb, grado, materia)
        if not nombre_hoja: 
            wb.close()
            return False
        ws = wb[nombre_hoja]
        
        # En caso de nueva_fecha, actualizar en Excel antes de procesar
        if nueva_fecha:
            from utils.date_helpers import normalizar_fecha_a_mm_dd
            trim_val = 1
            for t_idx in (1, 2, 3):
                for t_type in ("Diaria / Parcial", "Apreciación", "Examen"):
                    r_col = self._obtener_rango_columnas(ws, f"Trimestre {t_idx}", t_type)
                    if r_col[0] is not None and r_col[0] <= columna <= r_col[1]:
                        trim_val = t_idx
                        break
            nueva_fecha_normalized = normalizar_fecha_a_mm_dd(nueva_fecha, trimestre=trim_val)
            ws.cell(row=4, column=columna).value = nueva_fecha_normalized
            desc_val = ws.cell(row=self.fila_desc, column=columna).value
            if desc_val:
                desc_str = str(desc_val).replace('\n', ' ').strip()
                if desc_str.startswith("Examen "):
                    import re
                    nueva_desc_str = re.sub(r'\([\d\-/\s]+\)$', f'({nueva_fecha_normalized})', desc_str)
                    ws.cell(row=self.fila_desc, column=columna).value = nueva_desc_str
                    
        for id_estudiante, nota in dic_notas.items():
            fila_excel = 4 + (int(id_estudiante) % 100)
            try: ws.cell(row=fila_excel, column=columna).value = nota
            except AttributeError: continue
        try:
            desc = ws.cell(row=self.fila_desc, column=columna).value
            fecha = ws.cell(row=4, column=columna).value or "01-01"
            if desc:
                desc_limpia = str(desc).replace('\n', ' ').strip()
                cursor = self.db_conn.cursor()
                cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = ? AND modalidad = ?;", (grado, "A", self.modalidad))
                row_g = cursor.fetchone()
                if row_g:
                    grado_id = row_g[0]
                    cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
                    row_m = cursor.fetchone()
                    if row_m:
                        materia_id = row_m[0]
                        for id_estudiante, val_nota in dic_notas.items():
                            cursor.execute(
                                """SELECT id FROM notas WHERE estudiante_id = ? AND materia_id = ? AND descripcion = ?;""",
                                (str(id_estudiante), materia_id, desc_limpia)
                            )
                            row_n = cursor.fetchone()
                            if row_n:
                                cursor.execute(
                                    "UPDATE notas SET valor = ?, fecha = ? WHERE id = ?;",
                                    (float(val_nota) if val_nota is not None else None, str(fecha), row_n[0])
                                )
                            else:
                                cursor.execute(
                                    """INSERT INTO notas (estudiante_id, materia_id, trimestre, tipo, descripcion, valor, fecha)
                                    VALUES (?, ?, 1, 'Diaria / Parcial', ?, ?, ?);""",
                                    (str(id_estudiante), materia_id, desc_limpia, float(val_nota) if val_nota is not None else None, str(fecha))
                                )
                        self.db_conn.commit()
                        self._schedule_save()
        except Exception as e:
            logger.error(f"Error actualizando notas en SQLite: {e}")
        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        self.actualizar_resumen(grado)
        return True

    def _encontrar_hoja_asistencia(self, wb, grado):
        for sheet in wb.sheetnames:
            if "Asistencia" in sheet and (self.modalidad == "primaria" or grado.replace('°','') in sheet):
                return sheet
        return None

    def obtener_fechas_asistencia(self, grado, trimestre, wb=None):
        """Retorna las fechas distintas de asistencia para un grado y trimestre."""
        from utils.date_helpers import parse_fecha_segura
        if wb is not None or not os.path.basename(self.ruta).startswith("Registro_"):
            if wb is None:
                self._verificar_y_recargar_cache()
            if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None:
                return []
            should_close = not bool(self._wb_cache) if wb is None else False
            if wb is None:
                wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
            hoja = self._encontrar_hoja_asistencia(wb, grado)
            if not hoja:
                if should_close: wb.close()
                return []
            ws = wb[hoja]
            mapa_trimestres = {"Trimestre 1": 2, "Trimestre 2": 45, "Trimestre 3": 88}
            fila_fechas = mapa_trimestres.get(trimestre, 2)
            fechas = []
            for c in range(3, 61):
                val = ws.cell(row=fila_fechas, column=c).value
                if val:
                    fechas.append(str(val).strip())
            if should_close: wb.close()
            fechas.sort(key=parse_fecha_segura)
            return fechas

        cursor = self.db_conn.cursor()
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        if not row_g:
            return []
        grado_id = row_g[0]
        
        trim_num = int(str(trimestre).replace("Trimestre ", ""))
        cursor.execute("""
            SELECT DISTINCT fecha FROM asistencia a
            JOIN estudiantes e ON a.estudiante_id = e.id
            WHERE e.grado_id = ? AND a.trimestre = ?;
        """, (grado_id, trim_num))
        rows = cursor.fetchall()
        fechas = [r[0] for r in rows]
        fechas.sort(key=parse_fecha_segura)
        return fechas

    def buscar_asistencia_existente(self, grado, trimestre, fecha, wb=None):
        if wb is not None or not os.path.basename(self.ruta).startswith("Registro_"):
            if wb is None:
                self._verificar_y_recargar_cache()
            if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None: return None
            should_close = not bool(self._wb_cache) if wb is None else False
            if wb is None:
                wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
            hoja = self._encontrar_hoja_asistencia(wb, grado)
            if not hoja:
                if should_close: wb.close()
                return None
            ws = wb[hoja]
            mapa_trimestres = {"Trimestre 1": 2, "Trimestre 2": 45, "Trimestre 3": 88}
            fila_fechas = mapa_trimestres.get(trimestre, 2)
            col_encontrada = None
            for c in range(3, 61):
                val = ws.cell(row=fila_fechas, column=c).value
                if val and str(val).strip() == fecha.strip():
                    col_encontrada = c
                    break
            if not col_encontrada:
                if should_close: wb.close()
                return None
            asistencia = {}
            for id_est in range(1, 46): 
                val = ws.cell(row=fila_fechas + id_est, column=col_encontrada).value
                if val is not None: asistencia[id_est] = val
            if should_close: wb.close()
            return {"columna": col_encontrada, "asistencia": asistencia}

        cursor = self.db_conn.cursor()
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        if not row_g:
            return None
        grado_id = row_g[0]
        trim_num = int(str(trimestre).replace("Trimestre ", ""))
        
        cursor.execute("""
            SELECT a.estudiante_id, a.estado FROM asistencia a
            JOIN estudiantes e ON a.estudiante_id = e.id
            WHERE e.grado_id = ? AND a.trimestre = ? AND a.fecha = ?;
        """, (grado_id, trim_num, fecha))
        rows = cursor.fetchall()
        if not rows:
            return None
            
        asistencia = {}
        for est_id, estado in rows:
            try:
                idx_1based = int(est_id) % 100 if str(est_id).isdigit() else est_id
                asistencia[idx_1based] = estado
            except ValueError:
                pass
                
        return {"columna": 999, "asistencia": asistencia}

    def guardar_asistencia(self, grado, trimestre, fecha, dic_asistencia):
        from utils.date_helpers import normalizar_fecha_a_mm_dd
        trim_num = int(str(trimestre).replace("Trimestre ", ""))
        fecha = normalizar_fecha_a_mm_dd(fecha, trimestre=trim_num)
        if os.path.basename(self.ruta).startswith("Registro_"):
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
                row_g = cursor.fetchone()
                if not row_g:
                    return False, "Grado no encontrado."
                grado_id = row_g[0]
                
                cursor.execute("SELECT id FROM estudiantes WHERE grado_id = ? ORDER BY CAST(id AS INTEGER);", (grado_id,))
                student_ids = [r[0] for r in cursor.fetchall()]
                
                for id_estudiante, datos in dic_asistencia.items():
                    try:
                        idx_1based = int(id_estudiante) % 100 if str(id_estudiante).isdigit() else int(id_estudiante)
                        est_id = student_ids[idx_1based - 1]
                        cursor.execute(
                            """INSERT OR REPLACE INTO asistencia (estudiante_id, fecha, estado, trimestre)
                            VALUES (?, ?, ?, ?);""",
                            (str(est_id), str(fecha), datos["estado"], trim_num)
                        )
                    except (IndexError, ValueError) as e:
                        logger.error(f"Error mapping student in guardar_asistencia: {e}")
                self.db_conn.commit()
                self._schedule_save()
                self.actualizar_resumen(grado)
                return True, ""
            except Exception as e:
                logger.error(f"Error guardando asistencia en SQLite: {e}")
                return False, str(e)

        if not os.path.exists(self.ruta): return False, "El archivo Excel no existe."
        wb = openpyxl.load_workbook(self.ruta)
        hoja = self._encontrar_hoja_asistencia(wb, grado)
        if not hoja:
            wb.close()
            return False, f"No se encontró la hoja de Asistencia para {grado}."
        ws = wb[hoja]
        mapa_trimestres = {"Trimestre 1": 2, "Trimestre 2": 45, "Trimestre 3": 88}
        fila_fechas = mapa_trimestres.get(trimestre, 2)
        col_vacia = None
        for c in range(3, 61):
            if not ws.cell(row=fila_fechas, column=c).value:
                col_vacia = c
                break
        if not col_vacia:
            wb.close()
            return False, "No hay más columnas vacías para este trimestre."
        ws.cell(row=fila_fechas, column=col_vacia).value = fecha
        fuente_meduca = Font(name='Calibri', size=9, bold=True)
        for id_estudiante, datos in dic_asistencia.items():
            fila_excel = fila_fechas + (int(id_estudiante) % 100) 
            celda = ws.cell(row=fila_excel, column=col_vacia)
            celda.value = datos["estado"]
            celda.font = fuente_meduca

        try:
            cursor = self.db_conn.cursor()
            trim_num = int(str(trimestre).replace("Trimestre ", ""))
            for id_estudiante, datos in dic_asistencia.items():
                cursor.execute(
                    """INSERT OR REPLACE INTO asistencia (estudiante_id, fecha, estado, trimestre)
                    VALUES (?, ?, ?, ?);""",
                    (str(id_estudiante), str(fecha), datos["estado"], trim_num)
                )
            self.db_conn.commit()
            self._schedule_save()
        except Exception as e:
            logger.error(f"Error guardando asistencia en SQLite: {e}")

        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        self.actualizar_resumen(grado)
        return True, ""

    def actualizar_asistencia(self, grado, trimestre, columna, dic_asistencia, nueva_fecha=None):
        if os.path.basename(self.ruta).startswith("Registro_"):
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
                row_g = cursor.fetchone()
                if not row_g:
                    return False
                grado_id = row_g[0]
                
                cursor.execute("SELECT id FROM estudiantes WHERE grado_id = ? ORDER BY CAST(id AS INTEGER);", (grado_id,))
                student_ids = [r[0] for r in cursor.fetchall()]
                
                trim_num = int(str(trimestre).replace("Trimestre ", ""))
                
                fecha = None
                if isinstance(columna, str):
                    fecha = columna.strip()
                else:
                    cursor.execute("""
                        SELECT DISTINCT fecha FROM asistencia a
                        JOIN estudiantes e ON a.estudiante_id = e.id
                        WHERE e.grado_id = ? AND a.trimestre = ? ORDER BY fecha;
                    """, (grado_id, trim_num))
                    fechas = [r[0] for r in cursor.fetchall()]
                    idx = columna - 3
                    if 0 <= idx < len(fechas):
                        fecha = fechas[idx]
                
                if not fecha:
                    return False
                
                fecha_final = fecha
                if nueva_fecha:
                    from utils.date_helpers import normalizar_fecha_a_mm_dd
                    fecha_final = normalizar_fecha_a_mm_dd(nueva_fecha, trimestre=trim_num)
                    if fecha_final != fecha:
                        cursor.execute("""
                            DELETE FROM asistencia WHERE estudiante_id IN (
                                SELECT id FROM estudiantes WHERE grado_id = ?
                            ) AND fecha = ? AND trimestre = ?;
                        """, (grado_id, str(fecha), trim_num))
                    
                for id_estudiante, datos in dic_asistencia.items():
                    try:
                        idx_1based = int(id_estudiante) % 100 if str(id_estudiante).isdigit() else int(id_estudiante)
                        est_id = student_ids[idx_1based - 1]
                        cursor.execute(
                            """INSERT OR REPLACE INTO asistencia (estudiante_id, fecha, estado, trimestre)
                            VALUES (?, ?, ?, ?);""",
                            (str(est_id), str(fecha_final), datos["estado"], trim_num)
                        )
                    except (IndexError, ValueError) as e:
                        logger.error(f"Error mapping student in actualizar_asistencia: {e}")
                self.db_conn.commit()
                self._schedule_save()
                self.actualizar_resumen(grado)
                return True
            except Exception as e:
                logger.error(f"Error actualizando asistencia en SQLite: {e}")
                return False

        if not os.path.exists(self.ruta): return False
        wb = openpyxl.load_workbook(self.ruta)
        hoja = self._encontrar_hoja_asistencia(wb, grado)
        if not hoja:
            wb.close()
            return False
        ws = wb[hoja]
        mapa_trimestres = {"Trimestre 1": 2, "Trimestre 2": 45, "Trimestre 3": 88}
        fila_fechas = mapa_trimestres.get(trimestre, 2)
        fuente_meduca = Font(name='Calibri', size=9, bold=True)
        
        old_fecha = ws.cell(row=fila_fechas, column=columna).value
        fecha_final = old_fecha or "01-01"
        trim_num = int(str(trimestre).replace("Trimestre ", ""))
        
        if nueva_fecha:
            from utils.date_helpers import normalizar_fecha_a_mm_dd
            fecha_final = normalizar_fecha_a_mm_dd(nueva_fecha, trimestre=trim_num)
            ws.cell(row=fila_fechas, column=columna).value = fecha_final
            
        for id_estudiante, datos in dic_asistencia.items():
            fila_excel = fila_fechas + (int(id_estudiante) % 100)
            celda = ws.cell(row=fila_excel, column=columna)
            celda.value = datos["estado"]
            celda.font = fuente_meduca

        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
            row_g = cursor.fetchone()
            if row_g:
                grado_id = row_g[0]
                if nueva_fecha and old_fecha and str(old_fecha) != str(fecha_final):
                    cursor.execute("""
                        DELETE FROM asistencia WHERE estudiante_id IN (
                            SELECT id FROM estudiantes WHERE grado_id = ?
                        ) AND fecha = ? AND trimestre = ?;
                    """, (grado_id, str(old_fecha), trim_num))
                    
                for id_estudiante, datos in dic_asistencia.items():
                    cursor.execute(
                        """INSERT OR REPLACE INTO asistencia (estudiante_id, fecha, estado, trimestre)
                        VALUES (?, ?, ?, ?);""",
                        (str(id_estudiante), str(fecha_final), datos["estado"], trim_num)
                    )
                self.db_conn.commit()
                self._schedule_save()
        except Exception as e:
            logger.error(f"Error actualizando asistencia en SQLite: {e}")

        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True

    def actualizar_datos_generales(self, nombre_docente, ano_lectivo):
        if not os.path.exists(self.ruta): return False
        wb = openpyxl.load_workbook(self.ruta)
        nombre_maestro = self._encontrar_hoja_maestro(wb)
        ws_m = wb[nombre_maestro]
        titulo_actual = str(ws_m.cell(row=1, column=1).value or "")
        nuevo_titulo = re.sub(r'20\d{2}', str(ano_lectivo), titulo_actual)
        self._safe_set_value(ws_m, 1, 1, nuevo_titulo)
        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True

    def obtener_consejero_actual(self, grado, wb=None):
        if wb is None:
            self._verificar_y_recargar_cache()
        if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None: return "No asignado"
        should_close = not bool(self._wb_cache) if wb is None else False
        if wb is None:
            wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
        grado_num = grado.replace("°", "")
        consejero = "No asignado"
        for sheet in wb.sheetnames:
            if "PLANILLA" in sheet.upper() and (self.modalidad == "primaria" or grado_num in sheet):
                ws = wb[sheet]
                for r in range(1, 15):
                    for c in range(1, 15):
                        val = str(ws.cell(row=r, column=c).value or "").upper()
                        if "CONSEJERO" in val or "CONSEJERA" in val:
                            if len(val) > 25: 
                                consejero = val.split(":")[-1].strip()
                            else:
                                consejero = str(ws.cell(row=r, column=c+2).value or "").strip()
                            if should_close: wb.close()
                            return consejero
        if should_close: wb.close()
        return consejero

    def actualizar_consejero(self, grado, nuevo_consejero):
        if not os.path.exists(self.ruta): return False
        wb = openpyxl.load_workbook(self.ruta)
        grado_num = grado.replace("°", "")
        for sheet in wb.sheetnames:
            sheet_upper = sheet.upper()
            if "PROM" in sheet_upper and (self.modalidad == "primaria" or grado_num in sheet):
                ws = wb[sheet]
                for r in range(1, 15):
                    for c in range(1, 15):
                        try:
                            celda = ws.cell(row=r, column=c)
                            val = str(celda.value or "").upper()
                            if "CONSEJERO" in val or "CONSEJERA" in val:
                                celda.value = f"PROF. CONSEJERO: {nuevo_consejero.upper()}"
                        except AttributeError: pass
            
            elif "PLANILLA" in sheet_upper and (self.modalidad == "primaria" or grado_num in sheet):
                ws = wb[sheet]
                for r in range(1, 15):
                    for c in range(1, 15):
                        try:
                            celda = ws.cell(row=r, column=c)
                            val = str(celda.value or "").upper()
                            if "CONSEJERO" in val or "CONSEJERA" in val:
                                if len(val) > 20: celda.value = f"PROF. CONSEJERO (A): {nuevo_consejero.upper()}"
                                else:
                                    ws.cell(row=r, column=c+1).value = nuevo_consejero.upper()
                                    ws.cell(row=r, column=c+2).value = nuevo_consejero.upper()
                        except AttributeError: pass
        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True

    def agregar_grado(self, nuevo_grado, consejero, jornada):
        if not os.path.exists(self.ruta): return False, "Archivo no encontrado"
        wb = openpyxl.load_workbook(self.ruta)
        hoja_base_asist = None
        for sheet in wb.sheetnames:
            if "Asistencia" in sheet:
                hoja_base_asist = sheet
                break
        if not hoja_base_asist:
            wb.close()
            return False, "No se encontró hoja de Asistencia base para clonar."

        nueva_hoja_asist = f"Asistencia ({nuevo_grado})"
        if nueva_hoja_asist in wb.sheetnames:
            wb.close()
            return False, f"El grado {nuevo_grado} ya existe."

        ws_nueva = wb.copy_worksheet(wb[hoja_base_asist])
        ws_nueva.title = nueva_hoja_asist
        for r in range(2, 60):
            for c in range(3, 100):
                try: ws_nueva.cell(row=r, column=c).value = None
                except AttributeError: pass

        if self.modalidad == "primaria":
            hoja_base_maestro = None
            for sheet in wb.sheetnames:
                if "MAESTRO" in sheet.upper():
                    hoja_base_maestro = sheet
                    break
            if hoja_base_maestro:
                nueva_hoja_maestro = f"MAESTRO ({nuevo_grado})"
                if nueva_hoja_maestro not in wb.sheetnames:
                    ws_maestro_nueva = wb.copy_worksheet(wb[hoja_base_maestro])
                    ws_maestro_nueva.title = nueva_hoja_maestro
                    # Clean the names and keep the layout
                    for r in range(5, 50):
                        try:
                            self._safe_clear_value(ws_maestro_nueva, r, 2)
                        except AttributeError: pass
                    self._safe_set_value(ws_maestro_nueva, 3, 2, f"GRADO {nuevo_grado}")
        else:
            nombre_maestro = self._encontrar_hoja_maestro(wb)
            ws_m = wb[nombre_maestro]
            col_vacia = None
            for c in range(2, 40, 2):
                if not ws_m.cell(row=4, column=c).value:
                    col_vacia = c
                    break
            if col_vacia:
                self._safe_set_value(ws_m, 3, col_vacia, f"GRADO {nuevo_grado}")
                self._safe_set_value(ws_m, 4, col_vacia, "NOMBRES Y APELLIDOS")
                if self.modalidad == "premedia":
                    self._safe_set_value(ws_m, 4, col_vacia+1, "N° DE CÉDULA")
                    
        hoja_base_resumen = None
        for sheet in wb.sheetnames:
            if "RESUMEN" in sheet.upper():
                hoja_base_resumen = sheet
                break
                
        if hoja_base_resumen:
            ws_resumen = wb.copy_worksheet(wb[hoja_base_resumen])
            ws_resumen.title = f"RESUMEN ({nuevo_grado})"
            for r in range(4, 50):
                for c in range(1, 30):
                    try:
                        celda = ws_resumen.cell(row=r, column=c)
                        self._safe_clear_value(ws_resumen, r, c)
                    except AttributeError: pass
        
        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True, "Grado creado exitosamente."

    def eliminar_grado(self, grado):
        if not os.path.exists(self.ruta): return False
        wb = openpyxl.load_workbook(self.ruta)
        hojas_borrar = []
        grado_limpio = grado.replace("°", "")
        for sheet in wb.sheetnames:
            if f"({grado})" in sheet or f" {grado_limpio})" in sheet or f" {grado})" in sheet:
                hojas_borrar.append(sheet)

        if not hojas_borrar:
            wb.close()
            return False

        for h in hojas_borrar: del wb[h]

        nombre_maestro = self._encontrar_hoja_maestro(wb)
        ws_m = wb[nombre_maestro]
        for c in range(1, 40):
            val = str(ws_m.cell(row=3, column=c).value or "")
            if grado in val:
                for r in range(3, 51):
                    try:
                        self._safe_clear_value(ws_m, r, c)
                        self._safe_clear_value(ws_m, r, c+1)
                    except AttributeError: pass
                break
        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True

    def clonar_materia(self, grado, materia_origen, nueva_materia, jornada):
        if not os.path.exists(self.ruta): return False, "Archivo no encontrado"
        wb = openpyxl.load_workbook(self.ruta)
        hoja_prom_origen = self._encontrar_hoja_prom(wb, grado, materia_origen)
        hoja_plan_origen = None
        grado_num = grado.replace("°", "")

        consejero = "No asignado"
        for sheet in wb.sheetnames:
            if "PLANILLA" in sheet.upper() and (self.modalidad == "primaria" or grado_num in sheet):
                ws = wb[sheet]
                for r in range(1, 15):
                    for c in range(1, 15):
                        val = str(ws.cell(row=r, column=c).value or "").upper()
                        if "CONSEJERO" in val or "CONSEJERA" in val:
                            if len(val) > 25: consejero = val.split(":")[-1].strip()
                            else: consejero = str(ws.cell(row=r, column=c+2).value or "").strip()
                            break

        if self.modalidad == "premedia":
            for sheet in wb.sheetnames:
                if "PLANILLA" in sheet.upper() and materia_origen.lower().replace(" ", "") in sheet.lower().replace(" ", "") and grado_num in sheet:
                    hoja_plan_origen = sheet
                    break

        if not hoja_prom_origen:
            wb.close()
            return False, "No se encontró materia base para clonar."

        nueva_prom = wb.copy_worksheet(wb[hoja_prom_origen])
        if self.modalidad == "primaria": nueva_prom.title = f"PROM ({nueva_materia.title()})"
        else: nueva_prom.title = f"PROM ({nueva_materia.title()} {grado})"
        
        for r in range(1, 15):
            for c in range(1, 15):
                try:
                    val = str(nueva_prom.cell(row=r, column=c).value or "").upper()
                    if "ASIGNATURA" in val and len(val) < 20: self._safe_set_value(nueva_prom, r, c, f"ASIGNATURA: {nueva_materia.upper()}")
                    if "CONSEJERO" in val and len(val) < 30: self._safe_set_value(nueva_prom, r, c, f"PROF. CONSEJERO: {consejero.upper()}")
                    if "AULA" in val and len(val) < 15: self._safe_set_value(nueva_prom, r, c, f"AULA: {grado}")
                    if "JORNADA" in val and len(val) < 20: self._safe_set_value(nueva_prom, r, c, f"JORNADA: {jornada.upper()}")
                except AttributeError: pass

        for r in range(4, 50):
            for c in range(3, 100):
                try:
                    celda = nueva_prom.cell(row=r, column=c)
                    self._safe_clear_value(nueva_prom, r, c)
                except AttributeError: pass

        if hoja_plan_origen:
            nueva_plan = wb.copy_worksheet(wb[hoja_plan_origen])
            nueva_plan.title = f"Planilla ({nueva_materia.title()} {grado})"
            for r in range(1, 15):
                for c in range(1, 10):
                    try:
                        val = str(nueva_plan.cell(row=r, column=c).value or "").upper()
                        if "ASIGNATURA:" in val: self._safe_set_value(nueva_plan, r, c+2, nueva_materia.upper())
                        if "GRUPO:" in val: self._safe_set_value(nueva_plan, r, c+2, grado)
                        if "CONSEJERO" in val:
                            if len(val) > 20: self._safe_set_value(nueva_plan, r, c, f"PROF. CONSEJERO (A): {consejero.upper()}")
                            else: self._safe_set_value(nueva_plan, r, c+2, consejero.upper())
                    except AttributeError: pass
            
            for r in range(15, 60):
                for c in range(5, 20):
                    try:
                        celda = nueva_plan.cell(row=r, column=c)
                        self._safe_clear_value(nueva_prom, r, c)
                    except AttributeError: pass

        hoja_resumen = None
        for sheet in wb.sheetnames:
            if "RESUMEN" in sheet.upper() and grado_num in sheet:
                hoja_resumen = sheet
                break
                
        if hoja_resumen:
            ws_res = wb[hoja_resumen]
            fila_materias = None
            for r in range(10, 20):
                for c in range(2, 15):
                    if materia_origen.upper() in str(ws_res.cell(row=r, column=c).value or "").upper():
                        fila_materias = r
                        break
                if fila_materias: break
                
            if fila_materias:
                for c in range(2, 20):
                    if not ws_res.cell(row=fila_materias, column=c).value:
                        self._safe_set_value(ws_res, fila_materias, c, nueva_materia.upper())
                        break

        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True, "Materia clonada y agregada al Resumen."

    def exportar_todo_a_excel(self, ruta_destino):
        try:
            import shutil
            import datetime
            shutil.copy2(self.ruta, ruta_destino)
            wb = openpyxl.load_workbook(ruta_destino)
            
            grados = self.obtener_grados_activos()
            
            for grado in grados:
                cursor = self.db_conn.cursor()
                cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
                row_g = cursor.fetchone()
                if not row_g:
                    continue
                grado_id = row_g[0]
                
                # Exportar Notas
                materias = self.obtener_materias_por_grado(grado)
                for materia in materias:
                    nombre_hoja = self._encontrar_hoja_prom(wb, grado, materia)
                    if not nombre_hoja:
                        continue
                    ws = wb[nombre_hoja]
                    
                    cursor.execute("SELECT id FROM materias WHERE nombre = ? AND grado_id = ?;", (materia, grado_id))
                    row_m = cursor.fetchone()
                    if not row_m:
                        continue
                    materia_id = row_m[0]
                    
                    for trim_idx in (1, 2, 3):
                        trim_str = f"Trimestre {trim_idx}"
                        for tipo in ("Diaria / Parcial", "Apreciación", "Examen"):
                            col_inicio, col_fin = self._obtener_rango_columnas(ws, trim_str, tipo)
                            if col_inicio is None or col_fin is None:
                                continue
                                
                            cursor.execute("""
                                SELECT DISTINCT descripcion, fecha FROM notas
                                WHERE materia_id = ? AND trimestre = ? AND tipo = ?;
                            """, (materia_id, trim_idx, tipo))
                            cols_data = cursor.fetchall()
                            
                            c = col_inicio
                            for desc, fecha in cols_data:
                                if c > col_fin:
                                    break
                                    
                                try: ws.cell(row=4, column=c).value = fecha
                                except AttributeError: pass
                                
                                try:
                                    cell_desc = ws.cell(row=self.fila_desc, column=c)
                                    cell_desc.value = desc
                                    cell_desc.alignment = Alignment(wrapText=True, horizontal='center', vertical='center', textRotation=90)
                                    cell_desc.font = Font(name='Calibri', size=11, bold=True)
                                except AttributeError: pass
                                
                                cursor.execute("""
                                    SELECT estudiante_id, valor FROM notas
                                    WHERE materia_id = ? AND trimestre = ? AND tipo = ? AND descripcion = ?;
                                """, (materia_id, trim_idx, tipo, desc))
                                for est_id, val in cursor.fetchall():
                                    fila_excel = 4 + (int(est_id) % 100)
                                    try: ws.cell(row=fila_excel, column=c).value = val
                                    except AttributeError: pass
                                c += 1
                                
                # Exportar Asistencia
                hoja_asist = self._encontrar_hoja_asistencia(wb, grado)
                if hoja_asist:
                    ws_asist = wb[hoja_asist]
                    mapa_trimestres = {"Trimestre 1": 2, "Trimestre 2": 45, "Trimestre 3": 88}
                    fuente_meduca = Font(name='Calibri', size=9, bold=True)
                    
                    for trim_idx in (1, 2, 3):
                        trim_str = f"Trimestre {trim_idx}"
                        fila_fechas = mapa_trimestres.get(trim_str, 2)
                        
                        cursor.execute("""
                            SELECT DISTINCT a.fecha FROM asistencia a
                            JOIN estudiantes e ON a.estudiante_id = e.id
                            WHERE e.grado_id = ? AND a.trimestre = ? ORDER BY a.fecha;
                        """, (grado_id, trim_idx))
                        fechas = [r[0] for r in cursor.fetchall()]
                        
                        c = 3
                        for fecha in fechas:
                            if c > 60:
                                break
                            ws_asist.cell(row=fila_fechas, column=c).value = fecha
                            
                            cursor.execute("""
                                SELECT e.id, a.estado FROM asistencia a
                                JOIN estudiantes e ON a.estudiante_id = e.id
                                WHERE e.grado_id = ? AND a.fecha = ? AND a.trimestre = ?;
                            """, (grado_id, fecha, trim_idx))
                            for est_id, estado in cursor.fetchall():
                                fila_excel = fila_fechas + (int(est_id) % 100)
                                celda = ws_asist.cell(row=fila_excel, column=c)
                                celda.value = estado
                                celda.font = fuente_meduca
                            c += 1
            
            wb.save(ruta_destino)
            wb.close()
            
            wb_res = openpyxl.load_workbook(ruta_destino)
            for grado in grados:
                self.actualizar_resumen(grado, wb=wb_res, force_excel=True)
            wb_res.save(ruta_destino)
            wb_res.close()
            
            return True
        except Exception as e:
            logger.error(f"Error en exportar_todo_a_excel: {e}")
            return False

    def formatear_reportes_excel(self, grado):
        """
        Exporta los datos de la base de datos SQLite de vuelta al Excel de trabajo temporal,
        asegurando que los reportes impresos contengan los datos más recientes.
        """
        import tempfile
        import shutil
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_name = tmp.name
        try:
            # Primero copiamos la plantilla/archivo actual al temporal
            shutil.copy2(self.ruta, tmp_name)
            # Poblamos el archivo temporal con los datos de SQLite
            self.exportar_todo_a_excel(tmp_name)
            # Sobrescribimos el archivo original de trabajo con el poblado
            shutil.copy2(tmp_name, self.ruta)
            # Recargar el cache para que openpyxl use el nuevo contenido
            self._verificar_y_recargar_cache()
            return True
        except Exception as e:
            logger.error(f"Error al formatear reportes de Excel: {e}", exc_info=True)
            return False
        finally:
            if os.path.exists(tmp_name):
                try:
                    os.remove(tmp_name)
                except Exception:
                    pass

    def actualizar_resumen(self, grado, wb=None, force_excel=False):
        if not force_excel and os.path.basename(self.ruta).startswith("Registro_"):
            return True
        if wb is None:
            self._verificar_y_recargar_cache()
        if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None: return False
        should_close = not bool(self._wb_cache) if wb is None else False
        if wb is None:
            wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
        hoja_res = None
        grado_num = grado.replace("°", "")
        for sheet in wb.sheetnames:
            if "RESUMEN" in sheet.upper() and (self.modalidad == "primaria" or grado_num in sheet):
                hoja_res = sheet
                break

        if not hoja_res:
            if should_close: wb.close()
            return False

        ws_res = wb[hoja_res]

        # Primero vamos a identificar las columnas de Promedio (T1, T2, T3) y Nota Final/Anual
        cols_promedios = []
        col_anual = None
        col_estado = None

        for c in range(5, 40):
            val = str(ws_res.cell(row=9, column=c).value or "").upper()
            val2 = str(ws_res.cell(row=8, column=c).value or "").upper()
            if "PROMEDIO" in val or "T.1" in val or "T.2" in val or "T.3" in val:
                cols_promedios.append(c)
            if "ANUAL" in val or "FINAL" in val or "ANUAL" in val2 or "FINAL" in val2:
                col_anual = c
            if "ESTADO" in val or "ESTADO" in val2:
                col_estado = c

        if col_anual and col_estado:
            wb_write = openpyxl.load_workbook(self.ruta)
            ws_write = wb_write[hoja_res]

            for r in range(10, 50):
                # Calcular anual
                notas = []
                for c in cols_promedios:
                    try:
                        valido, nota, _ = validar_nota_meduca(ws_res.cell(row=r, column=c).value)
                        if valido:
                            notas.append(nota)
                    except (ValueError, TypeError): pass

                if notas:
                    anual = sum(notas) / len(notas)
                    anual = round(anual, 1)
                    try:
                        ws_write.cell(row=r, column=col_anual).value = anual
                        ws_write.cell(row=r, column=col_estado).value = "Aprobado" if anual >= 3.0 else "Reprobado"
                    except AttributeError: pass

            self._save_wb(wb_write)
            wb_write.close()

        if should_close:
            wb.close()
            self._cargar_en_memoria()
        return True

    def eliminar_materia(self, grado, materia):
        if not os.path.exists(self.ruta): return False
        wb = openpyxl.load_workbook(self.ruta)
        hojas_a_borrar = []
        materia_clean = materia.lower().replace(" ", "").replace(".", "")
        grado_num = grado.replace("°", "")

        for sheet in wb.sheetnames:
            sheet_clean = sheet.lower().replace(" ", "").replace(".", "")
            if materia_clean in sheet_clean and (self.modalidad == "primaria" or grado_num in sheet):
                if "PROM" in sheet.upper() or "PLANILLA" in sheet.upper():
                    hojas_a_borrar.append(sheet)

        if not hojas_a_borrar:
            wb.close()
            return False

        for hoja in hojas_a_borrar: del wb[hoja]

        hoja_resumen = None
        for sheet in wb.sheetnames:
            if "RESUMEN" in sheet.upper() and grado_num in sheet:
                hoja_resumen = sheet
                break
                
        if hoja_resumen:
            ws_res = wb[hoja_resumen]
            for r in range(10, 20):
                for c in range(2, 20):
                    if materia.upper() in str(ws_res.cell(row=r, column=c).value or "").upper():
                        self._safe_clear_value(ws_res, r, c)

        self._save_wb(wb)
        wb.close()
        self._cargar_en_memoria()
        return True

    def obtener_estadisticas_asistencia(self, grado, trimestre, id_estudiante, wb=None):
        if wb is None:
            self._verificar_y_recargar_cache()
        if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None:
            return {"total_dias": 0, "ausencias": 0, "tardanzas": 0, "excusas": 0}
        should_close = not bool(self._wb_cache) if wb is None else False
        if wb is None:
            wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
        hoja = self._encontrar_hoja_asistencia(wb, grado)
        if not hoja:
            if should_close: wb.close()
            return {"total_dias": 0, "ausencias": 0, "tardanzas": 0, "excusas": 0}
        ws = wb[hoja]
        mapa_trimestres = {"Trimestre 1": 2, "Trimestre 2": 45, "Trimestre 3": 88}
        fila_fechas = mapa_trimestres.get(trimestre, 2)
        
        total_dias = 0
        ausencias = 0
        tardanzas = 0
        excusas = 0
        
        for c in range(3, 61):
            fecha_val = ws.cell(row=fila_fechas, column=c).value
            if not fecha_val:
                continue
            val = ws.cell(row=fila_fechas + (int(id_estudiante) % 100), column=c).value
            if val is not None and str(val).strip():
                total_dias += 1
                if val == "-":
                    ausencias += 1
                elif val == "T":
                    tardanzas += 1
                elif val == "E":
                    excusas += 1
                
        if should_close: wb.close()
        return {
            "total_dias": total_dias,
            "ausencias": ausencias,
            "tardanzas": tardanzas,
            "excusas": excusas
        }

    def obtener_tareas_sin_nota(self, grado, trimestre, id_estudiante, wb=None):
        cursor = self.db_conn.cursor()
        cursor.execute("SELECT id FROM grados WHERE nombre = ? AND seccion = 'A' AND modalidad = ?;", (grado, self.modalidad))
        row_g = cursor.fetchone()
        
        sql_has_data = False
        if row_g and os.path.basename(self.ruta).startswith("Registro_"):
            grado_id = row_g[0]
            cursor.execute("""
                SELECT COUNT(*) FROM notas n
                JOIN materias m ON n.materia_id = m.id
                WHERE m.grado_id = ? AND n.trimestre = ?;
            """, (grado_id, int(str(trimestre).replace("Trimestre ", ""))))
            sql_has_data = (cursor.fetchone()[0] > 0)
            
        if not sql_has_data:
            if wb is None:
                self._verificar_y_recargar_cache()
            if not os.path.exists(self.ruta) and wb is None and self._wb_cache is None:
                return {"tareas_vacias_por_materia": {}, "total_vacias": 0}
            should_close = not bool(self._wb_cache) if wb is None else False
            if wb is None:
                wb = self._wb_cache if self._wb_cache else openpyxl.load_workbook(self.ruta, data_only=True)
                
            materias = self.obtener_materias_por_grado(grado, wb=wb)
            tareas_vacias = {}
            total_vacias = 0
            row_est = 4 + (int(id_estudiante) % 100)
            
            if not materias or materias == ["Sin materias registradas"]:
                materias = []
                grado_num = grado.replace("°", "")
                for sheet in wb.sheetnames:
                    if "PROM" in sheet.upper():
                        if self.modalidad == "premedia" and grado_num in sheet:
                            mat = sheet.upper().replace("PROM", "").replace("(", "").replace(")", "").replace(grado, "").replace(grado_num, "").replace("°", "").strip()
                            materias.append(mat.title())
                        elif self.modalidad == "primaria":
                            mat = sheet.upper().replace("PROM", "").replace("(", "").replace(")", "").strip()
                            materias.append(mat.title())
                materias = sorted(list(set(materias)))
            
            for mat in materias:
                if mat == "Sin materias registradas":
                    continue
                hoja = self._encontrar_hoja_prom(wb, grado, mat)
                if not hoja:
                    continue
                ws = wb[hoja]
                
                cnt_mat = 0
                for tipo_nota in ["Diaria / Parcial", "Apreciación", "Examen"]:
                    col_inicio, col_fin = self._obtener_rango_columnas(ws, trimestre, tipo_nota)
                    if col_inicio is None or col_fin is None:
                        continue
                    for c in range(col_inicio, col_fin + 1):
                        desc = ws.cell(row=self.fila_desc, column=c).value
                        if desc and str(desc).strip():
                            nota = ws.cell(row=row_est, column=c).value
                            if nota is None or str(nota).strip() == "":
                                cnt_mat += 1
                if cnt_mat > 0:
                    tareas_vacias[mat] = cnt_mat
                    total_vacias += cnt_mat
                    
            if should_close: wb.close()
            return {
                "tareas_vacias_por_materia": tareas_vacias,
                "total_vacias": total_vacias
            }

        grado_id = row_g[0]
        cursor.execute("SELECT id, nombre FROM materias WHERE grado_id = ?;", (grado_id,))
        materias = cursor.fetchall()
        
        tareas_vacias = {}
        total_vacias = 0
        trim_num = int(str(trimestre).replace("Trimestre ", ""))
        
        for mat_id, mat_nombre in materias:
            cursor.execute("""
                SELECT DISTINCT descripcion, tipo, fecha 
                FROM notas 
                WHERE materia_id = ? AND trimestre = ?;
            """, (mat_id, trim_num))
            tareas_registradas = cursor.fetchall()
            
            cnt_mat = 0
            for desc, tipo, fecha in tareas_registradas:
                cursor.execute("""
                    SELECT valor 
                    FROM notas 
                    WHERE estudiante_id = ? AND materia_id = ? AND trimestre = ? AND descripcion = ? AND tipo = ? AND fecha = ?;
                """, (str(id_estudiante), mat_id, trim_num, desc, tipo, fecha))
                row_val = cursor.fetchone()
                if not row_val or row_val[0] is None:
                    cnt_mat += 1
                    
            if cnt_mat > 0:
                tareas_vacias[mat_nombre] = cnt_mat
                total_vacias += cnt_mat
                
        return {
            "tareas_vacias_por_materia": tareas_vacias,
            "total_vacias": total_vacias
        }