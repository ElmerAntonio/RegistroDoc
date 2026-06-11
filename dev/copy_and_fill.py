import shutil
import openpyxl
import random
import sys
sys.path.append("src")
from rddata import DataEngine

def populate_premedia():
    print("Copying Registro_2026_FINAL_v2.xlsx to Registro_2026.xlsx...")
    shutil.copy2("Registro_2026_FINAL_v2.xlsx", "Registro_2026.xlsx")
    
    engine = DataEngine("Registro_2026.xlsx", "premedia")
    wb = openpyxl.load_workbook("Registro_2026.xlsx")
    
    ws_maestro = wb["MAESTRO"]
    active_rows = []
    for r in range(5, 46):
        if ws_maestro.cell(row=r, column=2).value:
            active_rows.append(r)
            
    print(f"Found {len(active_rows)} students in Premedia.")
    
    # Premedia periods and types
    trimestres = ["Trimestre 1", "Trimestre 2", "Trimestre 3"]
    tipos = ["Diaria / Parcial", "Apreciación", "Examen"]
    
    for sheetname in wb.sheetnames:
        if sheetname.startswith("PROM"):
            ws = wb[sheetname]
            print(f"Populating grades in {sheetname}...")
            
            for t_name in trimestres:
                for tipo in tipos:
                    # Resolve range
                    col_i, col_f = engine._obtener_rango_columnas(ws, t_name, tipo)
                    if col_i is not None and col_f is not None:
                        # Write a few random notes in this range
                        for c in range(col_i, col_f + 1):
                            # Set description to make it active if empty
                            desc_cell = ws.cell(row=engine.fila_desc, column=c)
                            if type(desc_cell).__name__ != 'MergedCell':
                                if not desc_cell.value:
                                    desc_cell.value = f"Tarea {c}"
                            
                            for row in active_rows:
                                cell = ws.cell(row=row, column=c)
                                if type(cell).__name__ != 'MergedCell':
                                    cell.value = round(random.uniform(3.0, 5.0), 1)

    wb.save("Registro_2026.xlsx")
    wb.close()
    print("Premedia population complete!")

def populate_primaria():
    print("Populating Primaria...")
    engine = DataEngine("Registro_Primaria.xlsx", "primaria")
    wb = openpyxl.load_workbook("Registro_Primaria.xlsx")
    
    # 1. Add students to MAESTRO 1°
    ws_maestro = wb["MAESTRO 1°"]
    students = [
        ("BEJERANO, Jose", "1-234-567"),
        ("GONZALEZ, Maria", "4-789-012"),
        ("PINTO, Luis", "9-345-678"),
        ("SANTOS, Ana", "8-901-234"),
        ("TUGRI, Carlos", "12-345-678"),
        ("ZURDO, Rosa", "4-111-222"),
        ("JIMENEZ, Juan", "8-444-555"),
        ("MONTERO, Sofia", "9-888-999"),
        ("SALDANA, Pedro", "2-222-333"),
        ("SANTOS, Lucas", "8-777-888")
    ]
    
    for idx, (name, cedula) in enumerate(students):
        row = 5 + idx
        ws_maestro.cell(row=row, column=1).value = idx + 1 # N°
        ws_maestro.cell(row=row, column=2).value = name
        ws_maestro.cell(row=row, column=5).value = cedula # Cédula
        
    print(f"Added {len(students)} students to Primaria.")
    
    # 2. Add grades in PROM (Ingles 1°)
    ws = wb["PROM (Ingles 1°)"]
    active_rows = list(range(5, 5 + len(students)))
    
    trimestres = ["Trimestre 1", "Trimestre 2", "Trimestre 3"]
    tipos = ["Diaria / Parcial", "Apreciación"]
    
    for t_name in trimestres:
        for tipo in tipos:
            col_i, col_f = engine._obtener_rango_columnas(ws, t_name, tipo)
            if col_i is not None and col_f is not None:
                # Fill first 2 columns in the range
                for c in range(col_i, col_i + 3):
                    if c <= col_f:
                        desc_cell = ws.cell(row=engine.fila_desc, column=c)
                        if type(desc_cell).__name__ != 'MergedCell':
                            if not desc_cell.value:
                                desc_cell.value = f"Actividad {c}"
                        
                        for row in active_rows:
                            cell = ws.cell(row=row, column=c)
                            if type(cell).__name__ != 'MergedCell':
                                cell.value = round(random.uniform(3.0, 5.0), 1)

    wb.save("Registro_Primaria.xlsx")
    wb.close()
    print("Primaria population complete!")

if __name__ == "__main__":
    populate_premedia()
    populate_primaria()
