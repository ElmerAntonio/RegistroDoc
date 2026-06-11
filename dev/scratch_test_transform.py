import openpyxl

wb = openpyxl.load_workbook("Registro_2026.xlsx")
ws = wb["Asistencia (9\u00b0)"]

# 1. Delete rows 87 and 88
ws.delete_rows(87, 2)

# 2. Delete rows 44 and 45
ws.delete_rows(44, 2)

# 3. Insert 2 rows at 42
ws.insert_rows(42, 2)

print("Formulas in Column B after transformation:")
for r in range(1, 100):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    if val_b or val_a:
        print(f"Row {r:03d} -> Col A: {val_a}, Col B: {val_b}")

wb.save("Registro_2026_test2.xlsx")
wb.close()
