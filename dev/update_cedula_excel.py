import openpyxl
import os

files = ["Registro_2026.xlsx", "Registro Primaria.xlsx"]
for fn in files:
    if os.path.exists(fn):
        print(f"Modifying {fn}...")
        wb = openpyxl.load_workbook(fn)
        if "Portada" in wb.sheetnames:
            ws = wb["Portada"]
            old_val = ws["H16"].value
            print(f"Old cell H16 value: {old_val}")
            ws["H16"] = "8-000-0000"
            wb.save(fn)
            print(f"Successfully updated H16 to 8-000-0000 in {fn}")
        else:
            print(f"Warning: 'Portada' sheet not found in {fn}")
    else:
        print(f"File {fn} not found.")
