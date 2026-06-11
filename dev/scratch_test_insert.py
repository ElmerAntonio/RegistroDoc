import openpyxl

wb = openpyxl.load_workbook("Registro_2026.xlsx")
ws = wb["Asistencia (9\u00b0)"]

print("Before insertion:")
print("Row 42:", [ws.cell(row=42, column=c).value for c in range(1, 10)])
print("Row 43:", [ws.cell(row=43, column=c).value for c in range(1, 10)])
print("Row 44:", [ws.cell(row=44, column=c).value for c in range(1, 10)])
print("Row 45:", [ws.cell(row=45, column=c).value for c in range(1, 10)])

# Insert 2 rows at row 42
ws.insert_rows(42, 2)

print("\nAfter insertion:")
print("Row 42:", [ws.cell(row=42, column=c).value for c in range(1, 10)])
print("Row 43:", [ws.cell(row=43, column=c).value for c in range(1, 10)])
print("Row 44:", [ws.cell(row=44, column=c).value for c in range(1, 10)])
print("Row 45:", [ws.cell(row=45, column=c).value for c in range(1, 10)])
print("Row 46:", [ws.cell(row=46, column=c).value for c in range(1, 10)])

wb.save("Registro_2026_test.xlsx")
wb.close()
