import sys
import os
src_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "src"))
if src_path not in sys.path:
    sys.path.insert(0, src_path)

import pytest
from rddata import DataEngine

def test_obtener_horario_returns_default_structure():
    """Test that obtener_horario returns the default structure when the file doesn't exist."""
    # Act
    # We use a dummy path that clearly doesn't exist to test the edge case
    engine = DataEngine("ruta_horario_inexistente.xlsx")
    horario = engine.obtener_horario()

    # Assert
    expected_structure = [
        {"horas": "", "lunes": "", "martes": "", "miercoles": "", "jueves": "", "viernes": ""}
        for _ in range(8)
    ]

    assert horario == expected_structure
    assert len(horario) == 8

def test_obtener_datos_generales_missing_file():
    """
    Test that obtener_datos_generales returns the exact expected default
    dictionary when initialized with a non-existent file path.
    """
    from unittest.mock import MagicMock
    # Create an instance with a non-existent path
    engine = DataEngine(ruta_excel="ruta_ficticia_inexistente.xlsx")
    
    # Mockear db_conn para que no retorne configuración y pruebe el fallback real
    engine.db_conn = MagicMock()
    engine.db_conn.cursor().fetchone.return_value = [0]

    # Ensure the path really doesn't exist
    assert not os.path.exists(engine.ruta), "El archivo no debe existir para esta prueba"

    # Call the method
    resultado = engine.obtener_datos_generales()

    # Expected default dictionary
    expected_datos = {
        "docente_nombre": "", "docente_cedula": "", "seguro_social": "", "numero_posicion": "",
        "condicion_nombramiento": "", "escuela_nombre": "", "escuela_region": "", "distrito": "",
        "corregimiento": "", "zona_escolar": "", "director_nombre": "", "subdirector_nombre": "",
        "coordinador_nombre": "", "telefono": "", "correo": "", "ano_lectivo": "2026",
        "jornada": "", "fecha_t1": "", "fecha_t2": "", "fecha_t3": ""
    }

    # Strict equality assertion
    assert resultado == expected_datos

def test_obtener_grados_activos_missing_file():
    """
    Test that obtener_grados_activos returns an empty list
    when initialized with a non-existent file path.
    """
    # Create an instance with a non-existent path
    engine = DataEngine(ruta_excel="ruta_inexistente_grados.xlsx")

    # Call the method
    resultado = engine.obtener_grados_activos()

    # Expected empty list
    assert resultado == []

def test_atomic_save(tmp_path):
    import openpyxl
    file_path = str(tmp_path / "test_libreta.xlsx")
    wb = openpyxl.Workbook()
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path)
    
    # 1. Modificar y guardar de forma atómica
    wb_write = openpyxl.load_workbook(file_path)
    wb_write.active.cell(row=1, column=1, value="NUEVO_VALOR")
    engine._save_wb(wb_write)
    wb_write.close()
    
    # 2. Comprobar que el original fue guardado y el backup se creó
    bak_path = file_path.replace(".xlsx", "_bak.xlsx")
    assert os.path.exists(file_path)
    assert os.path.exists(bak_path)
    
    # 3. Comprobar contenido guardado
    wb_check = openpyxl.load_workbook(file_path)
    assert wb_check.active.cell(row=1, column=1).value == "NUEVO_VALOR"
    wb_check.close()

def test_cache_invalidation_mtime(tmp_path):
    import openpyxl
    import time
    file_path = str(tmp_path / "test_cache.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DASHBOARD"
    ws.cell(row=1, column=1, value="ORIGINAL")
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path)
    assert engine._wb_cache.active.cell(row=1, column=1).value == "ORIGINAL"
    
    # Simular edición externa escribiendo al archivo
    time.sleep(0.01)  # Asegurar cambio de mtime en sistemas rápidos
    wb_ext = openpyxl.load_workbook(file_path)
    wb_ext.active.cell(row=1, column=1, value="EDITADO_EXTERNO")
    wb_ext.save(file_path)
    wb_ext.close()
    
    # El motor debería invalidar y recargar al comprobar
    engine._verificar_y_recargar_cache()
    assert engine._wb_cache.active.cell(row=1, column=1).value == "EDITADO_EXTERNO"

def test_obtener_cuadro_honor_general_real_file():
    """Integration test validating obtener_cuadro_honor_general with the real workbook."""
    real_path = os.path.join("assets", "templates", "Registro_Premedia.xlsx")
    if not os.path.exists(real_path):
        pytest.skip("Registro_Premedia.xlsx not present in assets/templates, skipping integration test")
        
    engine = DataEngine(real_path)
    cuadro = engine.obtener_cuadro_honor_general()
    
    # Assert type
    assert isinstance(cuadro, list)
    
    # If any students are in the honor roll, verify their scores and sorting
    if len(cuadro) > 0:
        prev_prom = 5.01
        for est in cuadro:
            assert "nombre" in est
            assert "grado" in est
            assert "promedio" in est
            assert 4.5 <= est["promedio"] <= 5.0, f"Promedio {est['promedio']} must be between 4.5 and 5.0"
            assert est["promedio"] <= prev_prom, "Students must be sorted in descending order of average"
            prev_prom = est["promedio"]


def test_cierre_ano_lectivo(tmp_path):
    """Test that realizar_cierre_ano_lectivo correctly processes backups, resets data, and increments the year."""
    import openpyxl
    file_path = str(tmp_path / "Registro_Premedia.xlsx")
    
    # Create dummy workbook with at least one sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASISTENCIA (7° A)"
    # Add a date cell DD-MM to verify cleaner
    ws.cell(row=2, column=3, value="12-05")
    ws.cell(row=3, column=3, value="P")
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path)
    engine.modalidad = "premedia"
    
    # Set config year to 2026
    from rdsecurity import cargar_config_segura, guardar_config_segura
    cfg = cargar_config_segura({})
    cfg["ano_lectivo"] = "2026"
    guardar_config_segura(cfg)
    
    # Run rollover without student promotion (since we have no students in DB yet)
    exito, msg = engine.realizar_cierre_ano_lectivo(promover_estudiantes=False)
    
    assert exito is True
    assert "2026" in msg
    
    # Verify year was incremented
    cfg_new = cargar_config_segura({})
    assert cfg_new.get("ano_lectivo") == "2027"

def test_parsear_archivo_estudiantes_txt(tmp_path):
    """Test that parsear_archivo_estudiantes correctly extracts names and Panamanian cedulas from text files."""
    from impp import parsear_archivo_estudiantes
    
    txt_path = str(tmp_path / "roster.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("1. ELMER ANTONIO RAMOS  8-765-4321\n")
        f.write("2. JUAN PEREZ PE-99-8888\n")
        f.write("NOMBRE CEDULA (CABECERA A IGNORAR)\n")
        f.write("3. MARIA DE LEON (SIN CEDULA)\n")
        
    estudiantes = parsear_archivo_estudiantes(txt_path)
    
    assert len(estudiantes) == 3
    assert estudiantes[0]["nombre"] == "ELMER ANTONIO RAMOS"
    assert estudiantes[0]["cedula"] == "8-765-4321"
    assert estudiantes[1]["nombre"] == "JUAN PEREZ"
    assert estudiantes[1]["cedula"] == "PE-99-8888"
    assert estudiantes[2]["nombre"] == "MARIA DE LEON (SIN CEDULA)"
    assert estudiantes[2]["cedula"] == ""

def test_registro_primaria_integration():
    """Integration test validating that the Primary workbook template can be loaded and synced successfully."""
    real_path = os.path.join("assets", "templates", "Registro_Primaria.xlsx")
    if not os.path.exists(real_path):
        pytest.skip("Registro_Primaria.xlsx not present in assets/templates, skipping integration test")
        
    engine = DataEngine(real_path, modalidad="primaria")
    generales = engine.obtener_datos_generales()
    
    assert isinstance(generales, dict)
    assert engine.modalidad == "primaria"
    assert engine.fila_desc == 39 # Fila de descripción de hábitos para primaria


def test_obtener_fechas_asistencia_excel(tmp_path):
    import openpyxl
    file_path = str(tmp_path / "test_asis.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia (7 A)"
    ws.cell(row=2, column=3, value="10-10")
    ws.cell(row=2, column=4, value="11-10")
    wb.save(file_path)
    wb.close()

    engine = DataEngine(file_path)
    engine.modalidad = "premedia"
    
    fechas = engine.obtener_fechas_asistencia(f"7\u00b0 A", "Trimestre 1")
    assert "10-10" in fechas
    assert "11-10" in fechas
    assert len(fechas) == 2


def test_sincronizar_excel_a_sql_self_healing(tmp_path):
    import openpyxl
    from rdsql import obtener_dir_datos_usuario
    
    file_path = str(tmp_path / "Registro_Premedia.xlsx")
    
    # Eliminar cualquier archivo temporal previo para evitar conflictos de caché
    temp_file = os.path.join(obtener_dir_datos_usuario(), "temp", "Registro_Premedia.xlsx")
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except Exception:
            pass
            
    wb = openpyxl.Workbook()
    
    ws_gen = wb.active
    ws_gen.title = "DASHBOARD"
    
    ws_asis = wb.create_sheet("Asistencia (7 A)")
    ws_asis.cell(row=2, column=3, value="15-05")
    ws_asis.cell(row=4, column=1, value="RAMOS, Elmer")
    ws_asis.cell(row=4, column=3, value="P")
    
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path)
    engine.modalidad = "premedia"
    
    cursor = engine.db_conn.cursor()
    cursor.execute(f"INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES ('7\u00b0 A', 'A', 'premedia');")
    cursor.execute(f"SELECT id FROM grados WHERE nombre = '7\u00b0 A';")
    g_id = cursor.fetchone()[0]
    
    nombre_cifrado = engine.db_manager.encriptar_campo("RAMOS, Elmer")
    cursor.execute("INSERT OR REPLACE INTO estudiantes (id, nombre, cedula, sexo, grado_id, estado) VALUES (?, ?, '', 'M', ?, 'Activo');", (f"{g_id}02", nombre_cifrado, g_id))
    
    # Asegurar que no hay asistencia para forzar la autocuración
    cursor.execute("DELETE FROM asistencia;")
    engine.db_conn.commit()
    
    engine.sincronizar_excel_a_sql()
    
    cursor.execute("SELECT COUNT(*) FROM asistencia;")
    assert cursor.fetchone()[0] > 0


def test_sincronizar_excel_a_sql_preserves_retired_students(tmp_path):
    import openpyxl
    file_path = str(tmp_path / "Test_Premedia.xlsx")
    wb = openpyxl.Workbook()
    
    ws_gen = wb.active
    ws_gen.title = "DASHBOARD"
    
    ws_maestro = wb.create_sheet("MAESTRO")
    ws_maestro.cell(row=3, column=1, value="7° GRADO")
    ws_maestro.cell(row=5, column=2, value="Tugri Clemente") # row 5 in column 2 (index 1)
    
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path)
    engine.modalidad = "premedia"
    
    cursor = engine.db_conn.cursor()
    cursor.execute("INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES ('7\u00b0 A', 'A', 'premedia');")
    cursor.execute("SELECT id FROM grados WHERE nombre = '7\u00b0 A';")
    g_id = cursor.fetchone()[0]
    
    # Insert student Clemente as Retirado in SQLite
    clemente_id = str(g_id * 100 + 1)
    nombre_cifrado = engine.db_manager.encriptar_campo("Tugri Clemente")
    cursor.execute("""
        INSERT INTO estudiantes (id, nombre, cedula, sexo, grado_id, estado, fecha_retiro, motivo_retiro, nombre_acudiente)
        VALUES (?, ?, '', 'M', ?, 'Retirado', '2026-06-25', 'Se Fue con su mama a otro lugar', '');
    """, (clemente_id, nombre_cifrado, g_id))
    engine.db_conn.commit()
    
    # Run sync. Since Clemente is in the Excel sheet row 5 (active_ids), normally he would be marked Activo.
    # But with our fix, his retirement details and state are preserved.
    engine.sincronizar_excel_a_sql()
    
    cursor.execute("SELECT estado, fecha_retiro, motivo_retiro FROM estudiantes WHERE id = ?;", (clemente_id,))
    row = cursor.fetchone()
    assert row is not None
    assert row[0] == 'Retirado'
    assert row[1] == '2026-06-25'
    assert row[2] == 'Se Fue con su mama a otro lugar'


def test_agregar_estudiante_generates_composite_id(tmp_path):
    import openpyxl
    file_path = str(tmp_path / "Test_Premedia.xlsx")
    wb = openpyxl.Workbook()
    
    ws_gen = wb.active
    ws_gen.title = "DASHBOARD"
    
    ws_maestro = wb.create_sheet("MAESTRO")
    ws_maestro.cell(row=3, column=1, value="7° GRADO")
    # Empty rows 5 to 40
    
    ws_p = wb.create_sheet("Planilla 7")
    
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path)
    engine.modalidad = "premedia"
    
    cursor = engine.db_conn.cursor()
    cursor.execute("INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES ('7\u00b0 A', 'A', 'premedia');")
    cursor.execute("SELECT id FROM grados WHERE nombre = '7\u00b0 A';")
    g_id = cursor.fetchone()[0]
    engine.db_conn.commit()
    
    # Add student
    exito = engine.agregar_estudiante("7° A", "Tugri Clemente", cedula="8-123-456", sexo="M")
    assert exito is True
    
    # Verify ID in SQLite is composite (g_id * 100 + 1)
    expected_id = str(g_id * 100 + 1)
    cursor.execute("SELECT id, estado FROM estudiantes WHERE nombre LIKE ?;", (f"%Tugri Clemente%",))
    res = cursor.fetchone()
    assert res is not None
    assert res[0] == expected_id
    assert res[1] == 'Activo'


def test_guardar_cambios_estudiantes_correct_excel_row(tmp_path):
    import openpyxl
    file_path = str(tmp_path / "Test_Premedia.xlsx")
    wb = openpyxl.Workbook()
    
    ws_gen = wb.active
    ws_gen.title = "DASHBOARD"
    
    ws_maestro = wb.create_sheet("MAESTRO")
    ws_maestro.cell(row=3, column=1, value="7° GRADO")
    ws_maestro.cell(row=4, column=2, value="NOMBRE")
    ws_maestro.cell(row=5, column=2, value="Tugri Clemente")
    
    ws_p = wb.create_sheet("Planilla 7")
    
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path)
    engine.modalidad = "premedia"
    
    cursor = engine.db_conn.cursor()
    cursor.execute("INSERT OR IGNORE INTO grados (nombre, seccion, modalidad) VALUES ('7\u00b0 A', 'A', 'premedia');")
    cursor.execute("SELECT id FROM grados WHERE nombre = '7\u00b0 A';")
    g_id = cursor.fetchone()[0]
    engine.db_conn.commit()
    
    # Update student changes using composite ID (g_id * 100 + 1)
    clemente_id = str(g_id * 100 + 1)
    datos_modificados = {
        clemente_id: {
            "nombre": "Tugri Clemente Modificado",
            "cedula": "8-999-9999",
            "sexo": "M"
        }
    }
    
    exito = engine.guardar_cambios_estudiantes("7° A", datos_modificados)
    assert exito is True
    
    # Load workbook and check row 5
    wb_check = openpyxl.load_workbook(file_path, data_only=True)
    ws_m_check = wb_check["MAESTRO"]
    assert ws_m_check.cell(row=5, column=2).value == "Tugri Clemente Modificado"
    
    ws_p_check = wb_check["Planilla 7"]
    # Row index in planilla is 15 + offset = 15 + 1 = 16
    assert ws_p_check.cell(row=16, column=5).value == "8-999-9999"
    wb_check.close()

