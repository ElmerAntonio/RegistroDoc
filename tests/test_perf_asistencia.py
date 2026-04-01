import time
import openpyxl
import os
import random
from src.rddata import DataEngine

def create_dummy_file(ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASIST (1° A)"

    # Trimestre 1: row 2
    for c in range(3, 61):
        ws.cell(row=2, column=c, value=f"2023-10-{c:02d}")

    for r in range(3, 48):
        for c in range(3, 61):
            ws.cell(row=r, column=c, value=random.choice(["P", "T", "A"]))

    wb.save(ruta)
    wb.close()

def test_performance_buscar_asistencia():
    test_file = "test_asist_bench.xlsx"
    if not os.path.exists(test_file):
        create_dummy_file(test_file)

    engine = DataEngine(ruta_excel=test_file)
    engine.modalidad = "premedia"

    # 1. Uncached Performance Baseline
    engine._wb_cache = None  # Ensure no cache

    # Warm up disk cache slightly
    engine.buscar_asistencia_existente("1° A", "Trimestre 1", "2023-10-15")

    t0 = time.time()
    iterations = 50
    for _ in range(iterations):
        engine._wb_cache = None
        engine.buscar_asistencia_existente("1° A", "Trimestre 1", "2023-10-15")
    t1 = time.time()
    uncached_total = t1 - t0
    uncached_avg = uncached_total / iterations

    # 2. Cached Performance Measurement
    engine._cargar_en_memoria() # Load cache

    t0 = time.time()
    for _ in range(iterations):
        engine.buscar_asistencia_existente("1° A", "Trimestre 1", "2023-10-15")
    t1 = time.time()
    cached_total = t1 - t0
    cached_avg = cached_total / iterations

    speedup = uncached_avg / cached_avg if cached_avg > 0 else float('inf')

    print(f"\n⚡ Performance Improvement: {speedup:.2f}x faster!")
    print(f"Uncached Average: {uncached_avg:.4f}s")
    print(f"Cached Average: {cached_avg:.6f}s")

    # We assert that the cache optimization provides at least a 100x speedup
    assert speedup > 100, f"Expected cache to be much faster, got {speedup:.2f}x speedup"

    if os.path.exists(test_file):
        os.remove(test_file)
