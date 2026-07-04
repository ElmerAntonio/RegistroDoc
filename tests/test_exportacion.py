import os
import sys
import time
import shutil
import tempfile
import pytest
import openpyxl

# Add src/ to python path
src_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "src"))
if src_path not in sys.path:
    sys.path.insert(0, src_path)

from rddata import DataEngine
from rdsql import SQLDatabaseManager

@pytest.fixture
def clean_db():
    """Setup a clean SQLite database environment for testing."""
    import rdsql  # Importar en runtime para que el monkeypatch de conftest.py sea efectivo
    db_manager = rdsql.SQLDatabaseManager()
    conn = db_manager.conectar()
    # Clean tables first to avoid conflict if already exists
    cursor = conn.cursor()
    for tbl in ["notas", "asistencia", "habitos", "estudiantes", "materias", "grados"]:
        cursor.execute(f"DROP TABLE IF EXISTS {tbl};")
    db_manager.inicializar_tablas(conn)
    yield
    db_manager.cerrar()

def test_exportar_todo_a_excel_and_benchmark(clean_db, tmp_path):
    # 1. Setup a dummy Excel file matching our templates
    template_src = os.path.join("assets", "templates", "Registro_Premedia.xlsx")
    if not os.path.exists(template_src):
        pytest.skip("Plantilla Registro_Premedia.xlsx no encontrada en assets/templates/")

    test_excel_path = str(tmp_path / "Registro_Premedia.xlsx")
    shutil.copy(template_src, test_excel_path)

    # Initialize SQLDatabaseManager and connector
    import rdsql as _rdsql
    db_manager = _rdsql.SQLDatabaseManager()
    conn = db_manager.conectar()
    cursor = conn.cursor()

    # Seed data directly to SQLite
    grado = "7°"
    seccion = "A"
    materia = "Ingles"
    trimestre_val = 1
    trimestre_str = "Trimestre 1"

    # Insert grado
    cursor.execute("INSERT INTO grados (nombre, seccion, modalidad) VALUES (?, ?, ?)", (grado, seccion, "premedia"))
    grado_id = cursor.lastrowid
    
    # Insert students (using sequential index strings as IDs, matching internal structure)
    estudiantes_mock = [
        {"id": "1", "cedula": "8-001-0001", "nombre": "ESTUDIANTE UNO", "sexo": "M", "grado_id": grado_id},
        {"id": "2", "cedula": "8-002-0002", "nombre": "ESTUDIANTE DOS", "sexo": "F", "grado_id": grado_id},
    ]
    for est in estudiantes_mock:
        nom_enc = db_manager.encriptar_campo(est["nombre"])
        ced_enc = db_manager.encriptar_campo(est["cedula"])
        
        cursor.execute(
            "INSERT INTO estudiantes (id, nombre, cedula, sexo, grado_id) VALUES (?, ?, ?, ?, ?)",
            (est["id"], nom_enc, ced_enc, est["sexo"], grado_id)
        )

    # Insert materia
    cursor.execute("INSERT INTO materias (nombre, grado_id) VALUES (?, ?)", (materia, grado_id))
    materia_id = cursor.lastrowid

    # Benchmark SQLite note saves (100 operations)
    t0 = time.time()
    for i in range(100):
        # Insert notes
        cursor.execute(
            "INSERT INTO notas (estudiante_id, materia_id, trimestre, tipo, descripcion, valor, puntos_obtenidos, puntos_maximos, fecha) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("1", materia_id, trimestre_val, "Diaria / Parcial", f"Tarea {i}", 4.5, 90, 100, "15-06")
        )
    conn.commit()
    t1 = time.time()
    sqlite_duration = t1 - t0
    print(f"\n[PERF] Tiempo de guardado en SQLite (100 notas): {sqlite_duration:.4f} segundos")

    # Insert attendance
    cursor.execute(
        "INSERT INTO asistencia (estudiante_id, fecha, estado, motivo, trimestre) VALUES (?, ?, ?, ?, ?)",
        ("1", "2026-06-15", ".", "P", trimestre_val)
    )
    cursor.execute(
        "INSERT INTO asistencia (estudiante_id, fecha, estado, motivo, trimestre) VALUES (?, ?, ?, ?, ?)",
        ("2", "2026-06-15", "-", "A", trimestre_val)
    )
    conn.commit()

    # Pre-populate student names in the source Excel file in column B (rows 5 and 6)
    # matching how students are loaded in Excel.
    wb_pre = openpyxl.load_workbook(test_excel_path)
    sheet_names = wb_pre.sheetnames
    
    # Write student names to the Ingles sheet
    ws_ingles = next(wb_pre[name] for name in sheet_names if "Ingles" in name)
    ws_ingles.cell(row=5, column=2, value="ESTUDIANTE UNO")
    ws_ingles.cell(row=6, column=2, value="ESTUDIANTE DOS")

    # Write student names to the Asistencia sheet
    ws_asist_pre = next(wb_pre[name] for name in sheet_names if "Asistencia" in name)
    # Column B is names in maestro sheet, but let's check. In premedia, column B of Asistencia has names
    ws_asist_pre.cell(row=3, column=2, value="ESTUDIANTE UNO")
    ws_asist_pre.cell(row=4, column=2, value="ESTUDIANTE DOS")

    wb_pre.save(test_excel_path)
    wb_pre.close()

    # Initialize engine pointing to the test file
    engine = DataEngine(ruta_excel=test_excel_path)
    engine.modalidad = "premedia"

    # 2. Call exportar_todo_a_excel
    export_dest = str(tmp_path / "Registro_Premedia_Exportado.xlsx")
    
    t_exp_0 = time.time()
    exito = engine.exportar_todo_a_excel(export_dest)
    t_exp_1 = time.time()
    export_duration = t_exp_1 - t_exp_0
    
    print(f"[PERF] Tiempo de exportacion en lote a Excel: {export_duration:.4f} segundos")

    # Assertions
    assert exito is True
    assert os.path.exists(export_dest)

    # Open the exported excel and verify grades and attendance
    wb = openpyxl.load_workbook(export_dest, data_only=True)
    sheet_names = wb.sheetnames
    assert any("Ingles" in name for name in sheet_names)
    assert any("Asistencia" in name for name in sheet_names)

    # Verify that the students list is written in Column B of the sheets
    ws_esp = next(wb[name] for name in sheet_names if "Ingles" in name)
    
    # Let's find student name written
    student_found = False
    for r in range(1, 100):
        val = ws_esp.cell(row=r, column=2).value
        if val == "ESTUDIANTE UNO":
            student_found = True
            break
            
    assert student_found, "El estudiante no se exporto a la hoja de calificaciones de Excel"

    # Verify grade value is written in the correct row (row 5) and column (column 3 or later)
    # The first Daily/Partial column for T1 starts around column 3/4. Let's find it.
    found_grade = False
    for col in range(3, 40):
        val = ws_esp.cell(row=5, column=col).value
        if val == 4.5:
            found_grade = True
            break
    assert found_grade, "La nota del estudiante no se exporto correctamente"

    wb.close()
