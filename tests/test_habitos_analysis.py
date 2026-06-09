import pytest
import os
import openpyxl
from rddata import DataEngine


def test_obtener_estadisticas_asistencia_missing_file():
    """Verifica que obtener_estadisticas_asistencia devuelva ceros si el archivo no existe."""
    engine = DataEngine("no_existente.xlsx")
    res = engine.obtener_estadisticas_asistencia("7\u00b0", "Trimestre 1", 1)
    assert res == {"total_dias": 0, "ausencias": 0, "tardanzas": 0, "excusas": 0}


def test_obtener_tareas_sin_nota_missing_file():
    """Verifica que obtener_tareas_sin_nota devuelva total cero si el archivo no existe."""
    engine = DataEngine("no_existente.xlsx")
    res = engine.obtener_tareas_sin_nota("7\u00b0", "Trimestre 1", 1)
    assert res == {"tareas_vacias_por_materia": {}, "total_vacias": 0}


def test_obtener_estadisticas_asistencia_mock(tmp_path):
    """Verifica el conteo de asistencia con un archivo de Excel mock estructurado."""
    file_path = str(tmp_path / "test_asistencia_mock.xlsx")
    wb = openpyxl.Workbook()
    
    # Crear hoja de asistencia (usando 'Asistencia' en Title Case to match case-sensitive check)
    ws = wb.active
    ws.title = "Asistencia (7)"
    
    # Trimestre 1 inicia en la fila 2
    # Fila 2: Fechas/Días registrados (columnas 3 a 60)
    ws.cell(row=2, column=3, value="10-06")
    ws.cell(row=2, column=4, value="11-06")
    ws.cell(row=2, column=5, value="12-06")
    ws.cell(row=2, column=6, value="13-06")
    
    # Fila del estudiante: 2 + id_estudiante. Para id=1 -> fila 3
    # Estudiante 1 asistencia:
    ws.cell(row=3, column=3, value=".")  # Presente
    ws.cell(row=3, column=4, value="-")  # Ausente
    ws.cell(row=3, column=5, value="T")  # Tardanza
    ws.cell(row=3, column=6, value="E")  # Excusa
    
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path, modalidad="premedia")
    
    # Debug info
    wb_read = openpyxl.load_workbook(file_path)
    print("DEBUG: sheetnames =", wb_read.sheetnames)
    hoja_found = engine._encontrar_hoja_asistencia(wb_read, "7\u00b0")
    print("DEBUG: found sheet =", hoja_found)
    wb_read.close()
    
    res = engine.obtener_estadisticas_asistencia("7\u00b0", "Trimestre 1", 1)
    
    assert res["total_dias"] == 4
    assert res["ausencias"] == 1
    assert res["tardanzas"] == 1
    assert res["excusas"] == 1


def test_obtener_tareas_sin_nota_mock(tmp_path):
    """Verifica el análisis de notas vacías con un archivo de Excel mock estructurado."""
    file_path = str(tmp_path / "test_notas_mock.xlsx")
    wb = openpyxl.Workbook()
    
    # Crear hoja PROM de materia ESP (7)
    ws = wb.active
    ws.title = "PROM ESP (7)"
    
    # Fila 3: Cabeceras de tipo de notas
    ws.cell(row=3, column=10, value="PARCIAL 1")
    ws.cell(row=3, column=11, value="PARCIAL 2")
    ws.cell(row=3, column=12, value="PROMEDIO PARCIAL")  # Detiene la búsqueda de Diaria/Parcial
    ws.cell(row=3, column=13, value="APRECIACIÓN 1")
    ws.cell(row=3, column=14, value="PROMEDIO APRECIACIÓN") # Detiene la búsqueda de Apreciación
    
    # Fila de descripción de la nota (fila 42 para premedia)
    ws.cell(row=42, column=10, value="Taller de Lectura")
    ws.cell(row=42, column=11, value="Exposición")
    ws.cell(row=42, column=13, value="Cuaderno")
    
    # Fila del estudiante: 4 + id_estudiante. Para id=1 -> fila 5
    # Nota del estudiante:
    ws.cell(row=5, column=10, value=4.5)  # Tiene nota
    ws.cell(row=5, column=11, value=None)  # Nota vacía (Exposición)
    ws.cell(row=5, column=13, value="")    # Nota vacía (Cuaderno)
    
    wb.save(file_path)
    wb.close()
    
    engine = DataEngine(file_path, modalidad="premedia")
    res = engine.obtener_tareas_sin_nota("7\u00b0", "Trimestre 1", 1)
    
    # ESP (7) debería tener 2 tareas vacías
    assert "Esp" in res["tareas_vacias_por_materia"]
    assert res["tareas_vacias_por_materia"]["Esp"] == 2
    assert res["total_vacias"] == 2
