import openpyxl

def limpiar_y_ajustar_todo(ruta, modalidad):
    wb = openpyxl.load_workbook(ruta)
    print(f"--- Procesando: {ruta} ({modalidad}) ---")

    # 1. Limpiar MAESTRO (Nombres y Cédulas)
    ws_m = wb["MAESTRO"]
    if modalidad == "primaria":
        for r in range(5, 51):
            ws_m.cell(row=r, column=2).value = None # Nombres en B
            ws_m.cell(row=r, column=5).value = None # Cédulas en E
    else: # Premedia
        for col in [2, 4, 6]: # B (7°), D (8°), F (9°)
            for r in range(5, 51):
                ws_m.cell(row=r, column=col).value = None

    # 2. Limpiar PROM (Fechas, Notas y Descripciones)
    fila_desc = 39 if modalidad == "primaria" else 42
    for sheet in wb.sheetnames:
        if "PROM" in sheet:
            ws = wb[sheet]
            # Limpiar fila 4 (Fechas) y fila de Descripciones
            for c in range(3, 80):
                ws.cell(row=4, column=c).value = None
                ws.cell(row=fila_desc, column=c).value = None
            
            # Limpiar Notas (respetando fórmulas)
            for r in range(5, 42):
                for c in range(3, 80):
                    cell = ws.cell(row=r, column=c)
                    if cell.data_type != 'f': # Si no es fórmula, borrar
                        cell.value = None

    # 3. Limpiar ASISTENCIA (Vertical 3-46-89 para AMBOS)
    hoja_asist = "Asistencia " if modalidad == "primaria" else "ASISTENCIA"
    if hoja_asist in wb.sheetnames:
        ws_a = wb[hoja_asist]
        bloques = [(3, 41), (46, 86), (89, 130)]
        for inicio, fin in bloques:
            for r in range(inicio, fin + 1):
                for c in range(3, 61): # C a BH
                    ws_a.cell(row=r, column=c).value = None

    # 4. Limpiar PLANILLA (Celda B7 y Notas Trimestrales)
    for sheet in wb.sheetnames:
        if "Planilla" in sheet:
            ws_p = wb[sheet]
            ws_p["B7"] = None # Limpiar basura en B7
            for r in range(16, 52): # Rango de estudiantes
                for c in [6, 7, 8]: # F, G, H (T1, T2, T3)
                    ws_p.cell(row=r, column=c).value = None

    nombre_salida = f"PLANTILLA_LIMPIA_{modalidad.upper()}.xlsx"
    wb.save(nombre_salida)
    print(f"✅ Archivo generado: {nombre_salida}")

# Ejecutar limpieza
# limpiar_y_ajustar_todo("Registro Primaria.xlsx", "primaria")
# limpiar_y_ajustar_todo("Registro_2026.xlsx", "premedia")