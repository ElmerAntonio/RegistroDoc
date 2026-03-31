import re

with open("src/happ.py", "r") as f:
    content = f.read()

# Replace export_excel with the logic to copy to "Registro_Oficial_MEDUCA.xlsx"
export_pattern = r"""    def export_excel\(self\):
        messagebox\.showinfo\("Información", "Módulo de exportación Excel en desarrollo\."\)"""

new_export = """    def export_excel(self):
        # Tomar datos de pantalla y pegarlos en la plantilla oficial.
        # Infiriendo basado en columnas Cédula, Nombre, Trimestre 1, etc.
        import os
        import openpyxl

        base = os.path.dirname(os.path.abspath(__file__))
        plantilla_ruta = os.path.join(base, "..", "Registro_Oficial_MEDUCA.xlsx")

        if not os.path.exists(plantilla_ruta):
            messagebox.showerror("Error", f"No se encontró la plantilla oficial en: {plantilla_ruta}")
            return

        try:
            wb_plantilla = openpyxl.load_workbook(plantilla_ruta)
            ws_plantilla = wb_plantilla.active # Assuming the active sheet is the destination

            # Extract data from engine and paste carefully without moving rows
            grados = self.engine.obtener_grados_activos()
            if not grados:
                messagebox.showwarning("Aviso", "No hay grados para exportar.")
                return

            fila_inicio = 5 # Asumimos que los datos empiezan en la fila 5
            for grado in grados:
                estudiantes = self.engine.obtener_estudiantes_completos(grado)
                for i, est in enumerate(estudiantes):
                    # Infiere nombres de columnas comunes
                    fila_destino = fila_inicio + i
                    ws_plantilla.cell(row=fila_destino, column=2).value = est.get("nombre", "") # B - Nombre
                    ws_plantilla.cell(row=fila_destino, column=3).value = est.get("cedula", "") # C - Cédula
                    # Just setting general info, leaving the rest untouched for stability

            out_file = os.path.join(base, "..", "Exportado_Registro_Oficial_MEDUCA.xlsx")
            wb_plantilla.save(out_file)
            wb_plantilla.close()
            messagebox.showinfo("Éxito", f"Datos exportados a: {out_file}")

        except Exception as e:
            messagebox.showerror("Error de Exportación", f"Hubo un error al exportar:\\n{str(e)}")"""

content = re.sub(export_pattern, new_export, content, flags=re.DOTALL)

with open("src/happ.py", "w") as f:
    f.write(content)
