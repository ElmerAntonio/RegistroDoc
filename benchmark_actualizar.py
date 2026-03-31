import time
import os
import openpyxl
from src.rddata import DataEngine

def setup_test_excel(ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROM (Matematica 7)"
    ws.cell(row=3, column=1).value = "PARCIAL"
    # Create some dummy students
    for r in range(5, 46):
        ws.cell(row=r, column=1).value = f"Estudiante {r-4}"
    wb.save(ruta)
    return ruta

def benchmark():
    ruta = "test_perf.xlsx"
    if os.path.exists(ruta):
        os.remove(ruta)
    setup_test_excel(ruta)

    engine = DataEngine(ruta, modalidad="premedia")

    grado = "7°"
    materia = "Matematica"
    columna = 3
    dic_notas = {i: 4.5 for i in range(1, 41)}

    # Warm up
    engine.actualizar_notas_existentes(grado, materia, columna, dic_notas)

    start_time = time.time()
    iterations = 5
    for i in range(iterations):
        engine.actualizar_notas_existentes(grado, materia, columna, dic_notas)
    end_time = time.time()

    avg_time = (end_time - start_time) / iterations
    print(f"Average time per call: {avg_time:.4f} seconds")

    if os.path.exists(ruta):
        os.remove(ruta)

if __name__ == "__main__":
    benchmark()
